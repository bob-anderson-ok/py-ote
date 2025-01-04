"""
Created on Sat May 20 15:32:13 2017

@author: Bob Anderson
"""
import copy
import pickle
import string
import glob
import zipfile
from contextlib import contextmanager
import zlib

# from pyoteapp.likelihood_calculations import cum_loglikelihood, aicc

# TODO Comment these lines out when not investigating memory usage issues
# from pympler import muppy, summary
# from resource import getrusage, RUSAGE_SELF
# import time
import shutil

# This and the memory stats printing method will not run on Apple silicon (m1 chip)
# so this has been commented out and the psutil requirement removed from setup.py
# import psutil
# from memory_profiler import profile

import gc
# import copy
import math
import subprocess
from pathlib import Path

from dataclasses import dataclass, field

from scipy import interpolate
from scipy.signal import savgol_filter

# from numba import jit
MIN_SIGMA = 0.05

import datetime
import os
import sys
import platform


# We add the current working directory to sys.path so that modules referenced by pyote can be found.
root, _ = os.path.split(__file__)
sys.path.insert(0, root)

print( "==========================")
print("pyoteapp modules will be sourced from ...")
print(root)
print( "==========================")

# for path in sys.path:
#     print(path)
# print( "==========================")


from openpyxl import load_workbook

import pandas as pd

from pyote_modelling_utility_functions import LightcurveParameters, generalizedDiffraction
from pyote_modelling_utility_functions import decide_model_to_use
from pyote_modelling_utility_functions import demo_event
from pyote_modelling_utility_functions import demo_diffraction_field
from pyote_modelling_utility_functions import plot_diffraction
from pyote_modelling_utility_functions import illustrateDiskOnDiskEvent, illustrateEdgeOnDiskEvent
from pyote_modelling_utility_functions import asteroid_km_from_mas, asteroid_mas_from_km

from solverUtils import aicModelValue

from math import trunc, floor
import matplotlib.pyplot as plt
import matplotlib

matplotlib.use('Qt5Agg')

from showVideoFrames import readAviFile
from showVideoFrames import readSerFile
from showVideoFrames import readFitsFile
from showVideoFrames import readAavFile

from false_positive import compute_drops, noise_gen_jit, simple_convolve, calc_sigma_lines
import numpy as np
import pyqtgraph as pg
import pyqtgraph.exporters as pex
# import scipy
import PyQt5
from PyQt5 import QtCore, QtWidgets
from PyQt5 import QtGui
from PyQt5.QtCore import QSettings, QPoint, QSize
from PyQt5.QtCore import pyqtSignal
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QDialog
from pyqtgraph import PlotWidget

import version
import fixedPrecision as fp
import gui
import timestampDialog
import helpDialog
import exponentialEdgeUtilities as ex

from checkForNewerVersion import getLatestPackageVersion
# from pyoteapp.checkForNewerVersion import upgradePyote
from csvreader import readLightCurve
from errorBarUtils import ciBars
from errorBarUtils import createDurDistribution
from errorBarUtils import edgeDistributionGenerator
from noiseUtils import getCorCoefs
from solverUtils import candidateCounter, solver, subFrameAdjusted
from timestampUtils import convertTimeStringToTime
from timestampUtils import convertTimeToTimeString
from timestampUtils import getTimeStepAndOutliers
from timestampUtils import manualTimeStampEntry
from blockIntegrateUtils import mean_std_versus_offset
from iterative_logl_functions import locate_event_from_d_and_r_ranges
from iterative_logl_functions import find_best_event_from_min_max_size
from iterative_logl_functions import find_best_r_only_from_min_max_size
from iterative_logl_functions import find_best_d_only_from_min_max_size
# from pyoteapp.subframe_timing_utilities import generate_underlying_lightcurve_plots, fresnel_length_km
# from pyoteapp.subframe_timing_utilities import time_correction, intensity_at_time


cursorAlert = pyqtSignal()

# The gui module was created by typing
#    !pyuic5 pyote.ui -o gui.py
# in the IPython console while in pyoteapp directory

# The timestampDialog module was created by typing
#    !pyuic5 timestamp_dialog_alt.ui -o timestampDialog.py
# in the IPython console while in pyoteapp directory

# The help-dialog module was created by typing
#    !pyuic5 helpDialog.ui -o helpDialog.py
# in the IPython console while in pyoteapp directory

# Status of points and associated dot colors ---
MISSING = 5   # (211, 211, 211) light gray
EVENT = 4     # (155, 150, 100) sick green
SELECTED = 3  # (255, 000, 000) red
BASELINE = 2  # (255, 150, 100) salmon
INCLUDED = 1  # (000, 032, 255) blue with touch of green
EXCLUDED = 0  # no dot
DEFECT_HIT = 6  # Add circle or something blindingly obvious

MISSING_COLOR = (211, 211, 211)  # light gray
EVENT_COLOR = (155, 150, 100)    # sick yellowish green
SELECTED_COLOR = (255, 0, 0)     # red
BASELINE_COLOR = (255, 150,100)  # salmon
INCLUDED_COLOR = (0, 32, 255)    # blue + green tinge
DEFECT_HIT_COLOR = (255, 0, 0)   # red

LINESIZE = 2

acfCoefThreshold = 0.05  # To match what is being done in R-OTE 4.5.4+

@dataclass
class BestFit:
    thisPassMetric: float = None
    metricAtStartOfPass: float = None
    modelTimeOffset: float = None
    chordTime: float = None
    missDistance: float = None
    Dangle: float = None
    Rangle: float = None


@dataclass
class FitStatus:
    improvementPassCompleted: bool = False
    currentMetric: float = None
    metricInUse: str = field(default='?')   # 'd' | 'r' | 'both'
    modelTimeOffset: float = None
    chordTime: float = None
    missDistance: float = None
    Dangle: float = None
    Rangle: float = None
    angleToOptimize: str = field(default='?')  # 'D' | 'R' when edge-on-disk being solved
    currentDelta: float = None
    edgeDelta: float = None
    chordDelta: float = None
    missDelta: float = None
    DangleDelta: float = None
    RangleDelta: float = None
    fitComplete: bool = False
    failureCount = 0
    beingChanged: str = field(default='?')  # 'edgeTime' | 'chord' | 'Dangle' | 'Rangle'

# There is a bug in pyqtgraph ImageExporter, probably caused by new versions of PyQt5 returning
# float values for image rectangles.  Those floats were being given to numpy to create a matrix,
# and that was raising an exception.  Below is my 'cure', effected by overriding the internal
# methods of ImageExporter that manipulate width and height


class TimestampAxis(pg.AxisItem):

    def tickStrings(self, values, scale, spacing):
        return [self.getTimestampString(val) for val in values]

    def setFetcher(self, timestampFetch):
        self.getTimestampString = timestampFetch


class TSdialog(QDialog, timestampDialog.Ui_manualTimestampDialog):

    def __init__(self):
        super(TSdialog, self).__init__()
        self.setupUi(self)


class FixedImageExporter(pex.ImageExporter):
    def __init__(self, item):
        pex.ImageExporter.__init__(self, item)

    def makeWidthHeightInts(self):
        self.params['height'] = int(self.params['height'] + 1)  # The +1 is needed
        self.params['width'] = int(self.params['width'] + 1)

    def widthChanged(self):
        sr = self.getSourceRect()
        ar = float(sr.height()) / sr.width()
        self.params.param('height').setValue(int(self.params['width'] * ar),
                                             blockSignal=self.heightChanged)

    def heightChanged(self):
        sr = self.getSourceRect()
        ar = float(sr.width()) / sr.height()
        self.params.param('width').setValue(int(self.params['height'] * ar),
                                            blockSignal=self.widthChanged)


class Signal:
    def __init__(self):
        self.__subscribers = []

    def emit(self, *args, **kwargs):
        for subs in self.__subscribers:
            subs(*args, **kwargs)

    def connect(self, func):
        self.__subscribers.append(func)


mouseSignal = Signal()


class CustomViewBox(pg.ViewBox):
    def __init__(self, *args, **kwds):
        pg.ViewBox.__init__(self, *args, **kwds)
        self.setMouseMode(self.RectMode)

    # re-implement right-click to zoom out
    def mouseClickEvent(self, ev):
        if ev.button() == QtCore.Qt.MouseButton.RightButton:
            self.autoRange()
            mouseSignal.emit()

    def mouseDragEvent(self, ev, axis=None):
        if ev.button() == QtCore.Qt.MouseButton.RightButton:
            ev.ignore()
        else:
            pg.ViewBox.mouseDragEvent(self, ev, axis)
            mouseSignal.emit()


class HelpDialog(QDialog, helpDialog.Ui_Dialog):
    def __init__(self):
        super(HelpDialog, self).__init__()
        self.setupUi(self)


# class SimplePlot(QtGui.QMainWindow, gui.Ui_MainWindow):
class SimplePlot(PyQt5.QtWidgets.QMainWindow, gui.Ui_mainWindow):  # noqa
    def __init__(self, csv_file):
        super(SimplePlot, self).__init__()

        self.ext = None  # Holds extension of file just read.

        self.getListOfVizieRlightcurves()

        self.camCorrection = 0.0  # pcs

        self.noiseSigmaDistance = 0.0

        self.consistentLcpPresent = False

        self.metricLimitLeft = None
        self.metricLimitRight = None

        self.suppressReport = False

        self.firstLightCurveDisplayed = True
        self.availableLightCurvesForDisplay = []

        self.homeDir = os.path.split(__file__)[0]

        self.keepRunning = True
        self.allCoreElementsEntered = False

        self.selectedPoints = {}

        self.csvFilePath = None

        self.fitStatus = None
        self.bestFit = None

        # The following variable only takes on values of
        # 'edgeTime' | 'chord' | 'Dangle' | 'Rangle' | '' and is only used when
        # the checkBox "Single optimization"  on the "Model selection" tab is checked
        self.singleOptimizationName = ''

        self.pauseFitRequested = False

        self.suppressParameterChange = False
        self.parameterChangeEntryCount = 0

        self.yTimes = []

        self.VizieRdict = None

        self.userDataSetAdditions = []

        self.dataLen = None
        self.yFrame = None
        self.left = None
        self.right = None

        self.yValues = None            # observation data

        # self.editMode = False

        self.chordSizeSecondsEdited = False
        self.chordSizeKmEdited = False

        self.starSizeMasEdited = False
        self.starSizeKmEdited = False

        self.modelXkm = None           # model x values (in km)
        self.modelY = None             # model y values (ADU)
        self.modelDedgeKm = None       # model D edge location (in km)
        self.modelRedgeKm = None       # model R edge location (in km)

        self.firstContactKm = None
        self.firstContactSecs = None
        self.lastContactKm = None
        self.lastContactSecs = None

        self.firstContactRdgValue = None
        self.lastContactRdgValue = None

        self.modelDedgeSecs = None     # timestamp of D edge in observation
        self.modelRedgeSecs = None     # timestamp of R edge in observation

        self.modelPtsY = None          # the computed model lightcurve (possibly trimmed)
        self.modelPtsXrdgNum = None    # with x values in reading number units (starting at 0.0)
        self.modelPtsXsec = None       # and with x values in time units

        self.modelYsamples = None

        self.modelDuration = None      # duration of model lightcurve (seconds)

        self.modelMetric = None        # sum((y[i] - model[i])**2) / n

        # This variable controls where the left edge of the computed model lightcurve
        # starts relative to the beginning of the observation
        self.modelTimeOffset = None     # time offset of model lightcurve from obs start

        # modelXvalues are in reading number units
        self.modelXvalues = None       # model x values extended or trimmed to match obs
        self.modelYvalues = None       # model y values extended or trimmed to match obs
        self.modelDedgeRdgValue = None    # D edge from model (in reading number units)
        self.modelRedgeRdgValue = None    # R edge from model (in reading number units)

        self.allowShowDetails = False

        # This is an externally supplied csv file path (probably from PyMovie)
        self.externalCsvFilePath = csv_file

        self.firstPassPenumbralFit = True

        self.homeDir = os.path.split(__file__)[0]

        # Change pyqtgraph plots to be black on white
        pg.setConfigOption('background', (255, 255, 255))  # Do before any widgets drawn
        pg.setConfigOption('foreground', 'k')  # Do before any widgets drawn
        pg.setConfigOptions(imageAxisOrder='row-major')

        self.setupUi(self)

        self.setWindowTitle('PYOTE  Version: ' + version.version())

        self.skipNormalization = False  # A flag used to prevent infinite recursion in self.redrawMainPlot

        self.suppressNormalization = False

        self.targetIndex = 0

        self.referenceKey = ''  # The key into self.fullDataDictionary for the reference curve used in normalization
        self.targetKey = ''     # The key into self.fullDataDictionary for the target curve that is to be normalized

        self.targetCheckBoxes = [self.targetCheckBox_1, self.targetCheckBox_2, self.targetCheckBox_3,
                                 self.targetCheckBox_4, self.targetCheckBox_5, self.targetCheckBox_6,
                                 self.targetCheckBox_7, self.targetCheckBox_8, self.targetCheckBox_9,
                                 self.targetCheckBox_10]

        self.showCheckBoxes = [self.showCheckBox_1, self.showCheckBox_2, self.showCheckBox_3,
                               self.showCheckBox_4, self.showCheckBox_5, self.showCheckBox_6,
                               self.showCheckBox_7, self.showCheckBox_8, self.showCheckBox_9,
                               self.showCheckBox_10]

        self.lightcurveTitles = [self.lightcurveTitle_1, self.lightcurveTitle_2, self.lightcurveTitle_3,
                                 self.lightcurveTitle_4, self.lightcurveTitle_5, self.lightcurveTitle_6,
                                 self.lightcurveTitle_7, self.lightcurveTitle_8, self.lightcurveTitle_9,
                                 self.lightcurveTitle_10]

        self.yOffsetSpinBoxes = [self.yOffsetSpinBox_1, self.yOffsetSpinBox_2, self.yOffsetSpinBox_3,
                                 self.yOffsetSpinBox_4, self.yOffsetSpinBox_5, self.yOffsetSpinBox_6,
                                 self.yOffsetSpinBox_7, self.yOffsetSpinBox_8, self.yOffsetSpinBox_9,
                                 self.yOffsetSpinBox_10]

        self.referenceCheckBoxes = [self.referenceCheckBox_1, self.referenceCheckBox_2, self.referenceCheckBox_3,
                                    self.referenceCheckBox_4, self.referenceCheckBox_5, self.referenceCheckBox_6,
                                    self.referenceCheckBox_7, self.referenceCheckBox_8, self.referenceCheckBox_9,
                                    self.referenceCheckBox_10]

        self.colorStyle = ['color: rgb(255,0,0)', 'color: rgb(160,32,255)', 'color: rgb(80,208,255)',
                           'color: rgb(96,255,128)', 'color: rgb(255,224,32)', 'color: rgb(255,160,16)',
                           'color: rgb(160,128,96)', 'color: rgb(64,64,64)', 'color: rgb(255,208,160)',
                           'color: rgb(0,128,0)']

        self.colorBlobs = [self.colorBlob0, self.colorBlob1, self.colorBlob2,
                           self.colorBlob3, self.colorBlob4, self.colorBlob5,
                           self.colorBlob6, self.colorBlob7, self.colorBlob8,
                           self.colorBlob9]

        self.xOffsetSpinBoxes = [self.xOffsetSpinBox_1, self.xOffsetSpinBox_2, self.xOffsetSpinBox_3,
                                 self.xOffsetSpinBox_4, self.xOffsetSpinBox_5, self.xOffsetSpinBox_6,
                                 self.xOffsetSpinBox_7, self.xOffsetSpinBox_8, self.xOffsetSpinBox_9,
                                 self.xOffsetSpinBox_10]

        self.initializeLightcurvePanel()

        self.smoothingLabel.installEventFilter(self)

        self.yOffsetLabel1.installEventFilter(self)
        self.yOffsetLabel2.installEventFilter(self)
        self.yOffsetLabel3.installEventFilter(self)
        self.yOffsetLabel4.installEventFilter(self)
        self.yOffsetLabel5.installEventFilter(self)
        self.yOffsetLabel6.installEventFilter(self)
        self.yOffsetLabel7.installEventFilter(self)
        self.yOffsetLabel8.installEventFilter(self)
        self.yOffsetLabel9.installEventFilter(self)
        self.yOffsetLabel10.installEventFilter(self)

        self.xOffsetLabel1.installEventFilter(self)
        self.xOffsetLabel2.installEventFilter(self)
        self.xOffsetLabel3.installEventFilter(self)
        self.xOffsetLabel4.installEventFilter(self)
        self.xOffsetLabel5.installEventFilter(self)
        self.xOffsetLabel6.installEventFilter(self)
        self.xOffsetLabel7.installEventFilter(self)
        self.xOffsetLabel8.installEventFilter(self)
        self.xOffsetLabel9.installEventFilter(self)
        self.xOffsetLabel10.installEventFilter(self)

        self.stepBy2radioButton.clicked.connect(self.processStepBy2)
        self.stepBy10radioButton.clicked.connect(self.processStepBy10)
        self.stepBy100radioButton.clicked.connect(self.processStepBy100)

        self.stepByButtonGroup = QtWidgets.QButtonGroup()
        self.stepByButtonGroup.addButton(self.stepBy2radioButton)
        self.stepByButtonGroup.addButton(self.stepBy10radioButton)
        self.stepByButtonGroup.addButton(self.stepBy100radioButton)

        self.yOffsetStep10radioButton.clicked.connect(self.processYoffsetStepBy10)
        self.yOffsetStep100radioButton.clicked.connect(self.processYoffsetStepBy100)
        self.yOffsetStep1000radioButton.clicked.connect(self.processYoffsetStepBy1000)

        self.offsetStepButtonGroup = QtWidgets.QButtonGroup()
        self.offsetStepButtonGroup.addButton(self.yOffsetStep10radioButton)
        self.offsetStepButtonGroup.addButton(self.yOffsetStep100radioButton)
        self.offsetStepButtonGroup.addButton(self.yOffsetStep1000radioButton)

        self.yOffsetStep10radioButton.setChecked(True)

        self.curveSelectionComboBox.activated.connect(self.handleDataSetSelection)
        self.curveSelectionComboBox.installEventFilter(self)
        self.availableCurvesLabel.installEventFilter(self)

        self.targetCheckBox_1.clicked.connect(self.processTargetSelection1)
        self.targetCheckBox_2.clicked.connect(self.processTargetSelection2)
        self.targetCheckBox_3.clicked.connect(self.processTargetSelection3)
        self.targetCheckBox_4.clicked.connect(self.processTargetSelection4)
        self.targetCheckBox_5.clicked.connect(self.processTargetSelection5)
        self.targetCheckBox_6.clicked.connect(self.processTargetSelection6)
        self.targetCheckBox_7.clicked.connect(self.processTargetSelection7)
        self.targetCheckBox_8.clicked.connect(self.processTargetSelection8)
        self.targetCheckBox_9.clicked.connect(self.processTargetSelection9)
        self.targetCheckBox_10.clicked.connect(self.processTargetSelection10)

        self.showCheckBox_1.clicked.connect(self.processShowSelection)
        self.showCheckBox_2.clicked.connect(self.processShowSelection)
        self.showCheckBox_3.clicked.connect(self.processShowSelection)
        self.showCheckBox_4.clicked.connect(self.processShowSelection)
        self.showCheckBox_5.clicked.connect(self.processShowSelection)
        self.showCheckBox_6.clicked.connect(self.processShowSelection)
        self.showCheckBox_7.clicked.connect(self.processShowSelection)
        self.showCheckBox_8.clicked.connect(self.processShowSelection)
        self.showCheckBox_9.clicked.connect(self.processShowSelection)
        self.showCheckBox_10.clicked.connect(self.processShowSelection)

        self.referenceCheckBox_1.clicked.connect(self.processReferenceSelection1)
        self.referenceCheckBox_2.clicked.connect(self.processReferenceSelection2)
        self.referenceCheckBox_3.clicked.connect(self.processReferenceSelection3)
        self.referenceCheckBox_4.clicked.connect(self.processReferenceSelection4)
        self.referenceCheckBox_5.clicked.connect(self.processReferenceSelection5)
        self.referenceCheckBox_6.clicked.connect(self.processReferenceSelection6)
        self.referenceCheckBox_7.clicked.connect(self.processReferenceSelection7)
        self.referenceCheckBox_8.clicked.connect(self.processReferenceSelection8)
        self.referenceCheckBox_9.clicked.connect(self.processReferenceSelection9)
        self.referenceCheckBox_10.clicked.connect(self.processReferenceSelection10)

        self.yOffsetSpinBox_1.valueChanged.connect(self.processYoffsetChange)
        self.yOffsetSpinBox_2.valueChanged.connect(self.processYoffsetChange)
        self.yOffsetSpinBox_3.valueChanged.connect(self.processYoffsetChange)
        self.yOffsetSpinBox_4.valueChanged.connect(self.processYoffsetChange)
        self.yOffsetSpinBox_5.valueChanged.connect(self.processYoffsetChange)
        self.yOffsetSpinBox_6.valueChanged.connect(self.processYoffsetChange)
        self.yOffsetSpinBox_7.valueChanged.connect(self.processYoffsetChange)
        self.yOffsetSpinBox_8.valueChanged.connect(self.processYoffsetChange)
        self.yOffsetSpinBox_9.valueChanged.connect(self.processYoffsetChange)
        self.yOffsetSpinBox_10.valueChanged.connect(self.processYoffsetChange)

        self.xOffsetSpinBox_1.valueChanged.connect(self.processXoffsetChange)
        self.xOffsetSpinBox_2.valueChanged.connect(self.processXoffsetChange)
        self.xOffsetSpinBox_3.valueChanged.connect(self.processXoffsetChange)
        self.xOffsetSpinBox_4.valueChanged.connect(self.processXoffsetChange)
        self.xOffsetSpinBox_5.valueChanged.connect(self.processXoffsetChange)
        self.xOffsetSpinBox_6.valueChanged.connect(self.processXoffsetChange)
        self.xOffsetSpinBox_7.valueChanged.connect(self.processXoffsetChange)
        self.xOffsetSpinBox_8.valueChanged.connect(self.processXoffsetChange)
        self.xOffsetSpinBox_9.valueChanged.connect(self.processXoffsetChange)
        self.xOffsetSpinBox_10.valueChanged.connect(self.processXoffsetChange)

        self.smoothingIntervalSpinBox.valueChanged.connect(self.newRedrawMainPlot)

        # Vizier export widgets

        self.fillFromNAxlsxFileButton.clicked.connect(self.fillFromNAxlsxFile)
        self.fillFromNAxlsxFileButton.installEventFilter(self)

        self.vizierShowPlotButton.installEventFilter(self)
        self.vizierShowPlotButton.clicked.connect(self.showVizieRplot)

        self.vizierExportButton.installEventFilter(self)
        self.vizierExportButton.clicked.connect(self.exportVizieRtextFile)

        # Special analysis widgets

        self.processDeepVtestButton.clicked.connect(self.processDeepVtest)
        self.processDeepVtestButton.installEventFilter(self)

        self.processSteppedLevelsTestButton.clicked.connect(self.processSteppedLevelsTest)
        self.processSteppedLevelsTestButton.installEventFilter(self)

        # Model lightcurves widgets

        self.beingOptimizedLabel.installEventFilter(self)
        self.beingOptimizedEdit.installEventFilter(self)

        self.currentEventLabel.installEventFilter(self)
        self.currentEventEdit.installEventFilter(self)
        self.currentEventEdit.editingFinished.connect(self.processNewCurrentEventEdit)

        self.saveCurrentEventButton.installEventFilter(self)
        self.saveCurrentEventButton.clicked.connect(self.saveCurrentEvent)

        self.clearEventDataButton.clicked.connect(self.clearEventDataEntries)
        self.clearEventDataButton.installEventFilter(self)

        self.pastEventsLabel.installEventFilter(self)
        self.pastEventsComboBox.installEventFilter(self)
        self.pastEventsComboBox.activated.connect(self.handlePastEventSelection)

        self.vzCoordsComboBox.activated.connect(self.handleSiteCoordSelection)
        self.vzCoordsComboBox.installEventFilter(self)

        self.deleteEventButton.installEventFilter(self)
        self.deleteEventButton.clicked.connect(self.deletePastEvent)

        self.baselineADUlabel.installEventFilter(self)
        self.baselineADUedit.installEventFilter(self)

        self.baselineADUbutton.installEventFilter(self)
        self.baselineADUbutton.clicked.connect(self.modelMarkBaselineRegion)

        self.calcBaselineADUbutton.clicked.connect(self.calcBaselineADU)
        self.calcBaselineADUbutton.installEventFilter(self)

        self.clearBaselineADUselectionButton.clicked.connect(self.clearBaselineRegions)
        self.clearBaselineADUselectionButton.installEventFilter(self)

        self.baselineADUedit.editingFinished.connect(self.processModelLightcurveCoreEdit)

        self.bottomADUlabel.installEventFilter(self)
        self.bottomADUedit.installEventFilter(self)

        self.magDropEdit.installEventFilter(self)
        self.magDropEdit.editingFinished.connect(self.processMagDropFinish)

        self.frameTimeLabel.installEventFilter(self)
        self.frameTimeEdit.installEventFilter(self)
        self.frameTimeEdit.editingFinished.connect(self.processModelLightcurveCoreEdit)

        self.missDistanceLabel.installEventFilter(self)
        self.missDistanceKmEdit.installEventFilter(self)
        self.missDistanceKmEdit.editingFinished.connect(self.processMissDistanceFinish)

        self.asteroidDiameterLabel.installEventFilter(self)
        self.asteroidDiameterKmLabel.installEventFilter(self)
        self.asteroidDiameterMasLabel.installEventFilter(self)

        self.asteroidDiameterKmEdit.installEventFilter(self)
        self.asteroidDiameterKmEdit.editingFinished.connect(self.processAsteroidDiameterKmFinish)
        self.asteroidDiameterMasEdit.installEventFilter(self)
        self.asteroidDiameterMasEdit.editingFinished.connect(self.processAsteroidDiameterMasFinish)

        self.asteroidSpeedLabel.installEventFilter(self)
        self.asteroidSpeedShadowEdit.installEventFilter(self)
        self.asteroidSpeedShadowLabel.installEventFilter(self)
        self.asteroidSpeedSkyEdit.installEventFilter(self)
        self.asteroidSpeedSkyLabel.installEventFilter(self)
        self.asteroidSpeedShadowEdit.editingFinished.connect(self.processAsteroidSpeedShadowFinish)
        self.asteroidSpeedSkyEdit.editingFinished.connect(self.processAsteroidSpeedSkyFinish)

        self.asteroidDistLabel.installEventFilter(self)
        self.asteroidDistAUedit.installEventFilter(self)
        self.asteroidDistAUlabel.installEventFilter(self)
        self.asteroidDistArcsecEdit.installEventFilter(self)
        self.asteroidDistArcsecLabel.installEventFilter(self)
        self.asteroidDistAUedit.editingFinished.connect(self.processAsteroidDistAUfinish)
        self.asteroidDistArcsecEdit.editingFinished.connect(self.processAsteroidDistArcsecFinish)

        self.wavelengthLabel.installEventFilter(self)
        self.wavelengthEdit.installEventFilter(self)
        self.wavelengthEdit.editingFinished.connect(self.processWavelengthFinish)

        self.limbAnglesLabel.installEventFilter(self)
        self.DdegreesEdit.installEventFilter(self)
        self.DdegreesEdit.editingFinished.connect(self.processDangleEditFinish)
        self.DdegreesLabel.installEventFilter(self)

        self.RdegreesEdit.installEventFilter(self)
        self.RdegreesEdit.editingFinished.connect(self.processRangleEditFinish)
        self.RdegreesLabel.installEventFilter(self)

        self.fresnelSizeLabel.installEventFilter(self)
        self.fresnelSizeKmEdit.installEventFilter(self)
        self.fresnelSizeKmLabel.installEventFilter(self)
        self.fresnelSizeSecondsEdit.installEventFilter(self)
        self.fresnelSizeSecondsLabel.installEventFilter(self)

        self.starSizeLabel.installEventFilter(self)
        self.starSizeKmLabel.installEventFilter(self)
        self.starSizeMasLabel.installEventFilter(self)
        self.starSizeMasEdit.installEventFilter(self)
        self.starSizeKmEdit.installEventFilter(self)
        self.starSizeMasEdit.editingFinished.connect(self.processStarSizeMasFinish)
        self.starSizeKmEdit.editingFinished.connect(self.processStarSizeKmFinish)

        self.chordSizeLabel.installEventFilter(self)
        self.chordSizeSecondsLabel.installEventFilter(self)
        self.chordSizeKmEdit.installEventFilter(self)
        self.chordSizeKmEdit.editingFinished.connect(self.processChordSizeKmFinish)
        self.chordSizeSecondsEdit.editingFinished.connect(self.processChordSizeSecondsFinish)

        self.modelToUseLabel.installEventFilter(self)
        self.diffractionRadioButton.installEventFilter(self)
        self.diffractionRadioButton.clicked.connect(self.handleModelSelectionRadioButtonClick)

        self.edgeOnDiskRadioButton.installEventFilter(self)
        self.edgeOnDiskRadioButton.clicked.connect(self.handleModelSelectionRadioButtonClick)

        self.diskOnDiskRadioButton.installEventFilter(self)
        self.diskOnDiskRadioButton.clicked.connect(self.handleModelSelectionRadioButtonClick)

        self.squareWaveRadioButton.installEventFilter(self)
        self.squareWaveRadioButton.clicked.connect(self.handleModelSelectionRadioButtonClick)

        self.fitLightcurveButton.installEventFilter(self)
        self.askAdviceButton.installEventFilter(self)
        self.showDiffractionButton.installEventFilter(self)

        self.plotFamilyButton.clicked.connect(self.plotFamilyOfLightcurves)
        self.plotFamilyButton.installEventFilter(self)

        self.setMetricLimitsButton.clicked.connect(self.markMetricRegion)
        self.setMetricLimitsButton.installEventFilter(self)

        self.askAdviceButton.clicked.connect(self.showModelChoiceAdvice)
        self.fitLightcurveButton.clicked.connect(self.fitModelLightcurveButtonClicked)
        self.showDiffractionButton.clicked.connect(self.plotDiffractionPatternOnGround)

        self.majorAxisEdit.editingFinished.connect(self.validateEllipseParameters)
        self.minorAxisEdit.editingFinished.connect(self.validateEllipseParameters)
        self.ellipseAngleEdit.editingFinished.connect(self.validateEllipseParameters)
        self.useUpperEllipseChord.clicked.connect(self.validateEllipseParameters)

        self.demoModelButton.installEventFilter(self)
        self.demoModelButton.clicked.connect(self.demoModel)

        self.demoGeometryButton.installEventFilter(self)
        self.demoGeometryButton.clicked.connect(self.demoGeometry)

        self.printEventParametersButton.installEventFilter(self)
        self.printEventParametersButton.clicked.connect(self.printEventParameters)

        self.majorAxisLabel.installEventFilter(self)
        self.minorAxisLabel.installEventFilter(self)
        self.ellipseAngleLabel.installEventFilter(self)
        self.useUpperEllipseChord.installEventFilter(self)

        self.showDetailsCheckBox.installEventFilter(self)
        self.versusTimeCheckBox.installEventFilter(self)
        self.showAnnotationsCheckBox.installEventFilter(self)
        self.showLegendsCheckBox.installEventFilter(self)

        self.fitMetricEdit.installEventFilter(self)
        self.fitMetricLabel.installEventFilter(self)
        self.fitMetricChangeEdit.installEventFilter(self)

        self.helpPdfButton.installEventFilter(self)
        self.helpPdfButton.clicked.connect(self.helpPdfButtonClicked)

        self.fitPrecisionLabel.installEventFilter(self)
        self.edgeTimePrecisionLabel.installEventFilter(self)
        self.chordDurationPrecisionLabel.installEventFilter(self)
        self.limbAnglePrecisionLabel.installEventFilter(self)
        self.missDistancePrecisionLabel.installEventFilter(self)

        self.pauseFitButton.installEventFilter(self)
        self.pauseFitButton.clicked.connect(self.pauseFitInProgress)

        self.edgeTimePrecisionEdit.installEventFilter(self)
        self.edgeTimePrecisionEdit.editingFinished.connect(self.processEdgeTimePrecisionFinish)

        self.chordDurationPrecisionEdit.installEventFilter(self)
        self.chordDurationPrecisionEdit.editingFinished.connect(self.processChordDurationPrecisionFinish)

        self.limbAnglePrecisionEdit.installEventFilter(self)
        self.limbAnglePrecisionEdit.editingFinished.connect(self.processLimbAnglePrecisionFinish)

        self.missDistancePrecisionEdit.installEventFilter(self)
        self.missDistancePrecisionEdit.editingFinished.connect(self.processMissDistancePrecisionFinish)

        self.LC1 = []
        self.LC2 = []
        self.LC3 = []
        self.LC4 = []
        self.yRefStar = []

        self.dotSizeSpinner.valueChanged.connect(self.newRedrawMainPlot)

        # This object is used to display tooltip help in a separate
        # modeless dialog box.
        self.helperThing = HelpDialog()

        self.helpButton.clicked.connect(self.helpButtonClicked)
        self.helpButton.installEventFilter(self)

        self.tutorialButton.clicked.connect(self.tutorialButtonClicked)
        self.tutorialButton.installEventFilter(self)

        self.lightcurvesHelpButton.clicked.connect(self.lightcurvesHelpButtonClicked)
        self.lightcurvesHelpButton.installEventFilter(self)

        self.minEventLabel.installEventFilter(self)
        self.maxEventLabel.installEventFilter(self)

        self.bkgndRegionLimits = []
        self.bkgndRegions = []

        self.ne3ExplanationButton.clicked.connect(self.ne3ExplanationClicked)
        self.ne3ExplanationButton.installEventFilter(self)

        self.yPositionLabel.installEventFilter(self)

        self.dnrOffRadioButton.installEventFilter(self)
        self.dnrLowRadioButton.installEventFilter(self)
        self.dnrMiddleRadioButton.installEventFilter(self)
        self.dnrHighRadioButton.installEventFilter(self)

        self.dnrLowDtcLabel.installEventFilter(self)
        self.dnrLowRtcLabel.installEventFilter(self)
        self.dnrMiddleDtcLabel.installEventFilter(self)
        self.dnrMiddleRtcLabel.installEventFilter(self)
        self.dnrHighDtcLabel.installEventFilter(self)
        self.dnrHighRtcLabel.installEventFilter(self)

        self.exponentialDtheoryPts = None
        self.exponentialRtheoryPts = None

        self.exponentialDinitialX = 0
        self.exponentialRinitialX = 0

        self.exponentialDedge = 0
        self.exponentialRedge = 0

        self.userDeterminedBaselineStats = False
        self.userDeterminedEventStats = False
        self.userTrimInEffect = False

        self.markBaselineRegionButton.clicked.connect(self.markBaselineRegion)
        self.markBaselineRegionButton.installEventFilter(self)

        self.markBaselineRegionLightcurvesButton.clicked.connect(self.markBaselineRegion)
        self.markBaselineRegionLightcurvesButton.installEventFilter(self)

        self.normMarkBaselineRegionButton.clicked.connect(self.markBaselineRegion)
        self.normMarkBaselineRegionButton.installEventFilter(self)

        self.clearBaselineRegionsButton.clicked.connect(self.clearBaselineRegions)
        self.clearBaselineRegionsButton.installEventFilter(self)

        self.clearBaselineRegionsLightcurvesButton.clicked.connect(self.clearBaselineRegions)
        self.clearBaselineRegionsLightcurvesButton.installEventFilter(self)

        self.clearMetricPointsButton.clicked.connect(self.clearModelsBaselineRegion)
        self.clearMetricPointsButton.installEventFilter(self)

        self.calcStatsFromBaselineRegionsButton.clicked.connect(self.calcBaselineStatisticsFromMarkedRegions)
        self.calcStatsFromBaselineRegionsButton.installEventFilter(self)

        self.calcStatsFromBaselineRegionsLightcurvesButton.clicked.connect(self.calcBaselineStatisticsFromMarkedRegions)
        self.calcStatsFromBaselineRegionsLightcurvesButton.installEventFilter(self)

        self.calcDetectabilityButton.clicked.connect(self.calcDetectability)
        self.calcDetectabilityButton.installEventFilter(self)

        self.eventDurLabel.installEventFilter(self)
        self.obsDurLabel.installEventFilter(self)
        self.durStepLabel.installEventFilter(self)
        self.detectMagDropLabel.installEventFilter(self)

        # Checkbox: Use manual timestamp entry
        # self.manualTimestampCheckBox.clicked.connect(self.toggleManualEntryButton)
        # self.manualTimestampCheckBox.installEventFilter(self)

        # Button: Manual timestamp entry
        self.manualEntryPushButton.clicked.connect(self.doManualTimestampEntry)
        self.manualEntryPushButton.installEventFilter(self)

        # Button: Info
        self.infoButton.clicked.connect(self.openHelpFile)
        self.infoButton.installEventFilter(self)

        # Button: Read light curve
        self.readData.clicked.connect(self.readDataFromFile)
        self.readData.installEventFilter(self)

        # Checkbox: Show timestamp errors
        self.showTimestampErrors.clicked.connect(self.toggleDisplayOfTimestampErrors)
        self.showTimestampErrors.installEventFilter(self)

        self.showTimestampsCheckBox.installEventFilter(self)

        # Checkbox: Show underlying lightcurve
        self.showUnderlyingLightcurveCheckBox.installEventFilter(self)
        self.showUnderlyingLightcurveCheckBox.clicked.connect(self.newRedrawMainPlot)

        self.showCameraResponseCheckBox.installEventFilter(self)
        self.showCameraResponseCheckBox.clicked.connect(self.newRedrawMainPlot)

        # Checkbox: Show error bars
        self.showErrBarsCheckBox.installEventFilter(self)
        self.showErrBarsCheckBox.clicked.connect(self.newRedrawMainPlot)

        # Checkbox: Show D and R edges
        self.showEdgesCheckBox.installEventFilter(self)
        self.showEdgesCheckBox.clicked.connect(self.newRedrawMainPlot)

        # Checkbox: Do OCR check
        self.showOCRcheckFramesCheckBox.installEventFilter(self)

        # line size
        self.lineWidthLabel.installEventFilter(self)
        self.lineWidthSpinner.valueChanged.connect(self.newRedrawMainPlot)

        # plotHelpButton
        self.plotHelpButton.clicked.connect(self.plotHelpButtonClicked)
        self.plotHelpButton.installEventFilter(self)

        # helpSqWaveButton
        self.helpSqWaveButton.clicked.connect(self.sqWaveHelpButtonClicked)
        self.helpSqWaveButton.installEventFilter(self)

        # detectabilityHelpButton
        self.detectabilityHelpButton.clicked.connect(self.detectabilityHelpButtonClicked)
        self.detectabilityHelpButton.installEventFilter(self)

        # Button: Trim/Select data points
        self.setDataLimits.clicked.connect(self.doTrim)
        self.setDataLimits.installEventFilter(self)

        self.vzTrimButton.clicked.connect(self.doTrim)
        self.vzTrimButton.installEventFilter(self)

        self.vzSaveSiteCoordButton.clicked.connect(self.saveSiteCoords)
        self.vzSaveSiteCoordButton.installEventFilter(self)

        self.vzSiteCoordNameEdit.installEventFilter(self)

        self.vzStartOverButton.clicked.connect(self.vizierStartOver)
        self.vzStartOverButton.installEventFilter(self)


        self.vzInfoButton.clicked.connect(self.vzInfoClicked)
        self.vzInfoButton.installEventFilter(self)

        self.vzSiteLatLabel.installEventFilter(self)
        self.vzSiteLongLabel.installEventFilter(self)

        self.vzWhereToSendButton.clicked.connect(self.vzWhereClicked)
        self.vzWhereToSendButton.installEventFilter(self)

        self.fillVzCoordsComboBox()

        # Button: Do block integration
        self.doBlockIntegration.clicked.connect(self.doIntegration)
        self.doBlockIntegration.installEventFilter(self)

        self.blockSizeEdit.installEventFilter(self)
        self.blockSizeLabel.installEventFilter(self)

        # Button: Accept integration
        self.acceptBlockIntegration.clicked.connect(self.applyIntegration)
        self.acceptBlockIntegration.installEventFilter(self)

        # Button: Validate a potential single point event
        self.singlePointDropButton.clicked.connect(self.validateSinglePointDrop)
        self.singlePointDropButton.installEventFilter(self)

        # Button: Mark D zone
        self.markDzone.clicked.connect(self.showDzone)
        self.markDzone.installEventFilter(self)

        # Button: Mark R zone
        self.markRzone.clicked.connect(self.showRzone)
        self.markRzone.installEventFilter(self)

        # Button: Mark E zone
        self.markEzone.clicked.connect(self.markEventRegion)
        self.markEzone.installEventFilter(self)

        # Button: Calc flash edge
        self.calcFlashEdge.clicked.connect(self.calculateFlashREdge)
        self.calcFlashEdge.installEventFilter(self)

        self.magDropSqwaveLabel.installEventFilter(self)
        self.magDropSqwaveEdit.installEventFilter(self)
        self.magDropSqwaveEdit.editingFinished.connect(self.processSqwaveMagDropEntry)

        # Edit box: min event
        self.minEventEdit.installEventFilter(self)

        # Edit box: max event
        self.maxEventEdit.installEventFilter(self)

        # Button: Locate event
        self.locateEvent.clicked.connect(self.findEventWithPlots)
        self.locateEvent.installEventFilter(self)
        self.locateEventFromLightcurves.clicked.connect(self.findEventNoPlots)
        self.locateEventFromLightcurves.installEventFilter(self)

        # Button: Cancel operation
        self.cancelButton.clicked.connect(self.requestCancel)

        self.clearFitMetricCsvButton.clicked.connect(self.clearFitMetricTxtFile)
        self.clearFitMetricCsvButton.installEventFilter(self)

        self.renameFitMetricCsvButton.clicked.connect(self.renameFitMetricFile)
        self.renameFitMetricCsvButton.installEventFilter(self)

        # Button: Copy results to Asteroid Occultation Report Form (... fill Excel report)
        self.fillExcelReportButton.installEventFilter(self)
        self.fillExcelReportButton.clicked.connect(self.fillExcelReport)

        # Button: View frame
        self.viewFrameButton.clicked.connect(self.viewFrame)
        self.viewFrameButton.installEventFilter(self)
        self.frameNumSpinBox.installEventFilter(self)
        self.fieldViewCheckBox.installEventFilter(self)
        self.flipYaxisCheckBox.installEventFilter(self)
        self.flipXaxisCheckBox.installEventFilter(self)

        # Button: Write error bar plot to file
        self.writeBarPlots.clicked.connect(self.exportBarPlots)
        self.writeBarPlots.installEventFilter(self)

        # Button: Write graphic to file
        self.writePlot.clicked.connect(self.exportGraphic)
        self.writePlot.installEventFilter(self)

        # Button: Write csv file
        self.writeCSVButton.clicked.connect(self.writeCSVfile)
        self.writeCSVButton.installEventFilter(self)

        # Button: Start over
        self.startOver.clicked.connect(self.restart)
        self.startOver.installEventFilter(self)

        # Set up handlers for clicks on table view of data
        self.table.cellClicked.connect(self.cellClick)
        self.table.verticalHeader().sectionClicked.connect(self.rowClick)
        self.table.installEventFilter(self)
        self.helpLabelForDataGrid.installEventFilter(self)

        # Re-instantiate mainPlot             Note: examine gui.py
        # to get this right after a re-layout !!!!  self.widget changes sometimes
        # as does horizontalLayout_?
        axesPen = pg.mkPen((0, 0, 0), width=3)

        self.timeAxis = TimestampAxis(orientation='bottom', pen=axesPen)
        self.timeAxis.setFetcher(self.getTimestampFromRdgNum)

        toptimeAxis = TimestampAxis(orientation='top', pen=axesPen)
        toptimeAxis.setFetcher(self.getTimestampFromRdgNum)

        leftAxis = pg.AxisItem(orientation='left', pen=axesPen)

        oldMainPlot = self.mainPlot
        self.mainPlot = PlotWidget(self.splitterTwo,
                                   viewBox=CustomViewBox(border=(255, 255, 255)),
                                   axisItems={'bottom': self.timeAxis, 'top': toptimeAxis, 'left': leftAxis},
                                   enableMenu=False, stretch=1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding,
                                           QtWidgets.QSizePolicy.Policy.Expanding)
        sizePolicy.setHorizontalStretch(1)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.mainPlot.sizePolicy().hasHeightForWidth())
        self.mainPlot.setSizePolicy(sizePolicy)
        self.mainPlot.setObjectName("mainPlot")
        self.mainPlot.getPlotItem().showAxis('bottom', True)
        self.mainPlot.getPlotItem().showAxis('top', True)
        self.mainPlot.getPlotItem().showAxis('left', True)
        self.splitterTwo.addWidget(self.mainPlot)

        oldMainPlot.setParent(None)

        self.plotItem = self.mainPlot.getPlotItem()

        self.mainPlot.scene().sigMouseMoved.connect(self.reportMouseMoved)
        self.verticalCursor = pg.InfiniteLine(angle=90, movable=False, pen=(0, 0, 0))
        self.mainPlot.addItem(self.verticalCursor)
        self.blankCursor = True
        self.mainPlot.viewport().setProperty("cursor", QtGui.QCursor(QtCore.Qt.CursorShape.BlankCursor))
        mouseSignal.connect(self.mouseEvent)

        # Set up handler for clicks on data plot
        self.mainPlot.scene().sigMouseClicked.connect(self.processClick)
        self.mainPlotViewBox = self.mainPlot.getViewBox()
        self.mainPlotViewBox.rbScaleBox.setPen(pg.mkPen((255, 0, 0), width=2))
        self.mainPlotViewBox.rbScaleBox.setBrush(pg.mkBrush(None))
        self.mainPlot.hideButtons()
        self.mainPlot.showGrid(y=True, alpha=.5)

        self.extra = []
        self.demoLightCurve = []  # Used for detectability demonstration
        self.minDetectableDurationRdgs = None  # Used for detectability demonstration
        self.minDetectableDurationSecs = None  # Used for detectability demonstration
        self.aperture_names = []
        self.fullDataDictionary = {}
        self.additionalDataSetNames = []
        self.initializeTableView()  # Mostly just establishes column headers

        # Open (or create) file for holding 'sticky' stuff
        self.settings = QSettings('pyote.ini', QSettings.IniFormat)
        self.settings.setFallbacksEnabled(False)

        self.edgeTimePrecisionEdit.setText(self.settings.value('edgeTimeFitPrecision', '0.010'))
        self.chordDurationPrecisionEdit.setText(self.settings.value('chordDurationFitPrecision', '0.010'))
        self.limbAnglePrecisionEdit.setText(self.settings.value('limbAngleFitPrecision', '1'))
        self.missDistancePrecisionEdit.setText(self.settings.value('missDistanceFitPrecision', '0.01'))

        dotSize = self.settings.value('vizierPlotDotSize', '4')
        self.vzDotSizeSpinner.setValue(int(dotSize))

        nagLevel = self.settings.value('vizierNagLevel', '1')
        self.vzNagLevelSpinbox.setValue(int(nagLevel))

        obsYear = self.settings.value('vizierObsYear', '2023')
        self.vzDateYearSpinner.setValue(int(obsYear))

        self.vizierZipButton.clicked.connect(self.zipVizieRdatFiles)
        self.vizierZipButton.installEventFilter(self)

        self.vzNagLevelSpinbox.installEventFilter(self)
        self.vzNagLevelLabel.installEventFilter(self)

        lineWidth = self.settings.value('lineWidth', '5')
        dotSize = self.settings.value('dotSize', '8')

        self.lineWidthSpinner.setValue(int(lineWidth))
        self.dotSizeSpinner.setValue(int(dotSize))

        showDetails = self.settings.value('showDetails', 'true')
        self.showDetailsCheckBox.setChecked(showDetails == 'true')

        versusTime = self.settings.value('versusTime', 'true')
        self.versusTimeCheckBox.setChecked(versusTime == 'true')

        showLegend = self.settings.value('showLegend', 'true')
        self.showLegendsCheckBox.setChecked(showLegend == 'true')

        showNotes = self.settings.value('showNotes', 'true')
        self.showAnnotationsCheckBox.setChecked(showNotes == 'true')

        allowNewVersionPopup = self.settings.value('allowNewVersionPopup', 'true')
        if allowNewVersionPopup == 'true':
            self.allowNewVersionPopupCheckbox.setChecked(True)
        else:
            self.allowNewVersionPopupCheckbox.setChecked(False)

        self.allowNewVersionPopupCheckbox.installEventFilter(self)

        tabNameList = self.settings.value('tablist', [])
        if tabNameList:
            self.redoTabOrder(tabNameList)
            # self.switchToTabNamed(tabNameList[0])

        # This is a 'hack' to override QtDesigner which has evolved somehow the ability to block my attempts
        # at setting reasonable size parameters in the drag-and drop Designer.
        self.resize(QSize(0, 0))
        self.setMinimumSize(QSize(0, 0))

        self.logFile = None
        self.detectabilityLogFile = None
        self.normalizationLogFile = None

        # Use 'sticky' settings to size and position the main screen
        self.resize(self.settings.value('size', QSize(800, 800)))
        self.move(self.settings.value('pos', QPoint(50, 50)))


        self.showErrBarsCheckBox.setChecked(self.settings.value('showErrorBars', 'true') == 'true')
        self.showUnderlyingLightcurveCheckBox.setChecked(self.settings.value('showUnderlyingLightCurve', 'true') == 'true')
        self.showCameraResponseCheckBox.setChecked(self.settings.value('showCameraResponse', 'false') == 'true')
        self.showEdgesCheckBox.setChecked(self.settings.value('showEdges', 'true') == 'true')
        self.showOCRcheckFramesCheckBox.setChecked(self.settings.value('doOCRcheck', 'true') == 'true')
        self.showTimestampsCheckBox.setChecked(self.settings.value('showTimestamps', 'true') == 'true')

        self.removeAddedDataSetsButton.clicked.connect(self.removeAddedDataSets)
        self.removeAddedDataSetsButton.installEventFilter(self)

        self.ne3NotInUseRadioButton.setChecked(self.settings.value('ne3NotInUse', 'false') == 'true')
        self.ne3NotInUseRadioButton.clicked.connect(self.clearNe3SolutionPoints)
        self.dnrOffRadioButton.setChecked(self.settings.value('dnrOff', 'false') == 'true')
        self.dnrLowRadioButton.setChecked(self.settings.value('dnrLow', 'false') == 'true')
        self.dnrMiddleRadioButton.setChecked(self.settings.value('dnrMiddle', 'false') == 'true')
        self.dnrHighRadioButton.setChecked(self.settings.value('dnrHigh', 'false') == 'true')

        self.dnrLowDspinBox.setValue(float(self.settings.value('dnrLowDtc', 0.50)))
        self.dnrLowRspinBox.setValue(float(self.settings.value('dnrLowRtc', 0.50)))

        self.dnrMiddleDspinBox.setValue(float(self.settings.value('dnrMiddleDtc', 1.00)))
        self.dnrMiddleRspinBox.setValue(float(self.settings.value('dnrMiddleRtc', 1.00)))

        self.dnrHighDspinBox.setValue(float(self.settings.value('dnrHighDtc', 4.50)))
        self.dnrHighRspinBox.setValue(float(self.settings.value('dnrHighRtc', 2.00)))

        self.dnrHighDspinBox.valueChanged.connect(self.processTimeConstantChange)
        self.dnrHighRspinBox.valueChanged.connect(self.processTimeConstantChange)
        self.dnrMiddleDspinBox.valueChanged.connect(self.processTimeConstantChange)
        self.dnrMiddleRspinBox.valueChanged.connect(self.processTimeConstantChange)
        self.dnrLowDspinBox.valueChanged.connect(self.processTimeConstantChange)
        self.dnrLowRspinBox.valueChanged.connect(self.processTimeConstantChange)

        # splitterOne is the vertical splitter in the lower panel.
        # splitterTwo is the vertical splitter in the upper panel
        # splitterThree is the horizontal splitter between the top and bottom panel

        if self.settings.value('splitterOne') is not None:
            self.splitterOne.restoreState(self.settings.value('splitterOne'))
            self.splitterTwo.restoreState(self.settings.value('splitterTwo'))
            self.splitterThree.restoreState(self.settings.value('splitterThree'))

        self.pymovieFileInUse = False

        self.droppedFrames = []
        self.cadenceViolation = []
        self.timeDelta = None

        self.selPts = []
        self.initializeVariablesThatDontDependOnAfile()

        self.pathToVideo = None
        self.cascadePosition = None
        self.cascadeDelta = 25
        self.frameViews = []

        self.fieldMode = False

        self.d_underlying_lightcurve = None
        self.r_underlying_lightcurve = None

        self.d_candidates = None
        self.d_candidate_entry_nums = None
        self.r_candidates = None
        self.r_candidate_entry_nums = None
        self.b_intensity = None
        self.a_intensity = None
        self.penumbral_noise = None
        self.penumbralDerrBar = None
        self.penumbralRerrBar = None
        self.lastDmetric = 0.0
        self.lastRmetric = 0.0

        self.xlsxDict = {}

        self.checkForNewVersion()

        # We are removing this procedure in version 4.8.0 and above
        # so that can use pipenv to install PyOTE without the need
        # for an Anaconda3 install or a pip install pyote==x.y.z
        # self.copy_desktop_icon_file_to_home_directory()

        # self.helperThing = HelpDialog()

        if self.externalCsvFilePath is not None:
            if os.path.exists(self.externalCsvFilePath):
                self.showMsg(f'We will read: {self.externalCsvFilePath}')
                self.readDataFromFile()
            else:
                self.showMsg(f'Could not find csv file specified: {self.externalCsvFilePath}')
                self.externalCsvFilePath = None



        # noinspection PyTypeChecker
        self.Lcp: LightcurveParameters = None

        self.newEventDataBeingEntered = False

        self.initializeModelLightcurvesPanel()

        self.showMsg(f'Home directory: {self.homeDir}', color='black', bold=True)
        try:
            self.copy_modelExamples_to_Documents()
        except FileNotFoundError as e:
            self.showInfo(f'{e}')


        if self.allowNewVersionPopupCheckbox.isChecked():
            self.showHelp(self.allowNewVersionPopupCheckbox)

        vizierFilesToSend = self.getListOfVizieRlightcurves()
        count = len(vizierFilesToSend)
        # self.showInfo(f'{count} VizieR files were found ready to send')
        if count >= self.vzNagLevelSpinbox.value():
            self.showInfo(f'You have {count} VizieR archive lightcurves\n'
                          f'that have not been zipped and sent!')

    # ====  New method entry point ===

    def TBD(self):
        self.showInfo(f'TBD')

    def processDeepVtest(self):

        if not len(self.yValues) == 1500:
            self.showInfo("The target light curve is NOT a deep V light curve!")
            return

        baseline_left = self.yValues[0:501]        # 0...500
        baseline_right = self.yValues[1000:1500]   # 1000...1499
        baseline_points = np.concatenate((baseline_left, baseline_right))
        baseline_mean = np.mean(baseline_points)


        # Calculate the v points

        segment = 250
        npts = 1500
        model = np.ndarray(npts)  # noqa
        for i in range(npts):
            if i < 2 * segment:  # 0 to 500 are baseline values (model[500] is set at the else: calculation)
                model[i] = 1.0
            elif i > 4 * segment:  # 1000 to 1499 are baseline values
                model[i] = 1.0
            else:
                model[i] = abs(
                    1 - (i - 2 * segment) / segment)  # 0 at 750  500 to 750 = left slope  750 to 1000 = right slope
        model *= baseline_mean

        v_model = model[501:1000]
        v_values = self.yValues[501:1000]
        v_deltas = v_values - v_model
        v_noise = np.std(v_deltas)
        v_delta_mean = np.mean(v_deltas)

        plt.figure(figsize=(12,12))
        ax1 = plt.subplot(2,1,1)
        ax2 = plt.subplot(2,1,2)

        if self.targetKey == '':
            light_curve_target = self.lightcurveTitles[0].text()
        else:
            light_curve_target = self.targetKey
        ax1.set_title(f'Light curve with V occultation fit ({light_curve_target})')

        ax1.set_ylabel('ADU')
        ax1.set_xlabel('reading number')
        ax1.plot(self.yValues, marker='.', markersize=4, linestyle='None')

        baseline_noise = np.std(baseline_points)
        if baseline_noise > 0.0:
            snr = baseline_mean / baseline_noise
        else:
            snr = 0.0

        s = f'baseline noise: {baseline_noise:0.2f}  snr: {snr:0.2f}\n'
        s += f'noise during V: {v_noise:0.2f}\n'
        s += f'avg difference from model during V: {v_delta_mean:0.2f}'

        ax1.plot(model, marker='.', markersize=1)
        ax1.text(0.01, 0.4, s, transform=ax1.transAxes, bbox=dict(facecolor='white', alpha=1), fontsize=10)

        ax2.set_title("Deviations from model")
        ax2.set_ylabel('ADU')
        ax2.set_xlabel('reading number')
        flattened_data = self.yValues - model
        ymin = np.min(flattened_data)
        ymax = np.max(flattened_data)
        ax2.plot(flattened_data, marker='.', markersize=3, linestyle='None')
        ax2.vlines(501, ymin=ymin, ymax=ymax, color='red', label='V start/end')
        ax2.vlines(999, ymin=ymin, ymax=ymax, color='red')
        ax2.vlines([750], ymin=ymin, ymax=ymax, color='green', label='V center')
        ax2.hlines(0, xmin=501, xmax=999, color='black')
        ax2.hlines(v_delta_mean, xmin=501, xmax=999, color='orange', label='avg diff from model during V')
        ax2.legend(loc='upper left')
        plt.show()

    def processSteppedLevelsTest(self):

        if not len(self.yValues) == 6000:
            self.showInfo("The target light curve is NOT a stepped levels light curve!")
            return

        level1_mean = np.mean(self.yValues[0:1000])
        level1_sigma = np.std(self.yValues[0:1000])

        level2_mean = np.mean(self.yValues[1000:2000])
        level2_sigma = np.std(self.yValues[1000:2000])

        level3_mean = np.mean(self.yValues[2000:3000])
        level3_sigma = np.std(self.yValues[2000:3000])

        level4_mean = np.mean(self.yValues[3000:4000])
        level4_sigma = np.std(self.yValues[3000:4000])

        level5_mean = np.mean(self.yValues[4000:5000])
        level5_sigma = np.std(self.yValues[4000:5000])

        level6_mean = np.mean(self.yValues[5000:6000])
        level6_sigma = np.std(self.yValues[5000:6000])

        Bvalues = [level1_mean, level2_mean, level3_mean, level4_mean, level5_mean, level6_mean]
        sigmaB = [level1_sigma, level2_sigma, level3_sigma, level4_sigma, level5_sigma, level6_sigma]

        if self.targetKey == '':
            curve_title = self.lightcurveTitles[0].text()
        else:
            curve_title = self.targetKey

        self.plotLinearFitAndResiduals(Bvalues, sigmaB, curve_title)

    @staticmethod
    def plotLinearFitAndResiduals(Bvalues, sigmaB, curve_title):

        # Compute weights to use in fit
        snr = []
        for i in range(len(Bvalues)):
            snr.append(Bvalues[i] / sigmaB[i])  # Weight by snr

        x = [1.0, 0.7, 0.3, 0.2, 0.1, 0.05]

        z = np.polyfit(x, Bvalues, 1, w=snr, full=True)
        f = np.poly1d(z[0])

        # calculate new x's and y's
        x_new = np.linspace(x[0], x[-1], 50)
        y_new = f(x_new)

        plt.figure(figsize=(14, 6))
        ax1 = plt.subplot(1, 2, 1)
        ax2 = plt.subplot(1, 2, 2)

        ax1.plot(x, Bvalues, 'o', x_new, y_new)
        ax1.set_title(f'{curve_title} linear fit')
        ax1.set_ylabel('ADU')
        ax1.set_xlabel('ratio expected')
        ax1.set_xlim([x[0] - 1, x[-1] + 1])
        ax1.grid()

        variance = 0.0
        deltas = []
        for i in range(len(x)):
            y_new = f(x[i])
            delta = y_new - Bvalues[i]
            deltas.append(delta)
            variance += delta ** 2

        ax2.plot(x, deltas, 'o')
        ax2.set_title(f'{curve_title} residuals')
        ax2.plot([0, 1], [0, 0], '-')
        ax2.set_xlim([x[0] - 1, x[-1] + 1])
        big_limit = max(max(deltas), abs(min(deltas))) + 2
        ax2.set_ylim(-big_limit, big_limit)
        ax2.set_ylabel('ADU')
        ax2.set_xlabel('ratio expected')
        s = f'standard deviation of residuals: {np.sqrt(variance / len(x)):0.2f}'
        ax2.text(0.2, 0.93, s, transform=ax2.transAxes, bbox=dict(facecolor='white', alpha=1), fontsize=12)

        ax2.grid()

        plt.show()
    # def setMetricLimits(self):
    #     leftLimit, done = QtWidgets.QInputDialog.getInt(self, ' ', 'Lefthand reading number of data to be used for metric:', min=0)
    #     if done:
    #         rightLimit, done = QtWidgets.QInputDialog.getInt(self, ' ',
    #                                                          'Righthand reading number of data to be used for metric:', min=0)
    #         if done:
    #             if 0 <= leftLimit < rightLimit:
    #                 self.metricLimitLeft = leftLimit
    #                 self.metricLimitRight = rightLimit
    #                 self.Lcp.metricLimitLeft = leftLimit
    #                 self.Lcp.metricLimitRight = rightLimit
    #                 self.showInfo(f'Metric will be calculated using data points from {leftLimit} to {rightLimit} inclusive.')
    #             else:
    #                 self.showInfo(f'The reading numbers supplied are invalid.')

    def plotFamilyOfLightcurves(self):
        def move_figure(f, x, y):  # noqa
            backend = matplotlib.get_backend()
            if backend == 'TkAgg':
                f.canvas.manager.window.wm_geometry("+%d+%d" % (x,y))
            elif backend == 'WXAgg':
                f.canvas.manager.window.SetPosition((x,y))
            else:
                f.canvas.manager.window.move(x,y)

        def plot_curve_set(calculate_curves, x, yOut):  # noqa
            fig = plt.figure(constrained_layout=False, figsize=(10, 8))
            fig.canvas.manager.set_window_title("Family plot")
            ax1 = fig.add_subplot(1, 1, 1)  # noqa lightcurve axes

            self.showMsg("", blankLine=False)
            if calculate_curves:
                self.showMsg(f'Initiating computations ...', bold=True, color='red', blankLine=False)
            for i in range(1, numCurves + 1):  # noqa
                if calculate_curves:
                    QtWidgets.QApplication.processEvents()
                    x, y, D_edge, R_edge = generalizedDiffraction(LCP=self.Lcp, wavelength1=None, wavelength2=None,  # noqa
                                                                  skip_central_calc=False)
                    if i == 1:
                        ax1.set_ylim(0, 1.1 * np.max(y))

                    yOut.append(y)
                else:
                    y = yOut[i-1]

                if param == 'miss distance km':
                    label = f'{i}: {param} {self.Lcp.miss_distance_km:0.4f} km'
                    if calculate_curves:
                        self.showMsg(f'Computed curve {label}', bold=True, color='red', blankLine=False)
                    ax1.plot(x, y, label=label)
                    plt.draw()
                    self.Lcp.miss_distance_km += stepValue
                elif param == 'ellipse angle degrees':
                    label = f'{i}: {param} {self.Lcp.ellipse_angle_degrees:0.4f} degrees'
                    if calculate_curves:
                        self.showMsg(f'Computed curve {label}', bold=True, color='red', blankLine=False)
                    ax1.plot(x, y, label=label)
                    self.Lcp.ellipse_angle_degrees += stepValue
                elif param == 'major axis km':
                    label = f'{i}: {param} {self.Lcp.asteroid_major_axis:0.4f} km'
                    if calculate_curves:
                        self.showMsg(f'Computed curve {label}', bold=True, color='red', blankLine=False)
                    ax1.plot(x, y, label=label)
                    self.Lcp.asteroid_major_axis += stepValue
                elif param == 'minor axis km':
                    label = f'{i}: {param} {self.Lcp.asteroid_minor_axis:0.4f} km'
                    if calculate_curves:
                        self.showMsg(f'Computed curve {label}', bold=True, color='red', blankLine=False)
                    ax1.plot(x, y, label=label)
                    self.Lcp.asteroid_minor_axis += stepValue
                elif param == 'asteroid diameter km':
                    label = f'{i}: {param} {self.Lcp.asteroid_diameter_km:0.4f} km'
                    if calculate_curves:
                        self.showMsg(f'Computed curve {label}', bold=True, color='red', blankLine=False)
                    ax1.plot(x, y, label=label)
                    self.Lcp.asteroid_diameter_km += stepValue
                else:
                    self.showInfo(f'Programming error: no such parameter named {param}!')
                    return

            self.Lcp = copy.deepcopy(savedLcp)  # Restore the original self.Lcp structure
            ax1.set_ylim(0, 1.1 * np.max(yVec[:]))
            ax1.set_xlabel("Kilometers")
            ax1.legend(loc='best', fontsize=8)
            return fig, ax1, x

        if not self.diffractionRadioButton.isChecked():
            self.showInfo(f'Use of this feature is restricted to diffraction models')
            return

        params = ['miss distance km', 'ellipse angle degrees', 'major axis km', 'minor axis km', 'asteroid diameter km']
        param, gotParam = QtWidgets.QInputDialog.getItem(
            self, ' ', 'Select parameter to change:', params, editable=False)

        if gotParam:
            centerValue, gotCenterValue = QtWidgets.QInputDialog.getDouble(
                self, ' ', 'Parameter value to use for center curve:', decimals=5)
            if gotCenterValue:
                stepValue, gotStepValue = QtWidgets.QInputDialog.getDouble(
                    self, ' ', 'Step change between curves:', decimals=5)
                if gotStepValue:
                    numSideCurves, gotNumSideCurves = QtWidgets.QInputDialog.getInt(
                        self, ' ', 'Number of curves on either side of center curve:', min=1)
                else:
                    self.showInfo('No step change between curves value given')
                    return
                if not gotNumSideCurves:
                    self.showInfo('The number of side curves wanted was not given')
                    return
            else:
                self.showInfo('No value given for center curve parameter')
                return
            startValue = centerValue - numSideCurves * stepValue
            endValue = centerValue + numSideCurves * stepValue
            numCurves = 2 * numSideCurves + 1
            reply = QMessageBox.question(
                self, "Family plot information",
                f'{param} will start at {startValue:0.4f} and go to {endValue:0.4f}  in {numCurves} steps.\n\n'
                f'The center value of {param} is {centerValue:0.4f}.\n\n'
                f'Is that OK?', QMessageBox.Yes, QMessageBox.No )
            if reply == QMessageBox.No:
                return
            self.showMsg('', blankLine=False)
            self.showMsg(f'{param} will start at {startValue:0.4f} and go to {endValue:0.4f} in {numCurves} steps '
                         f'of size {stepValue:0.4f}',
                         color='black', bold=True)
        else:
            self.showInfo('No parameter selection given')
            return


        savedLcp = copy.deepcopy(self.Lcp)

        self.setFamilyPlotStartValues(param, startValue)

        yVec = []
        fig_num, ax1, x = plot_curve_set(calculate_curves=True, x=None, yOut=yVec)
        move_figure(fig_num, 50, 50)
        plt.show()


        reply = QMessageBox.question(
            self, "Model comparisons", "Do you want to do a noise/model comparison", QMessageBox.Yes, QMessageBox.No
        )

        if reply == QMessageBox.No:
            plt.show()
            return
        else:

            sample_index_step = self.Lcp.shadow_speed * self.Lcp.frame_time / (x[1] - x[0])

            referenceCurveNumber = numCurves // 2 + 1

            noiseValues = []
            gotNoiseValue = True
            while gotNoiseValue:
                if not noiseValues:
                    noiseValue, gotNoiseValue = QtWidgets.QInputDialog.getDouble(
                        self, ' ', 'First noise level (sigmaB) to use:', decimals=5)
                else:
                    noiseValue, gotNoiseValue = QtWidgets.QInputDialog.getDouble(
                        self, ' ', '... another noise level to use (click cancel if done entering noise values):', decimals=5)
                if gotNoiseValue:
                    noiseValues.append(noiseValue)

            if not noiseValues:
                self.showMsg(f'No noise values provided --- finished.')
                return

            for value in noiseValues:
                if value < 0:
                    self.showInfo(f'A noise level specified is negative --- aborting.')
                    return

            sampled_y = [[] for _ in range(numCurves)]
            xpts = []
            sample_index = -sample_index_step
            i_limit = len(yVec[0])
            for k in range(numCurves):
                while True:
                    sample_index += sample_index_step
                    i = round(sample_index)
                    if i >= i_limit:
                        break
                    sampled_y[k].append(yVec[k][i])
                    if k == 0:
                        xpts.append(x[i])
                sample_index = -sample_index_step  # set up for another run (next k - next curve)

            ref = np.array(sampled_y[referenceCurveNumber-1])
            xpts = np.array(xpts)

            noise_plot_num = 0
            for noiseToUse in noiseValues:
                noise_plot_num += 1
                noiseVec = np.random.normal(0.0, noiseToUse, xpts.size)

                noisedRef = ref + noiseVec  # For demo purposes on the plot

                # Repeat the base plot in a new figure
                self.setFamilyPlotStartValues(param, startValue)

                fig_num, ax1, x = plot_curve_set(calculate_curves=False, x=x, yOut=yVec)
                move_figure(fig_num, 50 + 40 * noise_plot_num, 50 + 40 * noise_plot_num)
                plt.show()

                ax1.plot(xpts, ref, ".", label="ref sample points")
                ax1.plot(xpts, noisedRef,"-", marker='.', label="example simulated observation",
                         color='gray', alpha=0.5)
                ax1.legend(loc='best', fontsize=8)

                # Compute model comparison stats
                numTrials=50000
                cum_aic_weights = None

                for j in range(numTrials):
                    # re-noise the reference and comparison light curves
                    noiseVec = np.random.normal(0.0, noiseToUse, xpts.size)
                    noisedObs = ref + noiseVec  # re-create the observation signal using ref and noiseVec

                    # Calculate the metrics and relative probabilities
                    rss = []
                    aic = []
                    for k in range(numCurves):
                        rss.append(np.sum(((noisedObs - np.array(sampled_y[k]))/noiseToUse)**2) / xpts.size)
                        aic.append(xpts.size * np.log(rss[-1]))
                    aic = np.array(aic)
                    min_aic = np.min(aic)
                    delta_aic = aic - min_aic
                    denom = np.sum(np.exp(-delta_aic/2.0))
                    aic_weights = np.exp(-delta_aic/2.0) / denom

                    if cum_aic_weights is None:
                        cum_aic_weights = np.exp(-delta_aic/2.0) / denom
                    else:
                        cum_aic_weights += aic_weights

                avg_aic_weight = cum_aic_weights / numTrials
                text_for_plot = f'Noise (sigmaB) used: {noiseToUse:0.1f}  SNR = {self.Lcp.baseline_ADU/noiseToUse:0.2f}\n'
                self.showMsg(f'{text_for_plot}', bold=True, blankLine=False)

                for i in range(numCurves):
                    text_line = f'curve {i+1} relative probability: {avg_aic_weight[i]:0.2f}'
                    self.showMsg(f'{text_line}', bold=True, blankLine=False)
                    text_for_plot += text_line + '\n'
                text_for_plot = text_for_plot[:-1]  # Get rid of trailing \n

                ax1.text(0.5, 0.03, text_for_plot, transform=ax1.transAxes,
                         bbox=dict(facecolor='white', alpha=1), fontsize=10)

            plt.show()

    def setFamilyPlotStartValues(self, param, startValue):
        if param == 'miss distance km':
            self.Lcp.miss_distance_km = startValue
        elif param == 'ellipse angle degrees':
            self.Lcp.ellipse_angle_degrees = startValue
        elif param == 'major axis km':
            self.Lcp.asteroid_major_axis = startValue
        elif param == 'minor axis km':
            self.Lcp.asteroid_minor_axis = startValue
        elif param == 'asteroid diameter km':
            self.Lcp.asteroid_diameter_km = startValue
        else:
            self.showInfo(f'Programming error: no such parameter named {param}!')
            # return

    def removeAddedDataSets(self):

        # If lightcurves came from Tangra, we must not remove LC2, LC3, or LC4 because they cannot be
        # replaced - they are not in the curveSelectionComboBox widget.
        if self.lightcurveTitle_1.text() == 'LC1':
            return

        self.userDataSetAdditions = []
        # for i in range(len(self.aperture_names), 10):
        for i in range(1, 10):
            self.lightcurveTitles[i].clear()
            if i == 1:
                self.LC2 = np.array([])
                continue
            if i == 2:
                self.LC3 = np.array([])
                continue
            if i == 3:
                self.LC4 = np.array([])
                continue
            else:
                self.extra = []

        self.lightcurveTitles[0].clear()
        self.LC1 = np.array([])

        for i in range(1, 10):
            self.targetCheckBoxes[i].setChecked(False)
            self.targetCheckBoxes[i].setEnabled(False)
            self.showCheckBoxes[i].setChecked(False)
            self.showCheckBoxes[i].setEnabled(False)
            self.referenceCheckBoxes[i].setChecked(False)
            self.referenceCheckBoxes[i].setEnabled(False)
            self.yOffsetSpinBoxes[i].setEnabled(False)
            self.yOffsetSpinBoxes[i].setValue(0)
            self.xOffsetSpinBoxes[i].setEnabled(False)
            self.xOffsetSpinBoxes[i].setValue(0)

        self.targetCheckBoxes[0].setChecked(True)
        self.referenceCheckBoxes[0].setChecked(False)
        self.showCheckBoxes[0].setChecked(True)

        self.userDataSetAdditions = ['ClearAll']
        self.initializeTableView()

        self.newRedrawMainPlot()

    def handleDataSetSelection(self):
        dataSetSelected = self.curveSelectionComboBox.currentText()
        self.userDataSetAdditions.append(dataSetSelected)
        self.initializeTableView()
        self.checkForBothCorrectAndTargetPresent()

    def isValidInput(self, valueStr='', valueName='', entryType='int', negativeAllowed=True, allowEmpty=False):
        # entryType: 'int' | 'float'

        if not allowEmpty and len(valueStr) == 0:
            self.showInfo(f'Please enter a value for {valueName}')
            return False

        try:
            if entryType == 'int':
                value = int(valueStr)
                if not negativeAllowed and value < 0:
                    self.showInfo(f'{valueName} cannot be negative.')
                    return False
                return True
            elif entryType == 'float':
                value = float(valueStr)
                if not negativeAllowed and value < 0:
                    self.showInfo(f'{valueName} cannot be negative.')
                    return False
                return True
            else:
                self.showInfo(f'entryType of {entryType} is not supported - parameter error.')
                return False
        except ValueError:
            self.showInfo(f'The input string for {valueName} was {valueStr}\n\n'
                          f' This is NOT a valid {entryType}')
            return False

    @staticmethod
    def isNegZero(value):
        if value == 0.0:
            return math.copysign(1, value) < 0
        else:
            return False

    def showVizieRplot(self):
        self.processVizieRdataInput(plotWanted=True)

    @staticmethod
    @contextmanager
    def changeWorkingDirectory(newDir):
        oldpwd = os.getcwd()
        os.chdir(newDir)
        try:
            yield
        finally:
            os.chdir(oldpwd)

    def getListOfVizieRlightcurves(self):
        VizieRdir = self.getVizieRdirectory()
        if sys.platform == 'darwin' or sys.platform == 'linux':
            filePaths = glob.glob(f'{VizieRdir}/*.dat')
        else:
            filePaths = glob.glob(f'{VizieRdir}\\*.dat')
        # print(filePaths)
        fileNames = []
        for file in filePaths:
            fileNames.append(os.path.basename(file))
        return fileNames

    def getPathsOfVizieRlightcurves(self):
        VizieRdir = self.getVizieRdirectory()
        if sys.platform == 'darwin' or sys.platform == 'linux':
            filePaths = glob.glob(f'{VizieRdir}/*.dat')
        else:
            filePaths = glob.glob(f'{VizieRdir}\\*.dat')
        return filePaths

    def zipVizieRdatFiles(self):
        targetDir = self.getDocumentsDirectory()
        datFileNames = self.getListOfVizieRlightcurves()
        if len(datFileNames) == 0:
            self.showInfo(f'You have no VizieR files to be sent.')
            return


        # Get the crc32 value of all the files that are to be zipped
        # We have to read all of the files.
        datFilePaths = self.getPathsOfVizieRlightcurves()
        code = 0
        for filepath in datFilePaths:
            with open(filepath, mode='r') as file:
                all_of_it = file.read()
                code = zlib.crc32(all_of_it.encode("utf8"), code)
        crc32 = f'{code:8x}'

        userName = self.getUserName()
        today = str(datetime.date.today())
        archiveName = f'Archive_of_VizieR_lightcurves_{today}_{userName}_{crc32}.zip'
        archiveName = os.path.join(targetDir, archiveName)

        with self.changeWorkingDirectory(self.getVizieRdirectory()):
            with zipfile.ZipFile(archiveName, mode='w') as archive:
                for filename in datFileNames:
                    archive.write(filename)
                    os.rename(filename, filename + ".addedtoziparchive")
                self.showInfo(f'Your VizieR archive of lightcurves has been written to:\n\n'
                              f'{archiveName}')
                self.showHelp(self.vizierLabel)

    def processVizieRdataInput(self, plotWanted=False, fileWanted=False):
        if self.VizieRdict is None:
            self.showInfo(f'There is no data available.')
            return

        if self.left == 0 and self.right == (self.dataLen - 1) and fileWanted:
            title = self.tr("Have you forgotten to trim the lightcurve ???")
            query = self.tr(
                f"The lightcurve has not been trimmed.\n\n"
                f"Only the occultation event and enough points on either side\n"
                f"of the event to allow baseline noise to be well respresented\n"
                f"is needed. Typically around a hundred points on either side\n"
                f"will be sufficient.\n\n"
                f"Do you wish to apply a 'trim' ?"
            )
            reply = QMessageBox.question(
                self,
                title,
                query,
                QMessageBox.Yes,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                return

        # We have to get correspondence between VizieR yValues (with inserted dropped readings)
        # and self.yValues so that we can calculate numReadings with the inclusion of filled-in
        # dropped readings.

        # Create corrspondence table that maps self.yValue indices into VizieR["yValues"] indices
        yValueIndicesToVizierIndices = []
        k = 0  # index into self.yValues
        for i in range(len(self.VizieRdict["yValues"])):
            if not self.isNegZero(self.VizieRdict["yValues"][i]):
                yValueIndicesToVizierIndices.append(k)
            k += 1

        vizierLeft = yValueIndicesToVizierIndices[self.left]
        try:
            vizierRight = yValueIndicesToVizierIndices[self.right]
        except IndexError:
            vizierRight = self.right

        if plotWanted:
            vizierY = self.VizieRdict["yValues"][vizierLeft:vizierRight + 1]

            # Compute scale factor
            maxValue = np.max(vizierY)
            scaleFactor = 9524 / maxValue

            for i, value in enumerate(vizierY):
                if not self.isNegZero(value):
                    scaledValue = round(value * scaleFactor)
                    vizierY[i] = scaledValue

            # Extract the plot segments created by dropped readings
            plotYseg = [[]]
            plotXseg = [[]]
            missingX = []
            missingY = []
            for i, reading in enumerate(vizierY):
                if self.isNegZero(reading):
                    missingX.append(i)
                    missingY.append(0)
                    if not plotYseg[-1] == []:  # We need to start a new segment
                        # Start a new segment
                        plotYseg.append([])
                        plotXseg.append([])
                else:
                    plotYseg[-1].append(reading)
                    plotXseg[-1].append(i)

            fig = plt.figure(constrained_layout=False, figsize=(8,4))
            lowVal = min(-1000, np.min(vizierY)-1000)
            plt.ylim(lowVal, 10000)
            plt.plot([0,len(vizierY)], [0, 0], ls='-', color='lightgray')
            fig.canvas.manager.set_window_title("VizieR archive lightcurve")
            k = 0
            while k < len(plotYseg):
                plt.plot(plotXseg[k], plotYseg[k], ls='-', linewidth=1, color='black', marker='')
                plt.plot(plotXseg[k], plotYseg[k], ls='', color='blue', marker='o',
                         markersize=self.vzDotSizeSpinner.value())
                k += 1

            if not missingX == []:
                plt.plot(missingX, missingY, ls='', color='red', marker='o',
                         markersize=self.vzDotSizeSpinner.value(), label='dropped reading')
                plt.legend()

            plt.show()
            return

        # If we're not simply doing a plot, we fall through to this code which prepares to produce
        # the VizieR lightcurve file

        initialTimestamp = self.VizieRdict["timestamps"][vizierLeft]
        finalTimestamp = self.VizieRdict["timestamps"][vizierRight]
        deltaTime = convertTimeStringToTime(finalTimestamp) - convertTimeStringToTime(initialTimestamp)
        time = convertTimeStringToTime(initialTimestamp)

        # pcs adjust for camera and VTI time correction
        time += self.camCorrection

        roundedTimeStr = round(time * 100) / 100
        vizierTimestamp = convertTimeToTimeString(float(roundedTimeStr))[1:-3]


        year = self.vzDateYearSpinner.value()

        month = self.vzDateMonthSpinner.value()

        day = self.vzDateDaySpinner.value()

        # pcs added "+1" to next line to correct number of points in the file
        # numReadings = vizierRight - vizierLeft
        numReadings = vizierRight - vizierLeft + 1

        dateText = f'Date: {year}-{month}-{day} {vizierTimestamp}: {deltaTime:0.2f}: {numReadings}'


        emptyStarFields = 0
        UCAC4 = self.vzStarUCAC4Edit.text()
        Tycho2 = self.vzStarTycho2Edit.text()
        hipparcos = self.vzStarHipparcosEdit.text()

        if UCAC4 == '':
            emptyStarFields += 1

        if Tycho2 == '':
            emptyStarFields += 1

        if hipparcos == '':
            emptyStarFields += 1

        if emptyStarFields == 3:
            self.showInfo("You have not entered a star catalog number. This is acceptable IF intentional.\n\n"
                          "It may be intentional because the involved star does not have a supported catalog number type.\n\n"
                          "VizieR accepts a no-star entry, so it is not a problem to leave all star fields empty.\n\n"
                          "Particularly in the case of a G star, the proper course is to leave the star number empty.\n\n"
                          "Best practice is to use the star designation from the Occult4 prediction whenever possible "
                          "- if there were a reliable correlation between a G star designation and a Hipparcos, UCAC4, or Tycho2 "
                          "designation, it would have been found and supplied by Occult4.")

        if emptyStarFields == 1 or emptyStarFields == 0:
            self.showInfo("Please use a single star number.\n\n"
                          "Best practice is to use the star designation from the Occult4 prediction whenever possible.")
            return

        if hipparcos == '':
            hipparcos = '0'
        else:
            if not self.isValidInput(hipparcos, 'Hipparcos'):
                return

        # The following star catalogue numbers are in the VizieR format, but are used for lunars or
        # are no longer supported by Occult 4

        # SAO = self.vzStarSAOedit.text()
        # if SAO == '':
        SAO = '0'
        # XZ80Q = self.vzStarXZ80Qedit.text()
        # if XZ80Q == '':
        XZ80Q = '0'
        # Kepler2 = self.vzStarKeplerEdit.text()
        # if Kepler2 == '':
        Kepler2 = '0'

        if Tycho2 == '':
            Tycho2 = '0-0-1'

        if UCAC4 == '':
            UCAC4 = '0-0'
        else:
            parts = UCAC4.split('-')
            if not len(parts) == 2:
                self.showInfo(f'UCAC4 star designation has incorrect format.\n\n'
                              f'The correct form is: xxx-xxxxxx')
                return
            elif len(parts[0]) > 3 or len(parts[1]) > 6:
                self.showInfo(f'UCAC4 star designation has incorrect format.\n\n'
                              f'The correct form is: xxx-xxxxxx\n\n'
                              f'There are too many digits in one of the fields.')
                return
            else:
                if not (self.isValidInput(parts[0],'UCAC4') and self.isValidInput(parts[1],'UCAC4')):
                    return

        parts = Tycho2.split('-')
        if not len(parts) == 3:
            self.showInfo(f'Tycho2 star designation has incorrect format.\n\n'
                          f'The correct form is: xxxx-xxxxx-x')
            return
        elif len(parts[0]) > 4 or len(parts[1]) > 5 or not len(parts[2]) == 1:
            self.showInfo(f'Tycho2 star designation has incorrect format.\n\n'
                          f'The correct form is: xxxx-xxxxx-x\n\n'
                          f'There are too many digits in one of the fields.')
            return
        else:
            if not (self.isValidInput(parts[0], 'Tycho2')
                    and self.isValidInput(parts[1], 'Tycho2') and self.isValidInput(parts[2], 'Tycho2')):
                return

        starText = f'Star: {hipparcos}: {SAO}: {XZ80Q}: {Kepler2}: {Tycho2}: {UCAC4}'

        longDeg = self.vzSiteLongDegEdit.text()
        if not self.isValidInput(longDeg, "Site longitude (deg)"):
            return
        if not longDeg.startswith('-') and not longDeg.startswith('+'):
            longDeg = '+' + longDeg

        longMin = self.vzSiteLongMinEdit.text()
        if not self.isValidInput(longMin, "Site longitude (min)", negativeAllowed=False):
            return

        longSec = self.vzSiteLongSecsEdit.text()
        if not self.isValidInput(longSec, "Site longitude (sec)", negativeAllowed=False, entryType='float'):
            return

        latDeg = self.vzSiteLatDegEdit.text()
        if not self.isValidInput(latDeg, "Site latitude (deg)"):
            return
        if not latDeg.startswith('-') and not latDeg.startswith('+'):
            latDeg = '+' + latDeg

        latMin = self.vzSiteLatMinEdit.text()
        if not self.isValidInput(latMin, "Site latitude (min)", negativeAllowed=False):
            return

        latSec = self.vzSiteLatSecsEdit.text()
        if not self.isValidInput(latSec, "Site latitude (sec)", negativeAllowed=False, entryType='float'):
            return

        altitude = self.vzSiteAltitudeEdit.text()
        if not self.isValidInput(altitude, "altitude"):
            return

        observer = self.vzObserverNameEdit.text()
        if len(observer) == 0:
            self.showInfo(f'Please enter an observer name')
            return

        locationText = f'Observer: {longDeg}:{longMin}:{longSec}: {latDeg}:{latMin}:{latSec}: '
        locationText += f'{altitude}: {self.vzObserverNameEdit.text()}'

        asteroidNumber = self.vzAsteroidNumberEdit.text()
        if not self.isValidInput(asteroidNumber, "asteroid number"):
            return
        if len(asteroidNumber) > 6:
            self.showInfo(f'Asteroid number is restricted to a max of 6 digits')
            return

        asteroidName = self.vzAsteroidNameEdit.text()
        if len(asteroidName) == 0:
            self.showInfo(f'Please enter an asteroid name')
            return

        objectText = f'Object: Asteroid: {asteroidNumber}: {asteroidName}'

        vizierY = self.VizieRdict["yValues"][vizierLeft:vizierRight+1]
        numDroppedReadings = 0
        for value in vizierY:
            if self.isNegZero(value):
                numDroppedReadings += 1

        # Compute scale factor
        maxValue = np.max(vizierY)
        scaleFactor = 9524 / maxValue

        valuesText = "Values"
        for i, value in enumerate(vizierY):
            if self.isNegZero(value):
                valuesText += ": "
            else:
                scaledValue = round(value * scaleFactor)
                valuesText += f':{scaledValue}'
                vizierY[i] = scaledValue

        if fileWanted:

            # Compose file name
            parts1 = vizierTimestamp.split(':')
            if len(parts1) < 3:
                self.showInfo(f'There is no timestamp data in this file.')
                return
            hh = f'{parts1[0]}'
            mm = f'{parts1[1]}'
            parts2 = parts1[2].split('.')
            sss = f'{parts2[0]:>2}_{parts2[1]:>2}'

            filename = f'({asteroidNumber})_{year}{month:>02}{day:>02}_{hh}{mm}{sss}.dat'
            dest_dir = self.getVizieRdirectory()

            if not os.path.exists(dest_dir):
                # We create the directory
                os.makedirs(dest_dir)

            vizierFilePath = os.path.join(dest_dir, filename)

            if sys.platform == 'darwin' or sys.platform == 'linux':
                CRLF = f'\r\n'
            else:
                # We must be on a Windows machine
                CRLF = f'\n'

            with open(vizierFilePath, 'w') as fileObject:
                fileObject.write(dateText + CRLF)
                fileObject.write(starText + CRLF)
                fileObject.write(locationText + CRLF)
                fileObject.write(objectText + CRLF)
                fileObject.write(valuesText + CRLF)

            self.showInfo(f'Your VizieR lightcurve file was written to:\n\n{vizierFilePath}')

            self.showHelp(self.vizierLabel)

            self.clearVizieRinputs()

    def clearVizieRinputs(self):
        self.vzStarUCAC4Edit.clear()
        self.vzStarTycho2Edit.clear()
        self.vzStarHipparcosEdit.clear()

        self.vzSiteLongDegEdit.clear()
        self.vzSiteLongMinEdit.clear()
        self.vzSiteLongSecsEdit.clear()

        self.vzSiteLatDegEdit.clear()
        self.vzSiteLatMinEdit.clear()
        self.vzSiteLatSecsEdit.clear()

        self.vzSiteAltitudeEdit.clear()
        self.vzObserverNameEdit.clear()

        self.vzAsteroidNumberEdit.clear()
        self.vzAsteroidNameEdit.clear()

    @staticmethod
    def getVizieRdirectory():
        if sys.platform == 'darwin' or sys.platform == 'linux':
            dest_dir = f"{os.environ['HOME']}{r'/Documents/VizieR_lightcurves'}"
        else:
            # We must be on a Windows machine
            dest_dir = f"{os.environ.get('userprofile')}\\Documents\\VizieR_lightcurves"
        return dest_dir

    @staticmethod
    def getUserName():
        if sys.platform == 'darwin' or sys.platform == 'linux':
            homePath = f"{os.environ['HOME']}"
            parts = homePath.split("/")
        else:
            # We must be on a Windows machine
            homePath = f"{os.environ.get('userprofile')}"
            parts = homePath.split("\\")

        return parts[-1]

    @staticmethod
    def getDocumentsDirectory():
        if sys.platform == 'darwin' or sys.platform == 'linux':
            dest_dir = f"{os.environ['HOME']}{r'/Documents/'}"
        else:
            # We must be on a Windows machine
            dest_dir = f"{os.environ.get('userprofile')}\\Documents\\"
        return dest_dir

    def exportVizieRtextFile(self):
        self.processVizieRdataInput(fileWanted=True)

    @ staticmethod
    def validFilename(file_name):
        valid_filename_chars = "-_.() %s%s" % (string.ascii_letters, string.digits)
        for char in file_name:
            if char not in valid_filename_chars:
                return False
        return True

    def deletePastEvent(self):
        currentSelection = self.pastEventsComboBox.currentText()
        title = self.tr("Please confirm ...")
        query = self.tr(
            f"Are you sure you want to delete {currentSelection} ?"
        )
        reply = QMessageBox.question(
            self,
            title,
            query,
            QMessageBox.Yes,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            LCPdir = os.path.dirname(self.csvFilePath)
            if sys.platform == 'darwin' or sys.platform == 'linux':
                filepath = f'{LCPdir}/LCP_{currentSelection}.p'
            else:
                filepath = f'{LCPdir}\\LCP_{currentSelection}.p'
            if os.path.exists(filepath):
                os.remove(filepath)
                self.showInfo(f'Deleted: {currentSelection}')
                # Refill the combo list to show deletion effect
                self.pastEventsComboBox.clear()
                self.fillPastEventsComboBox()
            else:
                self.showInfo(f"{filepath} does not exist")

    def printEventParameters(self):
        self.printFinalReport(onlyLcpValuesWanted = True)

    def printFinalReport(self, onlyLcpValuesWanted=False):
        if self.Lcp is not None:
            self.showMsg('Lightcurve Model Parameters...',
                         color='black', bold=True, blankLine=True)

            for line in self.Lcp.document(
                    suppressLimbAngles=not (self.edgeOnDiskRadioButton.isChecked())):
                self.showMsg(line, color='black', bold=True, blankLine=False)
            self.showMsg('', blankLine=False)

            if onlyLcpValuesWanted:
                if self.firstContactRdgValue is not None:
                    self.showMsg(f"first contact rdg num: {self.firstContactRdgValue:0.3f}", color='black', bold=True, blankLine=False)

            # if self.diffractionRadioButton.isChecked():
            #     self.showMsg('Model used: diffraction', color='black', bold=True)
            # elif self.edgeOnDiskRadioButton.isChecked():
            #     self.showMsg('Model used: edge on disk', color='black', bold=True)
            # elif self.diskOnDiskRadioButton.isChecked():
            #     self.showMsg('Model used: disk on disk', color='black', bold=True)

            if self.modelDedgeSecs is not None and not onlyLcpValuesWanted:
                self.showMsg('', blankLine=False)
                self.printEdgeOrMissReport()

            if self.edgeOnDiskRadioButton.isChecked():
                self.showMsg(f'D limb angle: {self.Lcp.D_limb_angle_degrees:0.1f}',
                             color='black', bold=True, blankLine=False)
                self.showMsg(f'R limb angle: {self.Lcp.R_limb_angle_degrees:0.1f}',
                             color='black', bold=True, blankLine=False)

            if math.isclose(self.Lcp.chord_length_km, self.Lcp.asteroid_diameter_km) and not onlyLcpValuesWanted:
                self.showInfo(f'The chord length is at its upper limit compared to the asteroid size.\n\n'
                              f'If the fit is not very good, it may be because the fit routine tried a \n'
                              f'bigger chord, but was constrained by the asteroid diameter.\n\n'
                              f'Consider trying a fit with a different asteroid diameter.')

    def processDangleEditFinish(self):
        if self.DdegreesEdit.text() == '':
            return

        try:
            Dangle = float(self.DdegreesEdit.text())
            if 0 <= Dangle <= 90:
                if self.Lcp is not None:
                    self.Lcp.set("D_limb_angle_degrees", Dangle)
                else:
                    self.showInfo(f'D angle must be > 0 and <= 90')
                    self.DdegreesEdit.clear()
        except ValueError as e:
            self.showInfo(f'DdegreesEdit: {e}')
            self.DdegreesEdit.clear()

    def processRangleEditFinish(self):
        try:
            Rangle = float(self.RdegreesEdit.text())
            if 0 <= Rangle <= 90:
                if self.Lcp is not None:
                    self.Lcp.set("R_limb_angle_degrees", Rangle)
            else:
                self.showInfo(f'R angle must be > 0 and <= 90')
                self.RdegreesEdit.clear()
        except ValueError as e:
            self.showInfo(f'RdegreesEdit: {e}')
            self.RdegreesEdit.clear()

    def processEdgeTimePrecisionFinish(self):
        edgeTimeDelta = None
        try:
            edgeTimeDelta = float(self.edgeTimePrecisionEdit.text())
            if not edgeTimeDelta >= 0.0001:
                self.showInfo(f'The minimum precision for edgeTime is 0.0001')
                self.edgeTimePrecisionEdit.clear()
        except ValueError as e:
            self.showInfo(f'Error in Edge time precision.\n\n'
                          f'{e}')
            self.edgeTimePrecisionEdit.clear()

        if self.modelTimeOffset is not None:
            self.modelTimeOffset = self.modelTimeOffset - self.modelTimeOffset % edgeTimeDelta

    def processChordDurationPrecisionFinish(self):
        try:
            value = float(self.chordDurationPrecisionEdit.text())
            if value < 0:
                self.showInfo(f'Chord duration precision cannot be negative.')
                self.chordDurationPrecisionEdit.clear()
        except ValueError as e:
            self.showInfo(f'Error in Chord duration precision.\n\n'
                          f'{e}')
            self.chordDurationPrecisionEdit.clear()

    def processMissDistancePrecisionFinish(self):
        try:
            value = float(self.missDistancePrecisionEdit.text())
            if value < 0:
                self.showInfo(f'Miss distance precision cannot be negative.')
                self.missDistancePrecisionEdit.clear()
        except ValueError as e:
            self.showInfo(f'Error in Miss distance precision.\n\n'
                          f'{e}')
            self.missDistancePrecisionEdit.clear()

    def processLimbAnglePrecisionFinish(self):
        try:
            value = float(self.limbAnglePrecisionEdit.text())
            if value < 0:
                self.showInfo(f'Limb angle precision cannot be negative.')
                self.limbAnglePrecisionEdit.clear()
        except ValueError as e:
            self.showInfo(f'Error in Limb angle precision.\n\n'
                          f'{e}')
            self.limbAnglePrecisionEdit.clear()

    def processSqwaveMagDropEntry(self):
        try:
            magDrop = float(self.magDropSqwaveEdit.text())
            if not magDrop > 0:
                self.showInfo(f'square wave expected magDrops must be greater than zero\n\n'
                              f'Clearing the input.')
                self.magDropSqwaveEdit.clear()
        except ValueError as e:
            self.showInfo(f'Error in expected magDrop entry\n\n'
                          f'{e}\n\n'
                          f'... so clearing the input.')
            self.magDropSqwaveEdit.clear()

    def demoModel(self):
        if self.Lcp is None:
            self.showInfo(f'There is no model lightcurve defined.')
        else:
            if self.Lcp is None:
                self.showInfo(f'Model data has not been loaded.')
                return
            else:
                a_none_value_was_found, missing = self.Lcp.check_for_none()
                if a_none_value_was_found:
                    self.showInfo(f'{missing} has not been set.\n\nThere may be others.')
                    return
                self.computeModelLightcurve(demo=True)

    def demoGeometry(self):
        if self.Lcp is None:
            self.showInfo(f'There is no model lightcurve defined.')
        else:
            if self.Lcp is None:
                self.showInfo(f'Model data has not been loaded.')
                return
            else:
                a_none_value_was_found, missing = self.Lcp.check_for_none()
                if a_none_value_was_found:
                    self.showInfo(f'{missing} has not been set.\n\nThere may be others.')
                    return

            if self.diskOnDiskRadioButton.isChecked():
                fig = plt.figure(constrained_layout=False, figsize=(6, 6))
                fig.canvas.manager.set_window_title('Disk on disk geometry')
                axes = fig.add_subplot(1, 1, 1)
                illustrateDiskOnDiskEvent(LCP=self.Lcp, showLegend=True, showNotes=True, axes=axes)
                plt.show()
            elif self.edgeOnDiskRadioButton.isChecked():
                fig = plt.figure(constrained_layout=False, figsize=(6, 6))
                fig.canvas.manager.set_window_title('Disk on disk geometry')
                axes = fig.add_subplot(1, 1, 1)
                illustrateEdgeOnDiskEvent(LCP=self.Lcp, showLegend=True, showNotes=True, axes=axes)
                plt.show()
            elif self.diffractionRadioButton.isChecked():
                majorAxis = self.Lcp.asteroid_major_axis
                minorAxis = self.Lcp.asteroid_minor_axis
                thetaDegrees = self.Lcp.ellipse_angle_degrees
                upperChordWanted = self.Lcp.use_upper_chord
                plot_diffraction(x=None, y=None, first_wavelength=self.Lcp.wavelength_nm,
                                 last_wavelength=self.Lcp.wavelength_nm, LCP=self.Lcp, title='Diffraction geometry',
                                 showLegend=True, showNotes=True, plot_versus_time=True, zoom=1,
                                 majorAxis=majorAxis, minorAxis=minorAxis, thetaDegrees=thetaDegrees,
                                 upperChordWanted=upperChordWanted)
            else:
                self.showInfo(f'Invalid model button selection in demoGeometry()')

    def convertDandRrdgValueToTimestamp(self):
        tObsStart = convertTimeStringToTime(self.yTimes[0])
        tDedge = tObsStart + self.modelDedgeRdgValue * self.timeDelta
        tRedge = tObsStart + self.modelRedgeRdgValue * self.timeDelta
        return convertTimeToTimeString(tDedge), convertTimeToTimeString(tRedge)

    def pauseFitInProgress(self):
        self.pauseFitRequested = True
        self.keepRunning = False

    def processAsteroidDiameterKmFinish(self):
        try:
            if self.asteroidDiameterKmEdit.text() == "":
                return
            value = float(self.asteroidDiameterKmEdit.text())
            if value < 0:
                self.showInfo(f'Asteroid diameter cannot be negative.')
                self.asteroidDiameterKmEdit.clear()
                return
        except ValueError as e:
            self.showInfo(f'asteroidDiameterKmEdit: {e}')
            return

        if self.Lcp is not None:  # We are in edit mode after a full Lcp has been created
            self.Lcp.set("asteroid_diameter_km", None)
            self.Lcp.set("asteroid_diameter_mas", None)
            self.Lcp.set("asteroid_diameter_km", value)
            self.asteroidDiameterMasEdit.setText(f"{self.Lcp.asteroid_diameter_mas:0.4f}")
            return

        self.processModelLightcurveCoreEdit()  # No Lcp yet, so check for done with core entries

    def processAsteroidDiameterMasFinish(self):
        try:
            if self.asteroidDiameterMasEdit.text() == "":
                return
            value = float(self.asteroidDiameterMasEdit.text())
            if value < 0:
                self.showInfo(f'Asteroid diameter cannot be negative.')
                self.asteroidDiameterMasEdit.clear()
                return
        except ValueError as e:
            self.showInfo(f'asteroidDiameterMasEdit: {e}')
            return

        if self.Lcp is not None:  # We are in edit mode after a full Lcp has been created
            self.Lcp.set("asteroid_diameter_km", None)
            self.Lcp.set("asteroid_diameter_mas", None)
            self.Lcp.set("asteroid_diameter_mas", value)
            self.asteroidDiameterKmEdit.setText(f"{self.Lcp.asteroid_diameter_km:0.4f}")
            return

        self.processModelLightcurveCoreEdit()

    def processAsteroidSpeedShadowFinish(self):
        try:
            if self.asteroidSpeedShadowEdit.text() == "":
                return
            value = float(self.asteroidSpeedShadowEdit.text())
            if value < 0:
                self.showInfo(f'Asteroid speed cannot be negative.')
                self.asteroidSpeedShadowEdit.clear()
                return
        except ValueError as e:
            self.showInfo(f'asteroidSpeedShadowEdit: {e}')
            return

        if self.Lcp is not None:  # We are in edit mode after a full Lcp has been created
            # self.Lcp.set("shadow_speed", None)
            # self.Lcp.set("sky_motion_mas_per_sec", None)
            self.Lcp.set("shadow_speed", value)

            sky_motion_mas_per_sec = asteroid_mas_from_km(self.Lcp.shadow_speed, self.Lcp.asteroid_distance_AU)

            self.Lcp.set("sky_motion_mas_per_sec", sky_motion_mas_per_sec)

            self.asteroidSpeedSkyEdit.setText(f"{sky_motion_mas_per_sec:0.4f}")
            return

        self.processModelLightcurveCoreEdit()

    def processAsteroidSpeedSkyFinish(self):
        try:
            if self.asteroidSpeedSkyEdit.text() == "":
                return
            value = float(self.asteroidSpeedSkyEdit.text())
            if value < 0:
                self.showInfo(f'Asteroid speed cannot be negative.')
                self.asteroidSpeedSkyEdit.clear()
                return
        except ValueError as e:
            self.showInfo(f'asteroidSpeedSkyEdit: {e}')
            return

        if self.Lcp is not None:  # We are in edit mode after a full Lcp has been created
            # self.Lcp.set("shadow_speed", None)
            # self.Lcp.set("sky_motion_mas_per_sec", None)
            self.Lcp.set("sky_motion_mas_per_sec", value)

            shadow_speed = asteroid_km_from_mas(self.Lcp.sky_motion_mas_per_sec, self.Lcp.asteroid_distance_AU)

            self.asteroidSpeedSkyEdit.setText(f"{self.Lcp.sky_motion_mas_per_sec:0.4f}")
            self.Lcp.set("shadow_speed", shadow_speed)
            self.asteroidSpeedShadowEdit.setText(f"{shadow_speed:0.4f}")
            return

        self.processModelLightcurveCoreEdit()

    def processChordSizeSecondsFinish(self):
        if self.Lcp is not None:
            if self.Lcp.miss_distance_km > 0:
                if self.chordSizeSecondsEdit.text() == '0.0000':
                    return
                self.starSizeMasEdit.setFocus()
                self.chordSizeKmEdit.setText('0.00000')
                self.chordSizeSecondsEdit.setText('0.00000')
                self.showInfo(f'When Miss distance (km) is greater than 0,\n\n'
                              f'Chord estimates must remain at 0.00000')
        self.chordSizeSecondsEdited = True
        self.chordSizeKmEdited = False
        self.processModelParameterChange()

    def processChordSizeKmFinish(self):
        if self.Lcp is not None:
            if self.Lcp.miss_distance_km > 0:
                if self.chordSizeKmEdit.text() == '0.00000':
                    return
                self.starSizeMasEdit.setFocus()
                self.chordSizeKmEdit.setText('0.00000')
                self.chordSizeSecondsEdit.setText('0.00000')
                self.showInfo(f'When Miss distance (km) is greater than 0,\n\n'
                              f'Chord estimates must remain at 0.00000')
        self.chordSizeSecondsEdited = False
        self.chordSizeKmEdited = True
        self.processModelParameterChange()

    def processStarSizeMasFinish(self):
        self.starSizeMasEdited = True
        self.starSizeKmEdited = False
        self.processModelParameterChange()

    def processStarSizeKmFinish(self):
        self.starSizeMasEdited = False
        self.starSizeKmEdited = True
        self.processModelParameterChange()

    def processMagDropFinish(self):
        if self.Lcp is not None:
            if self.Lcp.baseline_ADU is None:
                self.showInfo(f'Baseline ADU needs to be set before a '
                              f'MagDrop value can be entered.')
                return
            else:
                if not self.magDropEdit.text() == '':
                    self.Lcp.set("magDrop", float(self.magDropEdit.text()))
                    self.processModelParameterChange()
                return

        self.magDropEdit.clear()
        self.showInfo(f'An "event" has not been defined yet so '
                      f'the MagDrop entry has been cleared.')

    def processMissDistanceFinish(self):
        try:
            if self.missDistanceKmEdit.text() == "":
                return
            value = float(self.missDistanceKmEdit.text())
            if value < 0:
                self.showInfo(f'Miss distance cannot be negative.')
                if self.Lcp is None:
                    self.missDistanceKmEdit.clear()
                else:
                    self.missDistanceKmEdit.setText(f"{self.Lcp.miss_distance_km:0.3f}")
                return
            if value > 0.0:
                self.chordSizeSecondsEdit.setText("0.0")
                self.chordSizeKmEdit.setText("0.0")
                if self.Lcp is not None:
                    self.Lcp.set("chord_length_km", None)
                    self.Lcp.set("chord_length_sec", None)
                    self.Lcp.set("chord_length_km", 0.0)
        except ValueError as e:
            self.showInfo(f'missDistanceKmEdit: {e}')
            return

        self.processModelLightcurveCoreEdit()

    def processAsteroidDistAUfinish(self):
        try:
            if self.asteroidDistAUedit.text() == "":
                return
            value = float(self.asteroidDistAUedit.text())
            if value < 0:
                self.showInfo(f'Asteroid distance cannot be negative.')
                self.asteroidDistAUedit.clear()
                return
        except ValueError as e:
            self.showInfo(f'asteroidDistAUedit: {e}')
            return

        self.processModelLightcurveCoreEdit()

    def processAsteroidDistArcsecFinish(self):
        try:
            if self.asteroidDistArcsecEdit.text() == "":
                return
            value = float(self.asteroidDistArcsecEdit.text())
            if value < 0:
                self.showInfo(f'Asteroid distance cannot be negative.')
                self.asteroidDistArcsecEdit.clear()
                return
        except ValueError as e:
            self.showInfo(f'asteroidDistArcsecEdit: {e}')
            return

        self.processModelLightcurveCoreEdit()

    def processWavelengthFinish(self):
        try:
            value = float(self.wavelengthEdit.text())
            if value < 0:
                self.showInfo(f'Wavelength (nm) cannot be negative.')
                self.wavelengthEdit.clear()
                return
        except ValueError as e:
            self.showInfo(f'wavelengthEdit: {e}')
            return

        self.processModelLightcurveCoreEdit()

    def sampleModelLightcurve(self):
        # Build interpolation function. This is done to deal with the finite resolution of
        # the 2048 point lightcurve
        interpolator = interpolate.interp1d(self.modelPtsXsec, self.modelY)

        tObsStart = convertTimeStringToTime(self.yTimes[0])
        sample_time = tObsStart

        x_vals = []
        y_vals = []

        while sample_time <= self.modelPtsXsec[-1]:
            if sample_time >= self.modelPtsXsec[0]:
                # Compute x_vals as reading number
                x_vals.append(round((sample_time - tObsStart) / self.Lcp.frame_time))
                y_vals.append(interpolator(sample_time))
            sample_time += self.Lcp.frame_time
            # This keeps us from from sampling the model curve past
            # the end of the observation
            if len(x_vals) >= self.dataLen:
                break

        if len(x_vals) == 0:
            breakpoint()

        self.modelXsamples = np.array(x_vals)
        self.modelYsamples = np.array(y_vals)

        del x_vals
        del y_vals
        gc.collect()

    # @jit(nopython=True)
    def calcModelFitMetric(self, showData=True, suppressPrint=True):
        self.extendAndDrawModelLightcurve()  # This fills self.modelPtsXsec
        self.sampleModelLightcurve()
        self.newRedrawMainPlot()

        # Calculate the time at the chord center
        center_time = (self.modelDedgeSecs + self.modelRedgeSecs) / 2.0
        tObsStart = convertTimeStringToTime(self.yTimes[0])
        center_frame = int((center_time - tObsStart) / self.Lcp.frame_time)

        modelSigmaB = self.Lcp.sigmaB
        matchingObsYvalues = []
        modelYsamples = []
        sigmaValues = []  # We will use these for the likelihood calculations
        i = self.modelXsamples[0]  # Units: frame number
        k = 0
        while i < len(self.yValues) and k < self.modelYsamples.size:
            try:
                # This 'try' can fail if an older LCP is in use that did not have these fields
                self.metricLimitLeft = self.Lcp.metricLimitLeft
                self.metricLimitRight = self.Lcp.metricLimitRight
            except Exception:  # noqa
                pass
            # This code added to limit region used in metric calculation
            if self.metricLimitLeft is not None:
                if self.metricLimitLeft <= i <= self.metricLimitRight:
                    matchingObsYvalues.append(self.yValues[i])
                    modelYsamples.append(self.modelYsamples[k])
                    sigmaValues.append(modelSigmaB)
            else:
                matchingObsYvalues.append(self.yValues[i])
                modelYsamples.append(self.modelYsamples[k])
                sigmaValues.append(modelSigmaB)

            i += 1
            k += 1

        if self.metricLimitLeft is not None and len(matchingObsYvalues) == 0:
            self.showInfo(f'Given the metric limits supplied, no metric points were found.\n\n'
                          f'Did you accidentally use frame numbers rather than reading numbers to specify your choices?')
            return 0.0

        # yVals = np.array(matchingObsYvalues)
        # mVals = np.array(modelYsamples)
        # sigVals = np.array(sigmaValues)
        # log_likelihood = cum_loglikelihood(yVals, mVals, sigVals, 0, len(sigmaValues)-1)
        # Figure out if this is appropriate
        # I'm setting k to 1 because only a single parameter is varied - all others are given
        # aiccValue = aicc(log_likelihood, n=len(sigmaValues), k=1)
        # Remove this print statement
        # self.showMsg(f'log_likelihood: {log_likelihood:0.4f}  aicc: {aiccValue:0.4f}')

        matchingObsYvalues = np.array(matchingObsYvalues)
        modelYsamples = np.array(modelYsamples)

        # We produce these values as a side effect - only used during edge-on-disk (penumbral) fits
        self.dMetric = np.sum(((modelYsamples[0:center_frame] - matchingObsYvalues[0:center_frame]) / modelSigmaB)**2) / modelYsamples.size
        self.rMetric = np.sum(((modelYsamples[center_frame:] - matchingObsYvalues[center_frame:]) / modelSigmaB)**2) / modelYsamples.size

        # totalMetric = np.sum(((modelYsamples - matchingObsYvalues) / modelSigmaB)**2) / modelYsamples.size
        totalMetric = self.dMetric + self.rMetric

        del matchingObsYvalues
        del modelYsamples
        gc.collect()

        ans = f'timeOffset: {self.modelTimeOffset:0.5f} chord: {self.Lcp.chord_length_sec:0.5f} '
        if self.modelMetric is None:
            ans += 'modelMetric: None...  '
            self.modelMetric = totalMetric
            ans += f'newMetric: {totalMetric:0.5f} change: None'
        else:
            ans += f'modelMetric: {self.modelMetric:0.5f} totalMetric: {totalMetric:0.5f} '
            # Note: self.modelMetric is the current (i.e, previous) metric
            if self.fitStatus.metricInUse == 'd':
                change = self.dMetric - self.modelMetric
            elif self.fitStatus.metricInUse == 'r':
                change = self.rMetric - self.modelMetric
            else:
                change = totalMetric - self.modelMetric
            if not suppressPrint:
                print(f'changing-> {self.fitStatus.beingChanged:8}  '
                      f'edgeTime: {self.modelTimeOffset:>10.5f}  '
                      f'chord: {self.Lcp.chord_length_sec:>10.5f}  '
                      f'dAngle: {self.Lcp.D_limb_angle_degrees:>10.5f}  '
                      f'rAngle: {self.Lcp.R_limb_angle_degrees:>10.5f}')
                print(f'edgeTime: {self.modelTimeOffset:>10.5f} '
                      f'totalMetric: {totalMetric:>10.5f}  '
                      f'dMetric: {self.dMetric:>10.5f}  '
                      f'rMetric: {self.rMetric:>10.5f}  '
                      f'modelMetric: {self.modelMetric:>10.5f}\n')
            if change < 0:
                ans += f'change: {change:<8.5f} '
            else:
                ans += f'change: +{change:<8.5f} '

            if change <= 0:
                ans += 'better'
                self.fitMetricChangeEdit.setStyleSheet("background: green")
            else:
                ans += 'worse'
                self.fitMetricChangeEdit.setStyleSheet("background: red")

            if self.fitStatus.metricInUse == 'd':
                self.modelMetric = self.dMetric
            elif self.fitStatus.metricInUse == 'r':
                self.modelMetric = self.rMetric
            else:
                self.modelMetric = totalMetric
            self.fitMetricChangeEdit.setText(f'{change:0.6f}')

        if self.fitStatus.metricInUse == 'd':
            self.fitMetricEdit.setText(f'{self.dMetric:0.6f}')
        elif self.fitStatus.metricInUse == 'r':
            self.fitMetricEdit.setText(f'{self.rMetric:0.6f}')
        else:
            self.fitMetricEdit.setText(f'{totalMetric:0.6f}')

        if showData:
            self.showMsg(f'{ans}', bold=True, blankLine=False)

        QtWidgets.QApplication.processEvents()  # to force display of the edit box changes
        return self.modelMetric

    def clearNe3SolutionPoints(self):
        try:
            self.exponentialDtheoryPts = None
            self.exponentialRtheoryPts = None
            self.newRedrawMainPlot()
        except Exception as e:
            self.showInfo(f'{e}')

    def modelMarkBaselineRegion(self):
        self.baselineADUedit.clear()
        self.clearBaselineADUselectionButton.setEnabled(True)
        self.markBaselineRegion()
        self.calcBaselineADUbutton.setStyleSheet("background-color: lightblue")

    def calcBaselineADU(self):
        if self.baselinePointsMarked():
            self.clearBaselineADUselectionButton.setStyleSheet("background-color: lightblue")

            self.calcBaselineStatisticsFromMarkedRegions()
            if self.sigmaB < 1.0:
                self.sigmaB = 1.0
            if self.B is not None:
                self.baselineADUedit.setText(f'{self.B:0.1f}')
                self.Lcp.set('baseline_ADU', self.B)
                self.Lcp.set('sigmaB', self.sigmaB)
                self.processModelParameterChange()
            self.clearModelsBaselineRegion()
        else:
            self.showInfo('Less than 3 baseline points are selected.')

    @staticmethod
    def calcBottomADU(baselineADU, magDrop):
        return 10.0**(np.log10(baselineADU) - magDrop / 2.5)

    def showModelChoiceAdvice(self):
        self.showInfo(decide_model_to_use(self.Lcp))

    def validateEllipseParameters(self):
        failure = (None, None, None)
        try:
            if not self.majorAxisEdit.text() == '':
                majorAxis = float(self.majorAxisEdit.text())
            else:
                majorAxis = None
        except ValueError as e:
            self.showInfo(f'error in majorAxis input: {e}')
            return failure

        try:
            if not self.minorAxisEdit.text() == '':
                minorAxis = float(self.minorAxisEdit.text())
            else:
                minorAxis = None
        except ValueError as e:
            self.showInfo(f'error in minorAxis input: {e}')
            return failure

        try:
            if not self.ellipseAngleEdit.text() == '':
                ellipseAngle = float(self.ellipseAngleEdit.text())
            else:
                ellipseAngle = None
        except ValueError as e:
            self.showInfo(f'error in ellipse angle input: {e}')
            return failure

        if minorAxis is not None or majorAxis is not None or ellipseAngle is not None:
            if minorAxis is None or majorAxis is None or ellipseAngle is None:
                self.showInfo(f'Incomplete specification of asteroid ellipse shape')
                return failure

        self.Lcp.use_upper_chord = self.useUpperEllipseChord.isChecked()
        self.Lcp.asteroid_minor_axis = minorAxis
        self.Lcp.asteroid_major_axis = majorAxis
        self.Lcp.ellipse_angle_degrees = ellipseAngle

        return majorAxis, minorAxis, ellipseAngle

    def plotDiffractionPatternOnGround(self):
        self.showDiffractionButton.setText('... computation in progress')
        QtWidgets.QApplication.processEvents()
        try:
            self.validateEllipseParameters()

            demo_diffraction_field(self.Lcp, title_adder=self.currentEventEdit.text())
        except ValueError as e:
            self.showInfo(f'plotDiffractionPatternOnGround(): {e}')
        self.showDiffractionButton.setText('Show diffraction pattern on the ground')

    def optimizePosition(self):

        # We set this because it is used in various printouts
        self.fitStatus.beingChanged = 'x position'
        self.beingOptimizedEdit.setText('x position')

        self.clearColoredParameters()
        self.beingOptimizedEdit.setStyleSheet("background-color: lightblue")

        eodAlgorithm = False

        # Get the starting value of the metric
        self.calcModelFitMetric(showData=False)

        edgeTimeMetric = self.dMetric if eodAlgorithm else self.modelMetric

        self.fitStatus.failureCount = 0

        while True:   # We loop until there is a need to execute a return

            if not self.keepRunning:
                self.clearColoredParameters()
                return "paused"

            self.makeAnEdgeTimeStep()  # Make a step

            self.calcModelFitMetric(showData=False)
            newMetric = self.dMetric if eodAlgorithm else self.modelMetric

            # Diagnose whether the step taken was an improvement or not
            if newMetric < edgeTimeMetric:
                edgeTimeMetric = newMetric
                self.bestFit.modelTimeOffset = self.fitStatus.modelTimeOffset
                self.fitStatus.failureCount = 0  # We have an improvement, so reset failure count
                continue
            else:
                self.fitStatus.failureCount += 1
                if self.fitStatus.failureCount >= 2:
                    self.fitStatus.failureCount = 0
                    # We return to best known values and change search direction
                    self.fitStatus.edgeDelta *= -1  # Change step direction
                    self.modelTimeOffset = self.bestFit.modelTimeOffset
                    self.fitStatus.modelTimeOffset = self.bestFit.modelTimeOffset
                    if self.fitStatus.edgeDelta > 0:
                        # We do one more call to calcModelFitMetric() for the side-effect
                        # of updating the self.modelDedgeSecs and self.modelRedgeSecs variables
                        self.calcModelFitMetric(showData=False)
                        self.bestFit.thisPassMetric = edgeTimeMetric
                        self.clearColoredParameters()
                        return 'success'
                    continue
                continue

    def optimizeDangle(self):
        # Allow a zero (or negative) angle delta to terminate improvement cycle
        if float(self.limbAnglePrecisionEdit.text()) <= 0:
            self.clearColoredParameters()
            return "skipped"

        self.clearColoredParameters()
        self.DdegreesEdit.setStyleSheet("background-color: lightblue")

        self.fitStatus.metricInUse = 'd'
        bestMetricSoFar = self.dMetric
        stepPositiveAtEntry = self.fitStatus.DangleDelta >= 0
        self.beingOptimizedEdit.setText('D angle')
        failureCount = 0

        while True:
            if not self.keepRunning:
                self.clearColoredParameters()
                return "paused"
            # Make an angle change, recompute the model lightcurve and get the new metric
            stepWasTaken = self.makeDangleStep()
            if stepWasTaken:
                self.computeModelLightcurve()
                self.calcModelFitMetric(showData=False)
            if (self.dMetric < bestMetricSoFar) and stepWasTaken:
                failureCount = 0
                bestMetricSoFar = self.dMetric
                self.bestFit.Dangle = self.fitStatus.Dangle
                continue
            else:
                failureCount += 1
                if not stepWasTaken:
                    failureCount = 99
                if failureCount >= 2:
                    # We return to best known values and change search direction
                    self.fitStatus.DangleDelta *= -1  # Change step direction
                    self.returnToBestDangleFit()
                    stepPositiveNow = self.fitStatus.DangleDelta >= 0
                    if stepPositiveNow == stepPositiveAtEntry:  # We have tried both directions
                        self.clearColoredParameters()
                        return "success"
            continue

    def optimizeRangle(self):
        # Allow a zero (or negative) angle delta to terminate improvement cycle
        if float(self.limbAnglePrecisionEdit.text()) <= 0:
            self.clearColoredParameters()
            return "skipped"

        self.clearColoredParameters()
        self.RdegreesEdit.setStyleSheet("background-color: lightblue")

        self.fitStatus.metricInUse = 'r'
        bestMetricSoFar = self.rMetric
        failureCount = 0
        stepPositiveAtEntry = self.fitStatus.RangleDelta >= 0
        self.beingOptimizedEdit.setText('R angle')

        while True:
            if not self.keepRunning:
                self.clearColoredParameters()
                return "paused"
            # Make an angle change, recompute the model lightcurve and get the new metric
            stepWasTaken = self.makeRangleStep()
            if stepWasTaken:
                self.computeModelLightcurve()
                self.calcModelFitMetric(showData=False)

            if (self.rMetric < bestMetricSoFar) and stepWasTaken:
                self.failureCount = 0
                bestMetricSoFar = self.rMetric
                self.bestFit.Rangle = self.fitStatus.Rangle
                continue
            else:
                failureCount += 1
                if not stepWasTaken:
                    failureCount = 99
                if failureCount >= 2:
                    # We return to best known values and change search direction
                    self.fitStatus.RangleDelta *= -1  # Change step direction
                    self.returnToBestRangleFit()
                    stepPositiveNow = self.fitStatus.RangleDelta >= 0
                    if stepPositiveNow == stepPositiveAtEntry:  # We have tried both directions
                        self.clearColoredParameters()
                        return "success"
            continue

    def optimizeChord(self):
        # Allow a zero (or negative) chord delta to terminate improvement cycle
        if float(self.chordDurationPrecisionEdit.text()) <= 0:
            return "skipped"

        self.clearColoredParameters()
        self.chordSizeSecondsEdit.setStyleSheet("background-color: lightblue")

        self.fitStatus.metricInUse = 'both'
        bestMetricSoFar = self.dMetric + self.rMetric
        failureCount = 0
        stepPositiveAtEntry = self.fitStatus.chordDelta >= 0
        self.beingOptimizedEdit.setText('Chord secs')

        while True:
            if not self.keepRunning:
                self.clearColoredParameters()
                return "paused"
            # Make a chord change, recompute the model lightcurve and get the new metric
            stepWasTaken = self.makeAchordTimeStep()
            if stepWasTaken:
                self.computeModelLightcurve()
                self.calcModelFitMetric(showData=False)
            if (self.modelMetric < bestMetricSoFar) and stepWasTaken:
                failureCount = 0
                bestMetricSoFar = self.modelMetric
                self.bestFit.chordTime = self.fitStatus.chordTime
            else:
                failureCount += 1
                if not stepWasTaken:
                    failureCount = 99
                if failureCount >= 2:
                    failureCount = 0
                    # We return to best known values and change search direction
                    self.fitStatus.chordDelta *= -1  # Change step direction
                    self.returnToBestChordFit()
                    stepPositiveNow = self.fitStatus.chordDelta >= 0
                    if stepPositiveNow == stepPositiveAtEntry:  # We have tried both directions
                        self.clearColoredParameters()
                        return "success"
            continue

    def optimizeMissDistance(self):
        # Allow a zero (or negative) miss distance delta to terminate improvement cycle
        if float(self.missDistancePrecisionEdit.text()) <= 0:
            self.clearColoredParameters()
            return "skipped"

        self.clearColoredParameters()
        self.missDistanceKmEdit.setStyleSheet("background-color: lightblue")

        self.fitStatus.metricInUse = 'both'
        self.modelMetric = self.dMetric + self.rMetric
        stepPositiveAtEntry = self.fitStatus.missDelta >= 0
        self.beingOptimizedEdit.setText('Miss Km')

        while True:
            if not self.keepRunning:
                self.clearColoredParameters()
                return "paused"
            # Make a miss distance change, recompute the model lightcurve and get the new metric
            stepWasTaken = self.makeAmissDistanceStep()
            if stepWasTaken:
                self.computeModelLightcurve()
                self.calcModelFitMetric(showData=False)
            if (self.modelMetric < self.bestFit.thisPassMetric) and stepWasTaken:
                self.fitStatus.failureCount = 0
                self.bestFit.thisPassMetric = self.modelMetric
                self.bestFit.missDistance = self.fitStatus.missDistance
            else:
                self.fitStatus.failureCount += 1
                if not stepWasTaken:
                    self.fitStatus.failureCount = 99
                if self.fitStatus.failureCount >= 2:
                    self.fitStatus.failureCount = 0
                    # We return to best known values and change search direction
                    self.fitStatus.missDelta *= -1  # Change step direction
                    self.returnToBestMissDistanceFit()
                    stepPositiveNow = self.fitStatus.missDelta >= 0
                    if stepPositiveNow == stepPositiveAtEntry:  # We have tried both directions
                        self.clearColoredParameters()
                        return "success"
            continue

    # @profile
    def fitImprovementControlCenter(self):
        # self.showMemoryStats()
        self.keepRunning = True

        if self.edgeOnDiskRadioButton.isChecked():
            if self.limbAnglePrecisionEdit.text() == '':
                self.showInfo(f'You need to specify a limb angle precision step.')
                return

        if self.fitStatus is None:
            self.showMsg(f'Fit improvement entered for first time.',
                         color='black', bold=True)

            # Create and initialize fitStatus
            self.fitStatus = FitStatus()
            self.fitStatus.currentMetric = None  # A signal to calcModelFitMetric that we're starting.
            self.fitStatus.currentMetric = self.calcModelFitMetric(showData=False)
            self.fitStatus.modelTimeOffset = self.modelTimeOffset
            self.fitStatus.chordTime = self.Lcp.chord_length_sec
            self.fitStatus.missDistance = self.Lcp.miss_distance_km
            self.fitStatus.Dangle = self.Lcp.D_limb_angle_degrees
            self.fitStatus.Rangle = self.Lcp.R_limb_angle_degrees
            self.fitStatus.edgeDelta = float(self.edgeTimePrecisionEdit.text())
            self.fitStatus.chordDelta = float(self.chordDurationPrecisionEdit.text())
            self.fitStatus.missDelta = float(self.missDistancePrecisionEdit.text())

            if not self.limbAnglePrecisionEdit.text() == '':
                self.fitStatus.DangleDelta = float(self.limbAnglePrecisionEdit.text())
                self.fitStatus.RangleDelta = float(self.limbAnglePrecisionEdit.text())
            else:
                self.fitStatus.DangleDelta = 0
                self.fitStatus.RangleDelta = 0

            self.bestFit = BestFit()
            self.bestFit.thisPassMetric = self.fitStatus.currentMetric
            self.bestFit.metricAtStartOfPass = self.bestFit.thisPassMetric
            self.bestFit.modelTimeOffset = self.fitStatus.modelTimeOffset
            self.bestFit.chordTime = self.fitStatus.chordTime
            self.bestFit.missDistance = self.fitStatus.missDistance
            self.bestFit.Dangle = self.fitStatus.Dangle
            self.bestFit.Rangle = self.fitStatus.Rangle

            self.fitLightcurveButton.setStyleSheet("background-color: red")
            self.fitLightcurveButton.setText('... best fit search in progress')

        else:
            self.showMsg(f'Fit improvement is being re-entered',
                         color='black', bold=True)
            self.fitStatus.currentMetric = self.calcModelFitMetric(showData=False)
            self.bestFit.metricAtStartOfPass = self.fitStatus.currentMetric
            self.bestFit.thisPassMetric = self.fitStatus.currentMetric
            self.fitStatus.improvementPassCompleted = False
            self.fitStatus.failureCount = 0
            self.fitLightcurveButton.setStyleSheet("background-color: red")
            self.fitLightcurveButton.setText('... best fit search in progress')

        QtWidgets.QApplication.processEvents()  # Updates main plot display
        gc.collect()

        if self.edgeOnDiskRadioButton.isChecked():
            self.keepRunning = True
            self.calcModelFitMetric(showData=False)
            bestMetricSoFar = self.dMetric + self.rMetric
            while self.keepRunning:

                self.optimizePosition()
                if not self.keepRunning:
                    continue
                QtWidgets.QApplication.processEvents()

                self.optimizeDangle()
                if not self.keepRunning:
                    continue
                QtWidgets.QApplication.processEvents()

                self.optimizePosition()
                if not self.keepRunning:
                    continue
                QtWidgets.QApplication.processEvents()

                self.optimizeRangle()
                if not self.keepRunning:
                    continue
                QtWidgets.QApplication.processEvents()

                self.optimizePosition()
                if not self.keepRunning:
                    continue
                QtWidgets.QApplication.processEvents()

                if self.Lcp.miss_distance_km > 0:
                    self.optimizeMissDistance()
                    if not self.keepRunning:
                        continue
                else:
                    self.optimizeChord()
                    if not self.keepRunning:
                        continue

                QtWidgets.QApplication.processEvents()

                if self.dMetric + self.rMetric >= bestMetricSoFar:
                    self.showFinalEdgePositionReport(paused=False)
                    self.fitLightcurveButton.setStyleSheet("background-color: yellow")
                    self.fitLightcurveButton.setText("Fit model to observation points")
                    self.beingOptimizedEdit.clear()
                    self.processFitImprovementCompleted()
                    return
                else:
                    bestMetricSoFar = self.dMetric + self.rMetric
                    continue

        if self.diskOnDiskRadioButton.isChecked() or self.diffractionRadioButton.isChecked():
            self.keepRunning = True
            self.calcModelFitMetric(showData=False)
            bestMetricSoFar = self.dMetric + self.rMetric
            while self.keepRunning:

                self.optimizePosition()
                if not self.keepRunning:
                    continue
                QtWidgets.QApplication.processEvents()

                if self.Lcp.miss_distance_km == 0:
                    self.optimizeChord()
                    if not self.keepRunning:
                        continue
                else:
                    self.optimizeMissDistance()
                    if not self.keepRunning:
                        continue
                QtWidgets.QApplication.processEvents()

                if self.dMetric + self.rMetric >= bestMetricSoFar:
                    self.showFinalEdgePositionReport(paused=False)
                    self.fitLightcurveButton.setStyleSheet("background-color: yellow")
                    self.fitLightcurveButton.setText("Fit model to observation points")
                    self.beingOptimizedEdit.clear()
                    self.processFitImprovementCompleted()
                    return
                else:
                    bestMetricSoFar = self.dMetric + self.rMetric
                    continue

        if self.keepRunning:
            # This code can only be reached if there is a code error
            self.showInfo(f'Programming error detected in fitImprovementControlCenter()')
        else:
            # We got here by a pause request
            self.fitLightcurveButton.setStyleSheet("background-color: yellow")
            self.fitLightcurveButton.setText("Fit model to observation points")
            self.pauseFitButton.setEnabled(False)
            self.beingOptimizedEdit.clear()
            self.processFitPauseRequest()
            self.showFinalEdgePositionReport(paused=True)

    def processFitImprovementCompleted(self):
        self.showMsg(f'Maximum improvement has been achieved with metric: '
                     f'{self.bestFit.thisPassMetric:0.5f}  ',
                     color='black', bold=True)
        if self.metricLimitLeft is not None:
            self.showMsg(f'The metric above was calculated using only readings {self.metricLimitLeft}'
                         f' to {self.metricLimitRight} inclusive.', color='black', bold=True)

        self.beingOptimizedEdit.clear()
        self.printFinalReport()
        self.fitLightcurveButton.setStyleSheet("background-color: yellow")
        self.fitLightcurveButton.setText("Fit model to observation points")
        self.fitMetricEdit.setText(f'{self.bestFit.thisPassMetric:0.5f}')
        self.fitMetricChangeEdit.clear()
        self.fitMetricChangeEdit.setStyleSheet(None)

    def processFitPauseRequest(self):
        self.fitStatus.fitComplete = True
        self.fitStatus.improvementPassCompleted = True
        self.pauseFitRequested = False
        self.fitLightcurveButton.setStyleSheet("background-color: yellow")
        self.fitLightcurveButton.setText("Fit model to observation points")
        self.showMsg(f'Fit of model to observation paused by request.',
                     color='black', bold=True)
        self.keepRunning = False

    def showFinalEdgePositionReport(self, paused=False):
        if paused:
            self.showMsg(f'Current state fit (during pause) ...',
                         color='black', bold=True)
        else:
            self.showMsg(f'Optimized fit found and is being displayed.',
                         color='black', bold=True)
        self.fitMetricEdit.setText(f'{self.bestFit.thisPassMetric:0.5f}')
        self.printEdgeOrMissReport()

    def printEdgeOrMissReport(self):
        if self.Lcp.miss_distance_km == 0:
            DrdgNum = int(self.modelDedgeRdgValue)
            RrdgNum = int(self.modelRedgeRdgValue)

            DfractionalRdgNum = self.modelDedgeRdgValue - DrdgNum
            RfractionalRdgNum = self.modelRedgeRdgValue - RrdgNum

            Dtime = convertTimeStringToTime(self.yTimes[DrdgNum])
            Rtime = convertTimeStringToTime(self.yTimes[RrdgNum])

            Dtime += DfractionalRdgNum * self.Lcp.frame_time
            Rtime += RfractionalRdgNum * self.Lcp.frame_time

            Dtimestamp = convertTimeToTimeString(Dtime)
            Rtimestamp = convertTimeToTimeString(Rtime)

            self.showMsg(f'D: {Dtimestamp}', color='black', bold=True)
            self.showMsg(f'R: {Rtimestamp}', color='black', bold=True)

        if self.firstContactRdgValue is not None:
            fcRdgNum = int(self.firstContactRdgValue)
            lcRdgNum = int(self.lastContactRdgValue)

            fcFractionalRdgNum = self.firstContactRdgValue - fcRdgNum
            lcFractionalRdgNum = self.lastContactRdgValue - lcRdgNum

            fcTime = convertTimeStringToTime(self.yTimes[fcRdgNum])
            lcTime = convertTimeStringToTime(self.yTimes[lcRdgNum])

            fcTime += fcFractionalRdgNum * self.Lcp.frame_time
            lcTime += lcFractionalRdgNum * self.Lcp.frame_time

            fcTimestamp = convertTimeToTimeString(fcTime)
            lcTimestamp = convertTimeToTimeString(lcTime)

            self.showMsg(f'first contact: {fcTimestamp}', color='black', bold=True)
            self.showMsg(f'last  contact: {lcTimestamp}', color='black', bold=True)

        else:
            self.showMsg(f'Miss distance: {self.Lcp.miss_distance_km:0.5f} km', color='black', bold=True)

    def reportMetrics(self, leader='????'):

        self.showMsg(f'{leader}: '
                     f'bestMetricThisPass: {self.bestFit.thisPassMetric:0.5f} '
                     f'---new metric: {self.modelMetric:0.5f}  '
                     f'dMetric: {self.dMetric:0.5f}  '
                     f'rMetric: {self.rMetric:0.5f}'
                     f'---failure count: {self.fitStatus.failureCount}')

    def returnToBestChordFit(self):
        bestChordTime = self.bestFit.chordTime
        self.fitStatus.chordTime = self.bestFit.chordTime
        try:
            self.Lcp.set('chord_length_km', None)
            self.Lcp.set('chord_length_sec', bestChordTime)
        except ValueError as e:
            self.showInfo(f'{e}')
            self.Lcp.set('chord_length_sec', None)
            self.Lcp.set('chord_length_km', self.Lcp.asteroid_diameter_km)
        self.chordSizeKmEdit.setText(f'{self.Lcp.chord_length_km:0.5f}')
        self.chordSizeSecondsEdit.setText(f'{self.Lcp.chord_length_sec:0.5f}')

        self.computeModelLightcurve()
        self.extendAndDrawModelLightcurve()
        self.sampleModelLightcurve()
        self.newRedrawMainPlot()

    def returnToBestMissDistanceFit(self):
        bestMissDistance = self.bestFit.missDistance
        self.fitStatus.missDistance = self.bestFit.missDistance
        self.Lcp.set('miss_distance_km', bestMissDistance)
        self.missDistanceKmEdit.setText(f'{self.Lcp.miss_distance_km:0.5f}')

        self.computeModelLightcurve()
        self.extendAndDrawModelLightcurve()
        self.sampleModelLightcurve()
        self.newRedrawMainPlot()

    def returnToBestDangleFit(self):
        bestDangle = self.bestFit.Dangle
        self.fitStatus.Dangle = self.bestFit.Dangle
        self.Lcp.set('D_limb_angle_degrees', None)
        self.Lcp.set('D_limb_angle_degrees', bestDangle)
        self.DdegreesEdit.setText(f'{self.Lcp.D_limb_angle_degrees:0.1f}')

        self.computeModelLightcurve()
        self.extendAndDrawModelLightcurve()
        self.sampleModelLightcurve()
        self.newRedrawMainPlot()

    def returnToBestRangleFit(self):
        bestRangle = self.bestFit.Rangle
        self.fitStatus.Rangle = self.bestFit.Rangle
        self.Lcp.set('R_limb_angle_degrees', None)
        self.Lcp.set('R_limb_angle_degrees', bestRangle)
        self.RdegreesEdit.setText(f'{self.Lcp.R_limb_angle_degrees:0.1f}')

        self.computeModelLightcurve()
        self.extendAndDrawModelLightcurve()
        self.sampleModelLightcurve()
        self.newRedrawMainPlot()

    def makeAchordTimeStep(self):
        if self.fitStatus.chordTime == abs(self.fitStatus.chordDelta) \
                and self.fitStatus.chordDelta < 0:
            return False  # because we are already at the smallest acceptable value

        current_chord_size = self.fitStatus.chordTime * self.Lcp.shadow_speed
        if (math.isclose(current_chord_size, self.Lcp.asteroid_diameter_km) and
                self.fitStatus.chordDelta > 0):
            return False  # because we already at the largest acceptable value

        # It's safe to make the step, but we will still need to test the resulting
        # chordTime to keep it it within acceptable bounds.
        self.fitStatus.chordTime += self.fitStatus.chordDelta

        # Check that chordTime may have become too large (or too small)
        chord_size = self.fitStatus.chordTime * self.Lcp.shadow_speed
        if chord_size > self.Lcp.asteroid_diameter_km:
            self.Lcp.set('chord_length_sec', None)
            self.Lcp.set('chord_length_km', self.Lcp.asteroid_diameter_km)
            self.fitStatus.chordTime = self.Lcp.chord_length_sec
        else:
            if self.fitStatus.chordTime < abs(self.fitStatus.chordDelta):
                self.fitStatus.chordTime = abs(self.fitStatus.chordDelta)

        self.Lcp.set('chord_length_km', None)
        self.Lcp.set('chord_length_sec', self.fitStatus.chordTime)
        self.chordSizeKmEdit.setText(f'{self.Lcp.chord_length_km:0.5f}')
        self.chordSizeSecondsEdit.setText(f'{self.Lcp.chord_length_sec:0.5f}')
        self.fitStatus.chordTime = self.Lcp.chord_length_sec
        return True

    def makeAmissDistanceStep(self):
        minMissDistance = abs(self.fitStatus.missDelta)
        if self.fitStatus.missDistance == minMissDistance \
                and self.fitStatus.missDelta < 0:
            return False  # because we are already at the smallest acceptable value

        # It's safe to make the step, but we will still need to test the resulting
        # missDistance to keep it it within acceptable bounds.
        self.fitStatus.missDistance += self.fitStatus.missDelta

        # Check that chordTime may have become too large (or too small)
        if self.fitStatus.missDistance < minMissDistance:
            self.fitStatus.missDistance = minMissDistance
        self.Lcp.set('miss_distance_km', self.fitStatus.missDistance)
        self.missDistanceKmEdit.setText(f'{self.Lcp.miss_distance_km:0.5f}')
        return True

    def makeDangleStep(self):
        if self.fitStatus.Dangle == abs(self.fitStatus.DangleDelta) \
                and self.fitStatus.DangleDelta < 0:
            return False  # because we are already at the smallest acceptable value

        current_angle_size = self.fitStatus.Dangle
        if (math.isclose(current_angle_size, 90) and
                self.fitStatus.DangleDelta > 0):
            return False  # because we already at the largest acceptable value

        # It's safe to make the step, but we will still need to test the resulting
        # Dangle after a step is taken to keep it within acceptable bounds.
        self.fitStatus.Dangle += self.fitStatus.DangleDelta

        # Check that Dangle may have become too large (or too small)
        if self.fitStatus.Dangle > 90:
            self.Lcp.set('D_limb_angle_degrees', None)
            self.Lcp.set('D_limb_angle_degrees', 90)
            self.fitStatus.Dangle = self.Lcp.D_limb_angle_degrees
        elif self.fitStatus.Dangle < abs(self.fitStatus.DangleDelta):
            self.fitStatus.Dangle = abs(self.fitStatus.DangleDelta)

        # Update Lcp and relevant edit box text
        self.Lcp.set('D_limb_angle_degrees', None)
        self.Lcp.set('D_limb_angle_degrees', self.fitStatus.Dangle)
        self.DdegreesEdit.setText(f'{self.Lcp.D_limb_angle_degrees:0.1f}')
        return True

    def makeRangleStep(self):
        if self.fitStatus.Rangle == abs(self.fitStatus.RangleDelta) \
                and self.fitStatus.RangleDelta < 0:
            return False  # because we are already at the smallest acceptable value

        current_angle_size = self.fitStatus.Rangle
        if (math.isclose(current_angle_size, 90) and
                self.fitStatus.RangleDelta > 0):
            return False  # because we already at the largest acceptable value

        # It's safe to make the step, but we will still need to test the resulting
        # Dangle after a step is taken to keep it within acceptable bounds.
        self.fitStatus.Rangle += self.fitStatus.RangleDelta

        # Check that Rangle may have become too large (or too small)
        if self.fitStatus.Rangle > 90:
            self.fitStatus.Rangle = 90
            self.Lcp.set('R_limb_angle_degrees', None)
            self.Lcp.set('R_limb_angle_degrees', self.fitStatus.Rangle)
        elif self.fitStatus.Rangle < abs(self.fitStatus.RangleDelta):
            self.fitStatus.Rangle = abs(self.fitStatus.RangleDelta)
            self.Lcp.set('R_limb_angle_degrees', None)
            self.Lcp.set('R_limb_angle_degrees', self.fitStatus.Rangle)
        else:
            self.Lcp.set('R_limb_angle_degrees', None)
            self.Lcp.set('R_limb_angle_degrees', self.fitStatus.Rangle)

        # Update relevant edit box text
        self.RdegreesEdit.setText(f'{self.Lcp.R_limb_angle_degrees:0.1f}')
        return True

    def clearColoredParameters(self):
        self.missDistanceKmEdit.setStyleSheet(None)
        self.chordSizeSecondsEdit.setStyleSheet(None)
        self.DdegreesEdit.setStyleSheet(None)
        self.RdegreesEdit.setStyleSheet(None)
        self.beingOptimizedEdit.setStyleSheet(None)

    def makeAnEdgeTimeStep(self):
        self.fitStatus.modelTimeOffset += self.fitStatus.edgeDelta
        self.modelTimeOffset = self.fitStatus.modelTimeOffset

    def fitModelLightcurveButtonClicked(self):
        self.redrawMainPlot()
        QtWidgets.QApplication.processEvents()

        if self.Lcp is None:
            self.showInfo(f'Model data has not been loaded.')
            return
        else:
            a_none_value_was_found, missing = self.Lcp.check_for_none()
            if a_none_value_was_found:
                self.showInfo(f'{missing} has not been set.\n\nThere may be others.')

        self.allowShowDetails = True
        if self.fitStatus is not None:
            self.fitStatus.fitComplete = False
            self.fitStatus.improvementPassCompleted = False
            if len(self.selectedPoints) == 1:
                self.computeInitialModelTimeOffset()
            self.restart()  # To clear possible results from a previous square-wave solution

            # We do this because the user may have changed parameters after a pause.
            # Such changes are placed in self.Lcp but not in self.fitStatus
            self.fitStatus.modelTimeOffset = self.modelTimeOffset
            self.fitStatus.chordTime = self.Lcp.chord_length_sec
            self.fitStatus.Dangle = self.Lcp.D_limb_angle_degrees
            self.fitStatus.Rangle = self.Lcp.R_limb_angle_degrees

            # To be consistent with the above substitutions
            self.bestFit.modelTimeOffset = self.fitStatus.modelTimeOffset
            self.bestFit.chordTime = self.fitStatus.chordTime
            self.bestFit.Dangle = self.fitStatus.Dangle
            self.bestFit.Rangle = self.fitStatus.Rangle

            # Update the precision deltas in case the user changed them during a pause
            self.fitStatus.edgeDelta = float(self.edgeTimePrecisionEdit.text())
            self.fitStatus.chordDelta = float(self.chordDurationPrecisionEdit.text())
            if not self.limbAnglePrecisionEdit.text() == '':
                self.fitStatus.DangleDelta = float(self.limbAnglePrecisionEdit.text())
                self.fitStatus.RangleDelta = float(self.limbAnglePrecisionEdit.text())

        self.computeModelLightcurve(computeOnly=False)

    def computeModelLightcurve(self, computeOnly=True, demo=False):
        if self.squareWaveRadioButton.isChecked():
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Question)
            msg.setText(f'The square wave model is selected.\n\n'
                        f'That model is handled in the SqWave model tab. '
                        f'Do you want to switch to that tab?')
            msg.setWindowTitle('Get latest version of PyOTE query')
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            retval = msg.exec_()
            if retval == QMessageBox.Yes:
                self.switchToTabNamed('SqWave model')
            return

        if demo and not len(self.selectedPoints) == 1:
            self.showInfo(f'Select a single point as close as possible to the\n'
                          f'center of the event to guide the\n'
                          f'initial placement of the model lightcurve')
            return

        if not len(self.selectedPoints) == 1 and self.modelTimeOffset is None:
            self.showInfo(f'Select a single point as close as possible to the\n'
                          f'center of the event to guide the\n'
                          f'initial placement of the model lightcurve')
            return

        showLegend = self.showLegendsCheckBox.isChecked()
        showNotes = self.showAnnotationsCheckBox.isChecked()
        versusTime = self.versusTimeCheckBox.isChecked()
        plots_wanted = self.showDetailsCheckBox.isChecked() and demo

        if self.diffractionRadioButton.isChecked():
            self.fitLightcurveButton.setStyleSheet("background-color: lightblue")
            self.fitLightcurveButton.setText('... calculating model lightcurve')
            QtWidgets.QApplication.processEvents()

            if demo:
                self.modelYsamples = None
                self.newRedrawMainPlot()
                QtWidgets.QApplication.processEvents()

            # majorAxis, minorAxis, thetaDegrees = self.validateEllipseParameters()
            try:
                self.modelXkm, self.modelY, self.modelDedgeKm, self.modelRedgeKm, \
                    self.firstContactKm, self.lastContactKm = \
                    demo_event(LCP=self.Lcp, model='diffraction', showLegend=showLegend,
                               title=self.currentEventEdit.text(),
                               showNotes=showNotes, plot_versus_time=versusTime,
                               plots_wanted=plots_wanted)
            except ValueError as e:
                self.showInfo(f'{e}')
                return

            self.pauseFitButton.setEnabled(True)

            self.computeInitialModelTimeOffset()
            if not demo:
                self.removePointSelections()

            self.extendAndDrawModelLightcurve()
            self.sampleModelLightcurve()

            QtWidgets.QApplication.processEvents()
            self.newRedrawMainPlot()

            self.setFitModelButtonColorAndText()

            # We need this test because the lightcurve needs to be repeatedly
            # recalculated during the operation of self.fitImprovementControlCenter()
            if not computeOnly:
                self.fitImprovementControlCenter()

            if demo:
                self.fitLightcurveButton.setStyleSheet("background-color: yellow")
                self.fitLightcurveButton.setText("Fit model to observation points")

            QtWidgets.QApplication.processEvents()

            return

        if self.edgeOnDiskRadioButton.isChecked():
            if self.Lcp.star_diameter_mas == 0:
                self.showInfo(f'An edge on disk model cannot be used when the star diameter is zero.')
                return

            self.fitLightcurveButton.setStyleSheet("background-color: lightblue")
            self.fitLightcurveButton.setText('... calculating model lightcurve')
            QtWidgets.QApplication.processEvents()

            if demo:
                self.modelYsamples = None
                self.newRedrawMainPlot()
                QtWidgets.QApplication.processEvents()

            self.modelXkm, self.modelY, self.modelDedgeKm, self.modelRedgeKm, \
                self.firstContactKm, self.lastContactKm = \
                demo_event(LCP=self.Lcp, model='edge-on-disk',
                           title=self.currentEventEdit.text(),
                           showLegend=showLegend, showNotes=showNotes,
                           plot_versus_time=versusTime,
                           plots_wanted=plots_wanted)

            self.pauseFitButton.setEnabled(True)

            self.computeInitialModelTimeOffset()
            if not demo:
                self.removePointSelections()
            QtWidgets.QApplication.processEvents()

            self.extendAndDrawModelLightcurve()
            self.sampleModelLightcurve()
            self.newRedrawMainPlot()

            self.setFitModelButtonColorAndText()

            # We need this test because the lightcurve needs to be repeatedly
            # recalculated during the operation of self.fitImprovementControlCenter()
            if not computeOnly:
                self.fitImprovementControlCenter()

            if demo:
                self.fitLightcurveButton.setStyleSheet("background-color: yellow")
                self.fitLightcurveButton.setText("Fit model to observation points")

            QtWidgets.QApplication.processEvents()

            return

        if self.diskOnDiskRadioButton.isChecked():
            if self.Lcp.star_diameter_mas == 0:
                self.showInfo(f'A disk on disk model cannot be used when the star diameter is zero.')
                return

            self.fitLightcurveButton.setStyleSheet("background-color: lightblue")
            self.fitLightcurveButton.setText('... calculating model lightcurve')
            QtWidgets.QApplication.processEvents()

            if demo:
                self.modelYsamples = None
                self.newRedrawMainPlot()
                QtWidgets.QApplication.processEvents()

            self.modelXkm, self.modelY, self.modelDedgeKm, self.modelRedgeKm,\
                self.firstContactKm, self.lastContactKm = \
                demo_event(LCP=self.Lcp, model='disk-on-disk',
                           title=self.currentEventEdit.text(),
                           showLegend=showLegend, showNotes=showNotes,
                           plot_versus_time=versusTime,
                           plots_wanted=plots_wanted)

            self.pauseFitButton.setEnabled(True)

            self.computeInitialModelTimeOffset()
            if not demo:
                self.removePointSelections()
            QtWidgets.QApplication.processEvents()

            self.extendAndDrawModelLightcurve()
            self.sampleModelLightcurve()
            self.newRedrawMainPlot()

            self.setFitModelButtonColorAndText()

            # We need this test because the lightcurve needs to be repeatedly
            # recalculated during the operation of self.fitImprovementControlCenter()
            if not computeOnly:
                self.fitImprovementControlCenter()

            if demo:
                self.fitLightcurveButton.setStyleSheet("background-color: yellow")
                self.fitLightcurveButton.setText("Fit model to observation points")

            QtWidgets.QApplication.processEvents()

            return

    def setFitModelButtonColorAndText(self):
        if self.fitStatus is not None:
            if not self.fitStatus.fitComplete:
                self.fitLightcurveButton.setText('... best fit search in progress')
                self.fitLightcurveButton.setStyleSheet("background-color: red")
        else:
            self.fitLightcurveButton.setText('Fit model to observation points')
            self.fitLightcurveButton.setStyleSheet("background-color: yellow")

        QtWidgets.QApplication.processEvents()

    def computeInitialModelTimeOffset(self):
        # Note: this routine centers the model D and R edges around the selected point, but
        # these will be offset from the camera response curve by the effect of frame time
        # and finite star disk integrations, so the visual effect may not be as expected.

        if len(self.selectedPoints) == 1:
            selIndex = [key for key, _ in self.selectedPoints.items()]
            tObsStart = convertTimeStringToTime(self.yTimes[0])
            timeAtPointSelected = convertTimeStringToTime(self.yTimes[selIndex[0]])
            relativeTime = timeAtPointSelected - tObsStart
            if relativeTime < 0:  # The lightcurve acquisition passed through midnight
                relativeTime += 24 * 60 * 60  # And a days worth of seconds

            tModelDur = (self.modelXkm[-1] - self.modelXkm[0]) / self.Lcp.shadow_speed  # in seconds
            self.modelDuration = tModelDur
            self.modelTimeOffset = relativeTime - self.modelDuration / 2

            # This should always succeed as we check this entry whenever the user
            # edits it.
            try:
                edgeTimeDelta = float(self.edgeTimePrecisionEdit.text())
            except ValueError as e:
                self.showInfo(f'Error in Edge time precision.\n\n'
                              f'edgeTimePrecisionEdit: {e}')
                return

            self.modelTimeOffset = self.modelTimeOffset - self.modelTimeOffset % edgeTimeDelta

    def handleModelSelectionRadioButtonClick(self):
        self.showDiffractionButton.setEnabled(self.diffractionRadioButton.isChecked())
        self.plotFamilyButton.setEnabled(self.diffractionRadioButton.isChecked())
        self.DdegreesEdit.setEnabled(self.edgeOnDiskRadioButton.isChecked())
        self.RdegreesEdit.setEnabled(self.edgeOnDiskRadioButton.isChecked())
        self.limbAnglePrecisionEdit.setEnabled(self.edgeOnDiskRadioButton.isChecked())
        if not self.edgeOnDiskRadioButton.isChecked():
            self.DdegreesEdit.clear()
            self.RdegreesEdit.clear()
            self.limbAnglePrecisionEdit.clear()
        else:
            if self.Lcp is not None:
                self.DdegreesEdit.setText(f'{self.Lcp.D_limb_angle_degrees:0.1f}')
                self.RdegreesEdit.setText(f'{self.Lcp.R_limb_angle_degrees:0.1f}')

        if self.Lcp is not None:
            self.setModelToUseFromRadioButtons()

        self.enableDisableEllipseEditBoxes(state=self.diffractionRadioButton.isChecked() or
                                                 self.diskOnDiskRadioButton.isChecked())

    def fillLightcurvePanelEditBoxes(self):
        # If no frame time is present, use the one from the ved Lcp
        if self.frameTimeEdit.text() == '':
            self.frameTimeEdit.setText(f'{self.Lcp.frame_time:0.5f}')
        else:  # Use the already entered frame_time
            self.Lcp.set('frame_time', float(self.frameTimeEdit.text()))

        self.missDistanceKmEdit.setText(f'{self.Lcp.miss_distance_km:0.5f}')

        self.asteroidDiameterKmEdit.setText(f'{self.Lcp.asteroid_diameter_km:0.5f}')
        self.asteroidDiameterMasEdit.setText(f'{self.Lcp.asteroid_diameter_mas:0.5f}')

        if self.Lcp.asteroid_major_axis is not None:
            self.majorAxisEdit.setText(f'{self.Lcp.asteroid_major_axis:0.3f}')
            self.minorAxisEdit.setText(f'{self.Lcp.asteroid_minor_axis:0.3f}')
            self.ellipseAngleEdit.setText(f'{self.Lcp.ellipse_angle_degrees:0.3f}')

        self.asteroidSpeedShadowEdit.setText(f'{self.Lcp.shadow_speed:0.5f}')
        self.asteroidSpeedSkyEdit.setText(f'{self.Lcp.sky_motion_mas_per_sec:0.5f}')

        self.asteroidDistAUedit.setText(f'{self.Lcp.asteroid_distance_AU:0.5f}')
        self.asteroidDistArcsecEdit.setText(f'{self.Lcp.asteroid_distance_arcsec:0.5f}')

        self.wavelengthEdit.setText(f'{self.Lcp.wavelength_nm}')

        self.baselineADUedit.setText(f'{self.Lcp.baseline_ADU:0.1f}')
        self.bottomADUedit.setText(f'{self.Lcp.bottom_ADU:0.1f}')
        if self.Lcp.magDrop is not None:
            self.magDropEdit.setText(f'{self.Lcp.magDrop:0.4f}')
        else:
            self.magDropEdit.clear()

        self.DdegreesEdit.setText(f'{self.Lcp.D_limb_angle_degrees:0.0f}')
        self.RdegreesEdit.setText(f'{self.Lcp.R_limb_angle_degrees:0.0f}')

        if self.Lcp.chord_length_sec is not None:
            self.chordSizeSecondsEdit.setText(f'{self.Lcp.chord_length_sec:0.5f}')
        else:
            self.chordSizeSecondsEdit.clear()

        if self.Lcp.chord_length_km is not None:
            self.chordSizeKmEdit.setText(f'{self.Lcp.chord_length_km:0.5f}')
        else:
            self.chordSizeKmEdit.clear()

        if self.Lcp.star_diameter_mas is not None:
            self.starSizeMasEdit.setText(f'{self.Lcp.star_diameter_mas:0.3}')
        else:
            self.starSizeMasEdit.clear()

        if self.Lcp.star_diameter_km is not None:
            self.starSizeKmEdit.setText(f'{self.Lcp.star_diameter_km:0.3f}')
        else:
            self.starSizeKmEdit.clear()

        self.fresnelSizeKmEdit.setText(f'{self.Lcp.fresnel_length_km:0.3f}')
        self.fresnelSizeSecondsEdit.setText(f'{self.Lcp.fresnel_length_sec:0.3f}')

        self.disablePrimaryEntryEditBoxes()

        self.baselineADUedit.setEnabled(False)
        self.bottomADUedit.setEnabled(False)
        self.magDropEdit.setEnabled(True)
        self.baselineADUbutton.setEnabled(True)
        self.clearBaselineADUselectionButton.setEnabled(True)

        self.calcBaselineADUbutton.setEnabled(True)

    def handlePastEventSelection(self):
        if self.csvFilePath is None:
            self.showInfo(f'You must select a csv file first!')
            return

        file_selected = self.pastEventsComboBox.currentText()

        self.fitLightcurveButton.setStyleSheet("background-color: yellow")
        self.fitLightcurveButton.setText("Fit model to observation points")

        LCPdir = os.path.dirname(self.csvFilePath)
        if sys.platform == 'darwin' or sys.platform == 'linux':
            full_name = f'{LCPdir}/LCP_{file_selected}.p'
        else:
            full_name = f'{LCPdir}\\LCP_{file_selected}.p'
        try:
            pickle_file = open(full_name, "rb")
            lcp_item = pickle.load(pickle_file)
            self.Lcp = lcp_item
            if 'modelToUse' in self.Lcp.name_list:
                modelToUse = self.Lcp.modelToUse
                if modelToUse == 'diffraction':
                    self.diffractionRadioButton.setChecked(True)
                elif modelToUse == 'edge on disk':
                    self.edgeOnDiskRadioButton.setChecked(True)
                elif modelToUse == 'disk on disk':
                    self.diskOnDiskRadioButton.setChecked(True)
                elif modelToUse == 'square wave':
                    self.squareWaveRadioButton.setChecked(True)
                else:
                    self.showInfo(f'Found invalid modelToUse: {modelToUse}')


            try:
                self.useUpperEllipseChord.setChecked(self.Lcp.use_upper_chord)
            except Exception:  # noqa
                pass
            self.fillLightcurvePanelEditBoxes()
            self.enableLightcurveButtons()
            self.handleModelSelectionRadioButtonClick()
            self.modelTimeOffset = None
            self.fitStatus = None
            self.currentEventEdit.setText(file_selected)
            eventSourceFile = self.Lcp.sourceFile
            self.baselineADUedit.setStyleSheet(None)
            self.baselineADUbutton.setStyleSheet(None)

            _, currentSourceFile = os.path.split(self.csvFilePath)
            if not currentSourceFile == eventSourceFile:
                self.showInfo(f'The source file: {eventSourceFile}\n\n'
                              f'used in the original creation of this event data does not match\n'
                              f'the current source file: {currentSourceFile}\n\n'
                              f'This may be deliberate, in which case doing an eventual\n'
                              f'"save Event" will force a match.')
                # Set metric limits to each end of the data (so that they are valid, if not the best)
                self.Lcp.set('metricLimitLeft', self.left)
                self.Lcp.set('metricLimitRight', self.right)
                self.promptForBaselineADUentry()
            else:
                # self.showInfo(f'event source file: {eventSourceFile}')
                # Show the current metric region points
                # self.showInfo(f'left metric: {self.Lcp.metricLimitLeft}    right metric: {self.Lcp.metricLimitRight}')
                showInfo = False
                if self.Lcp.metricLimitLeft is not None and self.Lcp.metricLimitLeft >= 0:    # Test for legacy settings
                    self.togglePointSelected(self.Lcp.metricLimitLeft)
                    showInfo = True
                if self.Lcp.metricLimitRight is not None and self.Lcp.metricLimitRight >= 0:  # Test for legacy settings
                    self.togglePointSelected(self.Lcp.metricLimitRight)
                    showInfo = True
                self.redrawMainPlot()
                if showInfo:
                    self.showInfo(f'The current metric region selection points are displayed.\n\n'
                                  f'They are at reading {self.Lcp.metricLimitLeft} and {self.Lcp.metricLimitRight}')
                    self.togglePointSelected(self.Lcp.metricLimitLeft)
                    self.togglePointSelected(self.Lcp.metricLimitRight)


            self.currentEventEdit.setEnabled(True)
            self.allCoreElementsEntered = True
        except FileNotFoundError:
            self.showInfo(f'{full_name} could not be found.')

    def clearEventDataEntries(self):
        self.initializeModelLightcurvesPanel()
        self.fitStatus = None
        self.handleModelSelectionRadioButtonClick()

    def handleSiteCoordSelection(self):
        file_selected = self.vzCoordsComboBox.currentText()
        if file_selected == '<preset coords>':
            return

        sourceDir = self.homeDir
        if sys.platform == 'darwin' or sys.platform == 'linux':
            full_name = f'{sourceDir}/SiteCoord_{file_selected}.p'
        else:
            full_name = f'{sourceDir}\\SiteCoord_{file_selected}.p'
        try:
            pickle_file = open(full_name, "rb")
            coorDict = pickle.load(pickle_file)
            # Fill in the fields
            self.vzSiteLongDegEdit.setText(coorDict["longitude deg"])
            self.vzSiteLongMinEdit.setText(coorDict["longitude min"])
            self.vzSiteLongSecsEdit.setText(coorDict["longitude sec"])
            self.vzSiteLatDegEdit.setText(coorDict["latitude deg"])
            self.vzSiteLatMinEdit.setText(coorDict["latitude min"])
            self.vzSiteLatSecsEdit.setText(coorDict["latitude sec"])
            self.vzSiteAltitudeEdit.setText(coorDict["altitude"])
            self.vzObserverNameEdit.setText(coorDict["observer"])
        except FileNotFoundError:
            self.showInfo(f'{full_name} could not be found.')

    def saveSiteCoords(self):
        if self.vzSiteCoordNameEdit.text() == '':
            self.showInfo('A site coordinates name is required before a "save" can be performed')
            return

        # Build a dictionary
        siteCoordDict = {
            "longitude deg": self.vzSiteLongDegEdit.text(),
            "longitude min": self.vzSiteLongMinEdit.text(),
            "longitude sec": self.vzSiteLongSecsEdit.text(),
            "latitude deg": self.vzSiteLatDegEdit.text(),
            "latitude min": self.vzSiteLatMinEdit.text(),
            "latitude sec": self.vzSiteLatSecsEdit.text(),
            "altitude": self.vzSiteAltitudeEdit.text(),
            "observer": self.vzObserverNameEdit.text(),
        }

        destDir = self.homeDir

        # We overwrite without warning an event file with the same name
        try:
            filename = f'SiteCoord_{self.vzSiteCoordNameEdit.text()}.p'
            filepath = os.path.join(destDir, filename)
            pickle.dump(siteCoordDict, open(filepath, 'wb'))  # noqa
        except Exception as e:
            self.showInfo(f'Attempt to write site coordinate data: {e}')

        # Update the site coordinates combo box
        self.vzCoordsComboBox.clear()
        self.fillVzCoordsComboBox()

    def saveCurrentEvent(self):
        if self.Lcp is None:
            self.showInfo('There is no event data to save')
            return

        if self.currentEventEdit.text() == '':
            self.showInfo('An event name is required before a "save" can be performed')
            return

        if not self.allCoreElementsEntered:
            self.showInfo(f'There are some core event parameters yet to be entered.\n\n'
                          f'The core event parameters are those between the first pair of black bars.')
        # This allows the current filename to be updated
        self.Lcp.sourceFile = os.path.split(self.csvFilePath)[1]

        LCPdirectory = os.path.dirname(self.csvFilePath)

        majorAxis, minorAxis, thetaDegrees = self.validateEllipseParameters()
        if majorAxis is not None:
            self.Lcp.asteroid_major_axis = majorAxis
            self.Lcp.asteroid_minor_axis = minorAxis
            self.Lcp.ellipse_angle_degrees = thetaDegrees
            self.Lcp.use_upper_chord = self.useUpperEllipseChord.isChecked()

        # We overwrite without warning an event file with the same name
        try:
            filename = f'LCP_{self.currentEventEdit.text()}.p'
            filepath = os.path.join(LCPdirectory, filename)
            pickle.dump(self.Lcp, open(filepath, 'wb'))  # noqa

            self.showInfo(f'The current event data was written to '
                          f'\n\n{filepath}\n\n')
        except Exception as e:
            self.showInfo(f'Attempt to write event data: {e}')

        # Update the past events combo box
        self.pastEventsComboBox.clear()
        self.fillPastEventsComboBox()

    def processNewCurrentEventEdit(self):
        if self.csvFilePath is None:
            self.showInfo(f'A csv file must be present before entering\n'
                          f'of event data is allowed.')
            return

        proposed_filename = self.currentEventEdit.text()
        if not self.validFilename(proposed_filename):
            self.showInfo(f'The event name contains a character that is invalid in a filename.')
            return

        # This allows us to change the name for a set of lightcurve data
        if self.Lcp is not None:
            return

        # msg = QMessageBox()
        # msg.setIcon(QMessageBox.Question)
        # msg.setText(f'It is best to select the "Model to use" before entering new parameter data.\n\n'
        #             f'Do you need to do that now?')
        # msg.setWindowTitle('Exit to allow "Model to use" radio button selection')
        # msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        # retval = msg.exec_()
        # if retval == QMessageBox.Yes:
        #     self.currentEventEdit.setText("")
        #     return

        self.handleModelSelectionRadioButtonClick()

        self.baselineADUedit.setEnabled(False)
        self.baselineADUbutton.setEnabled(True)
        self.calcBaselineADUbutton.setEnabled(True)

        self.bottomADUedit.setEnabled(False)
        self.magDropEdit.setEnabled(True)

        self.frameTimeEdit.setEnabled(True)

        self.missDistanceKmEdit.setEnabled(True)
        self.missDistanceKmEdit.clear()

        self.asteroidDiameterKmEdit.setEnabled(True)
        self.asteroidDiameterKmEdit.clear()
        self.asteroidDiameterMasEdit.setEnabled(True)
        self.asteroidDiameterMasEdit.clear()

        self.asteroidSpeedShadowEdit.setEnabled(True)
        self.asteroidSpeedShadowEdit.clear()
        self.asteroidSpeedSkyEdit.setEnabled(True)
        self.asteroidSpeedSkyEdit.clear()

        self.asteroidDistAUedit.setEnabled(True)
        self.asteroidDistAUedit.clear()
        self.asteroidDistArcsecEdit.setEnabled(True)
        self.asteroidDistArcsecEdit.clear()

        self.wavelengthEdit.setEnabled(True)
        self.wavelengthEdit.setText('540')

        self.newEventDataBeingEntered = True
        self.asteroidDiameterKmEdit.setFocus()
        QtWidgets.QApplication.processEvents()

    def processModelParameterChange(self):
        self.parameterChangeEntryCount += 1

        try:
            empty = ''

            if not self.DdegreesEdit.text() == empty:
                valueEntered = float(self.DdegreesEdit.text())
                if valueEntered >= 90.0:
                    self.DdegreesEdit.setText('90')
                    valueEntered = 90
                self.Lcp.set('D_limb_angle_degrees', valueEntered)

            if not self.RdegreesEdit.text() == empty:
                valueEntered = float(self.RdegreesEdit.text())
                if valueEntered >= 90.0:
                    self.RdegreesEdit.setText('90')
                    valueEntered = 90
                self.Lcp.set('R_limb_angle_degrees', valueEntered)

            if self.chordSizeSecondsEdited and not self.chordSizeSecondsEdit.text() == empty:
                self.Lcp.set('chord_length_km', None)
                self.Lcp.set('chord_length_sec', float(self.chordSizeSecondsEdit.text()))
                self.chordSizeKmEdit.setText(f'{self.Lcp.chord_length_km:0.5f}')
            elif not self.chordSizeKmEdit.text() == empty and self.chordSizeKmEdited:
                self.Lcp.set('chord_length_sec', None)
                self.Lcp.set('chord_length_km', float(self.chordSizeKmEdit.text()))
                self.chordSizeSecondsEdit.setText(f'{self.Lcp.chord_length_sec:0.5f}')

            if not self.starSizeMasEdit.text() == empty and self.starSizeMasEdited:
                self.Lcp.set('star_diameter_km', None)
                self.Lcp.set('star_diameter_mas', float(self.starSizeMasEdit.text()))
                self.starSizeKmEdit.setText(f'{self.Lcp.star_diameter_km:0.3f}')
            elif not self.starSizeKmEdit.text() == empty and self.starSizeKmEdited:
                self.Lcp.set('star_diameter_mas', None)
                self.Lcp.set('star_diameter_km', float(self.starSizeKmEdit.text()))
                self.starSizeMasEdit.setText(f'{self.Lcp.star_diameter_mas:0.3f}')

            if not self.magDropEdit.text() == empty:
                magDrop = float(self.magDropEdit.text())
                try:
                    bottomADU = self.calcBottomADU(
                        self.Lcp.baseline_ADU,
                        magDrop=magDrop)
                    self.bottomADUedit.setText(f'{bottomADU:0.5f}')
                    self.Lcp.set('bottom_ADU', bottomADU)
                    self.Lcp.set('magDrop', magDrop)
                except Exception as e:  # noqc
                    self.showInfo(f'{e}')

            QtWidgets.QApplication.processEvents()

            anUnsetParameterFound, _ = self.Lcp.check_for_none()

            if not anUnsetParameterFound:
                self.enableLightcurveButtons()
                self.consistentLcpPresent = True
                self.enablePrimaryEntryEditBoxes()
                self.enableSecondaryEditBoxes()

                self.fitLightcurveButton.setStyleSheet("background-color: yellow")
                if self.edgeOnDiskRadioButton.isChecked():
                    self.DdegreesEdit.setText(f'{self.Lcp.D_limb_angle_degrees:0.0f}')
                    self.RdegreesEdit.setText(f'{self.Lcp.R_limb_angle_degrees:0.0f}')

                self.newRedrawMainPlot()

        except ValueError as e:  # noqc
            self.showInfo(f'At end of processModelParameterChange(): {e}')

    def processModelLightcurveCoreEdit(self):
        # We do this frequently to be certain to pick up when to gray out limb angle stuff
        self.handleModelSelectionRadioButtonClick()

        try:
            empty = ''

            frameTimeEntered = not self.frameTimeEdit.text() == empty
            try:
                if self.frameTimeEdit.text() == empty:
                    return
                _ = float(self.frameTimeEdit.text())
            except ValueError as e:
                self.showInfo(f'frame time edit: {e}')
                return

            asteroidDiameterKmEntered = not self.asteroidDiameterKmEdit.text() == empty
            if asteroidDiameterKmEntered:
                try:
                    if self.asteroidDiameterKmEdit.text() == empty:
                        return
                    _ = float(self.asteroidDiameterKmEdit.text())
                except ValueError as e:
                    self.showInfo(f'asteroid diameter (km): {e}')
                    return
                self.asteroidDiameterMasEdit.setEnabled(False)

                # Fill in defaults for ellipse parameters
                self.majorAxisEdit.setText(self.asteroidDiameterKmEdit.text())
                self.minorAxisEdit.setText(self.asteroidDiameterKmEdit.text())
                self.ellipseAngleEdit.setText('0.0')


            asteroidDiameterMasEntered = not self.asteroidDiameterMasEdit.text() == empty
            if asteroidDiameterMasEntered:
                try:
                    if self.asteroidDiameterMasEdit.text() == empty:
                        return
                    _ = float(self.asteroidDiameterMasEdit.text())
                except ValueError as e:
                    self.showInfo(f'asteroid diameter (mas) {e}')
                    return
                # We used to do this, but now allow edits at any time
                # self.asteroidDiameterKmEdit.setEnabled(False)

            asteroidShadowSpeedEntered = not self.asteroidSpeedShadowEdit.text() == empty
            if asteroidShadowSpeedEntered:
                try:
                    if self.asteroidSpeedShadowEdit.text() == empty:
                        return
                    _ = float(self.asteroidSpeedShadowEdit.text())
                except ValueError as e:
                    self.showInfo(f'shadow speed: {e}')
                    return
                self.asteroidSpeedSkyEdit.setEnabled(False)

            asteroidSkySpeedEntered = not self.asteroidSpeedSkyEdit.text() == empty
            if asteroidSkySpeedEntered:
                try:
                    if self.asteroidSpeedSkyEdit.text() == empty:
                        return
                    _ = float(self.asteroidSpeedSkyEdit.text())
                except ValueError as e:
                    self.showInfo(f'sky speed: {e}')
                    return
                self.asteroidSpeedShadowEdit.setEnabled(False)

            asteroidDistAUentered = not self.asteroidDistAUedit.text() == empty
            if asteroidDistAUentered:
                try:
                    if self.asteroidDistAUedit.text() == empty:
                        return
                    _ = float(self.asteroidDistAUedit.text())
                except ValueError as e:
                    self.showInfo(f'asteroid distance (AU): {e}')
                    return
                self.asteroidDistArcsecEdit.setEnabled(False)

            asteroidDistArcsecEntered = not self.asteroidDistArcsecEdit.text() == empty
            if asteroidDistArcsecEntered:
                try:
                    if self.asteroidDistArcsecEdit.text() == empty:
                        return
                    _ = float(self.asteroidDistArcsecEdit.text())
                except ValueError as e:
                    self.showInfo(f'asteroid distance (arcsec): {e}')
                    return
                self.asteroidDistAUedit.setEnabled(False)

            wavelengthEntered = not self.wavelengthEdit.text() == empty
            if wavelengthEntered:
                try:
                    if self.wavelengthEdit.text() == empty:
                        return
                    _ = int(self.wavelengthEdit.text())
                except ValueError as e:
                    self.showInfo(f'wavelength (nm): {e}')
                    return

            missDistanceEntered = not self.missDistanceKmEdit.text() == empty
            if missDistanceEntered:
                try:
                    if self.missDistanceKmEdit.text() == empty:
                        return
                    _ = float(self.missDistanceKmEdit.text())
                except ValueError as e:
                    self.showInfo(f'miss distance (km): {e}')
                    return

            self.allCoreElementsEntered = frameTimeEntered
            self.allCoreElementsEntered = self.allCoreElementsEntered and wavelengthEntered
            self.allCoreElementsEntered = self.allCoreElementsEntered and (asteroidShadowSpeedEntered or asteroidSkySpeedEntered)
            self.allCoreElementsEntered = self.allCoreElementsEntered and (asteroidDistAUentered or asteroidDistArcsecEntered)
            self.allCoreElementsEntered = self.allCoreElementsEntered and (asteroidDiameterMasEntered or asteroidDiameterKmEntered)
            self.allCoreElementsEntered = self.allCoreElementsEntered and missDistanceEntered
            if not self.allCoreElementsEntered:
                return

            if self.Lcp is not None:
                if not self.missDistanceKmEdit.text() == empty:
                    self.Lcp.set('miss_distance_km', float(self.missDistanceKmEdit.text()))

                if self.Lcp.sigmaB is not None:
                    return  # We got here from edit mode and baseline data is available

            self.baselineADUedit.setStyleSheet("background-color: lightblue")
            self.baselineADUbutton.setStyleSheet("background-color: lightblue")

            # Make sure that the plot cursor is not in Arrow mode
            self.blankCursor = True
            self.mainPlot.viewport().setProperty("cursor",
                                                 QtGui.QCursor(QtCore.Qt.CursorShape.BlankCursor))

            # Set default values for baselineADU and bottomADU so that we
            # can create an Lcp.
            baselineADU = 100.0
            bottomADU = 0.0

            if self.edgeOnDiskRadioButton.isChecked():
                self.DdegreesEdit.setText('45')
                self.RdegreesEdit.setText('45')
            else:
                self.DdegreesEdit.setEnabled(False)
                self.RdegreesEdit.setEnabled(False)

            if asteroidShadowSpeedEntered:
                # noinspection PyTypeChecker
                self.Lcp = LightcurveParameters(
                    baseline_ADU=baselineADU,
                    bottom_ADU=bottomADU,
                    frame_time=float(self.frameTimeEdit.text()),
                    wavelength_nm=int(self.wavelengthEdit.text()),
                    shadow_speed=float(self.asteroidSpeedShadowEdit.text()),
                    sky_motion_mas_per_sec=None
                )

                if not self.missDistanceKmEdit.text() == empty:
                    self.Lcp.set('miss_distance_km', float(self.missDistanceKmEdit.text()))
                    if self.Lcp.miss_distance_km > 0:
                        self.chordSizeKmEdit.setText('0.0')
                        self.chordSizeSecondsEdit.setText('0.0')
                        self.chordSizeKmEdited = True

                _, filenameWithoutPath = os.path.split(self.csvFilePath)
                self.Lcp.set('sourceFile', filenameWithoutPath)

                if asteroidDistAUentered:
                    self.Lcp.set('asteroid_distance_AU', float(self.asteroidDistAUedit.text()))
                    self.asteroidDistArcsecEdit.setText(f'{self.Lcp.asteroid_distance_arcsec:0.5f}')
                else:
                    self.Lcp.set('asteroid_distance_arcsec', float(self.asteroidDistArcsecEdit.text()))
                    self.asteroidDistAUedit.setText(f'{self.Lcp.asteroid_distance_AU:0.5f}')

                if asteroidDiameterKmEntered:
                    self.Lcp.set('asteroid_diameter_km', float(self.asteroidDiameterKmEdit.text()))
                    self.asteroidDiameterMasEdit.setText(f'{self.Lcp.asteroid_diameter_mas:0.5f}')
                else:
                    self.Lcp.set('asteroid_diameter_mas', float(self.asteroidDiameterMasEdit.text()))
                    self.asteroidDiameterKmEdit.setText(f'{self.Lcp.asteroid_diameter_km:0.5f}')
                    # Fill in defaults for ellipse parameters
                    self.majorAxisEdit.setText(self.asteroidDiameterKmEdit.text())
                    self.minorAxisEdit.setText(self.asteroidDiameterKmEdit.text())
                    self.ellipseAngleEdit.setText('0.0')

                self.asteroidSpeedSkyEdit.setText(f'{self.Lcp.sky_motion_mas_per_sec:0.5f}')
                self.consistentLcpPresent = True

            else:
                # noinspection PyTypeChecker
                self.Lcp = LightcurveParameters(
                    baseline_ADU=baselineADU,
                    bottom_ADU=bottomADU,
                    frame_time=float(self.frameTimeEdit.text()),
                    wavelength_nm=int(self.wavelengthEdit.text()),
                    shadow_speed=None,
                    sky_motion_mas_per_sec=float(self.asteroidSpeedSkyEdit.text())
                )

                if not self.missDistanceKmEdit.text() == empty:
                    self.Lcp.set('miss_distance_km', float(self.missDistanceKmEdit.text()))
                    if self.Lcp.miss_distance_km > 0:
                        self.chordSizeKmEdit.setText('0.0')
                        self.chordSizeSecondsEdit.setText('0.0')
                        self.chordSizeKmEdited = True

                _, filenameWithoutPath = os.path.split(self.csvFilePath)
                self.Lcp.set('sourceFile', filenameWithoutPath)
                self.showInfo(f'file name: {filenameWithoutPath}')

                if asteroidDistAUentered:
                    self.Lcp.set('asteroid_distance_AU', float(self.asteroidDistAUedit.text()))
                    self.asteroidDistArcsecEdit.setText(f'{self.Lcp.asteroid_distance_arcsec:0.5f}')
                else:
                    self.Lcp.set('asteroid_distance_arcsec', float(self.asteroidDistArcsecEdit.text()))
                    self.asteroidDistAUedit.setText(f'{self.Lcp.asteroid_distance_AU:0.5f}')

                if asteroidDiameterKmEntered:
                    self.Lcp.set('asteroid_diameter_km', float(self.asteroidDiameterKmEdit.text()))
                    self.asteroidDiameterMasEdit.setText(f'{self.Lcp.asteroid_diameter_mas:0.5f}')
                else:
                    self.Lcp.set('asteroid_diameter_mas', float(self.asteroidDiameterMasEdit.text()))
                    self.asteroidDiameterKmEdit.setText(f'{self.Lcp.asteroid_diameter_km:0.5f}')

                self.asteroidSpeedShadowEdit.setText(f'{self.Lcp.shadow_speed:0.5f}')
                self.consistentLcpPresent = True

            self.setModelToUseFromRadioButtons()

            # Fill in defaults for ellipse parameters
            self.Lcp.set('asteroid_major_axis', float(self.asteroidDiameterKmEdit.text()))
            self.Lcp.set('asteroid_minor_axis', float(self.asteroidDiameterKmEdit.text()))
            self.Lcp.set('ellipse_angle_degrees', float(self.ellipseAngleEdit.text()))
            self.Lcp.set('metricLimitLeft', self.left)
            self.Lcp.set('metricLimitRight', self.right)

            self.disablePrimaryEntryEditBoxes()

            self.fresnelSizeKmEdit.setText(f'{self.Lcp.fresnel_length_km:0.3f}')
            self.fresnelSizeSecondsEdit.setText(f'{self.Lcp.fresnel_length_sec:0.3f}')

            # The following code has been commented out. The normal answer is NO, so I feel that the
            # question should not be asked. IF someone is doing a multi-chord analysis, they will also
            # know enough to save the data set at this point. And worst case is that they would have to
            # enter the basic parameters set repeatedly for each chord - not so burdensome.

            # msg = QMessageBox()
            # msg.setIcon(QMessageBox.Question)
            # msg.setText(f'Do you wish to save the event data entered so far?\n\n'
            #             f'Answer Yes IF you are going to analyze multiple chords from the same event and wish'
            #             f' to create a "template" to be inserted manually into the folder for each lightcurve.\n\n'
            #             f'For other than this specialty use, this question should answered with No\n\n'
            #             f'For normal use, finish entering the rest of the data and then click the Save Event '
            #             f'button!')
            # msg.setWindowTitle('Save core data')
            # msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            # retval = msg.exec_()
            # if retval == QMessageBox.Yes:
            #     self.saveCurrentEvent()
            #     return

            self.promptForBaselineADUentry()

        except ValueError as e:  # noqc
            self.showInfo(f'general error in CoreEdit(): {e}')

    def setModelToUseFromRadioButtons(self):

        # Check for legacy Lcp that does not have a modelToUse field
        if not 'modelToUse' in self.Lcp.name_list:
            return

        if self.diffractionRadioButton.isChecked():
            self.Lcp.set('modelToUse', 'diffraction')
            self.enableDisableEllipseEditBoxes(state=True)
        elif self.edgeOnDiskRadioButton.isChecked():
            self.Lcp.set('modelToUse', 'edge on disk')
        elif self.diskOnDiskRadioButton.isChecked():
            self.Lcp.set('modelToUse', 'disk on disk')
        elif self.squareWaveRadioButton.isChecked():
            self.Lcp.set('modelToUse', 'square wave')
        else:
            self.Lcp.set('modelToUse', None)
            self.showInfo(f'Invalid model radio button state')

    def promptForBaselineADUentry(self):
        self.showInfo(f'Set baselineADU by selecting points to be\n'
                      f'included in the calculation - multiple regions can be\n'
                      f'selected.\n\n'
                      f'When all regions have been selected, click on the Calc button.\n\n'
                      f'This will clear the selected points and '
                      f'put the entry cursor in the Predicted magDrop box '
                      f'ready for that entry.')

        self.baselineADUedit.setStyleSheet("background-color: lightblue")
        self.baselineADUbutton.setStyleSheet("background-color: lightblue")

    def enableSecondaryEditBoxes(self):
        if self.edgeOnDiskRadioButton.isChecked():
            self.DdegreesEdit.setEnabled(True)
            self.RdegreesEdit.setEnabled(True)
        self.chordSizeKmEdit.setEnabled(True)
        self.chordSizeSecondsEdit.setEnabled(True)
        self.magDropEdit.setEnabled(True)
        self.starSizeMasEdit.setEnabled(True)
        self.starSizeKmEdit.setEnabled(True)

    def enablePrimaryEntryEditBoxes(self):
        self.frameTimeEdit.setEnabled(True)
        self.missDistanceKmEdit.setEnabled(True)
        self.asteroidDiameterKmEdit.setEnabled(True)
        self.asteroidDiameterMasEdit.setEnabled(True)
        self.asteroidSpeedShadowEdit.setEnabled(True)
        self.asteroidSpeedSkyEdit.setEnabled(True)
        self.asteroidDistAUedit.setEnabled(False)
        self.asteroidDistArcsecEdit.setEnabled(False)
        self.wavelengthEdit.setEnabled(True)

    def disablePrimaryEntryEditBoxes(self):
        self.frameTimeEdit.setEnabled(False)
        self.missDistanceKmEdit.setEnabled(True)
        self.asteroidDiameterKmEdit.setEnabled(True)
        self.magDropEdit.setEnabled(True)
        self.asteroidDiameterMasEdit.setEnabled(True)
        self.asteroidSpeedShadowEdit.setEnabled(True)
        self.asteroidSpeedSkyEdit.setEnabled(True)
        self.asteroidDistAUedit.setEnabled(False)
        self.asteroidDistArcsecEdit.setEnabled(False)
        self.wavelengthEdit.setEnabled(False)

        self.enableSecondaryEditBoxes()

    def fillPastEventsComboBox(self):
        # self.pastEventsComboBox.addItem('<clear event data>')
        LCPdir = os.path.dirname(self.csvFilePath)
        if sys.platform == 'darwin' or sys.platform == 'linux':
            file_list = glob.glob(f'{LCPdir}/LCP_*.p')
        else:
            file_list = glob.glob(f'{LCPdir}\\LCP_*.p')
        for file in file_list:
            filename = os.path.basename(file)
            clean_name = filename[4:-2]
            self.pastEventsComboBox.addItem(clean_name)

    def fillVzCoordsComboBox(self):
        self.vzCoordsComboBox.addItem('<preset coords>')
        sourceDir = self.homeDir
        if sys.platform == 'darwin' or sys.platform == 'linux':
            file_list = glob.glob(f'{sourceDir}/SiteCoord_*.p')
        else:
            file_list = glob.glob(f'{sourceDir}\\SiteCoord_*.p')
        for file in file_list:
            filename = os.path.basename(file)
            clean_name = filename[10:-2]
            self.vzCoordsComboBox.addItem(clean_name)

    def initializeModelLightcurvesPanel(self):

        self.modelTimeOffset = 0.0
        self.fitStatus = {}
        self.allCoreElementsEntered = False

        self.allowShowDetails = False

        self.modelMetric = None

        self.Lcp = None

        self.currentEventEdit.clear()
        self.currentEventEdit.setEnabled(True)

        self.pastEventsComboBox.clear()
        if self.csvFilePath is not None:
            self.fillPastEventsComboBox()

        self.baselineADUedit.setEnabled(False)
        self.baselineADUedit.clear()
        self.baselineADUbutton.setEnabled(False)
        self.calcBaselineADUbutton.setEnabled(False)
        self.clearBaselineADUselectionButton.setEnabled(False)

        self.bottomADUedit.setEnabled(False)
        self.bottomADUedit.clear()

        self.magDropEdit.setEnabled(False)
        self.magDropEdit.clear()

        self.frameTimeEdit.setEnabled(False)
        self.frameTimeEdit.clear()

        self.missDistanceKmEdit.setEnabled(False)
        self.missDistanceKmEdit.clear()

        self.asteroidDiameterKmEdit.setEnabled(False)
        self.asteroidDiameterKmEdit.clear()

        self.asteroidDiameterMasEdit.setEnabled(False)
        self.asteroidDiameterMasEdit.clear()

        self.asteroidSpeedShadowEdit.setEnabled(False)
        self.asteroidSpeedShadowEdit.clear()
        self.asteroidSpeedSkyEdit.setEnabled(False)
        self.asteroidSpeedSkyEdit.clear()

        self.asteroidDistAUedit.setEnabled(False)
        self.asteroidDistAUedit.clear()
        self.asteroidDistArcsecEdit.setEnabled(False)
        self.asteroidDistArcsecEdit.clear()

        self.wavelengthEdit.setEnabled(False)
        self.wavelengthEdit.clear()

        self.DdegreesEdit.setEnabled(False)
        self.DdegreesEdit.clear()
        self.RdegreesEdit.setEnabled(False)
        self.RdegreesEdit.clear()

        self.chordSizeSecondsEdit.setEnabled(False)
        self.magDropEdit.setEnabled(False)
        self.chordSizeSecondsEdit.clear()
        self.chordSizeKmEdit.setEnabled(False)
        self.chordSizeKmEdit.clear()

        self.fresnelSizeKmEdit.clear()
        self.fresnelSizeKmEdit.setEnabled(False)
        self.fresnelSizeSecondsEdit.clear()
        self.fresnelSizeSecondsEdit.setEnabled(False)

        self.starSizeMasEdit.setEnabled(False)
        self.starSizeMasEdit.clear()
        self.starSizeKmEdit.setEnabled(False)
        self.starSizeKmEdit.clear()

        self.enableDisableEllipseEditBoxes()

        self.fitLightcurveButton.setEnabled(False)
        self.plotFamilyButton.setEnabled(False)
        self.setMetricLimitsButton.setEnabled(False)
        self.pauseFitButton.setEnabled(False)
        self.askAdviceButton.setEnabled(False)
        self.showDiffractionButton.setEnabled(False)
        self.plotFamilyButton.setEnabled(False)
        self.demoModelButton.setEnabled(False)
        self.demoGeometryButton.setEnabled(False)
        self.printEventParametersButton.setEnabled(False)

        self.modelXkm = None
        self.modelY = None

        self.handleModelSelectionRadioButtonClick()

        self.newRedrawMainPlot()

    def enableDisableEllipseEditBoxes(self, state=False):
        self.majorAxisEdit.setEnabled(state)
        self.minorAxisEdit.setEnabled(state)
        self.ellipseAngleEdit.setEnabled(state)
        self.useUpperEllipseChord.setEnabled(state)

    def enableLightcurveButtons(self):

        self.fitLightcurveButton.setEnabled(True)
        self.askAdviceButton.setEnabled(True)
        self.showDiffractionButton.setEnabled(self.diffractionRadioButton.isChecked())
        self.plotFamilyButton.setEnabled(self.diffractionRadioButton.isChecked())
        self.demoModelButton.setEnabled(True)
        self.demoGeometryButton.setEnabled(True)
        self.plotFamilyButton.setEnabled(True)
        self.setMetricLimitsButton.setEnabled(True)
        self.printEventParametersButton.setEnabled(True)

    def processYoffsetStepBy10(self):
        self.yOffsetStep10radioButton.repaint()
        for spinBox in self.yOffsetSpinBoxes:
            spinBox.setSingleStep(10)

    def processYoffsetStepBy100(self):
        self.yOffsetStep100radioButton.repaint()
        for spinBox in self.yOffsetSpinBoxes:
            spinBox.setSingleStep(100)

    def processYoffsetStepBy1000(self):
        self.yOffsetStep1000radioButton.repaint()
        for spinBox in self.yOffsetSpinBoxes:
            spinBox.setSingleStep(1000)

    def processStepBy2(self):
        self.smoothingIntervalSpinBox.setSingleStep(2)

    def processStepBy10(self):
        self.smoothingIntervalSpinBox.setSingleStep(10)

    def processStepBy100(self):
        self.smoothingIntervalSpinBox.setSingleStep(100)

    def initializeLightcurvePanel(self):
        for checkBox in self.targetCheckBoxes:
            checkBox.setChecked(False)
            checkBox.setEnabled(False)
        self.targetCheckBoxes[0].setChecked(True)

        for checkBox in self.showCheckBoxes:
            checkBox.setChecked(False)
            checkBox.setEnabled(False)
        self.showCheckBoxes[0].setChecked(True)

        for checkBox in self.referenceCheckBoxes:
            checkBox.setChecked(False)
            checkBox.setEnabled(False)

        for title in self.lightcurveTitles:
            title.setText('')

        for spinBox in self.yOffsetSpinBoxes:
            spinBox.setValue(0)
            spinBox.setEnabled(False)
        self.yOffsetSpinBoxes[0].setValue(0)
        self.yOffsetSpinBoxes[0].setEnabled(True)

        for spinBox in self.xOffsetSpinBoxes:
            spinBox.setValue(0)
            spinBox.setEnabled(False)

        self.recolorBlobs()

    def processYoffsetChange(self):
        self.newRedrawMainPlot()

    def processXoffsetChange(self):
        self.newRedrawMainPlot()

    def clearReferenceSelections(self):
        for checkBox in self.referenceCheckBoxes:
            checkBox.setChecked(False)
        for spinBox in self.xOffsetSpinBoxes:
            spinBox.setEnabled(False)

    def recolorBlobs(self):
        for i, colorBlob in enumerate(self.colorBlobs):
            colorBlob.setStyleSheet(self.colorStyle[i])
        for i, checkBox in enumerate(self.targetCheckBoxes):
            if checkBox.isChecked():
                self.colorBlobs[i].setStyleSheet('color: rgb(0,0,255)')  # Set target curve bright blue
                break
        for i, checkBox in enumerate(self.referenceCheckBoxes):
            if checkBox.isChecked():
                self.colorBlobs[i].setStyleSheet('color: rgb(0,255,0)')  # set reference curve bright green
                break

    def processReferenceSelection(self, i):
        # Undo any previous normalization
        self.fillTableViewOfData()
        self.smoothingIntervalSpinBox.setValue(0)
        self.referenceKey = ''

        if self.referenceCheckBoxes[i].isChecked():
            self.referenceKey = self.lightcurveTitles[i].text()
            self.showMsg(f'{self.referenceKey} is selected as the reference curve for normalization.')
            self.clearReferenceSelections()
            self.xOffsetSpinBoxes[i].setEnabled(True)
            self.referenceCheckBoxes[i].setChecked(True)
            self.showCheckBoxes[i].setChecked(True)
            if i == 0:
                self.yRefStar = self.LC1[:]
            elif i == 1:
                self.yRefStar = self.LC2[:]
            elif i == 2:
                self.yRefStar = self.LC3[:]
            elif i == 3:
                self.yRefStar = self.LC4[:]
            else:
                self.yRefStar = self.extra[i - 4][:]
            self.recolorBlobs()
            self.newRedrawMainPlot()
        else:
            self.xOffsetSpinBoxes[i].setEnabled(False)
            for i, checkBox in enumerate(self.referenceCheckBoxes):
                if checkBox.isChecked():
                    self.showMsg(
                        f'{self.lightcurveTitles[i].text()} is selected as the reference curve for normalization.')
                    return
            self.yRefStar = []
            self.newRedrawMainPlot()
            self.showMsg(f'The reference curve has been deselected so normalization is disabled.')

    def processReferenceSelection1(self):
        self.processReferenceSelection(0)

    def processReferenceSelection2(self):
        self.processReferenceSelection(1)

    def processReferenceSelection3(self):
        self.processReferenceSelection(2)

    def processReferenceSelection4(self):
        self.processReferenceSelection(3)

    def processReferenceSelection5(self):
        self.processReferenceSelection(4)

    def processReferenceSelection6(self):
        self.processReferenceSelection(5)

    def processReferenceSelection7(self):
        self.processReferenceSelection(6)

    def processReferenceSelection8(self):
        self.processReferenceSelection(7)

    def processReferenceSelection9(self):
        self.processReferenceSelection(8)

    def processReferenceSelection10(self):
        self.processReferenceSelection(9)

    def forceTargetToShow(self):
        for i, checkBox in enumerate(self.targetCheckBoxes):
            if checkBox.isChecked():
                self.showCheckBoxes[i].setChecked(True)

    def calcCorrectVersusTargetMetric(self, target_index, correct_index):
        if self.right is not None:
            right = min(self.dataLen, self.right + 1)
        else:
            right = self.dataLen
        if self.left is None:
            left = 0
        else:
            left = self.left

        if target_index == 0:
            yi = [self.LC1[i] for i in range(left, right)]
        elif target_index == 1:
            yi = [self.LC2[i] for i in range(left, right)]
        elif target_index == 2:
            yi = [self.LC3[i] for i in range(left, right)]
        elif target_index == 3:
            yi = [self.LC4[i] for i in range(left, right)]
        else:
            yi = [self.extra[target_index - 4][j] for j in range(left, right)]

        if correct_index == 0:
            yj = [self.LC1[i] for i in range(left, right)]
        elif correct_index == 1:
            yj = [self.LC2[i] for i in range(left, right)]
        elif correct_index == 2:
            yj = [self.LC3[i] for i in range(left, right)]
        elif correct_index == 3:
            yj = [self.LC4[i] for i in range(left, right)]
        else:
            yj = [self.extra[correct_index   - 4][i] for i in range(left, right)]

        metric = 0
        for i in range(len(yi)):
            metric += (yi[i] - yj[i])**2
        metric = np.sqrt(metric)
        self.showMsg(f'sqrt(sum((signal-correct - signal-target)**2)): {metric:0.3f}')
        # print(f'metric: {metric:0.3f}')

    def checkForBothCorrectAndTargetPresent(self):
        correct_lightcurve_showing = False
        correct_lightcurve_index = None
        target_lightcurve_showing = False
        target_lightcurve_index = None
        for i, title in enumerate(self.lightcurveTitles):
            if title.text().startswith('signal-correct'):
                if self.showCheckBoxes[i].isChecked():
                    correct_lightcurve_showing = True
                    correct_lightcurve_index = i
                    # self.showInfo(f'The theoretically correct lightcurve is present at index {i} and showing')
                else:
                    # self.showInfo(f'The theoretically correct lightcurve is present at index {i} but not showing')
                    pass
            elif title.text().startswith('signal-target'):
                if self.showCheckBoxes[i].isChecked():
                    target_lightcurve_showing = True
                    target_lightcurve_index = i
                    # self.showInfo(f'The target lightcurve is present at index {i} and showing')
                else:
                    # self.showInfo(f'The targetlightcurve is present at index {i} but not showing')
                    pass
        if correct_lightcurve_showing and target_lightcurve_showing:
            # self.showInfo(f'target and correct are both present and showing. We will calculate metric.')
            self.calcCorrectVersusTargetMetric(target_lightcurve_index, correct_lightcurve_index)


    def processShowSelection(self):
        self.forceTargetToShow()
        self.newRedrawMainPlot()
        self.checkForBothCorrectAndTargetPresent()

    def clearTargetSelections(self):
        for checkBox in self.targetCheckBoxes:
            checkBox.setChecked(False)
        # for yOffsetSpin in self.yOffsetSpinBoxes:
        #     yOffsetSpin.setEnabled(True)

    def noTargetSelected(self):
        for checkBox in self.targetCheckBoxes:
            if checkBox.isChecked():
                return False
        return True

    def processTargetSelection(self, i, redraw):

        self.targetKey = ''

        if self.targetCheckBoxes[i].isChecked():
            self.targetKey = self.lightcurveTitles[i].text()
            if not self.targetKey == '':
                self.showMsg(f'{self.targetKey} is the target curve.')
            self.clearTargetSelections()
            self.targetCheckBoxes[i].setChecked(True)
            self.showCheckBoxes[i].setChecked(True)
            self.yOffsetSpinBoxes[i].setEnabled(True)
            self.yOffsetSpinBoxes[i].setValue(0)
            if i == 0:
                self.yValues = self.LC1.copy()
            elif i == 1:
                self.yValues = self.LC2.copy()
            elif i == 2:
                self.yValues = self.LC3.copy()
            elif i == 3:
                self.yValues = self.LC4.copy()
            else:
                self.yValues = self.extra[i - 4].copy()
            if redraw:
                self.newRedrawMainPlot()
            self.recolorBlobs()
            self.VizieRdict = {
                "timestamps": None,
                "yValues": None,
                "yStatus": None,
            }
            self.timeDelta, self.droppedFrames, self.cadenceViolation, self.errRate, _ = \
                getTimeStepAndOutliers(self.yTimes, self.yValues, self.yStatus, VizieRdict=self.VizieRdict)
        else:
            if self.noTargetSelected():
                self.targetCheckBoxes[i].setChecked(True)
                self.showCheckBoxes[i].setChecked(True)
                self.yOffsetSpinBoxes[i].setEnabled(True)
                self.yOffsetSpinBoxes[i].setValue(0)

    def processTargetSelection1(self):
        self.processTargetSelection(0, redraw=True)

    def processTargetSelection2(self):
        self.processTargetSelection(1, redraw=True)

    def processTargetSelection3(self):
        self.processTargetSelection(2, redraw=True)

    def processTargetSelection4(self):
        self.processTargetSelection(3, redraw=True)

    def processTargetSelection5(self):
        self.processTargetSelection(4, redraw=True)

    def processTargetSelection6(self):
        self.processTargetSelection(5, redraw=True)

    def processTargetSelection7(self):
        self.processTargetSelection(6, redraw=True)

    def processTargetSelection8(self):
        self.processTargetSelection(7, redraw=True)

    def processTargetSelection9(self):
        self.processTargetSelection(8, redraw=True)

    def processTargetSelection10(self):
        self.processTargetSelection(9, redraw=True)

    def processTimeConstantChange(self):
        # We don't want to respond to Ne3 time constant changes until
        # the initial solution has been found, hence the following test
        # if self.solution is not None:
        #     self.showMsg('Called findEvent()')
        #     self.findEvent()
        pass

    def handlePymovieColumnChange(self):
        if self.pymovieFileInUse:
            self.externalCsvFilePath = self.csvFilePath
            self.readDataFromFile()

    def redoTabOrder(self, tabnames):

        def getIndexOfTabFromName(name):
            for i_local in range(self.tabWidget.count()):
                if self.tabWidget.tabText(i_local) == name:
                    return i_local
            return -1

        numTabs = self.tabWidget.count()
        if not len(tabnames) == numTabs:
            return

        for i in range(len(tabnames)):
            from_index = getIndexOfTabFromName(tabnames[i])
            to_index = i
            if from_index < 0:
                # self.showMsg(f'Could not locate {tabnames[i]} in the existing tabs')
                return
            else:
                self.tabWidget.tabBar().moveTab(from_index, to_index)

    def clearModelsBaselineRegion(self):
        self.baselineADUedit.setStyleSheet(None)
        self.baselineADUbutton.setStyleSheet(None)
        self.calcBaselineADUbutton.setStyleSheet(None)
        self.clearBaselineADUselectionButton.setStyleSheet(None)

        self.magDropEdit.setEnabled(True)
        self.magDropEdit.setFocus()
        if self.magDropEdit.text() == '':
            self.showInfo(f'Enter Predicted magDrop')

        self.clearBaselineDotsFromPlot()

    def clearBaselineRegions(self):
        self.clearBaselineRegionsButton.setEnabled(False)
        self.calcStatsFromBaselineRegionsButton.setEnabled(False)

        self.clearBaselineDotsFromPlot()

    def clearBaselineDotsFromPlot(self):
        if self.dataLen is None:
            return
        self.bkgndRegionLimits = []
        self.showMsg('Background regions cleared.')
        x = [i for i in range(self.dataLen) if self.yStatus[i] == BASELINE]
        for i in x:
            self.yStatus[i] = INCLUDED
        self.newRedrawMainPlot()

    def markEventRegion(self):
        if len(self.selectedPoints) == 0:
            x = [i for i in range(self.dataLen) if self.yStatus[i] == EVENT]
            for i in x:
                self.yStatus[i] = INCLUDED
            self.newRedrawMainPlot()
            # self.showInfo('No points were selected. Exactly two are required')
            return

        if len(self.selectedPoints) != 2:
            self.showInfo('Exactly two points must be selected for this operation.')
            return

        # Erase any previously selected event points
        x = [i for i in range(self.dataLen) if self.yStatus[i] == EVENT]
        for i in x:
            self.yStatus[i] = INCLUDED
        self.newRedrawMainPlot()

        selIndices = [key for key, _ in self.selectedPoints.items()]
        selIndices.sort()

        leftEdge = int(min(selIndices))
        rightEdge = int(max(selIndices))
        self.showMsg('Event region selected: ' + str([leftEdge, rightEdge]))

        # The following loop marks event points and removes the 2 red points
        # by overwriting then with the EVENT color
        self.selectedPoints = {}
        for i in range(leftEdge, rightEdge + 1):
            self.yStatus[i] = EVENT

        self.calcEventStatisticsFromMarkedRegion()

        self.newRedrawMainPlot()

    def markMetricRegion(self):
        # If the user has not selected any points, we ignore the request
        if len(self.selectedPoints) == 0:
            self.showInfo('No points were selected. Exactly two are required')
            return

        if len(self.selectedPoints) != 2:
            self.showInfo('Exactly two points must be selected for this operation.')
            return

        selIndices = [key for key, _ in self.selectedPoints.items()]
        selIndices.sort()

        if self.Lcp is not None:
            self.Lcp.set('metricLimitLeft', int(min(selIndices)))
            self.Lcp.set('metricLimitRight', int(max(selIndices)))
            self.showInfo(f'Fit metrics will be calculated using data values from readings\n\n'
                          f'{self.Lcp.metricLimitLeft} to {self.Lcp.metricLimitRight} inclusive.\n\n'
                          f'NOTE: You may want to save the event with this updated information.')
        else:
            self.showInfo('There is no event data structure defined yet.')

        # The following code removes the 2 red points by overwriting with the INCLUDED color
        self.selectedPoints = {}
        for i in range(self.left, self.right + 1):
            self.yStatus[i] = INCLUDED
        self.redrawMainPlot()

    def markBaselineRegion(self):
        # If the user has not selected any points, we ignore the request
        if len(self.selectedPoints) == 0:
            self.showInfo('No points were selected. Exactly two are required')
            return

        if len(self.selectedPoints) != 2:
            self.showInfo('Exactly two points must be selected for this operation.')
            return

        self.clearBaselineRegionsButton.setEnabled(True)
        self.calcStatsFromBaselineRegionsButton.setEnabled(True)

        selIndices = [key for key, _ in self.selectedPoints.items()]
        selIndices.sort()

        leftEdge = int(min(selIndices))
        rightEdge = int(max(selIndices))

        self.bkgndRegionLimits.append([leftEdge, rightEdge])

        # The following loop removes the 2 red points because they get overwritten by the BASELINE color
        self.selectedPoints = {}
        for i in range(leftEdge, rightEdge + 1):
            self.yStatus[i] = BASELINE

        self.showMsg('Background region selected: ' + str([leftEdge, rightEdge]))
        self.newRedrawMainPlot()

    def eventPointsMarked(self):
        y = [self.yValues[i] for i in range(self.dataLen) if self.yStatus[i] == EVENT]
        return len(y) > 2

    def baselinePointsMarked(self):
        y = [self.yValues[i] for i in range(self.dataLen) if self.yStatus[i] == BASELINE]
        return len(y) > 2

    def calcEventStatisticsFromMarkedRegion(self):
        y = [self.yValues[i] for i in range(self.dataLen) if self.yStatus[i] == EVENT]
        mean = np.mean(y)
        self.A = mean
        self.sigmaA = float(np.std(y))

        self.userDeterminedEventStats = True

        self.showMsg(f'mean event = {mean:0.2f}')
        self.showMsg(f'sigmaA: = {self.sigmaA:0.2f}')

    def calcBaselineStatisticsFromMarkedRegions(self):
        if self.dataLen is None:
            return
        xIndices = [i for i in range(self.dataLen) if self.yStatus[i] == BASELINE]
        y = [self.yValues[i] for i in range(self.dataLen) if self.yStatus[i] == BASELINE]
        mean = np.mean(y)

        raw_sigma = np.std(y)

        self.B = mean
        baselineXvals = []
        baselineYvals = []
        for i in xIndices:
            baselineXvals.append(i)
            baselineYvals.append(self.yValues[i])

        # Note: the getCorCoefs() routine uses a savgol linear filter of window 301 to remove trends during
        # the calculation of sigB
        self.newCorCoefs, self.numNApts, sigB = getCorCoefs(baselineXvals,
                                                            baselineYvals)

        if self.targetKey == '':
            self.targetKey = self.lightcurveTitles[0].text()

        self.showMsg(f'{self.targetKey} is the light curve undergoing "selected points" noise analysis.',
                     bold=True, blankLine=False)
        self.showMsg('...This noise analysis uses ' + str(self.numNApts) +
                     ' points')
        self.corCoefs = np.ndarray(shape=(len(self.newCorCoefs),))
        np.copyto(self.corCoefs, self.newCorCoefs)
        self.numPtsInCorCoefs = self.numNApts
        self.sigmaB = sigB
        if self.Lcp is not None:
            self.Lcp.set('sigmaB', self.sigmaB)

        self.userDeterminedBaselineStats = True

        self.prettyPrintCorCoefs()

        self.showMsg(f'statistics of selected points:', bold=True, blankLine=False)
        self.showMsg(f'...mean(selected points): {mean:0.2f}', blankLine=False)
        self.showMsg(f"...std(selected points) : {raw_sigma:0.3f}", blankLine=False)
        self.showMsg(f"...std(detrended points): {sigB:0.3f}  (Warning: this sets sigmaB for the remainder of this session)", blankLine=False)
        self.showMsg(f'...baseline snr: {mean / sigB:0.2f}  (using detrended std)')

    def getTimestampFromRdgNum(self, rdgNum):
        readingNumber = int(rdgNum)
        if readingNumber >= 0:
            if readingNumber < len(self.yTimes):
                tString = self.yTimes[readingNumber]
                if tString == '[]' or tString == '' or not self.showTimestampsCheckBox.isChecked():
                    return f'{readingNumber}'
                else:
                    return self.yTimes[readingNumber]
        return '???'

    def fillFromNAxlsxFile(self):
        # Open a file select dialog
        xlsxfilepath, _ = QFileDialog.getOpenFileName(
            self,  # parent
            "Select Asteroid Occultation Report form",  # title for dialog
            self.settings.value('lightcurvedir', ""),  # starting directory
            "Excel files (*.xlsx)")

        if xlsxfilepath:
            # noinspection PyBroadException

            # pcs added "data_only=True" to next line so it reads values, not equations
            # pcs this is required for the camera and VTI offset since it's a calculated
            # pcs quantity in the spreadsheet.  it does not appear pyote ever uses the
            # pcs equations, anyway.

            # wb = load_workbook(xlsxfilepath)
            wb = load_workbook(xlsxfilepath, data_only=True)

            try:
                sheet = wb['DATA']

                # pcs added next line so the camera and VTI correction can be read directly
                calcsheet = wb['Corrections Calculations']

                # Validate that a proper Asteroid Occultation Report Form was selected by reading the report header
                if not sheet['G1'].value == 'Asteroid Occultation Report Form':
                    self.showMsg(f'The xlsx file selected does not appear to be an Asteroid Occultation Report Form')
                    return

                Year = 'D5'
                year = sheet[Year].internal_value
                self.vzDateYearSpinner.setValue(year)

                Month = 'K5'
                monthStr = sheet[Month].internal_value
                months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August',
                          'September', 'October', 'November', 'December']
                month = None
                for i in range(len(months)):
                    if monthStr == months[i]:
                        month = i + 1
                        break
                if month is not None:
                    self.vzDateMonthSpinner.setValue(month)
                else:
                    raise Exception('Invalid month string in xlsx file')

                Day = 'P5'
                day = sheet[Day].internal_value
                self.vzDateDaySpinner.setValue(day)

                Longitude = 'N18'
                LongitudeEW = 'R18'
                longitudeStr = sheet[Longitude].internal_value
                longitudeEW = sheet[LongitudeEW].internal_value
                # print(f'Longitude: {longitudeStr} {longitudeEW}')

                longParts = longitudeStr.split(' ')
                if longitudeEW == 'W':
                    self.vzSiteLongDegEdit.setText(f'-{longParts[0]}')
                else:
                    self.vzSiteLongDegEdit.setText(f'+{longParts[0]}')

                self.vzSiteLongMinEdit.setText(longParts[1])
                self.vzSiteLongSecsEdit.setText(longParts[2])

                Latitude = 'E18'
                LatitudeNS = 'J18'
                latitudeStr = sheet[Latitude].internal_value
                latitudeNS = sheet[LatitudeNS].internal_value
                # print(f'Longitude: {latitudeStr} {latitudeNS}')

                latParts = latitudeStr.split(' ')
                if latitudeNS == 'S':
                    self.vzSiteLatDegEdit.setText(f'-{latParts[0]}')
                else:
                    self.vzSiteLatDegEdit.setText(f'+{latParts[0]}')

                self.vzSiteLatMinEdit.setText(latParts[1])
                self.vzSiteLatSecsEdit.setText(latParts[2])

                Altitude = 'V18'
                AltitudeUnits = 'W18'
                altitude = sheet[Altitude].internal_value
                altitudeUnits = sheet[AltitudeUnits].internal_value
                if altitudeUnits == 'm':
                    self.vzSiteAltitudeEdit.setText(f'{altitude}')
                else:
                    # pcs modified the next line to round to nearest integer in meters.  this is
                    # pcs needed because the code tries to convert the resulting altitude from a
                    # pcs string to an int and this fails if the string contains a decimal.  that
                    # pcs line also uses "int" which would truncate instead of rounding.  so this
                    # pcs fixes that, too.

                    # self.vzSiteAltitudeEdit.setText(f'{altitude * 0.3048}')
                    self.vzSiteAltitudeEdit.setText(f'{round(altitude * 0.3048)}') # pcs

                # print(f'Altitude: {altitude} {altitudeUnits}')

                Observer = 'D9'
                observer = sheet[Observer].internal_value
                self.vzObserverNameEdit.setText(observer)
                # print(f'Observer: {observer}')

                StarType = 'S7'
                StarNumber = 'X7'
                starType = sheet[StarType].internal_value
                starNumber = sheet[StarNumber].internal_value
                if type(starNumber) is int:
                    starNumber = f'{starNumber}'
                if starType.startswith('TYC'):      # pcs
                    self.vzStarTycho2Edit.setText(starNumber)
                elif starType.startswith('HIP'):    # pcs
                    self.vzStarHipparcosEdit.setText(starNumber)
                elif starType.startswith('UCAC4'):  # pcs
                    self.vzStarUCAC4Edit.setText(starNumber)

                # print(f'Star id: {starType} {starNumber}')

                AsteroidNumber = 'E7'
                AsteroidName = 'K7'
                asteroidNumber = sheet[AsteroidNumber].internal_value
                asteroidName = sheet[AsteroidName].internal_value
                self.vzAsteroidNameEdit.setText(asteroidName)
                self.vzAsteroidNumberEdit.setText(f'{asteroidNumber}')
                # print(f'{asteroidName}({asteroidNumber})')

                # pcs added next block to calculate camera and VTI correction.  for a camera
                # pcs with a global shutter, the D and R corrections are the same so the
                # pcs average will match either one.  for a camera with a rolling shutter,
                # pcs though, the two corrections can be different.  if this is a drift scan,
                # pcs the difference will vary linearly thoughout the entire video so, in that
                # pcs case, it would be more correct to linearly interpolate the correction to
                # pcs each time point.  essentially, that will amount to a slightly different
                # pcs event duration than what's calculated the normal way.  whether the event
                # pcs is from a drift scan could be read from the spreadsheet as well.  if this
                # pcs is a rolling shutter on a tracked scope, there should be no motion so the
                # pcs the D and R corrections would match and the average would be adequate.
                # pcs there could be some motion, though, from things like periodic error.  for
                # pcs short events, linear interpolation could be used as for drift scans but
                # pcs it's probably safer to simply use the average D and R correction.  if it
                # pcs was a long event and the motion was large, the target star could oscillate
                # pcs back and forth.  to accurately correct for that, the adjusted timesteps
                # pcs wouldn't be constant and thus couldn't be represented by the required
                # pcs output format.  so taking the average is probably the only alternative
                # pcs as linear would be incorrect in this case.  for a tracked scope, the errors
                # pcs due to ignoring the small motion across a few rows would be tiny and not
                # pcs worth trying to mmmodel so averaging the D and R corrections is all that
                # pcs really makes sense.  in fact, using the average for a drift scan is
                # pcs possibly the safest option there as well.  if an event is very short,
                # pcs scintillation or wind shake could cause the two rows to be quite
                # pcs different leading to "large", nonphysical corrections.  for the short
                # pcs subsets of the events used for the vizier files, the row correction is
                # pcs probably insignificant, even for the drift scan case.  thus the average
                # pcs is used in these mods for all cases.
                # pcs note that this does not check to see if there is both a D and R specified
                # pcs in the spreadsheet.  if one is missing, both corrections will still be
                # pcs calculated correctly for a camera with a global shutter.  if the camera
                # pcs has a rolling shutter, however, and thus the row number is not specified
                # pcs for for the D or R, a zero will be used in the calculated of either L45
                # pcs or L47.  thus, averaging the two corrections will result in a somewhat
                # pcs wrong value.  so this should really include a check for valid D and R
                # pcs values and ignore the other correction.

                DCamCorrection = 'L45'
                dCamCorrection = calcsheet[DCamCorrection].internal_value
                RCamCorrection = 'L47'
                rCamCorrection = calcsheet[RCamCorrection].internal_value
                self.camCorrection = 0.5 * (dCamCorrection + rCamCorrection)

                self.showMsg(
                    f'{self.camCorrection} second camera and VTI correction will be added to the output timestamp')

            # pcs end of the added block

            except Exception as e:
                self.showMsg(repr(e))
                self.showMsg(f'FAILED to extract needed data from Asteroid Occultation Report Form', color='red', bold=True)
                return

            self.showInfo(f'Excel spreadsheet Asteroid Report Form entries extracted successfully.')

    def clearFitMetricTxtFile(self):
        if self.csvFilePath is None:
            self.showInfo(f'A light curve has not been selected yet\n\n'
                          f'so there is no fit metrics file to remove.')
            return

        lightCurveDir = os.path.dirname(self.csvFilePath)  # This gets the folder where the light-curve.csv is located
        metric_file_path = os.path.join(lightCurveDir, 'fit_metrics.txt')
        if os.path.exists(metric_file_path):
            os.remove(metric_file_path)
            self.showInfo(f'The fit metrics file has been restarted.')
            self.addSourceFileToFitMetricTxtFile()

    def renameFitMetricFile(self):
        if self.csvFilePath is None:
            self.showInfo('No lightcurve file selected yet.')
            return

        lightCurveDir = os.path.dirname(self.csvFilePath)
        current_metric_filepath = os.path.join(lightCurveDir, 'fit_metrics.txt')

        if not os.path.exists(current_metric_filepath):
            self.showInfo(f'There is no metrics file yet.')
            return

        inp = QtWidgets.QInputDialog(self)
        inp.setInputMode(QtWidgets.QInputDialog.InputMode.TextInput)
        inp.setFixedSize(400, 200)
        p = inp.palette()
        p.setColor(inp.backgroundRole(), QtCore.Qt.GlobalColor.gray)
        inp.setPalette(p)
        inp.setWindowTitle('Rename fit metrics')
        inp.setLabelText('Enter name to be used for the .xlsx file: ')

        lightCurveDir = os.path.dirname(self.csvFilePath)
        head, tail = os.path.split(lightCurveDir)

        inp.setTextValue(tail)

        if inp.exec_() == QtWidgets.QDialog.DialogCode.Accepted:
            # This gets the folder where the light-curve.csv is located
            lightCurveDir = os.path.dirname(self.csvFilePath)
            current_metric_filepath = os.path.join(lightCurveDir, 'fit_metrics.txt')
            name_given = inp.textValue()
            if '.xlsx' in name_given:
                new_metric_filepath = os.path.join(lightCurveDir, f'{name_given}')
            else:
                # new_metric_filepath = lightCurveDir + f'\\{name_given}.xlsx'
                new_metric_filepath = os.path.join(lightCurveDir, f'{name_given}.xlsx' )

            # Make a Pandas data frame from the csv file
            df = pd.read_csv(current_metric_filepath)

            # Create a Pandas Excel writer using xlswriter as the engine
            writer = pd.ExcelWriter(new_metric_filepath, engine='xlsxwriter')

            # Convert the dataframe to an Xlsxwriter Excel object
            df.to_excel(writer, sheet_name='Sheet1', index=False)

            # Get the xlsxwriter workbook and workbook objects
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Create some cell formats
            formatNum4 = workbook.add_format({"num_format": "0.0000"})
            # formatNum3 = workbook.add_format({"num_format": "0.000"})
            formatNum2 = workbook.add_format({"num_format": "0.00"})
            formatNum1 = workbook.add_format({"num_format": "0.0"})

            # Set column widths and formats
            # worksheet.set_column(first_col, last_col, width, format)  width can be None and format can be missing
            worksheet.set_column( 0,  0, 30)                  # aperture name
            worksheet.set_column( 1,  1, 15, formatNum4)      # time err
            worksheet.set_column( 2,  2, None, formatNum2)    # DNR
            worksheet.set_column( 3,  3, None, formatNum2)    # FP metric
            worksheet.set_column( 4,  4, None, formatNum2)    # magDrop
            worksheet.set_column( 5,  5, 13, formatNum1)      # percent drop
            worksheet.set_column( 6,  6, 15, formatNum4)      # duration
            worksheet.set_column( 7,  8, 13)                  # D time  and R time
            worksheet.set_column( 9, 10, None, formatNum2)    # D frame  and R frame
            worksheet.set_column(11, 12, 13, formatNum2)      # D and A
            worksheet.set_column(13, 14, 13, formatNum2)      # sigmaB and sigmaA
            worksheet.set_column(15, 15, 15, formatNum1)      # observed drop
            worksheet.set_column(16, 16, None, formatNum1)    # observed drop
            worksheet.set_column(17, 17, None, formatNum1)    # FP margin
            writer.close()

        inp.deleteLater()

    def fillExcelReport(self):
        # Open a file select dialog
        xlsxfilepath, _ = QFileDialog.getOpenFileName(
            self,  # parent
            "Select Asteroid Occultation Report form",  # title for dialog
            self.settings.value('lightcurvedir', ""),  # starting directory
            "Excel files (*.xlsx)")

        if xlsxfilepath:
            # noinspection PyBroadException
            wb = load_workbook(xlsxfilepath)

            try:
                sheet = wb['DATA']

                # Validate that a proper Asteroid Occultation Report Form was selected by reading the report header
                if not sheet['G1'].value == 'Asteroid Occultation Report Form':
                    self.showMsg(f'The xlsx file selected does not appear to be an Asteroid Occultation Report Form')
                    return

                # We're going to ignore the named cell info and reference all the cells of interest in
                # their col/row coordinates (not all the cells of interest had names)

                Derr68 = 'L33'
                Derr95 = 'M33'
                Derr99 = 'N33'

                Rerr68 = 'L35'
                Rerr95 = 'M35'
                Rerr99 = 'N35'

                Dhour = 'F32'
                Dmin = 'H32'
                Dsec = 'J32'

                Rhour = 'F36'
                Rmin = 'H36'
                Rsec = 'J36'

                # Exposure = 'P25'
                OTA = 'O23'
                SNR = 'W40'
                Comment = 'D43'

                sheet[OTA].value = 'PYOTE'
                sheet[SNR].value = f'{self.snrB:0.2f}'

                if 'Comment' in self.xlsxDict:
                    sheet[Comment].value = self.xlsxDict['Comment']

                if 'Derr68' in self.xlsxDict:
                    # sheet[Derr68].value = f'{self.xlsxDict["Derr68"]:0.2f}'
                    sheet[Derr68].value = self.xlsxDict["Derr68"]

                if 'Derr95' in self.xlsxDict:
                    # sheet[Derr95].value = f'{self.xlsxDict["Derr95"]:0.2f}'
                    sheet[Derr95].value = self.xlsxDict["Derr95"]

                if 'Derr99' in self.xlsxDict:
                    # sheet[Derr99].value = f'{self.xlsxDict["Derr99"]:0.2f}'
                    sheet[Derr99].value = self.xlsxDict["Derr99"]

                if 'Rerr68' in self.xlsxDict:
                    # sheet[Rerr68].value = f'{self.xlsxDict["Rerr68"]:0.2f}'
                    sheet[Rerr68].value = self.xlsxDict["Rerr68"]

                if 'Rerr95' in self.xlsxDict:
                    # sheet[Rerr95].value = f'{self.xlsxDict["Rerr95"]:0.2f}'
                    sheet[Rerr95].value = self.xlsxDict["Rerr95"]

                if 'Rerr99' in self.xlsxDict:
                    # sheet[Rerr99].value = f'{self.xlsxDict["Rerr99"]:0.2f}'
                    sheet[Rerr99].value = self.xlsxDict["Rerr99"]

                if 'Dhour' in self.xlsxDict:
                    sheet[Dhour] = int(self.xlsxDict['Dhour'])  # noqa  expected str
                if 'Dmin' in self.xlsxDict:
                    sheet[Dmin] = int(self.xlsxDict['Dmin'])    # noqa  expected str
                if 'Dsec' in self.xlsxDict:
                    sheet[Dsec] = float(self.xlsxDict['Dsec'])  # noqa  expected str

                if 'Rhour' in self.xlsxDict:
                    sheet[Rhour] = int(self.xlsxDict['Rhour'])  # noqa  expected str
                if 'Rmin' in self.xlsxDict:
                    sheet[Rmin] = int(self.xlsxDict['Rmin'])    # noqa  expected str
                if 'Rsec' in self.xlsxDict:
                    sheet[Rsec] = float(self.xlsxDict['Rsec'])  # noqa  expected str

                # Overwriting the original file !!!
                wb.save(xlsxfilepath)

            except Exception as e:
                self.showMsg(repr(e))
                self.showMsg(f'FAILED to fill Asteroid Occultation Report Form', color='red', bold=True)
                self.showMsg(f'Is it possible that you have the file already open somewhere?', color='red', bold=True)
                return

            self.showMsg(f'Excel spreadsheet Asteroid Report Form entries made successfully.')

            # noinspection PyBroadException
            try:
                if sys.platform == 'darwin' or sys.platform == 'linux':
                    subprocess.call(['open', xlsxfilepath])
                elif platform.system() == 'Windows':
                    os.startfile(xlsxfilepath)
                else:
                    subprocess.call(['xdg-open', xlsxfilepath])
            except Exception as e:
                self.showMsg('Attempt to get host OS to open xlsx file failed.', color='red', bold=True)
                self.showMsg(repr(e))
        else:
            return

    def findTimestampFromFrameNumber(self, frame):
        # Currently PyMovie uses nn.00 for frame number
        # Limovie uses nn.0 for frame number
        # We use the 'starts with' flag so that we pick up both forms
        items = self.table.findItems(f'{frame:0.1f}', QtCore.Qt.MatchFlag.MatchStartsWith)
        for item in items:
            if item.column() == 0:  # Avoid a possible match from a data column
                ts = self.table.item(item.row(), 1).text()
                return ts
        return ''

    def showAnnotatedFrame(self, frame_to_show, annotation):

        frame_number = frame_to_show

        table_timestamp = self.findTimestampFromFrameNumber(frame_to_show)

        if not table_timestamp:
            table_timestamp = 'no timestamp found'

        if self.pathToVideo is None:
            return

        _, ext = os.path.splitext(self.pathToVideo)

        if ext == '.avi':
            ans = readAviFile(frame_number, full_file_path=self.pathToVideo)
            if not ans['success']:
                self.showMsg(f'Attempt to view frame returned errmsg: {ans["errmsg"]}')
                return
        elif ext == '.ser':
            ans = readSerFile(frame_number, full_file_path=self.pathToVideo)
            if not ans['success']:
                self.showMsg(f'Attempt to view frame returned errmsg: {ans["errmsg"]}')
                return
        elif ext == '':
            # We assume it's a FITS folder that we have been given
            ans = readFitsFile(frame_number, full_file_path=self.pathToVideo)
            if not ans['success']:
                self.showMsg(f'Attempt to view frame returned errmsg: {ans["errmsg"]}')
                return
        elif ext == '.aav':
            ans = readAavFile(frame_number, full_file_path=self.pathToVideo)
            if not ans['success']:
                self.showMsg(f'Attempt to view frame returned errmsg: {ans["errmsg"]}')
                return
        else:
            self.showMsg(f'Unsupported file extension: {ext}')
            return

        # Check to see if user has closed all frame views
        frame_visible = False
        for frame_view in self.frameViews:
            if frame_view and frame_view.isVisible():
                frame_visible = True

        if not frame_visible:
            self.cascadePosition = 100

        title = f'{annotation} {table_timestamp} @ frame {frame_number}'
        self.frameViews.append(pg.GraphicsWindow(title=title))

        cascade_origin = self.pos() + QPoint(self.cascadePosition, self.cascadePosition)

        self.frameViews[-1].move(cascade_origin)
        self.cascadePosition += self.cascadeDelta

        self.frameViews[-1].resize(1000, 600)
        layout = QtWidgets.QGridLayout()
        self.frameViews[-1].setLayout(layout)
        imv = pg.ImageView()
        layout.addWidget(imv, 0, 0)

        imv.ui.menuBtn.hide()
        imv.ui.roiBtn.hide()

        image = ans['image']

        if self.fieldViewCheckBox.isChecked():
            upper_field = image[0::2, :]
            lower_field = image[1::2, :]
            image = np.concatenate((upper_field, lower_field))

        if self.flipYaxisCheckBox.isChecked():
            image = np.flipud(image)

        if self.flipXaxisCheckBox.isChecked():
            image = np.fliplr(image)

        imv.setImage(image)

        for i, frame_view in enumerate(self.frameViews):
            if frame_view and not frame_view.isVisible():
                # User has closed the image.  Remove it so that garbage collection occurs.
                self.frameViews[i].close()
                self.frameViews[i] = None
            else:
                if frame_view:
                    frame_view.raise_()

    def viewFrame(self):
        if self.pathToVideo is None:
            return

        frame_to_show = self.frameNumSpinBox.value()
        self.showAnnotatedFrame(frame_to_show=frame_to_show, annotation='User selected frame:')

    def helpButtonClicked(self):
        self.showHelp(self.helpButton)

    def vzInfoClicked(self):
        self.showHelp(self.vzInfoButton)

    def vzWhereClicked(self):
        self.showHelp(self.vzWhereToSendButton)

    def helpPdfButtonClicked(self):
        self.openModelHelpFile()

    def ne3ExplanationClicked(self):
        self.showHelp(self.ne3ExplanationButton)

    def tutorialButtonClicked(self):
        self.showHelp(self.tutorialButton)

    def lightcurvesHelpButtonClicked(self):
        self.showHelp(self.lightcurvesHelpButton)

    def plotHelpButtonClicked(self):
        self.showHelp(self.plotHelpButton)

    def sqWaveHelpButtonClicked(self):
        self.showHelp(self.helpSqWaveButton)

    def detectabilityHelpButtonClicked(self):
        self.showHelp(self.detectabilityHelpButton)

    @staticmethod
    def htmlFixup(html):
        output = ''
        endIndex = len(html) - 1
        for i in range(len(html)):
            if not (html[i] == '.' or html[i] == ','):
                output += html[i]
            else:
                if i == endIndex:
                    output += html[i]
                    return output
                if html[i + 1] == ' ':
                    output += html[i] + '&nbsp;'
                else:
                    output += html[i]
        return output

    def showHelp(self, obj):

        if obj.toolTip():
            self.helperThing.raise_()
            self.helperThing.show()
            self.helperThing.textEdit.clear()
            stuffToShow = self.htmlFixup(obj.toolTip())
            self.helperThing.textEdit.insertHtml(stuffToShow)
            self.helperThing.setHidden(True)
            self.helperThing.setVisible(True)

    @staticmethod
    def processKeystroke(event):
        _ = event.key()  # Just to satisfy PEP8
        return False

    def eventFilter(self, obj, event):
        if self.allowNewVersionPopupCheckbox.isChecked():
            self.helperThing.raise_()
        if event.type() == QtCore.QEvent.Type.KeyPress:
            handled = self.processKeystroke(event)
            if handled:
                return True
            else:
                return super(SimplePlot, self).eventFilter(obj, event)

        if event.type() == QtCore.QEvent.Type.MouseButtonPress:
            if event.button() == QtCore.Qt.MouseButton.RightButton:
                if obj.toolTip():
                    self.helperThing.raise_()
                    self.helperThing.show()
                    self.helperThing.textEdit.clear()
                    stuffToShow = self.htmlFixup(obj.toolTip())
                    self.helperThing.textEdit.insertHtml(stuffToShow)
                    return True
            return super(SimplePlot, self).eventFilter(obj, event)

        if event.type() == QtCore.QEvent.Type.ToolTip:
            return True

        return super(SimplePlot, self).eventFilter(obj, event)

    def writeCSVfile(self):
        _, name = os.path.split(self.csvFilePath)
        name = self.removeCsvExtension(name)

        name += '.pyote.csv'

        myOptions = QFileDialog.Options()
        # myOptions |= QFileDialog.DontConfirmOverwrite
        myOptions |= QFileDialog.DontUseNativeDialog
        # myOptions |= QFileDialog.ShowDirsOnly

        starterFilePath = str(Path(self.settings.value('lightcurvedir', "") + '/' + name))

        self.csvFile, _ = QFileDialog.getSaveFileName(
            self,  # parent
            "Select directory/modify filename",  # title for dialog
            starterFilePath,  # starting directory
            "csv files (*.csv)", options=myOptions)

        if self.csvFile:
            if not self.csvFile.endswith('.csv'):
                self.csvFile += '.csv'
            with open(self.csvFile, 'w') as fileObject:
                if not self.aperture_names:
                    fileObject.write('# ' + 'PYOTE ' + version.version() + '\n')
                else:
                    fileObject.write('# PyMovie file written by ' + 'PYOTE ' + version.version() + '\n')

                for hdr in self.headers:
                    fileObject.write(f'#  {hdr}\n')
                if not self.aperture_names:
                    # Handle non-PyMovie csv file
                    columnHeadings = 'FrameNum,timeInfo,primaryData'
                    if len(self.LC2) > 0:
                        columnHeadings += ',LC2'
                    if len(self.LC3) > 0:
                        columnHeadings += ',LC3'
                    if len(self.LC4) > 0:
                        columnHeadings += ',LC4'
                else:
                    columnHeadings = 'FrameNum,timeInfo'
                    column_number = 2
                    while True:
                        column_header_item = self.table.horizontalHeaderItem(column_number)
                        if column_header_item is None:
                            break
                        else:
                            column_name = column_header_item.text()
                            columnHeadings += f',{column_name}'
                            column_number += 1

                if self.modelYsamples is not None:
                    columnHeadings += f',signal-model'

                fileObject.write(columnHeadings + '\n')

                k = None
                for i in range(self.table.rowCount()):  # i is reading number (0-based row in output table)
                    if self.left <= i <= self.right:
                        line = self.table.item(i, 0).text()
                        for j in range(1, self.table.columnCount()):
                            # Deal with empty columns
                            if self.table.item(i, j) is not None:
                                line += ',' + self.table.item(i, j).text()
                        if self.modelYsamples is not None:
                            if i == self.modelXsamples[0]:
                                k = 0  # Start writing model Y values
                            if k is not None:
                                line += ',' + f'{self.modelYsamples[k]:0.3f}'
                                k += 1
                                if k >= len(self.modelYsamples):
                                    k = None  # Terminate writing of model values
                            else:
                                if i < self.modelXsamples[0]:
                                    line += ',' + f'{self.modelYsamples[0]:0.3f}'
                                elif i > self.modelXsamples[-1]:
                                    line += ',' + f'{self.modelYsamples[-1]:0.3f}'
                                else:
                                    self.showInfo(f'We should not be here in writeCSVfile()')

                        fileObject.write(line + '\n')

    @staticmethod
    def getTimestampString(timeValue):
        hours = int(timeValue / 3600)
        timeValue -= hours * 3600
        minutes = int(timeValue / 60)
        timeValue -= minutes * 60
        return f'{hours:02d}:{minutes:02d}:{timeValue:0.4f}'

    def writeExampleLightcurveToFile(self, lgtCurve, timeDelta):
        _, name = os.path.split(self.csvFilePath)
        name = self.removeCsvExtension(name)

        name += 'pyote.example-lightcurve.csv'

        myOptions = QFileDialog.Options()
        # myOptions |= QFileDialog.DontConfirmOverwrite
        myOptions |= QFileDialog.DontUseNativeDialog
        myOptions |= QFileDialog.ShowDirsOnly

        starterFilePath = str(Path(self.settings.value('lightcurvedir', "") + '/' + name))

        self.csvFile, _ = QFileDialog.getSaveFileName(
            self,  # parent
            "Select directory/modify filename",  # title for dialog
            starterFilePath,  # starting directory
            "", options=myOptions)

        if self.csvFile:
            with open(self.csvFile, 'w') as fileObject:
                fileObject.write('# ' + 'PYOTE ' + version.version() + '\n')

                fileObject.write(f'#  Example lightcurve from detectability analysis\n')
                columnHeadings = 'FrameNum,timeInfo,primaryData'
                fileObject.write(columnHeadings + '\n')

                readingTime = 0.0
                for i in range(len(lgtCurve)):
                    ts = self.getTimestampString(readingTime)
                    line = f'{i},[{ts}],{lgtCurve[i]:0.2f}'
                    fileObject.write(line + '\n')
                    readingTime += timeDelta

    def copy_modelExamples_to_Documents(self):
        # Added in 5.0.3  Removed in 5.4.3
        # if not self.allowNewVersionPopupCheckbox.isChecked():
        #     return

        # The model-examples folder is part of the distribution and should be in the home directory
        source_dir = os.path.join(self.homeDir, 'model-examples')
        if os.path.exists(source_dir):
            if sys.platform == 'darwin' or sys.platform == 'linux':
                dest_dir = f"{os.environ['HOME']}{r'/Documents/model-examples'}"
            else:
                # We must be on a Windows machine because Mac version number was empty
                dest_dir = f"{os.environ.get('userprofile')}\\Documents\\model-examples"

            if os.path.exists(dest_dir):
                self.showMsg(f'We found {dest_dir} already present. Adding any new examples ...',
                             color='black', bold=True, blankLine=False)
                distExamples = [os.path.basename(file) for file in glob.glob(f'{source_dir}\\*.csv')]
                desktopExamples = [os.path.basename(file) for file in glob.glob(f'{dest_dir}\\*.csv')]
                for filename in distExamples:
                    if filename in desktopExamples:
                        pass
                    else:
                        self.showMsg(f'... adding {filename}',
                                     color='black', bold=True, blankLine=False)
                        source = f'{source_dir}\\{filename}'
                        destination = f'{dest_dir}\\{filename}'
                        shutil.copy(source, destination)

            else:
                # We write the entire folder to Documents (i.e., create the folder)
                shutil.copytree(source_dir, dest_dir)
                self.showMsg(f'We have copied the example csv files from the distribution '
                             f'into {dest_dir} (useful for training purposes).',
                             color='black', bold=True)
        else:
            self.showMsg(f'We could not find model-examples folder in the distribution', color='red', bold=True)

    @staticmethod
    def copy_desktop_icon_file_to_home_directory():
        if sys.platform == 'darwin' or sys.platform == 'linux':
            icon_dest_path = f"{os.environ['HOME']}{r'/Desktop/run-pyote'}"
            if not os.path.exists(icon_dest_path):
                # Here is where the .bat file will be when running an installed pyote
                icon_src_path = f"{os.environ['HOME']}" + r"/Anaconda3/Lib/site-packages/pyoteapp/run-pyote-mac.bat"
                if not os.path.exists(icon_src_path):
                    # But here is where the .bat file is during a development run
                    icon_src_path = os.path.join(os.path.split(__file__)[0], 'run-pyote-mac.bat')
                with open(icon_src_path) as src, open(icon_dest_path, 'w') as dest:
                    dest.writelines(src.readlines())
                os.chmod(icon_dest_path, 0o755)  # Make it executable
        else:
            # We must be on a Windows machine because Mac version number was empty
            icon_dest_path = r"C:\Anaconda3\PYOTE.bat"

            if not os.path.exists(icon_dest_path):
                # Here is where the .bat file will be when running an installed pyote
                icon_src_path = r"C:\Anaconda3\Lib\site-packages\pyoteapp\PYOTE.bat"
                if not os.path.exists(icon_src_path):
                    # But here is where the .bat file is during a development run
                    icon_src_path = os.path.join(os.path.split(__file__)[0], 'PYOTE.bat')
                with open(icon_src_path) as src, open(icon_dest_path, 'w') as dest:
                    dest.writelines(src.readlines())

    # def toggleManualEntryButton(self):
    #     if self.manualTimestampCheckBox.isChecked():
    #         self.manualEntryPushButton.setEnabled(True)
    #         self.doManualTimestampEntry()
    #     else:
    #         self.manualEntryPushButton.setEnabled(False)

    def openHelpFile(self):
        helpFilePath = os.path.join(os.path.split(__file__)[0], 'pyote-info.pdf')

        url = QtCore.QUrl.fromLocalFile(helpFilePath)
        fileOpened = QtGui.QDesktopServices.openUrl(url)

        if not fileOpened:
            self.showMsg('Failed to open pyote-info.pdf', bold=True, color='red', blankLine=False)
            self.showMsg('Location of pyote information file: ' + helpFilePath, bold=True, color='blue')

    def openModelHelpFile(self):
        helpFilePath = os.path.join(os.path.split(__file__)[0], 'model-help.pdf')

        url = QtCore.QUrl.fromLocalFile(helpFilePath)
        fileOpened = QtGui.QDesktopServices.openUrl(url)

        if not fileOpened:
            self.showMsg('Failed to open model-help.pdf', bold=True, color='red', blankLine=False)
            self.showMsg('Location of model-help.pdf file: ' + helpFilePath, bold=True, color='blue')

    def mouseEvent(self):

        if not self.blankCursor:
            # self.showMsg('Mouse event')
            self.blankCursor = True
            self.mainPlot.viewport().setProperty("cursor", QtGui.QCursor(QtCore.Qt.CursorShape.BlankCursor))

    def keyPressEvent(self, ev):
        if ev.key() == QtCore.Qt.Key.Key_Shift:
            if self.blankCursor:
                self.mainPlot.viewport().setProperty("cursor", QtGui.QCursor(QtCore.Qt.CursorShape.ArrowCursor))
                self.blankCursor = False
            else:
                self.blankCursor = True
                self.mainPlot.viewport().setProperty("cursor", QtGui.QCursor(QtCore.Qt.CursorShape.BlankCursor))

    @staticmethod
    def timestampListIsEmpty(alist):
        ans = True
        for item in alist:
            # Limovie = '[::]'   Tangra = ''  R-OTE = '[NA]' or 'NA'
            if item == '' or item == '[::]' or item == '[NA]' or item == 'NA':
                pass
            else:
                ans = False
                break
        return ans

    def getNumberOfUnnamedLightCurves(self):
        if len(self.LC1) == 0:
            return 0
        elif len(self.LC2) == 0:
            return 1
        elif len(self.LC3) == 0:
            return 2
        elif len(self.LC4) == 0:
            return 3
        elif len(self.extra) == 0:
            return 4
        else:
            return 4 + len(self.extra)

    def checkForNewVersion(self):
        latestVersion = getLatestPackageVersion('pyote')

        self.showMsg(f'Query to PyPI returned latest version of PyOTE as: {latestVersion}',
                     color='blue', bold=True)

        if latestVersion.startswith("none"):  # 'none' is returned when no Internet connection
            gotVersion = False
        else:
            gotVersion = True if len(latestVersion) > 2 else False

        if not gotVersion:
            self.showMsg(f"Failed to connect to PyPI. Possible Internet connection problem ??",
                         color='red', bold=True)
            return

        if gotVersion:
            if latestVersion <= version.version():
                self.showMsg('You are running the most recent version of PyOTE', color='red', bold=True)

            else:
                self.showMsg('Version ' + latestVersion + ' is available.  To get it:',
                             color='red', bold=True)
                self.showMsg(
                    f"==== for pip based installations, in a command window type: pip install pyote=={latestVersion} (note double = symbols)",
                    color='red', bold=True)
                # self.showMsg(
                #     f"==== for pipenv based installations, execute the ChangePyoteVersion.bat file.",
                #     color='red', bold=True)
        else:
            self.showMsg(f'latestVersion found: {latestVersion}')

    # @staticmethod
    # def queryWhetherNewVersionShouldBeInstalled():
    #     msg = QMessageBox()
    #     msg.setIcon(QMessageBox.Question)
    #     msg.setText('A newer version of PyOTE is available. Do you wish to install it?')
    #     msg.setWindowTitle('Get latest version of PyOTE query')
    #     msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    #     retval = msg.exec_()
    #     return retval

    @staticmethod
    def queryWhetherBlockIntegrationShouldBeAccepted():
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Question)
        msg.setText(
            'Do you want the pyote estimation of block integration parameters to be used'
            ' for block integration?')
        msg.setWindowTitle('Is auto-determined block integration ok')
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        retval = msg.exec_()
        return retval

    def reportMouseMoved(self, pos):
        mousePoint = self.mainPlotViewBox.mapSceneToView(pos)
        self.verticalCursor.setPos(round(mousePoint.x()))

    def writeDefaultGraphicsPlots(self, error_bar_plots_available=True):
        if not error_bar_plots_available:
            return

        lightCurveDir = os.path.dirname(self.csvFilePath)  # This gets the folder where the light-curve.csv is located
        self.graphicFile = os.path.join(lightCurveDir, 'FP-plots')
        if not os.path.exists(self.graphicFile):
            os.mkdir(self.graphicFile)
        head, tail = os.path.split(self.csvFilePath)
        name_to_use, _ = os.path.splitext(tail)

        self.graphicFile = os.path.join(self.graphicFile, name_to_use)

        exporter = FixedImageExporter(self.dBarPlotItem)
        exporter.makeWidthHeightInts()
        targetFileD = self.graphicFile + '.D_pyote.png'
        exporter.export(targetFileD)

        exporter = FixedImageExporter(self.durBarPlotItem)
        exporter.makeWidthHeightInts()
        targetFileDur = self.graphicFile + '.R-D_pyote.png'
        exporter.export(targetFileDur)

        exporter = FixedImageExporter(self.falsePositivePlotItem)
        exporter.makeWidthHeightInts()
        targetFileDur = self.graphicFile + '.noise_induce_event_pyote.png'
        exporter.export(targetFileDur)

        exporter = FixedImageExporter(self.mainPlot.getPlotItem())
        exporter.makeWidthHeightInts()
        targetFile = self.graphicFile + '.pyote.png'
        exporter.export(targetFile)

    def exportBarPlots(self):
        if self.dBarPlotItem is None:
            self.showInfo('No error bar plots available yet')
            return

        _, name = os.path.split(self.csvFilePath)
        name = self.removeCsvExtension(name)

        myOptions = QFileDialog.Options()
        myOptions |= QFileDialog.DontConfirmOverwrite
        myOptions |= QFileDialog.DontUseNativeDialog
        myOptions |= QFileDialog.ShowDirsOnly

        self.graphicFile, _ = QFileDialog.getSaveFileName(
            self,  # parent
            "Select directory/modify filename (png will be appended for you)",  # title for dialog
            self.settings.value('lightcurvedir', "") + '/' + name,  # starting directory
            # "csv files (*.csv)", options=myOptions)
            "png files (*.png)", options=myOptions)

        if self.graphicFile:
            self.graphicFile = self.removeCsvExtension(self.graphicFile)
            exporter = FixedImageExporter(self.dBarPlotItem)
            exporter.makeWidthHeightInts()
            targetFileD = self.graphicFile + '.D_pyote.png'
            exporter.export(targetFileD)

            exporter = FixedImageExporter(self.durBarPlotItem)
            exporter.makeWidthHeightInts()
            targetFileDur = self.graphicFile + '.R-D_pyote.png'
            exporter.export(targetFileDur)

            exporter = FixedImageExporter(self.falsePositivePlotItem)
            exporter.makeWidthHeightInts()
            targetFileDur = self.graphicFile + '.noise_induced_event_pyote.png'
            exporter.export(targetFileDur)

            self.showInfo('Wrote to: \r\r' + targetFileD + ' \r\r' + targetFileDur)

    @staticmethod
    def removeCsvExtension(path):
        base, ext = os.path.splitext(path)
        if ext == '.csv':
            return base
        else:
            return path

    def exportGraphic(self):

        _, name = os.path.split(self.csvFilePath)
        name = self.removeCsvExtension(name)

        myOptions = QFileDialog.Options()
        myOptions |= QFileDialog.DontConfirmOverwrite
        myOptions |= QFileDialog.DontUseNativeDialog
        myOptions |= QFileDialog.ShowDirsOnly

        self.graphicFile, _ = QFileDialog.getSaveFileName(
            self,  # parent
            "Select directory/modify filename (png will be appended for you)",  # title for dialog
            self.settings.value('lightcurvedir', "") + '/' + name,  # starting directory
            "png files (*.png)", options=myOptions)

        if self.graphicFile:
            self.graphicFile = self.removeCsvExtension(self.graphicFile)
            exporter = FixedImageExporter(self.mainPlot.getPlotItem())
            exporter.makeWidthHeightInts()
            targetFile = self.graphicFile + '.pyote.png'
            exporter.export(targetFile)
            self.showInfo('Wrote to: \r\r' + targetFile)

    def initializeVariablesThatDontDependOnAfile(self):

        self.left = None  # Used during block integration
        self.right = None  # "
        self.selPts = []  # "

        self.exponentialDtheoryPts = None
        self.exponentialRtheoryPts = None

        self.flashEdges = []
        self.normalized = False
        self.timesAreValid = True  # until we find out otherwise
        self.selectedPoints = {}  # Clear/declare 'selected points' dictionary
        self.baselineXvals = []
        self.baselineYvals = []
        self.underlyingLightcurveAns = None
        self.solution = None
        self.firstPassSolution = None
        self.secondPassSolution = None
        self.smoothSecondary = []
        self.corCoefs = []
        self.numPtsInCorCoefs = 0
        self.Doffset = 1  # Offset (in readings) between D and 'start of exposure'
        self.Roffset = 1  # Offset (in readings) between R and 'start of exposure'
        self.sigmaB = None
        self.sigmaA = None
        self.A = None
        self.B = None
        self.snrB = None
        self.snrA = None
        self.dRegion = None
        self.rRegion = None
        self.eRegion = None
        self.dLimits = []
        self.rLimits = []
        self.eLimits = []
        self.minEvent = None
        self.maxEvent = None
        self.solution = None
        self.eventType = 'none'
        self.cancelRequested = False
        self.deltaDlo68 = 0
        self.deltaDlo95 = 0
        self.deltaDhi68 = 0
        self.deltaDhi95 = 0
        self.deltaRlo68 = 0
        self.deltaRlo95 = 0
        self.deltaRhi68 = 0
        self.deltaRhi95 = 0
        self.deltaDurlo68 = 0
        self.deltaDurlo95 = 0
        self.deltaDurhi68 = 0
        self.deltaDurhi95 = 0
        self.plusD = None
        self.minusD = None
        self.plusR = None
        self.minusR = None
        self.dBarPlotItem = None
        self.durBarPlotItem = None
        self.errBarWin = None

    def requestCancel(self):
        self.cancelRequested = True

    def showDzone(self):
        # If the user has not selected any points, we remove any dRegion that may
        # have been present
        if len(self.selectedPoints) == 0:
            self.dRegion = None
            self.dLimits = None
            self.newRedrawMainPlot()
            return

        if len(self.selectedPoints) != 2:
            self.showInfo('Exactly two points must be selected for this operation.')
            return
        selIndices = [key for key, _ in self.selectedPoints.items()]
        selIndices.sort()

        leftEdge = int(min(selIndices))
        rightEdge = int(max(selIndices))

        # Added Jan 8 2024
        if rightEdge - leftEdge < 2:
            if rightEdge < self.right:
                rightEdge += 1
            else:
                if leftEdge > self.left:
                    leftEdge -= 1

        if self.rLimits:
            if rightEdge > self.rLimits[0] - 2:  # Enforce at least 1 'a' point
                rightEdge = self.rLimits[0] - 2

            if self.rLimits[1] < self.right:  # At least 1 'b' point is present
                if leftEdge < self.left:
                    leftEdge = self.left
            else:
                if leftEdge < self.left + 0:  # Changed from 1 to zero in version 5.5.2
                    leftEdge = self.left + 0  # Changed from 1 to zero in version 5.5.2
        else:
            if leftEdge < self.left + 0:  # Changed from 1 to zero in version 5.5.2
                leftEdge = self.left + 0  # Enforce at least 1 'b' point (changed to match above)
            pass

        if rightEdge < self.left or rightEdge <= leftEdge:
            self.removePointSelections()
            self.newRedrawMainPlot()
            return

        self.setDataLimits.setEnabled(False)

        self.locateEvent.setEnabled(True)

        self.dLimits = [leftEdge, rightEdge]
        self.minEventEdit.clear()
        self.maxEventEdit.clear()

        if self.rLimits:
            self.eventType = 'DandR'
        else:
            self.eventType = 'Donly'

        self.dRegion = pg.LinearRegionItem(
            [leftEdge, rightEdge], movable=False, brush=(0, 200, 0, 50))
        self.mainPlot.addItem(self.dRegion)

        self.showMsg('D zone selected: ' + str([leftEdge, rightEdge]))
        self.removePointSelections()
        self.newRedrawMainPlot()

    def showRzone(self):
        # If the user has not selected any points, we remove any rRegion that may
        # have been present
        if len(self.selectedPoints) == 0:
            self.rRegion = None
            self.rLimits = None
            self.newRedrawMainPlot()
            return

        if len(self.selectedPoints) != 2:
            self.showInfo('Exactly two points must be selected for this operation.')
            return
        selIndices = [key for key, _ in self.selectedPoints.items()]
        selIndices.sort()

        leftEdge = int(min(selIndices))
        rightEdge = int(max(selIndices))

        if self.dLimits:
            if leftEdge < self.dLimits[1] + 2:
                leftEdge = self.dLimits[1] + 2  # Enforce at least 1 'a' point
            if self.dLimits[0] == self.left:
                if rightEdge >= self.right:
                    rightEdge = self.right - 1  # Enforce at least 1 'b' point
            else:
                if rightEdge >= self.right:
                    rightEdge = self.right
        else:
            if rightEdge >= self.right - 1:
                rightEdge = self.right - 1  # Enforce 1 'a' (for r-only search)
            if leftEdge < self.left + 1:
                leftEdge = self.left + 1  # Enforce  1 'b' point

        if rightEdge <= leftEdge:
            self.removePointSelections()
            self.newRedrawMainPlot()
            return

        self.setDataLimits.setEnabled(False)

        self.locateEvent.setEnabled(True)

        self.rLimits = [leftEdge, rightEdge]
        self.minEventEdit.clear()
        self.maxEventEdit.clear()

        if self.dLimits:
            # self.DandR.setChecked(True)
            self.eventType = 'DandR'
        else:
            self.eventType = 'Ronly'
            # self.Ronly.setChecked(True)

        self.rRegion = pg.LinearRegionItem(
            [leftEdge, rightEdge], movable=False, brush=(200, 0, 0, 50))
        self.mainPlot.addItem(self.rRegion)

        self.showMsg('R zone selected: ' + str([leftEdge, rightEdge]))
        self.removePointSelections()
        self.newRedrawMainPlot()

    def calculateFlashREdge(self):
        if len(self.selectedPoints) != 2:
            self.showInfo(
                'Exactly two points must be selected for this operation.')
            return

        self.minEventEdit.setText('')
        self.maxEventEdit.setText('')

        selIndices = [key for key, _ in self.selectedPoints.items()]
        selIndices.sort()

        # We save self.left and self.right because we need to manipulate them for doing
        # a trim to make the flash edge look like a valid R-only event
        savedLeft = self.left
        savedRight = self.right

        leftEdge = int(min(selIndices))
        rightEdge = int(max(selIndices))

        # This effects the 'trim'
        self.left = leftEdge
        self.right = rightEdge

        if self.dLimits:
            if leftEdge < self.dLimits[1] + 2:
                leftEdge = self.dLimits[1] + 2  # Enforce at least 1 'a' point
            if self.dLimits[0] == self.left:
                if rightEdge >= self.right:
                    rightEdge = self.right - 1  # Enforce at least 1 'b' point
            else:
                if rightEdge >= self.right:
                    rightEdge = self.right
        else:
            if rightEdge >= self.right - 1:
                rightEdge = self.right - 1  # Enforce 1 'a' (for r-only search)
            if leftEdge < self.left + 1:
                leftEdge = self.left + 1  # Enforce  1 'b' point

        if rightEdge <= leftEdge:
            self.removePointSelections()
            self.newRedrawMainPlot()
            return

        self.locateEvent.setEnabled(True)

        self.rLimits = [leftEdge, rightEdge]

        if self.dLimits:
            self.eventType = 'DandR'
        else:
            self.eventType = 'Ronly'

        self.rRegion = pg.LinearRegionItem(
            [leftEdge, rightEdge], movable=False, brush=(200, 0, 0, 50))
        self.mainPlot.addItem(self.rRegion)

        self.showMsg('R zone selected: ' + str([leftEdge, rightEdge]))
        self.removePointSelections()
        # self.newRedrawMainPlot()

        # We need to suppress the automatic report run
        self.suppressReport = True

        if self.timeDelta == 0.0:
            self.timeDelta, gotTimeDelta = QtWidgets.QInputDialog.getDouble(
                self, ' ', 'Enter frame time (seconds):', decimals=6)
            if not gotTimeDelta or self.timeDelta == 0.0:
                self.left = savedLeft
                self.right = savedRight
                self.dRegion = None
                self.rRegion = None
                self.removePointSelections()
                self.newRedrawMainPlot()
                return

        self.newRedrawMainPlot()
        self.findEvent()

        self.left = savedLeft
        self.right = savedRight
        self.newRedrawMainPlot()

        if self.solution:
            frameDelta = float(self.yFrame[1]) - float(self.yFrame[0])
            frameZero = float(self.yFrame[0])
            flashFrame = self.solution[1] * frameDelta + frameZero
            self.flashEdges.append(flashFrame)
            self.flashEdges[-1] = '%0.2f' % self.flashEdges[-1]
            msg = 'flash edges (in frame units): %s' % str(self.flashEdges)
            self.showMsg(msg, bold=True, color='red')

    def newNormalize(self):
        self.newSmoothRefStar()  # Produces/updates self.smoothSecondary

        # Find the mean value in the smoothedSecondary curve and use it as ref (previously we had the user
        # click on the point to use, but as that had no effect on magdrop calculations, was never needed,)
        ref = np.mean(self.smoothSecondary)

        xOffset = 0
        # Look for time shift (xOffset)
        for i, checkBox in enumerate(self.referenceCheckBoxes):
            if checkBox.isChecked():
                xOffset = self.xOffsetSpinBoxes[i].value()

        # Reminder: the smoothSecondary[] only cover self.left to self.right inclusive,
        # hence the index manipulation in the following code
        maxK = len(self.smoothSecondary) - 1
        targetY = []
        referenceY = []
        for i in range(self.left, self.right + 1):
            k = i - xOffset - self.left
            if k < 0 or k > maxK:
                self.yStatus[i] = EXCLUDED
                continue
            else:
                if not self.yStatus[i] == BASELINE:
                    self.yStatus[i] = INCLUDED
                targetY.append(self.yValues[i])
                referenceY.append(self.smoothSecondary[k])
            try:
                self.yValues[i] = (ref * self.yValues[i]) / self.smoothSecondary[k]
            except Exception as e:
                self.showMsg(str(e))

        # Compute standard deviation of normalized target curve - minimize this metric
        yValuesInMetric = []
        baselineSelectionAvailable = False
        for i in range(len(self.yValues)):
            if self.yStatus[i] == BASELINE:
                baselineSelectionAvailable = True
                break
        if baselineSelectionAvailable:
            for i in range(len(self.yValues)):
                if self.yStatus[i] == BASELINE:
                    yValuesInMetric.append(self.yValues[i])
        else:
            for i in range(len(self.yValues)):
                if self.yStatus[i] == INCLUDED:
                    yValuesInMetric.append(self.yValues[i])

        targetStd = np.std(yValuesInMetric)
        self.showMsg(f'Flatness  (minimize this value): {targetStd:0.2f} '
                     f'(readings: {self.smoothingIntervalSpinBox.value()})  (X offset: {xOffset})',
                     color='green', bold=True, alternateLogFile=self.normalizationLogFile)

        self.fullDataDictionary[self.targetKey] = self.yValues
        self.fullDataDictionary[self.referenceKey] = self.yRefStar

        self.fillTableViewOfData()  # This should capture/write the effects of the normalization to the table

        self.normalized = True

    def newSmoothRefStar(self):
        if self.right is None:
            self.right = self.dataLen
        if self.left is None:
            self.left = 0

        if (self.right - self.left) < 2:
            self.showInfo('The smoothing algorithm requires a minimum data set of 3 points')
            return

        y = [self.yRefStar[i] for i in range(self.left, self.right + 1)]

        userSpecedWindow = self.smoothingIntervalSpinBox.value()

        try:
            if len(y) > userSpecedWindow:
                window = userSpecedWindow
            else:
                window = len(y)

            # Enforce the odd window size required by savgol_filter()
            if window % 2 == 0:
                window += 1

            if window == 1:
                self.smoothSecondary = np.array(y)
            else:
                # We do a double pass with a first order (straight line with slope) savgol filter
                filteredY = savgol_filter(np.array(y), window, 1)
                self.smoothSecondary = savgol_filter(filteredY, window, 1)

        except Exception as e:
            self.showMsg(str(e))

    def switchToTabNamed(self, title):
        tabCount = self.tabWidget.count()  # Returns number of tabs
        for i in range(tabCount):
            if self.tabWidget.tabText(i) == title:
                self.tabWidget.setCurrentIndex(i)
                return

        self.showInfo(f'Cannot find tab with title: {title}')

    def toggleDisplayOfTimestampErrors(self):
        self.newRedrawMainPlot()
        self.mainPlot.autoRange()

    def showInfo(self, stuffToSay):
        QMessageBox.information(self, 'General information', stuffToSay)

    def showQuery(self, question, title=''):
        msgBox = QMessageBox(self)
        msgBox.setIcon(QMessageBox.Question)
        msgBox.setText(question)
        msgBox.setWindowTitle(title)
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msgBox.setDefaultButton(QMessageBox.Yes)
        self.queryRetVal = msgBox.exec_()

    def showInfoNonModal(self, stuffToSay, title=''):
        msgBox = QMessageBox(self)
        msgBox.setIcon(QMessageBox.Information)
        msgBox.setText(stuffToSay)
        msgBox.setWindowTitle(title)
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msgBox.setDefaultButton(QMessageBox.Yes)
        # self.queryRetVal = msgBox.show()
        return msgBox

    def fillPrimaryAndRef(self):
        # Load self.yValues and sel.yRefStar with proper light curves as
        # indicated by the spinner values

        # Get indices of selected primary and reference light curves
        primary = 0
        for i, checkBox in enumerate(self.targetCheckBoxes):
            if checkBox.isChecked():
                primary = i
                break

        reference = None
        for i, checkBox in enumerate(self.referenceCheckBoxes):
            if checkBox.isChecked():
                reference = i
                break

        if primary == 0:
            self.yValues = self.LC1
        elif primary == 1:
            self.yValues = self.LC2
        elif primary == 2:
            self.yValues = self.LC3
        elif primary == 3:
            self.yValues = self.LC4
        else:
            self.yValues = self.extra[primary - 4]

        if reference is not None:
            if reference == 0:
                self.yRefStar = self.LC1
            elif reference == 1:
                self.yRefStar = self.LC2
            elif reference == 2:
                self.yRefStar = self.LC3
            elif reference == 3:
                self.yRefStar = self.LC4
            else:
                self.yRefStar = self.extra[reference - 4]
        else:
            self.yRefStar = []

        # noinspection PyUnusedLocal
        self.yStatus = [INCLUDED for _ in range(self.dataLen)]

    def doIntegration(self):

        specifiedBlockSize = None
        if len(self.selectedPoints) == 0:
            if not self.blockSizeEdit.text() == '':
                try:
                    specifiedBlockSize = int(self.blockSizeEdit.text())
                except ValueError as e:
                    self.showInfo(f'blockSizeEdit: {e}')
                    return
                self.showMsg(f'Analysis of specified block size of {specifiedBlockSize}\n'
                             f'to determine best offset', color='red', bold=True)
            else:
                self.showMsg('Analysis of all possible block integration sizes and offsets',
                             color='red', bold=True)
            notchList = []
            kList = []
            offsetList = []

            self.progressBar.setValue(0)
            self.progressBarLightcurves.setValue(0)
            progress = 0
            if specifiedBlockSize is None:
                integrationSizes = [2, 4, 8, 16, 32, 48, 64, 96, 128, 256]
            else:
                integrationSizes = [specifiedBlockSize]

            for k in integrationSizes:
                kList.append(k)
                ans = mean_std_versus_offset(k, self.yValues)
                progress += 1
                self.progressBar.setValue((progress // len(integrationSizes)) * 100)
                self.progressBarLightcurves.setValue((progress // len(integrationSizes)) * 100)

                QtWidgets.QApplication.processEvents()
                offsetList.append(np.argmin(ans))
                median = np.median(ans)
                notch = np.min(ans) / median
                notchList.append(notch)
                s = '%3d notch %0.2f [' % (k, notch)

                for item in ans:
                    s = s + '%0.1f, ' % item
                self.showMsg(s[:-2] + ']', blankLine=False)
                # QtGui.QApplication.processEvents()
                QtWidgets.QApplication.processEvents()

            self.progressBar.setValue(0)
            self.progressBarLightcurves.setValue(0)
            QtWidgets.QApplication.processEvents()

            best = int(np.argmin(notchList))
            blockSize = kList[best]
            offset = int(offsetList[best])
            self.showMsg(' ', blankLine=False)
            s = '\r\nBest integration estimate: blockSize: %d @ offset %d' % (blockSize, offset)
            self.showMsg(s, color='red', bold=True)

            brush1 = (0, 200, 0, 70)
            brush2 = (200, 0, 0, 70)

            leftEdge = offset - 0.5
            rightEdge = leftEdge + blockSize
            bFlag = True

            while rightEdge <= len(self.yValues):
                if bFlag:
                    bFlag = False
                    brushToUse = brush2
                else:
                    bFlag = True
                    brushToUse = brush1

                if bFlag:
                    self.mainPlot.addItem(pg.LinearRegionItem([leftEdge, rightEdge],
                                                              movable=False, brush=brushToUse))
                leftEdge += blockSize
                rightEdge += blockSize

            # Set the integration selection point indices
            self.bint_left = offset
            self.bint_right = offset + blockSize - 1
            self.selPts = [self.bint_left, self.bint_right]

            self.acceptBlockIntegration.setEnabled(True)

        elif len(self.selectedPoints) != 2:
            self.showInfo('Exactly two points must be selected for a block integration')
            return
        else:
            self.bint_left = None  # Force use of selectPoints in applyIntegration()
            # self.acceptBlockIntegration.setEnabled(False)
            self.applyIntegration()

    def validateSinglePointDrop(self):
        if not self.userDeterminedBaselineStats:
            self.showInfo(f'You need to select baseline points and calculate statistics first. '
                          f'\n\nGo to the Noise analysis tab to do this.')
            return

        if len(self.selectedPoints) != 1:
            self.showInfo('Exactly one point must be selected for this operation.')
            return
        selectedPoints = [key for key in self.selectedPoints.keys()]
        selectedPoint = selectedPoints[0]

        self.errBarWin = pg.GraphicsWindow(
            title='simulated drop distribution for single point event')
        self.errBarWin.resize(1200, 500)
        layout = QtWidgets.QGridLayout()
        self.errBarWin.setLayout(layout)
        drop = self.B - self.yValues[selectedPoint]

        # We have to 'normalize' observed_drop and sigma so that the NIE calculations don't have
        # to deal with numerical problems while doing numerical integrations and curve fitting.
        observed_drop = drop / self.B
        sigma = self.sigmaB / self.B

        pw, falsePositive, probability, *_ = self.falsePositiveReport(
            event_duration=1, num_trials=50000, observation_size=self.right - self.left + 1,
            observed_drop=observed_drop,
            posCoefs=self.corCoefs, sigma=sigma)
        layout.addWidget(pw, 0, 0)

        self.showMsg(f"Reading number {selectedPoint} with drop {drop:0.2f} was selected for validation", color='blue',
                     bold=True, blankLine=True)

        self.showMsg(f"noiseSigmaDistance for reading number {selectedPoint} with drop {drop:0.2f} is "
                     f"{self.noiseSigmaDistance:0.1f}", color='blue',
                     bold=True, blankLine=True)

        # self.showMsg(f'\n==== That single point event has a probability of {self.drop_nie_probability:0.6f} of being due to noise',
        #              color='blue', bold=True)

        # if falsePositive:
        #     self.showMsg(f'==== The single point event is NOT valid - it has non-zero chance of being due to noise',
        #                  color='red', bold=True)
        # else:
        #     self.showMsg(f'==== The single point event is valid - it is unlikely to stem from noise',
        #                  color='blue', bold=True)

    def applyIntegration(self):
        if self.bint_left is None:
            if self.droppedFrames:
                self.showInfo('This data set contains some erroneous time steps, which have ' +
                              'been marked with red lines.  Best practice is to ' +
                              'choose an integration block that is ' +
                              'positioned in an unmarked region, hopefully containing ' +
                              'the "event".  Block integration ' +
                              'proceeds to the left and then to the right of the marked block.')

            self.selPts = [key for key in self.selectedPoints.keys()]
            self.removePointSelections()
            self.bint_left = min(self.selPts)
            self.bint_right = max(self.selPts)

        # Time to do the work
        p0 = self.bint_left
        span = self.bint_right - self.bint_left + 1  # Number of points in integration block
        self.blockSize = span
        newFrame = []
        newTime = []
        newLC1 = []
        newLC2 = []
        newLC3 = []
        newLC4 = []
        newDemoLightCurve = []
        newExtra = [[] for _ in range(len(self.extra))]

        # Create a new (local) fullDataDictionary using self.fullDataDictionary as the template if it exists.
        if self.fullDataDictionary and self.pymovieFileInUse:
            emptyDataDictionary = self.fullDataDictionary.copy()
        else:
            # This is a Tangra or Tangra-like csv, so we have to create the dictionary from scratch.
            emptyDataDictionary = {}  # noqa
            emptyDataDictionary['FrameNum'] = []
            emptyDataDictionary['timeInfo'] = []
            emptyDataDictionary['LC1'] = []
            if len(self.LC2) > 0:
                emptyDataDictionary['LC2'] = []
            if len(self.LC3) > 0:
                emptyDataDictionary['LC3'] = []
            if len(self.LC4) > 0:
                emptyDataDictionary['LC4'] = []

        for key in emptyDataDictionary.keys():
            emptyDataDictionary[key] = []

        # emptyDataDictionary = dict.fromkeys(self.fullDataDictionary.keys(), [])

        if not self.blockSize % 2 == 0:
            self.showInfo(f'Blocksize is {self.blockSize}\n\nAn odd number for blocksize is likely an error!')

        p = p0 - span  # Start working toward the left
        while p > 0:

            if self.fullDataDictionary and self.pymovieFileInUse:  # TODO 5.6.4 fix
                for key in emptyDataDictionary.keys():
                    if not (key == 'timeInfo' or key == 'FrameNum'):
                        avg = np.mean(self.fullDataDictionary[key][p:(p + span)])
                        emptyDataDictionary[key].append(avg)

            # for key in emptyDataDictionary.keys():
            #     if not (key == 'timeInfo' or key == 'FrameNum'):
            #         avg = np.mean(self.fullDataDictionary[key][p:(p + span)])
            #         emptyDataDictionary[key].insert(0, avg)

            avg = np.mean(self.LC1[p:(p + span)])
            newLC1.insert(0, avg)

            if len(self.LC2) > 0:
                avg = np.mean(self.LC2[p:(p + span)])
                newLC2.insert(0, avg)

            if len(self.LC3) > 0:
                avg = np.mean(self.LC3[p:(p + span)])
                newLC3.insert(0, avg)

            if len(self.LC4) > 0:
                avg = np.mean(self.LC4[p:(p + span)])
                newLC4.insert(0, avg)

            if len(newExtra) > 0:
                for k, lc in enumerate(self.extra):
                    avg = np.mean(lc[p:(p + span)])
                    newExtra[k].insert(0, avg)

            if len(self.demoLightCurve) > 0:
                avg = np.mean(self.demoLightCurve[p:(p + span)])
                newDemoLightCurve.insert(0, avg)

            newFrame.insert(0, self.yFrame[p])
            newTime.insert(0, self.yTimes[p])

            # Fix for Tangra block integration
            if not self.fullDataDictionary:
                emptyDataDictionary['FrameNum'].insert(0, self.yFrame[p])
                emptyDataDictionary['timeInfo'].insert(0, self.yTimes[p])
            else:
                emptyDataDictionary['FrameNum'].insert(0, self.yFrame[p])  # TODO 5.6.4 fix
                # emptyDataDictionary['FrameNum'].insert(0, self.fullDataDictionary['FrameNum'][p])
                emptyDataDictionary['timeInfo'].insert(0, self.fullDataDictionary['timeInfo'][p])

            p = p - span

        p = p0  # Start working toward the right
        while p < self.dataLen - span:

            if self.fullDataDictionary and self.pymovieFileInUse:  # TODO 5.6.4 fix
                for key in emptyDataDictionary.keys():
                    if not (key == 'timeInfo' or key == 'FrameNum'):
                        avg = np.mean(self.fullDataDictionary[key][p:(p + span)])
                        emptyDataDictionary[key].append(avg)

            avg = np.mean(self.LC1[p:(p + span)])
            newLC1.append(avg)

            if len(self.LC2) > 0:
                avg = np.mean(self.LC2[p:(p + span)])
                newLC2.append(avg)

            if len(self.LC3) > 0:
                avg = np.mean(self.LC3[p:(p + span)])
                newLC3.append(avg)

            if len(self.LC4) > 0:
                avg = np.mean(self.LC4[p:(p + span)])
                newLC4.append(avg)

            if len(newExtra) > 0:
                for k, lc in enumerate(self.extra):
                    avg = np.mean(lc[p:(p + span)])
                    newExtra[k].append(avg)

            if len(self.demoLightCurve) > 0:
                avg = np.mean(self.demoLightCurve[p:(p + span)])
                newDemoLightCurve.append(avg)

            newFrame.append(self.yFrame[p])
            newTime.append(self.yTimes[p])

            # Fix for Tangra block integration
            if not self.fullDataDictionary:
                emptyDataDictionary['FrameNum'].append(self.yFrame[p])
                emptyDataDictionary['timeInfo'].append(self.yTimes[p])
            else:
                emptyDataDictionary['FrameNum'].append(self.yFrame[p])  # TODO 5.6.4 fix
                # emptyDataDictionary['FrameNum'].append(self.fullDataDictionary['FrameNum'][p])
                emptyDataDictionary['timeInfo'].append(self.fullDataDictionary['timeInfo'][p])

            p = p + span

        self.fullDataDictionary = emptyDataDictionary.copy()

        self.dataLen = len(newLC1)

        self.LC1 = np.array(newLC1)
        self.LC2 = np.array(newLC2)
        self.LC3 = np.array(newLC3)
        self.LC4 = np.array(newLC4)
        if len(newExtra) > 0:
            for k in range(len(newExtra)):
                self.extra[k] = np.array(newExtra[k])
        self.demoLightCurve = np.array(newDemoLightCurve)

        # auto-select all points
        self.left = 0
        self.right = self.dataLen - 1

        self.fillPrimaryAndRef()

        self.yTimes = newTime[:]
        self.yFrame = newFrame[:]
        self.fillTableViewOfData()

        self.selPts.sort()
        self.showMsg('Block integration started at entry ' + str(self.selPts[0]) +
                     ' with block size of ' + str(self.selPts[1] - self.selPts[0] + 1) + ' readings')

        self.VizieRdict = {
            "timestamps": None,
            "yValues": None,
            "yStatus": None,
        }
        self.timeDelta, self.droppedFrames, self.cadenceViolation, self.errRate, _ = \
            getTimeStepAndOutliers(self.yTimes, self.yValues, self.yStatus, VizieRdict=self.VizieRdict)
        self.showMsg('timeDelta: ' + fp.to_precision(self.timeDelta, 6) + ' seconds per block', blankLine=False)
        self.showMsg('timestamp error rate: ' + fp.to_precision(100 * self.errRate, 2) + '%')

        self.frameTimeEdit.setText(fp.to_precision(self.timeDelta, 6))

        self.illustrateTimestampOutliers()

        self.doBlockIntegration.setEnabled(False)
        self.acceptBlockIntegration.setEnabled(False)
        self.blockSizeEdit.setEnabled(False)

        self.newRedrawMainPlot()
        self.mainPlot.autoRange()

    def togglePointSelected(self, index):
        if self.yStatus[index] != 3:
            # Save current status for possible undo (a later click)
            self.selectedPoints[index] = self.yStatus[index]
            self.yStatus[index] = 3  # Set color to 'selected'
        else:
            # Restore previous status (when originally clicked)
            self.yStatus[index] = self.selectedPoints[index]
            del(self.selectedPoints[index])
        self.suppressNormalization = True
        self.newRedrawMainPlot()  # Redraw plot to show selection change
        self.suppressNormalization = False

    def processClick(self, event):
        # Don't allow mouse clicks to select points unless the cursor is blank
        if self.blankCursor and self.left is not None and self.right is not None:
            # This try/except handles case where user clicks in plot area before a
            # plot has been drawn.
            try:
                mousePoint = self.mainPlotViewBox.mapSceneToView(event.scenePos())
                index = round(mousePoint.x())
                if event.button() == 1:  # left button clicked?
                    if index < self.left:
                        index = self.left
                    if index > self.right:
                        index = self.right
                    self.togglePointSelected(index)
                    self.acceptBlockIntegration.setEnabled(False)
                    self.suppressNormalization = True
                # Move the table view of data so that clicked point data is visible
                self.table.setCurrentCell(index, 0)
            except AttributeError:
                pass

    def initializeTableView(self):
        self.table.clear()
        self.table.setRowCount(3)
        columnsToDisplay = []
        if not self.aperture_names:
            # Handle non-PyMovie csv file
            colLabels = ['FrameNum', 'timeInfo', 'LC1', 'LC2', 'LC3', 'LC4']
            if len(self.LC1) > 0:
                self.lightcurveTitles[0].setText('LC1')
                self.targetCheckBoxes[0].setEnabled(True)
                self.showCheckBoxes[0].setEnabled(True)
                self.yOffsetSpinBoxes[0].setEnabled(True)
                self.yOffsetSpinBoxes[0].setValue(0)
                self.referenceCheckBoxes[0].setEnabled(True)

            if len(self.LC2) > 0:
                self.lightcurveTitles[1].setText('LC2')
                self.targetCheckBoxes[1].setEnabled(True)
                self.showCheckBoxes[1].setEnabled(True)
                self.yOffsetSpinBoxes[1].setEnabled(True)
                self.yOffsetSpinBoxes[1].setValue(0)
                self.referenceCheckBoxes[1].setEnabled(True)

            if len(self.LC3) > 0:
                self.lightcurveTitles[2].setText('LC3')
                self.targetCheckBoxes[2].setEnabled(True)
                self.showCheckBoxes[2].setEnabled(True)
                self.yOffsetSpinBoxes[2].setEnabled(True)
                self.yOffsetSpinBoxes[2].setValue(0)
                self.referenceCheckBoxes[2].setEnabled(True)

            if len(self.LC4) > 0:
                self.lightcurveTitles[3].setText('LC4')
                self.targetCheckBoxes[3].setEnabled(True)
                self.showCheckBoxes[3].setEnabled(True)
                self.yOffsetSpinBoxes[3].setEnabled(True)
                self.yOffsetSpinBoxes[3].setValue(0)
                self.referenceCheckBoxes[3].setEnabled(True)

            self.table.setColumnCount(6)
        else:
            colLabels = ['FrameNum', 'timeInfo']
            dataColumnNames = self.fullDataDictionary.keys()

            if len(self.userDataSetAdditions) == 1 and self.userDataSetAdditions[0] == 'ClearAll':
                columnsToDisplay = [self.lightcurveTitle_1.text()]  # This will be ['']
                self.userDataSetAdditions = []
            else:
                if not self.userDataSetAdditions:
                    columnsToDisplay = [name for name in dataColumnNames if name.startswith('signal')]
                    # Only accept up to 10 columns because that is the max that can be in the light curves panel
                    if len(columnsToDisplay) > 10:
                        columnsToDisplay = columnsToDisplay[0:10]
                else:
                    columnsToDisplay = self.userDataSetAdditions
                    if len(self.userDataSetAdditions) == 1:
                        self.LC1 = np.array(self.fullDataDictionary[columnsToDisplay[0]])
                        # Trigger a display of the first light curve
                        self.processTargetSelection(0, redraw=True)
                        if not columnsToDisplay[0] == '':
                            self.showMsg(f'{columnsToDisplay[0]} is the target curve')

            if not columnsToDisplay[0] == '':
                self.table.setColumnCount(2 + len(columnsToDisplay))
            else:
                self.table.setColumnCount(2)

            self.additionalDataSetNames = [name for name in dataColumnNames if name.startswith('signal')]
            self.additionalDataSetNames += [name for name in dataColumnNames if name.startswith('appsum')]
            self.additionalDataSetNames += [name for name in dataColumnNames if name.startswith('avgbkg')]
            self.additionalDataSetNames += [name for name in dataColumnNames if name.startswith('stdbkg')]
            self.additionalDataSetNames += [name for name in dataColumnNames if name.startswith('nmask')]
            self.additionalDataSetNames += [name for name in dataColumnNames if name.startswith('hit-defect')]
            self.curveSelectionComboBox.clear()
            self.availableLightCurvesForDisplay = []
            for dataSetName in self.additionalDataSetNames:
                self.curveSelectionComboBox.addItem(dataSetName)
                self.availableLightCurvesForDisplay.append(dataSetName)

            self.yFrame = self.fullDataDictionary['FrameNum']
            self.yTimes = self.fullDataDictionary['timeInfo']
            if not columnsToDisplay[0] == '':
                k = 0
                if k < len(columnsToDisplay):
                    self.LC1 = np.array(self.fullDataDictionary[columnsToDisplay[k]])
                    k += 1
                if k < len(columnsToDisplay):
                    self.LC2 = np.array(self.fullDataDictionary[columnsToDisplay[k]])
                    k += 1
                if k < len(columnsToDisplay):
                    self.LC3 = np.array(self.fullDataDictionary[columnsToDisplay[k]])
                    k += 1
                if k < len(columnsToDisplay):
                    self.LC4 = np.array(self.fullDataDictionary[columnsToDisplay[k]])
                    k += 1
                if k < len(columnsToDisplay):
                    self.extra = []
                    self.extra.append(np.array(self.fullDataDictionary[columnsToDisplay[k]]))
                    k += 1
                if k < len(columnsToDisplay):
                    self.extra.append(np.array(self.fullDataDictionary[columnsToDisplay[k]]))
                    k += 1
                if k < len(columnsToDisplay):
                    self.extra.append(np.array(self.fullDataDictionary[columnsToDisplay[k]]))
                    k += 1
                if k < len(columnsToDisplay):
                    self.extra.append(np.array(self.fullDataDictionary[columnsToDisplay[k]]))
                    k += 1
                if k < len(columnsToDisplay):
                    self.extra.append(np.array(self.fullDataDictionary[columnsToDisplay[k]]))
                    k += 1
                if k < len(columnsToDisplay):
                    self.extra.append(np.array(self.fullDataDictionary[columnsToDisplay[k]]))
                    k += 1
            else:
                self.LC1 = np.array([])
                columnsToDisplay = []
                # self.yValues = None

            # print(f'{sequentialLightcurveList}')
            k = 0
            for column_name in columnsToDisplay:
                colLabels.append(column_name)
                if k < 10:
                    self.lightcurveTitles[k].setText(column_name)
                    self.targetCheckBoxes[k].setEnabled(True)
                    self.showCheckBoxes[k].setEnabled(True)
                    self.yOffsetSpinBoxes[k].setEnabled(True)
                    self.yOffsetSpinBoxes[k].setValue(0)
                    self.referenceCheckBoxes[k].setEnabled(True)
                    if k == len(columnsToDisplay) - 1:
                        # Only 'show' the first signal
                        if k == 0:
                            self.showCheckBoxes[k].setChecked(True)
                            self.newRedrawMainPlot()
                    k += 1

        self.table.setHorizontalHeaderLabels(colLabels)
        if self.dataLen and self.yFrame:
            self.fillTableViewOfData()

        if not self.firstLightCurveDisplayed:
            self.firstLightCurveDisplayed = True
            # lightCurveInfo = f'Only the first light curve in the csv file will be displayed initially.\n\n' \
            #                  f'Below is the list of all light curves available for selection and display:\n\n'
            #
            # if len(self.availableLightCurvesForDisplay) > 0:
            #     for entry in self.availableLightCurvesForDisplay:
            #         lightCurveInfo += f'{entry}\n'
            # else:
            #     lightCurveInfo += 'LC1 to LC4 (input file came from Tangra)\n'
            #
            # lightCurveInfo += f'\nSwitch to the Lightcurves tab and use the Add data set drop down selection widget' \
            #                   f' to add additional curves to the display.'
            # self.showInfo(lightCurveInfo)

        if not columnsToDisplay:
            self.yValues = None

    def closeEvent(self, event):
        # Open (or create) file for holding 'sticky' stuff
        self.settings = QSettings('pyote.ini', QSettings.IniFormat)

        self.settings.setValue('showDetails', self.showDetailsCheckBox.isChecked())
        self.settings.setValue('versusTime', self.versusTimeCheckBox.isChecked())
        self.settings.setValue('showLegend', self.showLegendsCheckBox.isChecked())
        self.settings.setValue('showNotes', self.showAnnotationsCheckBox.isChecked())

        self.settings.setValue('edgeTimeFitPrecision', self.edgeTimePrecisionEdit.text())
        self.settings.setValue('chordDurationFitPrecision', self.chordDurationPrecisionEdit.text())
        self.settings.setValue('limbAngleFitPrecision', self.limbAnglePrecisionEdit.text())
        self.settings.setValue('missDistanceFitPrecision', self.missDistancePrecisionEdit.text())

        self.settings.setValue('allowNewVersionPopup', self.allowNewVersionPopupCheckbox.isChecked())

        self.settings.setValue('lineWidth', self.lineWidthSpinner.value())
        self.settings.setValue('dotSize', self.dotSizeSpinner.value())

        self.settings.setValue('vizierPlotDotSize', self.vzDotSizeSpinner.value())
        self.settings.setValue('vizierNagLevel', self.vzNagLevelSpinbox.value())
        self.settings.setValue('vizierObsYear', self.vzDateYearSpinner.value())


        tabOrderList = []
        numTabs = self.tabWidget.count()
        for i in range(numTabs):
            tabName = self.tabWidget.tabText(i)
            tabOrderList.append(tabName)

        self.settings.setValue('tablist', tabOrderList)
        # Capture the close request and update 'sticky' settings
        self.settings.setValue('size', self.size())
        self.settings.setValue('pos', self.pos())
        self.settings.setValue('showErrorBars', self.showErrBarsCheckBox.isChecked())
        self.settings.setValue('showUnderlyingLightCurve', self.showUnderlyingLightcurveCheckBox.isChecked())
        self.settings.setValue('showEdges', self.showEdgesCheckBox.isChecked())
        self.settings.setValue('doOCRcheck', self.showOCRcheckFramesCheckBox.isChecked())
        self.settings.setValue('showTimestamps', self.showTimestampsCheckBox.isChecked())
        self.settings.setValue('showCameraResponse', self.showCameraResponseCheckBox.isChecked())

        self.settings.setValue('ne3NotInUse', self.ne3NotInUseRadioButton.isChecked())
        self.settings.setValue('dnrOff', self.dnrOffRadioButton.isChecked())
        self.settings.setValue('dnrLow', self.dnrLowRadioButton.isChecked())
        self.settings.setValue('dnrMiddle', self.dnrMiddleRadioButton.isChecked())
        self.settings.setValue('dnrHigh', self.dnrHighRadioButton.isChecked())

        self.settings.setValue('dnrLowDtc', self.dnrLowDspinBox.value())
        self.settings.setValue('dnrLowRtc', self.dnrLowRspinBox.value())

        self.settings.setValue('dnrMiddleDtc', self.dnrMiddleDspinBox.value())
        self.settings.setValue('dnrMiddleRtc', self.dnrMiddleRspinBox.value())

        self.settings.setValue('dnrHighDtc', self.dnrHighDspinBox.value())
        self.settings.setValue('dnrHighRtc', self.dnrHighRspinBox.value())

        self.helperThing.close()

        tabOrderList = []
        numTabs = self.tabWidget.count()
        for i in range(numTabs):
            tabName = self.tabWidget.tabText(i)
            tabOrderList.append(tabName)

        self.settings.setValue('tablist', tabOrderList)

        self.settings.setValue('splitterOne', self.splitterOne.saveState())
        self.settings.setValue('splitterTwo', self.splitterTwo.saveState())
        self.settings.setValue('splitterThree', self.splitterThree.saveState())

        if self.d_underlying_lightcurve:
            matplotlib.pyplot.close(self.d_underlying_lightcurve)

        if self.r_underlying_lightcurve:
            matplotlib.pyplot.close(self.r_underlying_lightcurve)

        for frame_view in self.frameViews:
            if frame_view:
                frame_view.close()

        curDateTime = datetime.datetime.today().ctime()
        self.showMsg('')
        self.showMsg('#' * 20 + ' Session ended: ' + curDateTime + '  ' + '#' * 20)

        if self.errBarWin:
            self.errBarWin.close()

        event.accept()

        print(f"\nThe program has exited normally. Any error messages involving QBasicTimer \n"
              f"that may be printed following this are harmless artifacts "
              f"of the order in which various GUI elements are closed.\n")

    def rowClick(self, row):
        self.highlightReading(row)

    def cellClick(self, row):
        self.togglePointSelected(row)

    def highlightReading(self, rdgNum):
        x = [rdgNum]
        y = [self.yValues[x]]
        self.newRedrawMainPlot()
        self.mainPlot.plot(x, y, pen=None, symbol='o', symbolPen=(255, 0, 0),
                           symbolBrush=(255, 255, 0), symbolSize=10)

    def showMsg(self, msg, color=None, bold=False, blankLine=True, alternateLogFile=None):
        """ show standard output message """
        htmlmsg = msg
        if color:
            htmlmsg = '<font color=' + color + '>' + htmlmsg + '</font>'
        if bold:
            htmlmsg = '<b>' + htmlmsg + '</b>'
        htmlmsg = htmlmsg + '<br>'
        self.textOut.moveCursor(QtGui.QTextCursor.MoveOperation.End)
        self.textOut.insertHtml(htmlmsg)
        if blankLine:
            self.textOut.insertHtml('<br>')
        self.textOut.ensureCursorVisible()

        if alternateLogFile is not None:
            fileObject = open(alternateLogFile, 'a')
            fileObject.write(msg + '\n')
            if blankLine:
                fileObject.write('\n')
            fileObject.close()
        elif self.logFile:
            fileObject = open(self.logFile, 'a')
            fileObject.write(msg + '\n')
            if blankLine:
                fileObject.write('\n')
            fileObject.close()

    def reportSpecialProcedureUsed(self):
        if self.blockSize == 1:
            self.showMsg('This light curve has not been block integrated.',
                         color='blue', bold=True, blankLine=False)
        else:
            self.showMsg('Block integration of size %d has been applied to '
                         'this light curve.' %
                         self.blockSize, color='blue', bold=True, blankLine=False)

        if self.normalized:
            self.showMsg('', blankLine=False)
            self.showMsg('This light curve has been normalized against a '
                         'reference star.',
                         color='blue', bold=True, blankLine=False)
            for i, checkBox in enumerate(self.referenceCheckBoxes):
                if checkBox.isChecked():
                    self.showMsg(f'... {self.lightcurveTitles[i].text()} was used for normalization with '
                                 f'X offset: {self.xOffsetSpinBoxes[i].value()}', color='blue', bold=True,
                                 blankLine=False)
                    self.showMsg(f'... {self.smoothingIntervalSpinBox.value()} readings were used for smoothing',
                                 color='blue', bold=True, blankLine=True)

        if not (self.left == 0 and self.right == self.dataLen - 1):
            self.showMsg('This light curve has been trimmed.',
                         color='blue', bold=True, blankLine=False)

        self.showMsg('', blankLine=False)

        if not self.ne3NotInUseRadioButton.isChecked():
            self.writeNe3UsageReport()

        self.showMsg('', blankLine=False)

    def computeErrorBarPair(self, deltaHi, deltaLo, edge):
        noiseAsymmetry = self.snrA / self.snrB
        if (noiseAsymmetry > 0.7) and (noiseAsymmetry < 1.3):
            plus = (deltaHi - deltaLo) / 2
            minus = plus
        else:
            if edge == 'D':
                plus = deltaHi
                minus = -deltaLo
            else:
                plus = -deltaLo  # Deliberate 'inversion'
                minus = deltaHi  # Deliberate 'inversion'

        return plus, minus

    def Dreport(self, deltaDhi, deltaDlo):
        D, _ = self.solution

        intD = int(D)  # So that we can do lookup in the data table

        plusD, minusD = self.computeErrorBarPair(deltaHi=deltaDhi, deltaLo=deltaDlo, edge='D')

        # Save these for the 'envelope' plotter
        self.plusD = plusD
        self.minusD = minusD

        frameNum = float(self.yFrame[intD])

        Dframe = (D - intD) * self.framesPerEntry() + frameNum
        self.Dframe = Dframe  # For fit metric use
        self.showMsg('D: %.2f {+%.2f,-%.2f} (frame number)' % (Dframe, plusD * self.framesPerEntry(),
                                                               minusD * self.framesPerEntry()),
                     blankLine=False)
        ts = self.yTimes[int(D)]
        time = convertTimeStringToTime(ts)
        adjTime = time + (D - int(D)) * self.timeDelta
        ts = convertTimeToTimeString(adjTime)
        self.showMsg('D: %s  {+%.4f,-%.4f} seconds' %
                     (ts, plusD * self.timeDelta, minusD * self.timeDelta)
                     )
        return adjTime

    def Rreport(self, deltaRhi, deltaRlo):
        _, R = self.solution

        plusR, minusR = self.computeErrorBarPair(deltaHi=deltaRhi, deltaLo=deltaRlo, edge='R')

        # Save these for the 'envelope' plotter
        self.plusR = plusR
        self.minusR = minusR

        intR = int(R)
        frameNum = float(self.yFrame[intR])
        Rframe = (R - intR) * self.framesPerEntry() + frameNum
        self.Rframe = Rframe  # For fit metric use
        self.showMsg('R: %.2f {+%.2f,-%.2f} (frame number)' % (Rframe, plusR * self.framesPerEntry(),
                                                               minusR * self.framesPerEntry()),
                     blankLine=False)

        ts = self.yTimes[int(R)]
        time = convertTimeStringToTime(ts)
        adjTime = time + (R - int(R)) * self.timeDelta
        ts = convertTimeToTimeString(adjTime)
        self.showMsg('R: %s  {+%.4f,-%.4f} seconds' %
                     (ts, plusR * self.timeDelta, minusR * self.timeDelta)
                     )
        return adjTime

    # noinspection PyStringFormat
    def confidenceIntervalReport(self, numSigmas, deltaDurhi, deltaDurlo, deltaDhi, deltaDlo,
                                 deltaRhi, deltaRlo):

        D, R = self.solution

        self.showMsg('B: %0.2f  {+/- %0.2f}  sigmaB: %0.2f' %
                     (self.B, numSigmas * self.sigmaB / np.sqrt(self.nBpts), self.sigmaB))
        self.showMsg('A: %0.2f  {+/- %0.2f}  sigmaA: %0.2f' %
                     (self.A, numSigmas * self.sigmaA / np.sqrt(self.nApts), self.sigmaA))

        self.magdropReport(numSigmas)

        self.showMsg('DNR: %0.2f' % self.snrB)

        if self.eventType == 'Donly':
            self.Dreport(deltaDhi, deltaDlo)
        elif self.eventType == 'Ronly':
            self.Rreport(deltaRhi, deltaRlo)
        elif self.eventType == 'DandR':
            Dtime = self.Dreport(deltaDhi, deltaDlo)
            Rtime = self.Rreport(deltaDhi, deltaDlo)
            plusDur = ((deltaDurhi - deltaDurlo) / 2)
            minusDur = plusDur
            self.showMsg('Duration (R - D): %.4f {+%.4f,-%.4f} readings' %
                         ((R - D) * self.framesPerEntry(),
                          plusDur * self.framesPerEntry(), minusDur * self.framesPerEntry()),
                         blankLine=False)
            plusDur = ((deltaDurhi - deltaDurlo) / 2) * self.timeDelta
            minusDur = plusDur
            self.showMsg('Duration (R - D): %.4f {+%.4f,-%.4f} seconds' %
                         (Rtime - Dtime, plusDur, minusDur))

    # noinspection PyStringFormat
    def penumbralConfidenceIntervalReport(self, numSigmas, deltaDurhi, deltaDurlo, deltaDhi, deltaDlo,
                                          deltaRhi, deltaRlo):

        D, R = self.solution

        self.showMsg('B: %0.2f  {+/- %0.2f}' % (self.B, numSigmas * self.sigmaB / np.sqrt(self.nBpts)))
        self.showMsg('A: %0.2f  {+/- %0.2f}' % (self.A, numSigmas * self.sigmaA / np.sqrt(self.nApts)))

        self.magdropReport(numSigmas)

        self.showMsg('DNR: %0.2f' % self.snrB)

        if self.eventType == 'Donly':
            self.Dreport(deltaDhi, deltaDlo)
        elif self.eventType == 'Ronly':
            self.Rreport(deltaRhi, deltaRlo)
        elif self.eventType == 'DandR':
            Dtime = self.Dreport(deltaDhi, deltaDlo)
            Rtime = self.Rreport(deltaDhi, deltaDlo)
            plusDur = ((deltaDurhi - deltaDurlo) / 2)
            minusDur = plusDur
            self.showMsg('Duration (R - D): %.4f {+%.4f,-%.4f} readings' %
                         ((R - D) * self.framesPerEntry(),
                          plusDur * self.framesPerEntry(), minusDur * self.framesPerEntry()),
                         blankLine=False)
            plusDur = ((deltaDurhi - deltaDurlo) / 2) * self.timeDelta
            minusDur = plusDur
            self.showMsg('Duration (R - D): %.4f {+%.4f,-%.4f} seconds' %
                         (Rtime - Dtime, plusDur, minusDur))

    def magDropString(self, B, A, numSigmas):
        """
        We always report a percentage drop, dealing with negative A (event bottom below 0)
        by issuing a report containing only a percent drop of 100.0

        If A is > 0, we calculate the nominal magDrop and it's error bar.

        If the error bar is unreasonable (i.e., greater than the magDrop) we don't
        output the error bar and attach a message indicating that the observation
        was too noisy for a valid error bar.

        If the error bar was reasonable, we report it as usual.
        """
        if A > 0:
            percentDrop = 100.0 * (1.0 - A / B)
        else:
            percentDrop = 100.0

        self.percentMagDrop = percentDrop  # Used for fit metrics report

        if not percentDrop == 100.0:
            # A was > 0.  Attempt error bar calculations
            stdB = self.sigmaB / np.sqrt(self.nBpts)
            stdA = self.sigmaA / np.sqrt(self.nApts)
            ratio = A / B
            ratioError = numSigmas * np.sqrt((stdB / B) ** 2 + (stdA / A) ** 2) * ratio
            lnError = ratioError / ratio
            magdroperr = (2.5 / np.log(10.0)) * lnError
            magDrop = (np.log10(B) - np.log10(A)) * 2.5
            self.unvettedMagDrop = magDrop

            # Check that error bar is unreasonable (greater than the calculated magDrop)
            if magdroperr < magDrop:
                if numSigmas == 1:
                    ciStr = '(0.68 ci)'
                elif numSigmas == 2:
                    ciStr = '(0.95 ci)'
                else:
                    ciStr = '(0.9973 ci)'
            else:
                ciStr = 'too much noise; cannot calculate error bars'
                return f'percentDrop: {percentDrop:0.1f}  magDrop: {magDrop:0.3f}  {ciStr}'
            return f'percentDrop: {percentDrop:0.1f}  magDrop: {magDrop:0.3f}  +/- {magdroperr:0.3f}  {ciStr}'
        else:
            self.unvettedMagDrop = 25.0  # Return an exceptionally high (meaningless) magDrop for the fit metric

            # A was <= 0, so we can only report a maximum percent drop of 100
            return f'percentDrop: {percentDrop:0.1f}  (magDrop cannot be calculated because A is negative)'

    def magdropReport(self, numSigmas):
        Anom = self.A
        Bnom = self.B

        self.showMsg(f'magDrop report: {self.magDropString(Bnom, Anom, numSigmas)}')

    # noinspection PyStringFormat
    # def finalReportPenumbral(self):
    #
    #     self.displaySolution(True)
    #     self.minusD = self.plusD = self.penumbralDerrBar  # This the 2 sigma (95% ci) value
    #     self.minusR = self.plusR = self.penumbralRerrBar
    #     self.drawEnvelope()  # Shows error bars at the 95% ci level
    #
    #     self.deltaDlo68 = self.deltaDhi68 = self.plusD / 2.0
    #     self.deltaDlo95 = self.deltaDhi95 = self.plusD
    #     self.deltaDlo99 = self.deltaDhi99 = 3.0 * self.plusD / 2.0
    #
    #     self.deltaRlo68 = self.deltaRhi68 = self.plusR / 2.0
    #     self.deltaRlo95 = self.deltaRhi95 = self.plusR
    #     self.deltaRlo99 = self.deltaRhi99 = 3.0 * self.plusR / 2.0
    #
    #     self.deltaDurhi68 = np.sqrt(self.deltaDhi68 ** 2 + self.deltaRhi68 ** 2)
    #     self.deltaDurlo68 = - self.deltaDurhi68
    #     self.deltaDurhi95 = 2.0 * self.deltaDurhi68
    #     self.deltaDurlo95 = - self.deltaDurhi95
    #     self.deltaDurhi99 = 3.0 * self.deltaDurhi68
    #     self.deltaDurlo99 = - self.deltaDurhi99
    #
    #     # Grab the D and R values found and apply our timing convention
    #     D, R = self.solution
    #
    #     if self.eventType == 'DandR':
    #         self.showMsg('Timestamp validity check ...')
    #         self.reportTimeValidity(D, R)
    #
    #     self.showMsg('================= 0.68 containment interval report =================')
    #
    #     self.penumbralConfidenceIntervalReport(1, self.deltaDurhi68, self.deltaDurlo68,
    #                                            self.deltaDhi68, self.deltaDlo68,
    #                                            self.deltaRhi68, self.deltaRlo68)
    #
    #     self.showMsg('=============== end 0.68 containment interval report ===============')
    #
    #     self.showMsg('================= 0.95 containment interval report =================')
    #
    #     self.penumbralConfidenceIntervalReport(2, self.deltaDurhi95, self.deltaDurlo95,
    #                                            self.deltaDhi95, self.deltaDlo95,
    #                                            self.deltaRhi95, self.deltaRlo95)
    #
    #     self.showMsg('=============== end 0.95 containment interval report ===============')
    #
    #     self.showMsg('================= 0.9973 containment interval report ===============')
    #
    #     self.penumbralConfidenceIntervalReport(3, self.deltaDurhi99, self.deltaDurlo99,
    #                                            self.deltaDhi99, self.deltaDlo99,
    #                                            self.deltaRhi99, self.deltaRlo99)
    #
    #     self.showMsg('=============== end 0.9973 containment interval report =============')
    #
    #     self.doDframeReport()
    #     self.doRframeReport()
    #     self.doDurFrameReport()
    #
    #     self.showMsg('=============== Summary report for Excel file =====================')
    #
    #     self.reportSpecialProcedureUsed()  # This includes use of asteroid distance/speed and star diameter
    #
    #     if not self.timesAreValid:
    #         self.showMsg("Times are invalid due to corrupted timestamps!",
    #                      color='red', bold=True)
    #
    #     self.magDropReportStr = f'magDrop report: {self.magDropString(self.B, self.A, 2)}'
    #     self.xlsxDict['Comment'] = self.magDropReportStr
    #     self.showMsg(self.magDropReportStr)
    #
    #     self.showMsg('DNR: %0.2f' % self.snrB)
    #
    #     self.doDtimeReport()
    #     self.doRtimeReport()
    #     self.doDurTimeReport()
    #
    #     self.showMsg(
    #         'Enter D and R error bars for each containment interval in Excel spreadsheet without + or - sign (assumed to be +/-)')
    #
    #     self.showMsg('=========== end Summary report for Excel file =====================')
    #
    #     self.showMsg("Solution 'envelope' in the main plot drawn using 0.95 containment interval error bars")
    #
    #     if self.targetKey == '':
    #         self.targetKey = self.lightcurveTitles[0].text()
    #
    #     self.showMsg(f'fit metrics for {self.targetKey}', blankLine=False, bold=True)
    #
    #     # time_uncertainty = (self.deltaDhi95 - self.deltaDlo95) / 2.0 * self.timeDelta
    #     time_uncertainty = max(abs(self.deltaDhi95), abs(self.deltaDlo95)) * self.timeDelta
    #
    #     stats_msg = f'\nfit metrics === DNR: {self.snrB:0.2f}  edge uncertainty (0.95 ci): +/- {time_uncertainty:0.4f} seconds'
    #     self.showMsg(stats_msg, blankLine=False, bold=True)
    #
    #     stats_msg = f'fit metrics === B: {self.B:0.2f}  A: {self.A:0.2f} sigmaB: {self.sigmaB:0.2f}  sigmaA: {self.sigmaA:0.2f}'
    #     self.showMsg(stats_msg, blankLine=False, bold=True)
    #
    #
    #     margin = self.observedDrop - self.threeSigmaLine
    #     if margin > 0:
    #         stats_msg = f'fit metrics === from noise-induced-event test (observed drop: ' \
    #                     f'{self.observedDrop:0.1f})  (three sigma drop from noise: {self.threeSigmaLine:0.1f}) gives margin: {margin:0.1f}'
    #         self.showMsg(stats_msg, blankLine=False, bold=True)
    #     else:
    #         stats_msg = f'fit metrics === from noise-induced-event test (observed drop: ' \
    #                     f'{self.observedDrop:0.1f})  (three sigma drop from noise: {self.threeSigmaLine:0.1f}) gives margin: {margin:0.1f}'
    #         self.showMsg(stats_msg, blankLine=False, bold=True, color='red')
    #
    #     stats_msg = f'fit metrics === observed drop has {self.reportDropProbability()} of being a noise-induced-event'
    #     self.showMsg(stats_msg, blankLine=False, bold=True)
    #
    #     stats_msg = f'fit metrics === {self.magDropReportStr}'
    #     self.showMsg(stats_msg, blankLine=False, bold=True)
    #
    #     if self.eventType == 'Donly' or self.eventType == 'DandR':
    #         stats_msg = f'\nfit metrics === D frame number: {self.Dframe:0.4f}'
    #         self.showMsg(stats_msg, blankLine=False, bold=True)
    #
    #     if self.eventType == 'Ronly' or self.eventType == 'DandR':
    #         stats_msg = f'\nfit metrics === R frame number: {self.Rframe:0.4f}'
    #         self.showMsg(stats_msg, blankLine=False, bold=True)
    #
    #     if self.eventType == 'Donly' or self.eventType == 'DandR':
    #         stats_msg = f'\nfit metrics === D time: {self.Dtimestring}'
    #         self.showMsg(stats_msg, blankLine=False, bold=True)
    #
    #     if self.eventType == 'Ronly' or self.eventType == 'DandR':
    #         stats_msg = f'\nfit metrics === R time: {self.Rtimestring}'
    #         self.showMsg(stats_msg, blankLine=False, bold=True)
    #
    #     self.showMsg("")
    #
    #     self.updateFitMetricTxtFile()


    def finalReport(self, error_bar_plots_available=True):

        self.xlsxDict = {}
        self.writeDefaultGraphicsPlots(error_bar_plots_available=error_bar_plots_available)

        # Grab the D and R values found and apply our timing convention
        D, R = self.solution

        if self.eventType == 'DandR':
            self.showMsg('Timestamp validity check ...')
            self.reportTimeValidity(D, R)

        self.calcNumBandApoints()

        self.showMsg('================= 0.68 containment interval report =================')

        self.confidenceIntervalReport(1, self.deltaDurhi68, self.deltaDurlo68,
                                      self.deltaDhi68, self.deltaDlo68,
                                      self.deltaRhi68, self.deltaRlo68)

        self.showMsg('=============== end 0.68 containment interval report ===============')

        self.showMsg('================= 0.95 containment interval report =================')

        self.confidenceIntervalReport(2, self.deltaDurhi95, self.deltaDurlo95,
                                      self.deltaDhi95, self.deltaDlo95,
                                      self.deltaRhi95, self.deltaRlo95)

        envelopePlusR = self.plusR
        envelopePlusD = self.plusD
        envelopeMinusR = self.minusR
        envelopeMinusD = self.minusD

        self.showMsg('=============== end 0.95 containment interval report ===============')

        self.showMsg('================= 0.9973 containment interval report ===============')

        self.confidenceIntervalReport(3, self.deltaDurhi99, self.deltaDurlo99,
                                      self.deltaDhi99, self.deltaDlo99,
                                      self.deltaRhi99, self.deltaRlo99)

        self.showMsg('=============== end 0.9973 containment interval report =============')

        # Set the values to be used for the envelope plot (saved during 0.95 ci calculations)
        self.plusR = envelopePlusR
        self.plusD = envelopePlusD
        self.minusR = envelopeMinusR
        self.minusD = envelopeMinusD

        self.doDframeReport()
        self.doRframeReport()
        self.doDurFrameReport()

        self.showMsg('=============== Summary report for Excel file =====================')

        self.reportSpecialProcedureUsed()  # This includes use of asteroid distance/speed and star diameter

        # New nie report
        # if false_positive:
        #     self.showMsg(f"This 'drop' has a {false_probability:0.5f} probability of being an artifact of noise.",
        #                  bold=True, color='red', blankLine=False)
        # else:
        #     self.showMsg(f"This 'drop' has a zero probability of being an artifact of noise.",
        #                  bold=True, color='green', blankLine=False)

        self.showMsg(f'The observed drop has {self.reportDropProbability()} of being a noise-induced-event',
                     bold=True, color='blue', blankLine=True)

        # self.showMsg(f">>>> probability > 0.0000 indicates the 'drop' may be spurious (a noise artifact)."
        #              f" Consult with an IOTA Regional Coordinator.", color='blue', blankLine=False)
        # self.showMsg(f">>>> probability = 0.0000 indicates the 'drop' is unlikely to be a noise artifact, but"
        #              f" does not prove that the 'drop' is due to an occultation", color='blue', blankLine=False)
        # self.showMsg(f">>>> Consider 'drop' shape, timing, mag drop, duration and other positive observer"
        #              f" chords before reporting the 'drop' as a positive.", color='blue')
        # End new nie report

        self.showMsg("All times are calculated/reported based on the assumption that timestamps are "
                     "start-of-exposure times.",
                     color='blue', bold=True)
        self.showMsg("It is critical that you make appropriate time adjustments when timestamps are "
                     "NOT start-of-exposure times.", color='red', bold=True, blankLine=False)
        self.showMsg("If you use the North American Excel Spreadsheet report, all times will be properly corrected",
                     color='red', bold=True, blankLine=False)
        self.showMsg("for camera delay and reported start-of-exposure times.",
                     color='red', bold=True, blankLine=False)
        self.showMsg("For other users worldwide, use the appropriate corrections documented in the North American "
                     "Spreadsheet report form - use",
                     color='red', bold=True, blankLine=False)
        self.showMsg("the documentation shown on the Corrections Tables tab.",
                     color='red', bold=True)

        if not self.timesAreValid:
            self.showMsg("Times are invalid due to corrupted timestamps!",
                         color='red', bold=True)

        if self.choleskyFailed:
            self.showMsg('Cholesky decomposition failed during error bar '
                         'calculations. '
                         'Noise has therefore been treated as being '
                         'uncorrelated.',
                         bold=True, color='red')

        self.magDropReportStr = f'magDrop report: {self.magDropString(self.B, self.A, 2)}'
        self.xlsxDict['Comment'] = self.magDropReportStr
        self.showMsg(self.magDropReportStr)

        # noinspection PyStringFormat
        self.showMsg('DNR: %0.2f' % self.snrB)

        self.doDtimeReport()
        self.doRtimeReport()
        self.doDurTimeReport()

        self.showMsg(
            'Enter D and R error bars for each containment interval in Excel spreadsheet without + or - sign (assumed to be +/-)')

        self.showMsg('=========== end Summary report for Excel file =====================')

        self.showMsg("Solution 'envelope' in the main plot drawn using 0.95 containment interval error bars")

        if self.targetKey == '':
            self.targetKey = self.lightcurveTitles[0].text()

        self.showMsg(f'fit metrics for {self.targetKey}', blankLine=False, bold=True)

        # Calculate noise-induced-event ratio
        # margin = self.observedDrop - self.maxNoiseInducedDrop
        # false_positive_ratio = margin / self.maxNoiseInducedDrop

        # false_positive_metric = self.observedDrop / self.threeSigmaLine - 1.0

        time_uncertainty = max(abs(self.deltaDhi95), abs(self.deltaDlo95)) * self.timeDelta

        stats_msg = f'\nfit metrics === noiseSigmaDistance: {self.noiseSigmaDistance:0.1f}'
        self.showMsg(stats_msg, blankLine=False, bold=True)

        stats_msg = f'\nfit metrics === time error bar: +/- {time_uncertainty:0.4f} seconds'
        self.showMsg(stats_msg, blankLine=False, bold=True)

        stats_msg = f'\nfit metrics === DNR: {self.snrB:0.2f}  '
        self.showMsg(stats_msg, blankLine=False, bold=True)

        stats_msg = f'fit metrics === B: {self.B:0.2f}  A: {self.A:0.2f} sigmaB: {self.sigmaB:0.2f}  sigmaA: {self.sigmaA:0.2f}'
        self.showMsg(stats_msg, blankLine=False, bold=True)

        # margin = self.observedDrop - self.threeSigmaLine
        # if margin > 0:
        #     stats_msg = f'fit metrics === from noise-induced-event test (observed drop: ' \
        #                 f'{self.observedDrop:0.1f})  (three sigma drop from noise: {self.threeSigmaLine:0.1f}) gives margin: {margin:0.1f}'
        #     self.showMsg(stats_msg, blankLine=False, bold=True)
        # else:
        #     stats_msg = f'fit metrics === from noise-induced-event test (observed drop: ' \
        #                 f'{self.observedDrop:0.1f})  (three sigma drop from noise: {self.threeSigmaLine:0.1f}) gives  margin: {margin:0.1f}'
        #     self.showMsg(stats_msg, blankLine=False, bold=True, color='red')
        #
        #
        # stats_msg = f'fit metrics === observed drop has {self.reportDropProbability()} of being a noise-induced-event'
        # self.showMsg(stats_msg, blankLine=False, bold=True)

        stats_msg = f'fit metrics === {self.magDropReportStr}'
        self.showMsg(stats_msg, blankLine=False, bold=True)

        if self.eventType == 'Donly' or self.eventType == 'DandR':
            stats_msg = f'\nfit metrics === D frame number: {self.Dframe:0.4f}'
            self.showMsg(stats_msg, blankLine=False, bold=True)

        if self.eventType == 'Ronly' or self.eventType == 'DandR':
            stats_msg = f'\nfit metrics === R frame number: {self.Rframe:0.4f}'
            self.showMsg(stats_msg, blankLine=False, bold=True)

        if self.eventType == 'DandR':
            time_at_R = convertTimeStringToTime(self.Rtimestring)
            time_at_D = convertTimeStringToTime(self.Dtimestring)
            duration = time_at_R - time_at_D
            # duration = (self.Rframe - self.Dframe) * (self.timeDelta / self.blockSize)
            stats_msg = f'\nfit metrics === duration: {duration:0.4f} seconds'
            self.showMsg(stats_msg, blankLine=False, bold=True)

        if self.eventType == 'Donly' or self.eventType == 'DandR':
            stats_msg = f'\nfit metrics === D time: {self.Dtimestring}'
            self.showMsg(stats_msg, blankLine=False, bold=True)

        if self.eventType == 'Ronly' or self.eventType == 'DandR':
            stats_msg = f'\nfit metrics === R time: {self.Rtimestring}'
            self.showMsg(stats_msg, blankLine=False, bold=True)

        self.showMsg("")

        self.updateFitMetricTxtFile()

        if error_bar_plots_available:
            self.showHelp(self.helpLabelForFalsePositive)  # Hidden label on SqWave model tab

    def addSourceFileToFitMetricTxtFile(self):
        lightCurveDir = os.path.dirname(self.csvFilePath)  # This gets the folder where the light-curve.csv is located
        metric_file_path = os.path.join(lightCurveDir, 'fit_metrics.txt')

        if os.path.exists(metric_file_path):
            # We will append the new metric_file_path to an existing fit_metric.csv file
            pass
        else:
            # Here we will create a new fit_metric.csv file' with a header line
            with open(metric_file_path, 'w') as fileObject:
                fileObject.writelines('aperture name,time err +/-secs,DNR,FP metric,magDrop,'
                                      'percent drop,duration (secs),D time,'
                                      'R time,D frame,R frame,B,A,sigmaB,sigmaA,'
                                      'observed drop,FP drop,FP margin\n')

        with open(metric_file_path, 'a') as fileObject:
            fileObject.writelines(f'\n')
            fileObject.writelines(f'Source file is {self.csvFilePath}\n')
            fileObject.writelines(f'\n')

    def updateFitMetricTxtFile(self):
        lightCurveDir = os.path.dirname(self.csvFilePath)  # This gets the folder where the light-curve.csv is located
        metric_file_path = os.path.join(lightCurveDir, 'fit_metrics.txt')

        if os.path.exists(metric_file_path):
            # self.showInfo('We will append to an existing fit_metric.csv file')
            pass
        else:
            # self.showInfo('We fill create a new fit_metric.csv file')
            with open(metric_file_path, 'w') as fileObject:
                fileObject.writelines('aperture name,time err +/-secs,DNR,FP metric,magDrop,'
                                      'percent drop,duration (secs),D time,'
                                      'R time,D frame,R frame,B,A,sigmaB,sigmaA,'
                                      'observed drop,FP drop,FP margin\n')

        with open(metric_file_path, 'a') as fileObject:
            # Start with the aperture name
            if self.targetKey == '':
                self.targetKey = self.lightcurveTitles[0].text()
            new_data = self.targetKey

            hit_name = ''
            if self.targetKey.startswith('signal') or self.targetKey.startswith('appsum'):
                parts = self.targetKey.split('-', 1)
                hit_name = 'hit-defect-' + parts[1]
            if hit_name:
                if hit_name in self.fullDataDictionary.keys():
                    hit_defect_flags = self.fullDataDictionary[hit_name]
                    # Count number of points that involved defect pixels in the sampline mask
                    num_no_defect_hits = hit_defect_flags.count(0.0)
                    num_defect_hits = len(hit_defect_flags) - num_no_defect_hits
                    if not num_defect_hits == 0:
                        msg = f'light curve: {self.targetKey} had {num_defect_hits} points that had defect ' \
                              f'pixels included in the sample mask.'
                        # self.showInfo(msg)
                        self.showMsg(msg, color='red', bold=True)
                    pass

            margin = self.observedDrop - self.threeSigmaLine
            false_positive_metric = self.observedDrop / self.threeSigmaLine - 1.0


            # This works even for an Ronly event
            time_uncertainty = (self.deltaDhi95 - self.deltaDlo95) / 2.0 * self.timeDelta
            new_data += f',{time_uncertainty:0.4f}'       # add time error bar
            new_data += f',{self.snrB:0.2f}'              # add DNR
            new_data += f',{false_positive_metric:0.3f}'  # add false positive metric


            # add magDrop data
            new_data += f',{self.unvettedMagDrop:0.4f}'  # This gets displayed even if magDrop report would have suppressed it.
            new_data += f',{self.percentMagDrop:0.1f}'

            if self.eventType == 'DandR':
                time_at_R = convertTimeStringToTime(self.Rtimestring)
                time_at_D = convertTimeStringToTime(self.Dtimestring)
                duration = time_at_R - time_at_D
            else:
                duration = 0.0
            new_data += f',{duration:0.4f}'            # add duration

            # add D and R timestamps
            if self.eventType == 'Donly' or self.eventType == 'DandR':
                new_data += f',{self.Dtimestring}'
            else:
                new_data += f',[00:00:0.0000]'

            if self.eventType == 'Ronly' or self.eventType == 'DandR':
                new_data += f',{self.Rtimestring}'
            else:
                new_data += f',[00:00:0.0000]'

            # add D and R frame values
            if self.eventType == 'Donly' or self.eventType == 'DandR':
                new_data += f',{self.Dframe:0.4f}'
            else:
                new_data += f',0'

            if self.eventType == 'Ronly' or self.eventType == 'DandR':
                new_data += f',{self.Rframe:0.4f}'
            else:
                new_data += f',0'

            # add B then A then sigmaB then sigmaA
            new_data += f',{self.B:0.2f}'
            new_data += f',{self.A:0.2f}'
            new_data += f',{self.sigmaB:0.2f}'
            new_data += f',{self.sigmaA:0.2f}'

            # add false positive data details
            new_data += f',{self.observedDrop:0.1f}'
            new_data += f',{self.threeSigmaLine:0.1f}'
            new_data += f',{margin:0.1f}'
            # new_data += f',{false_positive_metric:0.3f}'  # add false positive metric

            self.showMsg("")

            fileObject.writelines(f'{new_data}\n')

        return

    def doDframeReport(self):
        if self.eventType == 'DandR' or self.eventType == 'Donly':
            D, _ = self.solution
            entryNum = int(D)
            frameNum = float(self.yFrame[entryNum])

            Dframe = (D - int(D)) * self.framesPerEntry() + frameNum
            self.showMsg('D frame number: {0:0.2f}'.format(Dframe), blankLine=False)
            errBar = max(abs(self.deltaDlo68), abs(self.deltaDhi68)) * self.framesPerEntry()
            self.showMsg('D: 0.6800 containment intervals:  {{+/- {0:0.2f}}} (readings)'.format(errBar),
                         blankLine=False)
            errBar = max(abs(self.deltaDlo95), abs(self.deltaDhi95)) * self.framesPerEntry()
            self.showMsg('D: 0.9500 containment intervals:  {{+/- {0:0.2f}}} (readings)'.format(errBar),
                         blankLine=False)
            errBar = max(abs(self.deltaDlo99), abs(self.deltaDhi99)) * self.framesPerEntry()
            self.showMsg('D: 0.9973 containment intervals:  {{+/- {0:0.2f}}} (readings)'.format(errBar))

    def framesPerEntry(self):
        # Normally, there is 1 frame per reading (entry), but if the source file was recorded
        # in field mode, there is only 0.5 frames per reading (entry).  Here we make the correction.
        if self.fieldMode:
            return self.blockSize / 2
        else:
            return self.blockSize

    def doRframeReport(self):
        if self.eventType == 'DandR' or self.eventType == 'Ronly':
            _, R = self.solution
            entryNum = int(R)
            frameNum = float(self.yFrame[entryNum])

            Rframe = (R - int(R)) * self.framesPerEntry() + frameNum
            self.showMsg('R frame number: {0:0.2f}'.format(Rframe), blankLine=False)
            errBar = max(abs(self.deltaRlo68), abs(self.deltaRhi68)) * self.framesPerEntry()
            self.showMsg('R: 0.6800 containment intervals:  {{+/- {0:0.2f}}} (readings)'.format(errBar),
                         blankLine=False)
            errBar = max(abs(self.deltaRlo95), abs(self.deltaRhi95)) * self.framesPerEntry()
            self.showMsg('R: 0.9500 containment intervals:  {{+/- {0:0.2f}}} (readings)'.format(errBar),
                         blankLine=False)
            errBar = max(abs(self.deltaRlo99), abs(self.deltaRhi99)) * self.framesPerEntry()
            self.showMsg('R: 0.9973 containment intervals:  {{+/- {0:0.2f}}} (readings)'.format(errBar))

    def doDurFrameReport(self):
        if self.eventType == 'DandR':
            D, R = self.solution
            self.showMsg('Duration (R - D): {0:0.4f} readings'.format((R - D) * self.framesPerEntry()), blankLine=False)
            errBar = ((self.deltaDurhi68 - self.deltaDurlo68) / 2) * self.framesPerEntry()
            self.showMsg('Duration: 0.6800 containment intervals:  {{+/- {0:0.4f}}} (readings)'.format(errBar),
                         blankLine=False)
            errBar = ((self.deltaDurhi95 - self.deltaDurlo95) / 2) * self.framesPerEntry()
            self.showMsg('Duration: 0.9500 containment intervals:  {{+/- {0:0.4f}}} (readings)'.format(errBar),
                         blankLine=False)
            errBar = ((self.deltaDurhi99 - self.deltaDurlo99) / 2) * self.framesPerEntry()
            self.showMsg('Duration: 0.9973 containment intervals:  {{+/- {0:0.4f}}} (readings)'.format(errBar))

    def doDtimeReport(self):
        if self.eventType == 'DandR' or self.eventType == 'Donly':
            D, _ = self.solution
            ts = self.yTimes[int(D)]

            time = convertTimeStringToTime(ts)
            adjTime = time + (D - int(D)) * self.timeDelta
            self.Dtime = adjTime  # This is needed for the duration report (assumed to follow!!!)
            ts = convertTimeToTimeString(adjTime)

            self.Dtimestring = ts

            tsParts = ts[1:-1].split(':')
            self.xlsxDict['Dhour'] = tsParts[0]
            self.xlsxDict['Dmin'] = tsParts[1]
            self.xlsxDict['Dsec'] = tsParts[2]

            self.showMsg('D time: %s' % ts, blankLine=False)

            plusD, minusD = self.computeErrorBarPair(deltaHi=self.deltaDhi68, deltaLo=self.deltaDlo68, edge='D')
            errBar = max(abs(plusD), abs(minusD)) * self.timeDelta
            self.xlsxDict['Derr68'] = errBar
            self.showMsg('D: 0.6800 containment intervals:  {{+/- {0:0.4f}}} seconds'.format(errBar), blankLine=False)

            plusD, minusD = self.computeErrorBarPair(deltaHi=self.deltaDhi95, deltaLo=self.deltaDlo95, edge='D')
            errBar = max(abs(plusD), abs(minusD)) * self.timeDelta
            self.xlsxDict['Derr95'] = errBar
            self.showMsg('D: 0.9500 containment intervals:  {{+/- {0:0.4f}}} seconds'.format(errBar), blankLine=False)

            plusD, minusD = self.computeErrorBarPair(deltaHi=self.deltaDhi99, deltaLo=self.deltaDlo99, edge='D')
            errBar = max(abs(plusD), abs(minusD)) * self.timeDelta
            self.xlsxDict['Derr99'] = errBar
            self.showMsg('D: 0.9973 containment intervals:  {{+/- {0:0.4f}}} seconds'.format(errBar))

    def doRtimeReport(self):
        if self.eventType == 'DandR' or self.eventType == 'Ronly':
            _, R = self.solution
            ts = self.yTimes[int(R)]

            time = convertTimeStringToTime(ts)
            adjTime = time + (R - int(R)) * self.timeDelta
            self.Rtime = adjTime  # This is needed for the duration report (assumed to follow!!!)
            ts = convertTimeToTimeString(adjTime)

            self.Rtimestring = ts

            tsParts = ts[1:-1].split(':')
            self.xlsxDict['Rhour'] = tsParts[0]
            self.xlsxDict['Rmin'] = tsParts[1]
            self.xlsxDict['Rsec'] = tsParts[2]

            self.showMsg('R time: %s' % ts, blankLine=False)

            plusR, minusR = self.computeErrorBarPair(deltaHi=self.deltaDhi68, deltaLo=self.deltaDlo68, edge='R')
            errBar = max(abs(plusR), abs(minusR)) * self.timeDelta
            self.xlsxDict['Rerr68'] = errBar
            self.showMsg('R: 0.6800 containment intervals:  {{+/- {0:0.4f}}} seconds'.format(errBar), blankLine=False)

            plusR, minusR = self.computeErrorBarPair(deltaHi=self.deltaDhi95, deltaLo=self.deltaDlo95, edge='R')
            errBar = max(abs(plusR), abs(minusR)) * self.timeDelta
            self.xlsxDict['Rerr95'] = errBar
            self.showMsg('R: 0.9500 containment intervals:  {{+/- {0:0.4f}}} seconds'.format(errBar), blankLine=False)

            plusR, minusR = self.computeErrorBarPair(deltaHi=self.deltaDhi99, deltaLo=self.deltaDlo99, edge='R')
            errBar = max(abs(plusR), abs(minusR)) * self.timeDelta
            self.xlsxDict['Rerr99'] = errBar
            self.showMsg('R: 0.9973 containment intervals:  {{+/- {0:0.4f}}} seconds'.format(errBar))

    def doDurTimeReport(self):
        if self.eventType == 'DandR':
            dur = self.Rtime - self.Dtime
            if dur < 0:  # We have bracketed midnight
                dur = dur + 3600 * 24  # Add seconds in a day
            self.showMsg('Duration (R - D): {0:0.4f} seconds'.format(dur), blankLine=False)
            errBar = ((self.deltaDurhi68 - self.deltaDurlo68) / 2) * self.timeDelta
            self.showMsg('Duration: 0.6800 containment intervals:  {{+/- {0:0.4f}}} seconds'.format(errBar),
                         blankLine=False)
            errBar = ((self.deltaDurhi95 - self.deltaDurlo95) / 2) * self.timeDelta
            self.showMsg('Duration: 0.9500 containment intervals:  {{+/- {0:0.4f}}} seconds'.format(errBar),
                         blankLine=False)
            errBar = ((self.deltaDurhi99 - self.deltaDurlo99) / 2) * self.timeDelta
            self.showMsg('Duration: 0.9973 containment intervals:  {{+/- {0:0.4f}}} seconds'.format(errBar))

    def reportTimeValidity(self, D, R):
        intD = int(D)
        intR = int(R)
        dTime = convertTimeStringToTime(self.yTimes[intD])
        rTime = convertTimeStringToTime(self.yTimes[intR])

        # Here we check for a 'midnight transition'
        if rTime < dTime:
            rTime += 24 * 60 * 60
            self.showMsg('D and R enclose a transition through midnight')

        if self.timeDelta == 0:
            self.timesAreValid = False
            self.showMsg('Timestamps are corrupted in a manner that caused a '
                         'timeDelta of '
                         '0.0 to be estimated!', color='red', bold=True)
            self.showInfo('Timestamps are corrupted in a manner that caused a '
                          'timeDelta of '
                          '0.0 to be estimated!')
            return

        numEnclosedReadings = int(round((rTime - dTime) / self.timeDelta))
        self.showMsg(
            'From timestamps at D and R, calculated %d reading blocks.  From reading blocks, calculated %d blocks.' %
            (numEnclosedReadings, intR - intD))
        if numEnclosedReadings == intR - intD:
            self.showMsg('Timestamps appear valid @ D and R')
            self.timesAreValid = True
        else:
            self.timesAreValid = False
            self.showMsg('! There is something wrong with timestamps at D '
                         'and/or R or frames have been dropped !', bold=True,
                         color='red')

    def calculateVarianceFromHistogram(self, y, x):
        peakIndex = np.where(y == np.max(y))
        self.showMsg(f'peakIndex: {peakIndex[0][0]} xPeak: {x[peakIndex[0][0]]:0.2f}')

    def computeErrorBars(self, plots_wanted=True):

        x1 = x2 = 0
        pw = None
        pw2 = None

        self.falsePositivePlotItem = None

        if self.sigmaB == 0.0:
            self.sigmaB = MIN_SIGMA * 2

        if self.sigmaA == 0.0:
            self.sigmaA = MIN_SIGMA

        self.snrB = (self.B - self.A) / self.sigmaB
        self.snrA = (self.B - self.A) / self.sigmaA
        snr = max(self.snrB, 0.2)  # A more reliable number

        D = int(round(80 / snr ** 2 + 0.5))

        D = max(10, D)
        if self.corCoefs.size > 1:
            D = round(1.5 * D)
        numPts = 2 * (D - 1) + 1
        posCoefs = []
        for entry in self.corCoefs:
            if entry < acfCoefThreshold:
                break
            posCoefs.append(entry)

        # noinspection PyTypeChecker
        distGen = edgeDistributionGenerator(
            ntrials=50000, numPts=numPts, D=D, acfcoeffs=posCoefs,
            B=self.B, A=self.A, sigmaB=self.sigmaB, sigmaA=self.sigmaA)

        dist = None
        self.choleskyFailed = False
        for dist in distGen:
            if type(dist) == float:
                if dist == -1.0:
                    self.choleskyFailed = True
                    # if self.ne3NotInUseRadioButton.isChecked():
                    if True:
                        self.showInfo(
                            'The Cholesky-Decomposition routine has failed.  This may be because the light curve ' +
                            'required block integration and you failed to do either a manual block integration or to accept the '
                            'PyOTE estimated block integration.  Please '
                            'examine the light curve for that possibility.' +
                            '\nWe treat this situation as though there is no '
                            'correlation in the noise.')
                        self.showMsg('Cholesky decomposition has failed.  '
                                     'Proceeding by '
                                     'treating noise as being uncorrelated.',
                                     bold=True, color='red')
                self.progressBar.setValue(int(dist * 100))
                self.progressBarLightcurves.setValue(int(dist * 100))
                QtWidgets.QApplication.processEvents()
                if self.cancelRequested:
                    self.cancelRequested = False
                    self.showMsg('Error bar calculation was cancelled')
                    self.progressBar.setValue(0)
                    self.progressBarLightcurves.setValue(0)
                    return
            else:
                self.progressBar.setValue(0)
                self.progressBarLightcurves.setValue(0)

        # pickle.dump(dist, open('sample-dist.p', 'wb'))

        y, x = np.histogram(dist, bins=1000)
        self.loDbar68, _, self.hiDbar68, self.deltaDlo68, self.deltaDhi68 = ciBars(dist=dist, ci=0.6827, D=D)
        self.loDbar95, _, self.hiDbar95, self.deltaDlo95, self.deltaDhi95 = ciBars(dist=dist, ci=0.95, D=D)
        self.loDbar99, _, self.hiDbar99, self.deltaDlo99, self.deltaDhi99 = ciBars(dist=dist, ci=0.9973, D=D)

        self.deltaRlo95 = - self.deltaDhi95
        self.deltaRhi95 = - self.deltaDlo95

        self.deltaRlo99 = - self.deltaDhi99
        self.deltaRhi99 = - self.deltaDlo99

        self.deltaRlo68 = - self.deltaDhi68
        self.deltaRhi68 = - self.deltaDlo68

        if isinstance(dist, np.ndarray):
            durDist = createDurDistribution(dist)
        else:
            self.showInfo('Unexpected error: variable dist is not of type np.ndarray')
            return
        ydur, xdur = np.histogram(durDist, bins=1000)
        self.loDurbar68, _, self.hiDurbar68, self.deltaDurlo68, self.deltaDurhi68 = \
            ciBars(dist=durDist, ci=0.6827, D=0.0)
        self.loDurbar95, _, self.hiDurbar95, self.deltaDurlo95, self.deltaDurhi95 = \
            ciBars(dist=durDist, ci=0.95, D=0.0)
        self.loDurbar99, _, self.hiDurbar99, self.deltaDurlo99, self.deltaDurhi99 = \
            ciBars(dist=durDist, ci=0.9973, D=0.0)

        pg.setConfigOptions(antialias=True)
        pen = pg.mkPen((0, 0, 0), width=2)

        # Get rid of a previous errBarWin that may have been closed (but not properly disposed of) by the user.
        if self.errBarWin is not None:
            self.errBarWin.close()

        # _, false_positive, false_probability, self.observedDrop, self.threeSigmaLine, self.fourSigmaLine, \
        #     self.fiveSigmaLine =  self.doFalsePositiveReport(posCoefs, plots_wanted=False)  # noqa

        fp_plot, false_positive, false_probability, self.observedDrop, self.threeSigmaLine, \
            self.fourSigmaLine, self.fiveSigmaLine = self.doFalsePositiveReport(posCoefs)  # noqa


        if plots_wanted:
            self.errBarWin = pg.GraphicsWindow(
                title='Solution distributions with containment intervals marked --- Noise Induced Events distribution')
            self.errBarWin.resize(1200, 1000)
            layout = QtWidgets.QGridLayout()
            self.errBarWin.setLayout(layout)

            pw = PlotWidget(viewBox=CustomViewBox(border=(0, 0, 0)),
                            enableMenu=False, title='Distribution of edge (D) errors due to noise',
                            labels={'bottom': 'Reading blocks'})
            self.dBarPlotItem = pw.getPlotItem()
            pw.hideButtons()

            pw2 = PlotWidget(viewBox=CustomViewBox(border=(0, 0, 0)),
                             enableMenu=False, title='Distribution of duration (R - D) errors due to noise',
                             labels={'bottom': 'Reading blocks'})
            self.durBarPlotItem = pw2.getPlotItem()
            pw2.hideButtons()

            pw3 = fp_plot
            # pw3, false_positive, false_probability, self.observedDrop, self.threeSigmaLine, \
            #     self.fourSigmaLine, fiveSigmaLine = self.doFalsePositiveReport(posCoefs)  # noqa

            # Suppress this print statement
            # print(f'observed drop: {self.observedDrop:0.2f}  max noise-induced drop: {self.maxNoiseInducedDrop:0.2f}')
            self.falsePositivePlotItem = pw3.getPlotItem()

            layout.addWidget(pw, 0, 0)
            layout.addWidget(pw2, 0, 1)
            layout.addWidget(pw3, 1, 0, 1, 2)  # (pw3, row_start, col_start, n_rows_to_span, n_cols_to_span)

            pw.plot(x - D, y, stepMode=True, fillLevel=0, brush=(0, 0, 255, 150))
            pw.addLine(y=0, z=-10, pen=[0, 0, 255])
            pw.addLine(x=0, z=+10, pen=[255, 0, 0])

            yp = max(y) * 0.75
            x1 = self.loDbar68 - D
            pw.plot(x=[x1, x1], y=[0, yp], pen=pen)

            x2 = self.hiDbar68 - D
            pw.plot(x=[x2, x2], y=[0, yp], pen=pen)

            pw.addLegend()
            legend68 = '[%0.2f,%0.2f] @ 0.6827' % (x1, x2)
            pw.plot(name=legend68)

        self.showMsg("Error bar report based on 50,000 simulations (units are readings)...")

        self.showMsg('loDbar   @ .68 ci: %8.4f' % (x1 * self.framesPerEntry()), blankLine=False)
        self.showMsg('hiDbar   @ .68 ci: %8.4f' % (x2 * self.framesPerEntry()), blankLine=False)

        yp = max(y) * 0.25
        x1 = self.loDbar95 - D
        x2 = self.hiDbar95 - D

        if plots_wanted:
            pw.plot(x=[x1, x1], y=[0, yp], pen=pen)
            pw.plot(x=[x2, x2], y=[0, yp], pen=pen)

        self.showMsg('loDbar   @ .95 ci: %8.4f' % (x1 * self.framesPerEntry()), blankLine=False)
        self.showMsg('hiDbar   @ .95 ci: %8.4f' % (x2 * self.framesPerEntry()), blankLine=False)

        if plots_wanted:
            legend95 = '[%0.2f,%0.2f] @ 0.95' % (x1, x2)
            pw.plot(name=legend95)

        yp = max(y) * 0.15
        x1 = self.loDbar99 - D
        x2 = self.hiDbar99 - D

        if plots_wanted:
            pw.plot(x=[x1, x1], y=[0, yp], pen=pen)
            pw.plot(x=[x2, x2], y=[0, yp], pen=pen)

        self.showMsg('loDbar   @ .9973 ci: %8.4f' % (x1 * self.framesPerEntry()), blankLine=False)
        self.showMsg('hiDbar   @ .9973 ci: %8.4f' % (x2 * self.framesPerEntry()), blankLine=True)

        if plots_wanted:
            legend99 = '[%0.2f,%0.2f] @ 0.9973' % (x1, x2)
            pw.plot(name=legend99)

            pw.hideAxis('left')

            pw2.plot(xdur, ydur, stepMode=True, fillLevel=0, brush=(0, 0, 255, 150))
            pw2.addLine(y=0, z=-10, pen=[0, 0, 255])
            pw2.addLine(x=0, z=+10, pen=[255, 0, 0])

        yp = max(ydur) * 0.75
        x1 = self.loDurbar68
        x2 = self.hiDurbar68

        if plots_wanted:
            pw2.plot(x=[x1, x1], y=[0, yp], pen=pen)
            pw2.plot(x=[x2, x2], y=[0, yp], pen=pen)

            pw2.addLegend()
            legend68 = '[%0.2f,%0.2f] @ 0.6827' % (x1, x2)
            pw2.plot(name=legend68)

        self.showMsg('loDurBar @ .68 ci: %8.4f' % (x1 * self.framesPerEntry()), blankLine=False)
        self.showMsg('hiDurBar @ .68 ci: %8.4f' % (x2 * self.framesPerEntry()), blankLine=False)

        yp = max(ydur) * 0.25
        x1 = self.loDurbar95
        x2 = self.hiDurbar95

        if plots_wanted:
            pw2.plot(x=[x1, x1], y=[0, yp], pen=pen)
            pw2.plot(x=[x2, x2], y=[0, yp], pen=pen)

        self.showMsg('loDurBar @ .95 ci: %8.4f' % (x1 * self.framesPerEntry()), blankLine=False)
        self.showMsg('hiDurBar @ .95 ci: %8.4f' % (x2 * self.framesPerEntry()), blankLine=False)

        if plots_wanted:
            legend95 = '[%0.2f,%0.2f] @ 0.95' % (x1, x2)
            pw2.plot(name=legend95)

        yp = max(ydur) * 0.15
        x1 = self.loDurbar99
        x2 = self.hiDurbar99

        if plots_wanted:
            pw2.plot(x=[x1, x1], y=[0, yp], pen=pen)
            pw2.plot(x=[x2, x2], y=[0, yp], pen=pen)

        self.showMsg('loDurBar @ .9973 ci: %8.4f' % (x1 * self.framesPerEntry()), blankLine=False)
        self.showMsg('hiDurBar @ .9973 ci: %8.4f' % (x2 * self.framesPerEntry()), blankLine=True)

        if plots_wanted:
            legend99 = '[%0.2f,%0.2f] @ 0.9973' % (x1, x2)
            pw2.plot(name=legend99)

            pw2.hideAxis('left')

        self.writeBarPlots.setEnabled(True)

        if self.timestampListIsEmpty(self.yTimes):
            self.showMsg('Cannot produce final report because timestamps are missing.', bold=True, color='red')
        else:
            # self.finalReport(false_positive, false_probability, error_bar_plots_available=plots_wanted)
            self.finalReport(error_bar_plots_available=plots_wanted)
            self.fillExcelReportButton.setEnabled(True)

        self.newRedrawMainPlot()  # To add envelope to solution

    def reportDropProbability(self):

        # Expected usage: observed drop has <one of strings below> of being induced by noise
        if self.observedDrop < self.threeSigmaLine:
            return f'a less than 3 sigma probability level'
        elif self.observedDrop < self.fiveSigmaLine:
            return f'a probability of {self.drop_nie_probability:0.1e}'
        else:
            self.dropSigmaEstimate = 5.0 + (self.observedDrop - self.fiveSigmaLine) / (self.fiveSigmaLine - self.fourSigmaLine)
            return f'an estimated {self.dropSigmaEstimate:0.1f} sigma level'

    def calcDetectability(self):
        if self.timeDelta == 0:
            self.showInfo(f'Cannot use the detectibilty tool on a light curve without timestamps.')
            return

        if not self.userDeterminedBaselineStats:
            self.showInfo(f'Baseline statistics have not been extracted yet.')
            return

        posCoefs = self.newCorCoefs

        durText = self.observationDurEdit.text()
        if durText == '':
            self.showInfo(f'You need to fill in an observation duration (in seconds) for this operation.')
            return
        try:
            obs_duration_secs = float(durText)
            obs_duration = int(np.ceil(obs_duration_secs / self.timeDelta))
            if obs_duration < 1:
                self.showInfo(f'The obs duration: {obs_duration} is too small to use.')
                return
        except ValueError:
            self.showInfo(f'{durText} is invalid as a duration in seconds value')
            return

        durText = self.eventDurationEdit.text()
        if durText == '':
            self.showInfo(f'You need to fill in a duration (in seconds) for this operation.')
            return
        try:
            event_duration_secs = float(durText)

            # 4.6.2 change
            event_duration = int(np.ceil(event_duration_secs / self.timeDelta))  # In readings
            # event_duration = round((event_duration_secs / self.timeDelta))  # In readings

            if event_duration < 1:
                self.showInfo(f'The event duration: {event_duration} is too small to use.')
                return
            if obs_duration <= event_duration:
                self.showInfo(f'The event duration: {event_duration} '
                              f'is too large for an observation with {obs_duration} points.')
                return

        except ValueError:
            self.showInfo(f'{durText} is invalid as a duration in seconds value')
            return

        magDropText = self.detectabilityMagDropEdit.text()
        if magDropText == '':
            self.showInfo(f'You need to fill in a magDrop for this operation.')
            return
        try:
            event_magDrop = float(magDropText)
            if event_magDrop < 0:
                self.showInfo(f'magDrop must be a positive value.')
                return
        except ValueError:
            self.showInfo(f'{magDropText} is invalid as a magDrop value')
            return

        durStep = 0.0
        durStepText = self.durStepEdit.text()
        if not durStepText == '':
            try:
                durStep = float(durStepText)
                if durStep < 0:
                    self.showInfo(f'durStep must be a positive number.')
                    return
            except ValueError:
                self.showInfo(f'{durStepText} is invalid as a durSep value')
                return

        sigma = self.sigmaB
        # Use given magDrop to calculate observed drop
        ratio = 10 ** (event_magDrop / 2.5)
        self.A = self.B / ratio
        observed_drop = self.B - self.A
        num_trials = 50000

        self.minDetectableDurationRdgs = None
        self.minDetectableDurationSecs = None

        while True:
            i = 0
            for i, checkBox in enumerate(self.targetCheckBoxes):
                if checkBox.isChecked():
                    break
            self.showMsg(f'Using lightcurve: {self.lightcurveTitles[i].text()} ...',
                         blankLine=False, alternateLogFile=self.detectabilityLogFile)
            self.showMsg(f'... processing detectability of magDrop: {event_magDrop:0.2f} '
                         f'dur(secs): {event_duration_secs:0.3f} event', alternateLogFile=self.detectabilityLogFile,
                         blankLine=False)
            QtWidgets.QApplication.processEvents()

            drops = compute_drops(event_duration=event_duration, observation_size=obs_duration,
                                  noise_sigma=sigma, corr_array=np.array(posCoefs), num_trials=num_trials)

            title = (f'Noise Induced Event (brightness drops) found in correlated noise --- '
                     f'Event duration: {event_duration} readings ------ '
                     f'Event duration: {event_duration_secs:0.2f} seconds')

            pw = PlotWidget(viewBox=CustomViewBox(border=(0, 0, 0)),
                            enableMenu=False,
                            title=title,
                            labels={'bottom': 'drop size (ADU)',
                                    'left': 'relative number of times noise produced drop'})
            pw.hideButtons()

            y, x = np.histogram(drops, bins='auto', density=True, range=(0, np.max(drops)))

            sorted_drops = np.sort(drops)
            three_sigma_line = sorted_drops[int(.997 * drops.size)]

            blackDrop = np.max(x)
            redDrop = observed_drop
            # 4.6.2 change
            if event_duration_secs < self.timeDelta:
                redDrop = observed_drop * (event_duration_secs / self.timeDelta)
                self.showMsg(f'... drop reduced from {observed_drop:0.2f} to {redDrop:0.2f} due to sub-frame exposure',
                             alternateLogFile=self.detectabilityLogFile)
            else:
                self.showMsg('', alternateLogFile=self.detectabilityLogFile, blankLine=False)
            # end 4.6.2 change
            redMinusThreeSigmaLine = redDrop - three_sigma_line

            pw.plot(x, y, stepMode=True, fillLevel=0, brush=(0, 0, 255, 150))
            pw.plot(x=[redDrop, redDrop], y=[0, 1.5 * np.max(y)], pen=pg.mkPen([255, 0, 0], width=2))
            pw.plot(x=[np.max(x), np.max(x)], y=[0, 0.25 * np.max(y)], pen=pg.mkPen([0, 0, 0], width=2))
            pw.plot(x=[three_sigma_line, three_sigma_line], y=[0, 0.25 * np.max(y)], pen=pg.mkPen([0, 255, 0], width=2))

            pw.addLegend()
            pw.plot(
                name=f'red line @ {redDrop:0.1f} = location of a {event_magDrop} magDrop event in histogram of all detected events due to noise')
            pw.plot(name=f'black line @ {blackDrop:0.1f} = max drop found in {num_trials} trials with correlated noise')
            pw.plot(name=f'green line @ {three_sigma_line:0.1f} = three_sigma_line (99.7% of drops are left of this line')
            pw.plot(name=f'B: {self.B:0.2f}  A: {self.A:0.2f} (calculated from expected magDrop of event)')
            pw.plot(name=f'magDrop: {magDropText}')

            if redMinusThreeSigmaLine > 0:
                pw.plot(name=f'red - three_sigma_line = {redMinusThreeSigmaLine:0.2f}  An event of this duration and magDrop is detectable')
                self.minDetectableDurationRdgs = event_duration  # This is in 'readings'
                self.minDetectableDurationSecs = event_duration_secs  # This is in seconds
            else:
                pw.plot(name=f'red - three_sigma_line = {redMinusThreeSigmaLine:0.2f}  Detection of this event is unlikely')

            self.detectabilityWin = pg.GraphicsWindow(title='Detectability test: False positive distribution')
            self.detectabilityWin.resize(1700, 700)
            layout = QtWidgets.QGridLayout()
            self.detectabilityWin.setLayout(layout)
            layout.addWidget(pw, 0, 0)

            pw.getPlotItem().setFixedHeight(700)
            pw.getPlotItem().setFixedWidth(1700)

            lightCurveDir = os.path.dirname(self.csvFilePath)  # This gets the folder where the light-curve.csv is located
            detectibiltyPlotPath = lightCurveDir + '/DetectabilityPlots/'
            if not os.path.exists(detectibiltyPlotPath):
                os.mkdir(detectibiltyPlotPath)
            targetFile = detectibiltyPlotPath + f'plot.detectability-dur{event_duration_secs:0.3f}-magDrop{event_magDrop:0.2f}.pyote.png'

            exporter = FixedImageExporter(pw.getPlotItem())
            exporter.makeWidthHeightInts()

            # if durStep == 0.0:
                # Always write plot for single detectability requests
            exporter.export(targetFile)
            QtWidgets.QApplication.processEvents()

            if redMinusThreeSigmaLine < 0 and not durStep == 0.0:  # A failed detect during step down
                # Only write the final plot for "find minimum duration detectability" requests
                # exporter.export(targetFile)
                # QtWidgets.QApplication.processEvents()
                self.showMsg(
                    f'Undetectability reached at magDrop: {event_magDrop:0.2f}  duration=: {event_duration_secs:0.3f}',
                    alternateLogFile=self.detectabilityLogFile)
                break

            if durStep == 0.0:
                break
            else:
                event_duration_secs -= durStep  # Try next smaller duration

                # 4.6.2 change
                event_duration = int(np.ceil(event_duration_secs / self.timeDelta))
                # event_duration = round((event_duration_secs / self.timeDelta))  # readings

                if event_duration < 1:
                    # 4.6.1 change
                    event_duration_secs += durStep  # Last successful duration
                    event_duration = int(np.ceil(event_duration_secs / self.timeDelta))
                    # end 4.6.1 change

                    break

        if self.minDetectableDurationRdgs is None:
            self.showMsg(
                f'Undetectable event @ magDrop: {event_magDrop:0.2f}  duration: {event_duration_secs:0.3f}',
                color='red', bold=True
            )
        else:
            self.showMsg(
                f'An event of duration {self.minDetectableDurationSecs:0.3f} seconds with magDrop: {event_magDrop}'
                f' is likely detectable.', color='red', bold=True
            )

            # Show an example light-curve with the 'event' included.
            obs = noise_gen_jit(obs_duration, self.sigmaB)
            obs = simple_convolve(obs, np.array(self.newCorCoefs))
            title = (f'Example light curve at the minimum detectable duration found ---  '
                     f'Event duration: {event_duration} readings ------ '
                     f'Event duration:{event_duration_secs + durStep:0.3f} seconds ------ '
                     f'magDrop: {magDropText}'
                     )
            pw = PlotWidget(viewBox=CustomViewBox(border=(0, 0, 0)),
                            enableMenu=False,
                            title=title,
                            labels={'bottom': 'reading number', 'left': 'intensity'})
            pw.hideButtons()
            pw.getPlotItem().showAxis('bottom', True)
            pw.getPlotItem().showAxis('left', True)
            pw.showGrid(y=True, alpha=.5)
            pw.getPlotItem().setFixedHeight(700)
            pw.getPlotItem().setFixedWidth(1700)

            self.detectabilityWin = pg.GraphicsWindow(
                title='Detectability test: sample light curve with minimum duration event')
            self.detectabilityWin.resize(1700, 700)
            layout = QtWidgets.QGridLayout()
            self.detectabilityWin.setLayout(layout)
            layout.addWidget(pw, 0, 0)

            obs += self.B
            center = int(obs_duration / 2)
            eventStart = center - int(event_duration / 2)

            # 4.6.1 change
            # obs[eventStart:eventStart+event_duration+1] -= observed_drop
            obs[eventStart:eventStart + event_duration] -= redDrop

            pw.plot(obs)
            pw.plot(obs, pen=None, symbol='o', symbolBrush=(0, 0, 255), symbolSize=6)

            # left_marker = pg.InfiniteLine(pos=eventStart, pen='r')
            # pw.addItem(left_marker)
            # right_marker = pg.InfiniteLine(pos=eventStart + event_duration, pen='g')
            # pw.addItem(right_marker)

            # Plot the square wave
            x_vals = []
            y_vals = []
            x_vals.append(0)
            y_vals.append(self.B)

            x_vals.append(eventStart - 1)
            y_vals.append(self.B)

            x_vals.append(eventStart)
            y_vals.append(self.A)

            x_vals.append(eventStart+event_duration)
            y_vals.append(self.A)

            x_vals.append(eventStart+event_duration + 1)
            y_vals.append(self.B)

            x_vals.append(len(obs)-1)
            y_vals.append(self.B)

            pw.plot(x_vals, y_vals, pen=pg.mkPen([255, 0, 0], width=4))

            if self.writeExampleLightcurveCheckBox.isChecked():
                # Write the example lightcurve (obs) to a csv file
                self.writeExampleLightcurveToFile(obs, self.timeDelta)

        return

    def doFalsePositiveReport(self, posCoefs, plots_wanted=True):
        d, r = self.solution
        if self.eventType == 'Donly':
            event_duration = self.right - int(np.trunc(d))  # noqa
        elif self.eventType == 'Ronly':
            event_duration = int(np.ceil(r)) - self.left
        else:
            event_duration = int(np.ceil(r - d))

        observation_size = self.right - self.left + 1

        # sigma = max(self.sigmaA, self.sigmaB)  # Changed in version 5.5.2
        sigma = self.sigmaB                      # Changed in version 5.5.2

        observed_drop = self.B - self.A
        num_trials = 50000

        # We have to 'normalize' observed_drop and sigma so that the NIE calculations don't have
        # to deal with numerical problems while doing numerical integrations and curve fitting.
        observed_drop = observed_drop / self.B
        sigma = sigma / self.B

        return self.falsePositiveReport(event_duration, num_trials, observation_size, observed_drop,
                                        posCoefs, sigma, plots_wanted=plots_wanted)

    @staticmethod
    def centeredGaussian(x, xCenter, sigma):
        c = np.sqrt(2 * np.pi)
        return np.exp(-0.5 * ((x - xCenter) / sigma)**2)/sigma/c

    def falsePositiveReport(self, event_duration, num_trials, observation_size, observed_drop, posCoefs, sigma, plots_wanted=True):

        drops = compute_drops(event_duration=event_duration, observation_size=observation_size,
                              noise_sigma=sigma, corr_array=np.array(posCoefs), num_trials=num_trials)

        sorted_drops = np.sort(drops)

        counts, bins = np.histogram(drops, bins='auto')

        counts[0] = counts[1] / 2

        # Calculate the tail start drop as the drop that encloses tail_start_fraction of the counts
        tail_start_fraction = 0.85
        num_valid_trials = np.cumsum(counts)[-1]
        # tail_start_index = np.where(np.cumsum(counts) > tail_start_fraction * num_trials)[0][0]
        tail_start_index = np.where(np.cumsum(counts) > tail_start_fraction * num_valid_trials)[0][0]

        # Find all of the non-zero counts and their locations in the tail
        debug = False
        x_tail = []
        y = []
        zero_count = 0
        bin_delta = bins[1] - bins[0]
        if debug: print(f'bin_delta: {bin_delta:0.3f}')
        for i in range(tail_start_index, counts.size):
            if counts[i] > 0:
                new_x = bins[i] + bin_delta / 2
                x_tail.append(new_x)
                y.append(counts[i] / num_trials)
                if debug: print(f'x_tail: {new_x:6.3f}  count: {counts[i]:0.6f}')
            else:
                zero_count += 1
                if debug: print(f'zero')

        loglogy = np.log(-np.log(y))
        slope, y0, *_ = np.polyfit(x_tail, loglogy, 1)

        three_sigma_estimate = sorted_drops[int(.997 * drops.size)]

        # self.showInfo(f'Calculating the noise-induced-drop probability - this can take many seconds!\n\n'
        #               f'Be patient if/while the hourglass is showing.')

        self.showMsg(f'Calculating the noise-induced-drop probability - this can take many seconds!  ' 
                     f'Be patient.', color='red', bold=True)
        QtGui.QGuiApplication.processEvents()

        two_sigma_line, three_sigma_line, four_sigma_line, five_sigma_line, self.drop_nie_probability = \
            calc_sigma_lines(observed_drop=observed_drop, three_sigma_guess=three_sigma_estimate,
                             slope=slope, y0=y0, bin_delta=bin_delta,
                             debug=debug)

        if self.drop_nie_probability == 1.0:
            indices = np.where(sorted_drops > observed_drop)
            if indices:
                index_at_observed_drop = indices[0][0]
                self.drop_nie_probability = index_at_observed_drop / num_trials

        if plots_wanted:

            # We had to normalize observed_drop to avoid numerical instability during the NIE calculations.
            # That caused the scale of bins and the sigma lines to be changed. Here we reverse those
            # changes so that the plot remain unchanged.
            observed_drop = observed_drop * self.B
            bins = bins * self.B
            three_sigma_line = three_sigma_line * self.B
            four_sigma_line = four_sigma_line * self.B
            five_sigma_line = five_sigma_line * self.B

            pw = PlotWidget(viewBox=CustomViewBox(border=(0, 0, 0)),
                            enableMenu=False,
                            title=f'Noise Induced Events (brightness drops) found in correlated noise for event duration: {event_duration}',
                            labels={'bottom': 'brightness drop', 'left': 'number of times noise produced drop'})
            pw.hideButtons()

            # Plot drops histogram
            # pw.plot(bins, counts, stepMode=True, fillLevel=0, brush=(0, 0, 255, 150))
            pw.plot(bins, counts, stepMode=True, fillLevel=0, pen=pg.mkPen([0,0,255], width=1))

            smoothedCounts = savgol_filter(counts, 15, 3)
            smoothedBins = []
            for i in range(len(bins)-1):
                smoothedBins.append((bins[i]+bins[i+1])/ 2)
            pw.plot(smoothedBins, smoothedCounts, pen=pg.mkPen([0,0,255], width=2))
            indices = np.where(smoothedCounts == np.max(smoothedCounts))
            binsAtPeak = [smoothedBins[i] for i in indices[0]]
            noisePeakPosition = np.mean(binsAtPeak)
            pw.plot(x=[noisePeakPosition, noisePeakPosition], y=[0, 1.2 * np.max(smoothedCounts)], pen=pg.mkPen([0, 0, 255], width=2))


            stdBmean = self.sigmaB / np.sqrt(observation_size - event_duration)
            if self.sigmaA is None:
                stdAmean = self.sigmaB / np.sqrt(event_duration)
            else:
                stdAmean = self.sigmaA / np.sqrt(event_duration)
            sigmaDrop = np.sqrt(stdBmean ** 2 + stdAmean ** 2)
            xCenter = observed_drop
            xG = np.linspace(-5 * sigmaDrop + xCenter, 5 * sigmaDrop + xCenter, 1000)
            binDelta = bins[1] - bins[0]
            g = self.centeredGaussian(xG, xCenter, sigmaDrop) * binDelta * num_trials # Ready to plot g versus xG
            pw.plot(x=xG, y=g, pen=pg.mkPen([0, 255, 0], width=2))

            # Plot red observed drop line
            # pw.plot(x=[observed_drop, observed_drop], y=[0, 1.5 * np.max(counts)], pen=pg.mkPen([255, 0, 0], width=2))
            pw.plot(x=[observed_drop, observed_drop], y=[0, 1.2 * np.max(g)], pen=pg.mkPen([255, 0, 0], width=2))

            # Plot black max drop line
            # pw.plot(x=[max_drop, max_drop], y=[0, 0.1 * np.max(counts)], pen=pg.mkPen([0, 0, 0], width=2))

            # Plot green sigma lines
            # pw.plot(x=[three_sigma_line, three_sigma_line], y=[0, 0.5 * np.max(counts)], pen=pg.mkPen([0, 255, 0], width=3))
            # pw.plot(x=[four_sigma_line, four_sigma_line], y=[0, 0.3 * np.max(counts)], pen=pg.mkPen([0, 255, 0], width=3))
            # pw.plot(x=[five_sigma_line, five_sigma_line], y=[0, 0.2 * np.max(counts)], pen=pg.mkPen([0, 255, 0], width=3))

            # Undo the x axis values affected by the normalization required to keep the
            # NIE calculations stable.
            # scaled_x_tail = []
            # for value in x_tail:
            #     scaled_x_tail.append(value * self.B)
            # pw.plot(scaled_x_tail, tail_count_fraction(x_tail, slope, y0) * num_trials, pen=pg.mkPen([0, 0, 0], width=3),
            #         label='fit to tail')
            # End undo

            # tailValue = observed_drop / self.B
            # gumbelValueAtObservedDrop = None
            # if tailValue >= x_tail[0]:
            #     gumbelValueAtObservedDrop = tail_count_fraction(tailValue, slope,y0) * num_trials
            # else:
            #     print("Out of range of Gumbel tail fit.")
            #     for i, value in enumerate(bins):
            #         if value >= observed_drop:
            #             gumbelValueAtObservedDrop = counts[i-1]
            #             break

            # gaussCenterValue = self.centeredGaussian(xCenter, xCenter, sigmaDrop) * binDelta * num_trials
            sigmaDistance = (xCenter - noisePeakPosition) / sigmaDrop
            self.noiseSigmaDistance = sigmaDistance
            # relativeProbability = gaussCenterValue / gumbelValueAtObservedDrop
            # print(f"relative probability of event versus noise: {relativeProbability:0.1f}")

            # Plot a nice 0 line
            # right_edge = max(observed_drop * 1.1, np.max(bins) * 1.1)
            right_edge = max(5 * sigmaDrop + xCenter, np.max(bins) * 1.1)
            left_edge = np.min([-5 * sigmaDrop + xCenter, bins[0], 0.0])
            pw.plot(x=[left_edge, right_edge], y=[0, 0], pen=pg.mkPen([180, 180, 180], width=2))

            pw.addLegend()
            # pw.plot(name=f'red line: the observed drop (B - A) extracted from lightcurve has '
            #              f'a probability {self.drop_nie_probability:0.1e} of being noise induced')
            pw.plot(name=f'vertical red line: the observed drop (B - A) extracted from lightcurve')
            pw.plot(name=f'vertical blue line: peak of noise induced drops')
            # pw.plot(name=f'black curve: LSQ fit to drop data using the Gumbel Extreme Value distribution')
            pw.plot(name=f'green curve: observed drop distribution')
            pw.plot(name=f'blue histogram: noise induced drop distribution')
            # pw.plot(name=f"relative probability of observed drop versus noise induced drop: {relativeProbability:0.1f}")
            pw.plot(name=f'sigma distance between noisePeak and observed drop: {sigmaDistance:0.1f}')
            # pw.plot(name=f'green lines are at 3 sigma (99.7%), 4 sigma (99.9938%), and 5 sigma (99.99994%)')
            # pw.plot(name=f'PyOTE reports a "pass" if red line is to the right of the three_sigma_line,')
            # pw.plot(name=f'BUT !!! If red line is left of the 5 sigma line - refer to regional coordinator.')
            # pw.plot(name=f'If the red line is left of the 5 sigma line - refer to regional coordinator.')
        else:
            pw = None
            y, x = np.histogram(drops, bins=50)
            y[0] = y[1] / 2.0

        index_of_observed_drop_inside_sorted_drops = None
        for i, value in enumerate(sorted_drops):
            if value >= observed_drop:
                index_of_observed_drop_inside_sorted_drops = i
                break
        if index_of_observed_drop_inside_sorted_drops is None:
            false_probability = 0.0
            false_positive = False
        else:
            false_probability = 1.0 - index_of_observed_drop_inside_sorted_drops / drops.size
            # Change to criteria for noise-induced-event fail
            if observed_drop <= three_sigma_line:
                false_positive = True
            else:
                false_positive = False

        del sorted_drops
        del drops

        # Changed criteria for noise-induced-event fail
        # return pw, false_positive, false_probability, observed_drop, np.max(x)
        return pw, false_positive, false_probability, observed_drop, three_sigma_line, four_sigma_line, five_sigma_line

    def displaySolution(self, subframe=True):
        D, R = self.solution

        # D and R are floats and may be fractional because of sub-frame timing.
        # We have to remove the effects of sub-frame timing to calculate the D
        # and R transition points as integers.

        solMsg2 = ''
        frameConv = float(self.yFrame[0])
        DinFrameUnits = None
        RinFrameUnits = None
        if D and R:
            Dtransition = trunc(floor(self.solution[0]))
            Rtransition = trunc(floor(self.solution[1]))

            DinFrameUnits = Dtransition * self.framesPerEntry() + frameConv
            RinFrameUnits = Rtransition * self.framesPerEntry() + frameConv

            if subframe:
                solMsg = ('D: %d  R: %d  D(subframe): %0.4f  R(subframe): %0.4f' %
                          (Dtransition, Rtransition, D, R))

                solMsg2 = ('D: %d  R: %d  D(subframe): %0.4f  R(subframe): '
                           '%0.4f' %
                           (DinFrameUnits,
                            RinFrameUnits,
                            D * self.framesPerEntry() + frameConv, R * self.framesPerEntry() + frameConv))
            else:
                solMsg = ('D: %d  R: %d' % (D, R))
            self.showMsg('in entryNum units: ' + solMsg)
            if solMsg2:
                self.showMsg('in frameNum units: ' + solMsg2, bold=True)
        elif D:
            Dtransition = trunc(floor(self.solution[0]))
            DinFrameUnits = Dtransition * self.framesPerEntry() + frameConv

            if subframe:
                solMsg = ('D: %d  D(subframe): %0.4f' % (Dtransition, D))
                solMsg2 = ('D: %d  D(subframe): %0.4f' %
                           (DinFrameUnits, D * self.framesPerEntry() + frameConv))
            else:
                solMsg = ('D: %d' % D)
            self.showMsg('in entryNum units: ' + solMsg)
            if solMsg2:
                self.showMsg('in frameNum units: ' + solMsg2, bold=True)
        else:
            Rtransition = trunc(floor(self.solution[1]))
            RinFrameUnits = Rtransition * self.framesPerEntry() + frameConv

            if subframe:
                solMsg = ('R: %d  R(subframe): %0.4f' % (Rtransition, R))
                solMsg2 = ('R: %d  R(subframe): %0.4f' %
                           (RinFrameUnits, R * self.framesPerEntry() + frameConv))
            else:
                solMsg = ('R: %d' % R)

            self.showMsg('in entryNum units: ' + solMsg)
            if solMsg2:
                self.showMsg('in frameNum units: ' + solMsg2, bold=True)

        # This function is called twice: once without a subframe calculation and then again with
        # subframe calculations enabled.  We only want to display D and/or R frames at the end
        # of the second pass
        if subframe:
            if self.showOCRcheckFramesCheckBox.isChecked():
                if self.pathToVideo:
                    if DinFrameUnits:
                        self.showAnnotatedFrame(int(DinFrameUnits), "D edge:")
                    if RinFrameUnits:
                        self.showAnnotatedFrame(int(RinFrameUnits), 'R edge:')
                    return True
        return False

    def extract_noise_parameters_from_iterative_solution(self):

        D, R = self.solution
        # D and R are floats and may be fractional because of sub-frame timing.
        # Here we remove the effects of sub-frame timing to calculate the D and
        # R transition points as integers.
        if D:
            D = trunc(floor(D))
        if R:
            R = trunc(floor(R))

        if not self.userDeterminedBaselineStats:
            self.corCoefs = []

        if D and R:  # This is a funky way to test that neither D or R is None
            self.baselineXvals = []
            self.baselineYvals = []
            for i in range(self.left, D):  # Because of range(), this gets points at index self.left to D-1 inclusive
                self.baselineXvals.append(i)
                self.baselineYvals.append(self.yValues[i])

            if R == self.solution[1]:  # If R is not a transition point
                self.baselineYvals.append(self.yValues[R])
                self.baselineXvals.append(R)

            for i in range(R + 1, self.right + 1):  # This get points at index R+1 to self.right inclusive
                self.baselineXvals.append(i)
                self.baselineYvals.append(self.yValues[i])

            self.processBaselineNoiseFromIterativeSolution()

            if D == self.solution[0]:  # if D is not a transition point
                self.processEventNoiseFromIterativeSolution(D, R - 1)  # This updates self.sigmaA including the point at D
            else:
                self.processEventNoiseFromIterativeSolution(D + 1, R - 1)  # This updates self.sigmaA

            # Try to warn user about the possible need for block integration by testing the lag 1
            # and lag 2 correlation coefficients.  The tests are just guesses on my part, so only
            # warnings are given.  Later, the Cholesky-Decomposition may fail because block integration
            # was really needed.  That is a fatal error but is trapped and the user alerted to the problem

            if len(self.corCoefs) > 1:
                if self.corCoefs[1] >= 0.7 and self.ne3NotInUseRadioButton.isChecked():
                    self.showInfo(
                        'The auto-correlation coefficient at lag 1 is suspiciously large. '
                        'This may be because the light curve needs some degree of block integration. '
                        'Failure to do a needed block integration allows point-to-point correlations caused by '
                        'the camera integration to artificially induce non-physical correlated noise.')
                elif len(self.corCoefs) > 2:
                    if self.corCoefs[2] >= 0.3 and self.ne3NotInUseRadioButton.isChecked():
                        self.showInfo(
                            'The auto-correlation coefficient at lag 2 is suspiciously large. '
                            'This may be because the light curve needs some degree of block integration. '
                            'Failure to do a needed block integration allows point-to-point correlations caused by '
                            'the camera integration to artificially induce non-physical correlated noise.')

            if self.sigmaA is None:
                self.sigmaA = self.sigmaB
        elif D:
            self.baselineXvals = []
            self.baselineYvals = []
            for i in range(self.left, D):  # Because of range(), this gets points at index self.left to D-1 inclusive
                self.baselineXvals.append(i)
                self.baselineYvals.append(self.yValues[i])
            self.processBaselineNoiseFromIterativeSolution()

            self.processEventNoiseFromIterativeSolution(D + 1, self.right)
            obsValue = self.yValues[D]
            # self.A is correct if if D is a transition value
            if aicModelValue(
                    obsValue=obsValue, B=self.B, A=self.A, sigmaB=self.sigmaB, sigmaA=self.sigmaA) == obsValue:
                pass
            else:
                # get a new self.A starting from D
                self.processEventNoiseFromIterativeSolution(D, self.right)  # 5.5.7 test

        else:  # R only
            self.baselineXvals = []
            self.baselineYvals = []
            self.processEventNoiseFromIterativeSolution(self.left, R - 1)

            for i in range(R, self.right + 1):  # 5.5.7 test This get points at index R to self.right inclusive
                self.baselineXvals.append(i)
                self.baselineYvals.append(self.yValues[i])
            self.processBaselineNoiseFromIterativeSolution()
            obsValue = self.yValues[R]
            # self.B is correct if R is a baseline value
            if aicModelValue(
                    obsValue=obsValue, B=self.B, A=self.A, sigmaB=self.sigmaB, sigmaA=self.sigmaA) == self.B:
                pass
            else:
                # get a new self.B starting from R + 1 (because R was either a transition value or an event value
                self.baselineXvals = []
                self.baselineYvals = []
                for i in range(R + 1, self.right + 1):  # 5.5.7 test This get points at index R + 1 to self.right inclusive
                    self.baselineXvals.append(i)
                    self.baselineYvals.append(self.yValues[i])
                self.processBaselineNoiseFromIterativeSolution()

        self.prettyPrintCorCoefs()

        return

    def try_to_get_solution(self):
        self.solution = None
        self.newRedrawMainPlot()
        solverGen = solver(
            eventType=self.eventType, yValues=self.yValues,
            left=self.left, right=self.right,
            sigmaB=self.sigmaB, sigmaA=self.sigmaA,
            dLimits=self.dLimits, rLimits=self.rLimits,
            minSize=self.minEvent, maxSize=self.maxEvent)
        self.cancelRequested = False
        for item in solverGen:
            if item[0] == 'fractionDone':
                pass
            elif item[0] == 'no event present':
                self.showMsg('No event fitting search criteria could be found.')
                self.runSolver = False
                break
            else:
                self.solution = item[0]
                if not self.userDeterminedBaselineStats:
                    self.B = item[1]
                if self.userTrimInEffect:
                    self.B = item[1]
                if not self.userDeterminedEventStats:
                    self.A = item[2]

    def compareFirstAndSecondPassResults(self):
        D1, R1 = self.firstPassSolution
        D2, R2 = self.secondPassSolution
        if D1:
            D1 = trunc(floor(D1))
        if D2:
            D2 = trunc(floor(D2))
        if R1:
            R1 = trunc(floor(R1))
        if R2:
            R2 = trunc(floor(R2))

        if D1 == D2 and R1 == R2:
            return

        # There is a difference in the D and/or R transition points identified
        # in the first and second passes --- alert the user.
        self.showInfo('The D and/or R transition points identified in pass 1 '
                      'are different from those found in pass 2 (after '
                      'automatic noise analysis).  '
                      'It is recommended that you '
                      'rerun the light curve using the D and R values found in '
                      'this second pass to more accurately select points for '
                      'the initial noise analysis.')

    def extractBaselineAndEventData(self):
        if self.dLimits and self.rLimits:
            left_baseline_pts = self.yValues[self.left:self.dLimits[0]]
            right_baseline_pts = self.yValues[self.rLimits[1] + 1:self.right + 1]
            baseline_pts = np.concatenate((left_baseline_pts, right_baseline_pts))
            event_pts = self.yValues[self.dLimits[1] + 1:self.rLimits[0]]
        elif self.dLimits:
            baseline_pts = self.yValues[self.left:self.dLimits[0]]
            event_pts = self.yValues[self.dLimits[1] + 1:self.right + 1]
        elif self.rLimits:
            baseline_pts = self.yValues[self.rLimits[1] + 1:self.right + 1]
            event_pts = self.yValues[self.left:self.rLimits[0]]
        else:
            self.showInfo(f'No D or R region has been marked!')
            return None, None, None, None, None, None

        B = np.mean(baseline_pts)
        Bnoise = np.std(baseline_pts)
        numBpts = len(baseline_pts)
        A = np.mean(event_pts)
        Anoise = np.std(event_pts)
        numApts = len(event_pts)

        return B, Bnoise, numBpts, A, Anoise, numApts

    def findEventWithPlots(self):
        self.findEvent(plots_wanted=True)

    def findEventNoPlots(self):
        self.findEvent(plots_wanted=False)

    def findEvent(self, plots_wanted=False):

        self.squareWaveRadioButton.setChecked(True)

        if self.timeDelta == 0.0:
            self.showInfo(f'time per reading (timeDelta) has an invalid value of 0.0\n\nCannot proceed.')
            return

        need_to_invite_user_to_verify_timestamps = False

        if self.dLimits and self.rLimits:
            self.eventType = "DandR"
            self.showMsg('Locate a "D and R" event has been selected')
        elif self.dLimits:
            self.eventType = "Donly"
            self.showMsg('Locate a "D only" event has been selected')
        elif self.rLimits:
            self.eventType = "Ronly"
            self.showMsg('Locate an "R only" event has been selected')
        else:
            self.showMsg('Use min/max event to locate a "D and R" event has been selected')

        minText = self.minEventEdit.text().strip()
        maxText = self.maxEventEdit.text().strip()

        self.minEvent = None
        self.maxEvent = None

        if minText and not maxText:
            self.showInfo('If minEvent is filled in, so must be maxEvent')
            return

        if maxText and not minText:
            self.showInfo('If maxEvent is filled in, so must be minEvent')
            return

        if minText:
            if not minText.isnumeric():
                self.showInfo('Invalid entry for min event (rdgs)')
            else:
                self.minEvent = int(minText)
                if self.minEvent < 2:
                    self.showInfo('minEvent must be greater than 1.\n\nSetting it to 2.')
                    self.minEvent = 2
                    self.minEventEdit.setText('2')

        if maxText:
            if not maxText.isnumeric():
                self.showInfo('Invalid entry for max event (rdgs)')
            else:
                self.maxEvent = int(maxText)
                if self.maxEvent < self.minEvent:
                    self.showInfo('maxEvent must be >= minEvent')
                    return
                if self.maxEvent > self.right - self.left - 1:
                    self.showInfo('maxEvent is too large for selected points')
                    return

        if minText == '':
            minText = '<blank>'
        if maxText == '':
            maxText = '<blank>'

        if not minText == '<blank>':
            self.showMsg('minEvent: ' + minText + '  maxEvent: ' + maxText)

        if not minText == '<blank>' and not maxText == '<blank>':
            self.eventType = 'DandR'
            self.dLimits = []
            self.rLimits = []
        elif not self.dLimits and not self.rLimits:
            self.showInfo('No search criteria have been entered')
            return

        if not self.ne3NotInUseRadioButton.isChecked():
            yPosition = self.targetStarYpositionSpinBox.value()
            if yPosition == 0:
                self.showInfo("You need to set a valid value for the Night Eagle 3 target's Y position.")
                return

        candFrom, numCandidates = candidateCounter(eventType=self.eventType,
                                                   dLimits=self.dLimits, rLimits=self.rLimits,
                                                   left=self.left, right=self.right,
                                                   numPts=self.right - self.left + 1,
                                                   minSize=self.minEvent, maxSize=self.maxEvent)
        if numCandidates < 0:
            self.showInfo('Search parameters are not properly specified')
            return

        if candFrom == 'usedSize':
            self.showMsg('Number of candidate solutions: ' + str(numCandidates) +
                         ' (using event min/max entries)')
        else:
            self.showMsg(
                'Number of candidate solutions: ' + str(numCandidates) +
                ' (using D/R region selections)')

        self.runSolver = True
        self.fillExcelReportButton.setEnabled(False)

        if self.runSolver:
            if self.eventType == 'DandR':
                self.showMsg('New solver results...', color='blue', bold=True)

                if candFrom == 'usedSize':
                    solverGen = find_best_event_from_min_max_size(
                        self.yValues, self.left, self.right,
                        self.minEvent, self.maxEvent)
                else:
                    solverGen = locate_event_from_d_and_r_ranges(
                        self.yValues, self.left, self.right, self.dLimits[0],
                        self.dLimits[1], self.rLimits[0], self.rLimits[1])

            elif self.eventType == 'Ronly':
                self.showMsg('New solver results...', color='blue', bold=True)
                if candFrom == 'usedSize':
                    pass
                else:
                    self.minEvent = self.rLimits[0] - self.left
                    #self.minEvent = max(self.minEvent, 2)      Graem Remove
                    self.maxEvent = self.rLimits[1] - self.left

                # Graem temporary print message
                # self.showMsg('left, right, minevent, maxevent: %.2f %.2f %.2f %.2f ' %(self.left, self.right, self.minEvent, self.maxEvent))
                solverGen = find_best_r_only_from_min_max_size(
                    self.yValues, self.left, self.right, self.minEvent,
                    self.maxEvent
                )

            else:  # Donly
                self.showMsg('New solver results...', color='blue', bold=True)
                if candFrom == 'usedSize':
                    pass
                else:
                    self.minEvent = self.right - self.dLimits[1]
                    # self.minEvent = max(self.minEvent, 2)  Graem remove
                    self.minEvent = max(self.minEvent, 1)  # 5.5.3
                    # self.maxEvent = self.right - self.dLimits[0] - 1  Graem remove the "-1"
                    self.maxEvent = self.right - self.dLimits[0]

                # Graem temporary print message
                # self.showMsg('left, right, minevent, maxevent: %.2f %.2f %.2f %.2f ' %(self.left, self.right, self.minEvent, self.maxEvent))
                solverGen = find_best_d_only_from_min_max_size(
                    self.yValues, self.left, self.right, self.minEvent,
                    self.maxEvent
                )

            if solverGen is None:
                self.showInfo('Generator version not yet implemented')
                return

            self.cancelRequested = False

            d = r = -1
            b = a = 0.0
            sigmaB = 0.0
            # sigmaA = 0.0

            for item in solverGen:
                if item[0] == 1.0:
                    self.progressBar.setValue(int(item[1] * 100))
                    self.progressBarLightcurves.setValue(int(item[1] * 100))
                    QtWidgets.QApplication.processEvents()
                    if self.cancelRequested:
                        self.cancelRequested = False
                        self.runSolver = False
                        self.showMsg('Solution search was cancelled')
                        self.progressBar.setValue(0)
                        self.progressBarLightcurves.setValue(0)
                        return
                elif item[0] == -1.0:
                    self.showMsg(
                        'No event fitting search criteria could be found.')
                    self.progressBar.setValue(0)
                    self.progressBarLightcurves.setValue(0)
                    self.runSolver = False
                    return
                else:
                    _, _, d, r, b, a, sigmaB, sigmaA, metric = item
                    if d == -1.0:
                        d = None
                    if r == -1.0:
                        r = None
                    #self.showMsg('sigB:%.2f  sigA:%.2f B:%.2f A:%.2f' %(sigmaB, sigmaA, b, a),blankLine=False)  #graem temporary print message
                    self.solution = [d, r]
                    self.progressBar.setValue(0)
                    self.progressBarLightcurves.setValue(0)

            self.showMsg('Integer (non-subframe) solution...', blankLine=False)
            self.showMsg(f"B:{b:0.2f}  A:{a:0.2f}", blankLine=False)
            self.displaySolution(subframe=False)  # First solution

            # This fills in self.sigmaB and self.sigmaA and self.B and self.A Also, it is useful
            # because it tests correlation coefficients to warn of the need for block integration
            # We do this before subFrameAdjusted() is called (a few lines lower) so that the Akaike calculations
            # for determining transition points has good sigmaA and sigmaB values to work with.
            self.extract_noise_parameters_from_iterative_solution()  # This does proper exclusions at D and R

            DfitMetric = RfitMetric = 0.0

            if not self.ne3NotInUseRadioButton.isChecked():
                # We're being asked to perform an exponential edge fit for the Night Eagle 3 camera
                snr = b / sigmaB
                if snr < 4.0:
                    self.dnrOffRadioButton.setChecked(True)
                    self.showMsg(f'The dnr of {snr:0.1f} is too low to use an NE3 exponential solution...', color='red',
                                 blankLine=False)
                    self.showMsg(f'... NE3 DNR:Off has been automatically checked.', color='red')
                self.showUnderlyingLightcurveCheckBox.setChecked(True)
                resultFound, DfitMetric, RfitMetric = self.doExpFit(b, a)
                if not resultFound:
                    return

            # Here is where we get sigmaB and sigmaA with the transition points excluded
            if self.sigmaA == 0.0:
                self.sigmaA = MIN_SIGMA
            if self.sigmaB == 0.0:
                self.sigmaB = MIN_SIGMA * 2
            subDandR, new_b, new_a, newSigmaB, newSigmaA = subFrameAdjusted(
                eventType=self.eventType, cand=(d, r), B=self.B, A=self.A,
                sigmaB=self.sigmaB, sigmaA=self.sigmaA, yValues=self.yValues,
                left=self.left, right=self.right)

            if not self.ne3NotInUseRadioButton.isChecked():
                if not self.dnrOffRadioButton.isChecked():
                    subDandR = self.solution

            D = R = 0
            if self.eventType == 'Donly' or self.eventType == 'DandR':
                D = int(subDandR[0])
            if self.eventType == 'Ronly' or self.eventType == 'DandR':
                R = int(subDandR[1])

            if (self.eventType == 'Donly' or self.eventType == 'DandR') and not D == subDandR[0]:
                if self.exponentialDtheoryPts is None:
                    pass

            if (self.eventType == 'Ronly' or self.eventType == 'DandR') and not R == subDandR[1]:
                if self.exponentialRtheoryPts is None:
                    pass

            if self.exponentialDtheoryPts is None and self.exponentialRtheoryPts is None:
                self.solution = subDandR

            # New code which repeats B, A, sigmaB, and sigmaA calculations after subframe adjustment because
            # a new D and R may have been determined
            self.extract_noise_parameters_from_iterative_solution()  # This does proper exclusions at D and R
            new_b = self.B
            newSigmaB = self.sigmaB
            new_a = self.A
            newSigmaA = self.sigmaA

            self.showMsg('Subframe adjusted solution...', blankLine=False)
            self.showMsg(f"B:{new_b:0.2f}  A:{new_a:0.2f}", blankLine=False)
            # self.showMsg(
            #     'sigB:%.2f  sigA:%.2f B:%.2f A:%.2f' %
            #     (newSigmaB, newSigmaA, new_b, new_a),
            #     blankLine=False)

            need_to_invite_user_to_verify_timestamps = self.displaySolution()  # Adjusted solution

            if not self.userDeterminedBaselineStats:
                self.B = new_b
                self.sigmaB = newSigmaB
                pass
            else:
                self.showInfo(f'You have requested a "find event" afer determining baseline statistics.\n\n'
                              f'This may be what you want because of the need to exclude certain baseline regions '
                              f'for this event ...\n\n'
                              f'But if not, be aware that the baseline statistics are now "fixed" and will affect '
                              f'all subsequent "find event" --- this may not be what you want.\n\n'
                              f'To clear this state, it is necessary to start over completely.')
            if self.userTrimInEffect:
                self.B = new_b
            if not self.userDeterminedEventStats:
                self.A = new_a
                self.sigmaA = newSigmaA

            self.dRegion = None
            self.rRegion = None

            self.showMsg('... end New solver results', color='blue', bold=True)

            if not self.ne3NotInUseRadioButton.isChecked():
                self.showMsg(f'D edge fit metric: {DfitMetric:0.4f}   R edge fit metric: {RfitMetric:0.4f}')

        if self.runSolver and self.solution:
            D, R = self.solution  # type: int
            if D is not None:
                D = round(D, 4)
            if R is not None:
                R = round(R, 4)
            self.solution = [D, R]
            if self.eventType == 'DandR':
                # ans = '(%.2f,%.2f) B: %.2f  A: %.2f' % (D, R, self.B, self.A)
                # Check for solution search based on min max event limits
                if self.maxEvent is not None:
                    if (R - D) > self.maxEvent:
                        self.newRedrawMainPlot()
                        self.showMsg('Invalid solution: max event limit constrained solution', color='red', bold=True)
                        self.showInfo('The solution is likely incorrect because the max event limit' +
                                      ' was set too low.  Increase that limit and try again.')
                        return
                    if self.minEvent >= (R - D) and self.minEvent > 2:
                        self.newRedrawMainPlot()
                        self.showMsg('Invalid solution: min event limit constrained solution!', color='red', bold=True)
                        self.showInfo('The solution is likely incorrect because the min event limit' +
                                      ' was set too high.  Decrease that limit and try again.')
                        return
                pass
            elif self.eventType == 'Donly':
                # ans = '(%.2f,None) B: %.2f  A: %.2f' % (D, self.B, self.A)
                pass
            elif self.eventType == 'Ronly':
                # ans = '(None,%.2f) B: %.2f  A: %.2f' % (R, self.B, self.A)
                pass
            else:
                raise Exception('Undefined event type')
        elif self.runSolver:
            self.showMsg('Event could not be found')

        self.newRedrawMainPlot()

        if not self.suppressReport:
            # print('we are computing error bars')
            QtWidgets.QApplication.processEvents()
            # msgBox = self.showInfoNonModal('Computing error bars')
            # msgBox.show()
            self.computeErrorBars(plots_wanted=plots_wanted)
            QtWidgets.QApplication.processEvents()
            # msgBox.close()
            QtWidgets.QApplication.processEvents()

        self.suppressReport = False

        if need_to_invite_user_to_verify_timestamps:
            self.showInfo(f'The timing of the event found depends on the correctness '
                          f'of the timestamp assigned to the D and R frames.  Since '
                          f'OCR may have produced incorrect values, the relevant video frames have been found '
                          f'and displayed for your inspection.\n\n'
                          f'Please verify visually that the timestamp values are correct.\n\n'
                          f'If they are wrong, note the correct values and use manual timestamp entry '
                          f'to "rescue" the observation.')

    def getNe3TimeCorrection(self):
        yPosition = self.targetStarYpositionSpinBox.value()
        if yPosition == 0:
            self.showInfo("You need to set a valid value for the target's Y position.")
            return None
        # We assume that NE3 cameras always have a field time of 0.0167 seconds and a 640x480 format
        fieldTime = 0.0167
        correction = fieldTime - yPosition * (fieldTime / 480)
        return -correction

    def writeNe3UsageReport(self):
        self.showMsg('')
        self.showMsg('Night Eagle 3 exponential curve fitting was employed with parameters:',
                     color='blue', bold=True)
        if self.dnrOffRadioButton.isChecked():
            self.showMsg(f'==== DNR: OFF', color='black', bold=True, blankLine=False)
        elif self.dnrLowRadioButton.isChecked():
            tcD = self.dnrLowDspinBox.value()
            tcR = self.dnrLowRspinBox.value()
            msg = f'==== DNR:LOW    TC-D: {tcD:0.2f}    TC-R: {tcR:0.2f}'
            self.showMsg(msg, color='black', bold=True, blankLine=False)
        elif self.dnrMiddleRadioButton.isChecked():
            tcD = self.dnrMiddleDspinBox.value()
            tcR = self.dnrMiddleRspinBox.value()
            msg = f'==== DNR:MIDDLE   TC-D: {tcD:0.2f}    TC-R: {tcR:0.2f}'
            self.showMsg(msg, color='black', bold=True, blankLine=False)
        elif self.dnrHighRadioButton.isChecked():
            tcD = self.dnrHighDspinBox.value()
            tcR = self.dnrHighRspinBox.value()
            msg = f'==== DNR:HIGH    TC-D: {tcD:0.2f}    TC-R: {tcR:0.2f}'
            self.showMsg(msg, color='black', bold=True, blankLine=False)

        self.showMsg(f'==== y position of target: {self.targetStarYpositionSpinBox.value():4d}',
                     color='black', bold=True)

    def scoreExpFit(self, b, a, TC, edge):
        numTheoryPoints = 5

        if edge == 'D':
            D = int(self.solution[0])

            # Select a set of yValues to use for fitting
            p1 = D - 10
            p2 = D + 10
            if p1 < 0:
                p1 = 0
            if p2 > len(self.yValues) - 1:
                p2 = len(self.yValues) - 1

            actual = self.yValues[p1:p2 + 1]

            bestMatchPoint = ex.locateIndexOfBestMatchPoint(
                numTheoryPts=numTheoryPoints, B=b, A=a, offset=0, N=TC,
                actual=actual, edge=edge
            )

            bestOffset = ex.locateBestOffset(
                numTheoryPts=numTheoryPoints, B=b, A=a, N0=TC,
                actual=actual, matchPoint=bestMatchPoint, edge=edge
            )

            self.exponentialDtheoryPts = ex.getDedgePoints(
                numPoints=numTheoryPoints, B=b, A=a, offset=bestOffset, N=TC
            )
            self.exponentialDinitialX = bestMatchPoint + p1

            self.exponentialDedge = bestMatchPoint + bestOffset + p1 + numTheoryPoints

            DfitMetric = ex.scoreDedge(numTheoryPts=numTheoryPoints, B=b, A=a, offset=bestOffset,
                                       N=TC, actual=actual, matchPoint=bestMatchPoint)
            return DfitMetric

        if edge == 'R':
            R = int(self.solution[1])

            # Select a set of yValues to use for fitting
            p1 = R - 10
            p2 = R + 10
            if p1 < 0:
                p1 = 0
            if p2 > len(self.yValues) - 1:
                p2 = len(self.yValues) - 1

            actual = self.yValues[p1:p2 + 1]

            bestMatchPoint = ex.locateIndexOfBestMatchPoint(
                numTheoryPts=numTheoryPoints, B=b, A=a, offset=0, N=TC,
                actual=actual, edge=edge
            )

            bestOffset = ex.locateBestOffset(
                numTheoryPts=numTheoryPoints, B=b, A=a, N0=TC,
                actual=actual, matchPoint=bestMatchPoint, edge=edge
            )

            self.exponentialRtheoryPts = ex.getRedgePoints(
                numPoints=numTheoryPoints, B=b, A=a, offset=bestOffset, N=TC
            )
            self.exponentialRinitialX = bestMatchPoint + p1

            self.exponentialRedge = bestMatchPoint + bestOffset + p1 + numTheoryPoints

            RfitMetric = ex.scoreRedge(numTheoryPts=numTheoryPoints, B=b, A=a, offset=bestOffset,
                                       N=TC, actual=actual, matchPoint=bestMatchPoint)
            return RfitMetric

    def refineNe3TimeConstant(self, b, a, TC, edge):
        def doRefinementSteps(stepSize, timeConstant):
            oldScore = self.scoreExpFit(b=b, a=a, TC=timeConstant, edge=edge)
            direction = 1
            noImprovement = False
            while True:
                # Take a step
                timeConstant += direction * stepSize
                if timeConstant <= 0:
                    timeConstant = 0.1
                newScore = self.scoreExpFit(b=b, a=a, TC=timeConstant, edge=edge)
                if newScore < oldScore:
                    oldScore = newScore
                else:
                    if noImprovement:
                        break
                    direction *= -1
                    timeConstant += direction * stepSize
                    noImprovement = True
            return timeConstant - direction * stepSize

        newTC = doRefinementSteps(1.0, TC)
        newTC = doRefinementSteps(0.1, newTC)
        newTC = doRefinementSteps(0.01, newTC)

        return newTC

    def doExpFit(self, b, a):

        DfitMetric = RfitMetric = 0.0

        if self.userDeterminedEventStats:
            a = self.A

        if self.userDeterminedBaselineStats:
            b = self.B

        timeCorrection = self.getNe3TimeCorrection()
        if timeCorrection is None:
            return False, 0.0, 0.0

        # Convert timeCorrection to frame/field fraction
        deltaPosition = timeCorrection / self.timeDelta

        if self.eventType == 'Donly' or self.eventType == 'DandR':
            if self.dnrLowRadioButton.isChecked():
                DtimeConstant = self.dnrLowDspinBox.value()
            elif self.dnrMiddleRadioButton.isChecked():
                DtimeConstant = self.dnrMiddleDspinBox.value()
            elif self.dnrHighRadioButton.isChecked():
                DtimeConstant = self.dnrHighDspinBox.value()
            elif self.dnrOffRadioButton.isChecked():
                DtimeConstant = 0.001
            else:
                self.showInfo('Programming error - this point should never be reached')
                return False, 0.0, 0.0

            if not self.dnrOffRadioButton.isChecked():
                DtimeConstant = self.refineNe3TimeConstant(b=b, a=a, TC=DtimeConstant, edge='D')

                if self.dnrLowRadioButton.isChecked():
                    self.dnrLowDspinBox.setValue(DtimeConstant)
                elif self.dnrMiddleRadioButton.isChecked():
                    self.dnrMiddleDspinBox.setValue(DtimeConstant)
                elif self.dnrHighRadioButton.isChecked():
                    self.dnrHighDspinBox.setValue(DtimeConstant)

            self.showMsg(f'D timeconstant: {DtimeConstant:0.2f}', color='red', bold=True)

            DfitMetric = self.scoreExpFit(b=b, a=a, TC=DtimeConstant, edge='D')

            # Here we add self.timeDelta to follow the convention of the MLE solver that the output of the
            # camera shows what happened one timeDelta in the past and 'fudges' the solution so that
            # time-on-the-wire matches time-in-the-camera
            if self.fieldMode:
                self.solution[0] = self.exponentialDedge + 0.5
            else:
                self.solution[0] = self.exponentialDedge + 1.0

            self.solution[0] += deltaPosition

        if self.eventType == 'Ronly' or self.eventType == 'DandR':

            if self.dnrLowRadioButton.isChecked():
                RtimeConstant = self.dnrLowRspinBox.value()
            elif self.dnrMiddleRadioButton.isChecked():
                RtimeConstant = self.dnrMiddleRspinBox.value()
            elif self.dnrHighRadioButton.isChecked():
                RtimeConstant = self.dnrHighRspinBox.value()
            elif self.dnrOffRadioButton.isChecked():
                RtimeConstant = 0.001
            else:
                self.showInfo('Programming error - this point should never be reached')
                return False, 0.0, 0.0

            if not self.dnrOffRadioButton.isChecked():
                RtimeConstant = self.refineNe3TimeConstant(b=b, a=a, TC=RtimeConstant, edge='R')
                if self.dnrLowRadioButton.isChecked():
                    self.dnrLowRspinBox.setValue(RtimeConstant)
                elif self.dnrMiddleRadioButton.isChecked():
                    self.dnrMiddleRspinBox.setValue(RtimeConstant)
                elif self.dnrHighRadioButton.isChecked():
                    self.dnrHighRspinBox.setValue(RtimeConstant)

            self.showMsg(f'R timeconstant: {RtimeConstant:0.2f}', color='red', bold=True)

            RfitMetric = self.scoreExpFit(b=b, a=a, TC=RtimeConstant, edge='R')

            # Here we add self.timeDelta to follow the convention of the MLE solver that the output of the
            # camera shows what happened one timeDelta in the past and 'fudges' the solution so that
            # time-on-the-wire matches time-in-the-camera
            if self.fieldMode:
                self.solution[1] = self.exponentialRedge + 0.5
            else:
                self.solution[1] = self.exponentialRedge + 1.0

            self.solution[1] += deltaPosition

        return True, DfitMetric, RfitMetric

    def fillTableViewOfData(self):

        self.table.setRowCount(self.dataLen)
        self.table.setVerticalHeaderLabels([str(i) for i in range(self.dataLen)])
        self.table.setFont(QtGui.QFont('Arial', 10))

        if not self.yFrame:
            self.showInfo(f'yFrame is empty.  A problem with block integration?. Try specifying block size.')
            return

        min_frame = int(trunc(float(self.yFrame[0])))
        max_frame = int(trunc(float(self.yFrame[-1])))
        if self.frameNumSpinBox.isEnabled():
            self.frameNumSpinBox.setMinimum(min_frame)
            self.frameNumSpinBox.setMaximum(max_frame)

        for i in range(self.dataLen):
            if self.yValues is None:  # Handles a 'cancel' during file select
                return
            neatStr = fp.to_precision(self.yValues[i], 6)
            newitem = QtWidgets.QTableWidgetItem(str(neatStr))
            self.table.setItem(i, self.targetIndex + 2, newitem)
            newitem = QtWidgets.QTableWidgetItem(str(self.yTimes[i]))  # Add timestamps
            self.table.setItem(i, 1, newitem)  # Put timestamps in column 1
            frameNum = float(self.yFrame[i])
            if not np.ceil(frameNum) == np.floor(frameNum):  # noqa
                self.fieldMode = True
            newitem = QtWidgets.QTableWidgetItem(str(self.yFrame[i]))
            self.table.setItem(i, 0, newitem)  # Put frame numbers in column 0
            nextColumn = 2
            if len(self.LC1) > 0:
                neatStr = fp.to_precision(self.LC1[i], 6)
                newitem = QtWidgets.QTableWidgetItem(str(neatStr))
                if not nextColumn == self.targetIndex + 2:
                    self.table.setItem(i, 2, newitem)  # Put LC1 in column 2
                nextColumn += 1
            if len(self.LC2) > 0:
                neatStr = fp.to_precision(self.LC2[i], 6)
                newitem = QtWidgets.QTableWidgetItem(str(neatStr))
                if not nextColumn == self.targetIndex + 2:
                    self.table.setItem(i, 3, newitem)  # Put LC2 in column 3
                nextColumn += 1
            if len(self.LC3) > 0:
                neatStr = fp.to_precision(self.LC3[i], 6)
                newitem = QtWidgets.QTableWidgetItem(str(neatStr))
                if not nextColumn == self.targetIndex + 2:
                    self.table.setItem(i, 4, newitem)  # Put LC3 in column 4
                nextColumn += 1
            if len(self.LC4) > 0:
                neatStr = fp.to_precision(self.LC4[i], 6)
                newitem = QtWidgets.QTableWidgetItem(str(neatStr))
                if not nextColumn == self.targetIndex + 2:
                    self.table.setItem(i, 5, newitem)  # Put LC4 in column 5
                nextColumn += 1
            if len(self.extra) > 0:
                for k, lightcurve in enumerate(self.extra):
                    neatStr = fp.to_precision(lightcurve[i], 6)
                    newitem = QtWidgets.QTableWidgetItem(str(neatStr))
                    if not nextColumn == self.targetIndex + 2:
                        self.table.setItem(i, 6 + k, newitem)  # Put extras in order in next columns
                    nextColumn += 1
            if len(self.demoLightCurve) > 0:  # This is only used for detectability demonstration
                neatStr = fp.to_precision(self.demoLightCurve[i], 6)
                newitem = QtWidgets.QTableWidgetItem(str(neatStr))
                self.table.setItem(i, nextColumn, newitem)

        self.table.resizeColumnsToContents()
        self.writeCSVButton.setEnabled(True)

    def doManualTimestampEntry(self):

        self.solution = None
        self.newRedrawMainPlot()
        QtGui.QGuiApplication.processEvents()

        errmsg = ''
        while errmsg != 'ok':
            try:
                errmsg, manualTime, dataEntered, actualFrameCount, expectedFrameCount = \
                    manualTimeStampEntry(self.yFrame, TSdialog(), flashFrames=self.flashEdges, timeDelta=self.timeDelta)
            except AttributeError:
                self.showInfo('There is no csv data available yet.')
                return
            if errmsg != 'ok':
                if errmsg == 'cancelled':
                    return
                else:
                    self.showInfo(errmsg)
            else:
                self.showMsg(dataEntered, bold=True)
                if abs(actualFrameCount - expectedFrameCount) >= 0.12:
                    msg = (
                        f'Possible dropped readings !!!\n\n'
                        f'Reading count input: {actualFrameCount:.2f}  \n\n'
                        f'Reading count computed from frame rate: {expectedFrameCount:.2f}'
                    )
                    self.showMsg(msg, color='red', bold=True)
                    self.showInfo(msg)

                # If user cancelled out of timestamp entry dialog,
                # then manualTime will be an empty list.
                if manualTime:
                    self.yTimes = manualTime[:]
                    self.fullDataDictionary['timeInfo'] = manualTime[:]

                    self.VizieRdict = {
                        "timestamps": None,
                        "yValues": None,
                        "yStatus": None,
                    }
                    self.timeDelta, self.droppedFrames, self.cadenceViolation, self.errRate, _ = \
                        getTimeStepAndOutliers(self.yTimes, self.yValues, self.yStatus, VizieRdict=self.VizieRdict)
                    self.frameTimeEdit.setText(fp.to_precision(self.timeDelta, 6))

                    self.fillTableViewOfData()
                    self.newRedrawMainPlot()
                    self.showMsg(
                        'timeDelta: ' + fp.to_precision(self.timeDelta, 6) +
                        ' seconds per reading' +
                        ' (timeDelta calculated from manual input timestamps)',
                        blankLine=False)
                    self.showMsg(
                        'timestamp error rate: ' + fp.to_precision(100 *
                                                                   self.errRate,
                                                                   3) + '%')
                    self.fillTableViewOfData()
                    self.timeAxis.update()

    def enableDisableFrameViewControls(self, state_to_set):
        self.viewFrameButton.setEnabled(state_to_set)
        self.frameNumSpinBox.setEnabled(state_to_set)
        self.fieldViewCheckBox.setEnabled(state_to_set)
        self.flipXaxisCheckBox.setEnabled(state_to_set)
        self.flipYaxisCheckBox.setEnabled(state_to_set)

    def readDataFromFile(self):

        self.initializeVariablesThatDontDependOnAfile()
        self.blockSize = 1
        self.fieldMode = False

        self.metricLimitLeft = None
        self.metricLimitRight = None

        self.pathToVideo = None

        self.enableDisableFrameViewControls(state_to_set=False)

        self.disableAllButtons()
        self.mainPlot.clear()
        self.textOut.clear()
        self.initializeTableView()

        if self.externalCsvFilePath is None:
            # Open a file select dialog
            self.csvFilePath, _ = QFileDialog.getOpenFileName(
                self,  # parent
                "Select light curve csv file",  # title for dialog
                self.settings.value('lightcurvedir', ""),  # starting directory
                "Csv files (*.csv)")
        else:
            self.csvFilePath = self.externalCsvFilePath
            self.externalCsvFilePath = None

        if self.csvFilePath:
            self.targetKey = ''
            self.addSourceFileToFitMetricTxtFile()
            self.firstLightCurveDisplayed = False
            self.availableLightCurvesForDisplay = []
            self.curveSelectionComboBox.clear()

            # These are used to (try to) avoid exception when a new file is read while a reference curve is active
            self.smoothingIntervalSpinBox.setValue(0)
            self.yRefStar = []
            self.smoothSecondary = []

            QtGui.QGuiApplication.processEvents()

            self.initializeLightcurvePanel()

            self.initializeModelLightcurvesPanel()

            self.userDataSetAdditions = []

            # Get rid of any previously displayed model and displayed metric info
            self.modelY = None
            self.modelYsamples = None
            self.fitMetricEdit.clear()
            self.fitMetricChangeEdit.clear()
            self.fitMetricChangeEdit.setStyleSheet(None)
            self.fitStatus = None
            self.bestFit = None
            self.modelTimeOffset = None

            QtWidgets.QApplication.processEvents()

            self.userDeterminedBaselineStats = False
            self.userDeterminedEventStats = False
            self.setWindowTitle('PYOTE Version: ' + version.version() + '  File being processed: ' + self.csvFilePath)
            dirpath, _ = os.path.split(self.csvFilePath)
            self.logFile, _ = os.path.splitext(self.csvFilePath)
            self.logFile = self.logFile + '.pyote_log.txt'

            # self.detectabilityLogFile, _ = os.path.splitext(self.csvFilePath)
            # self.detectabilityLogFile = self.detectabilityLogFile + '.pyote_detectability.txt'

            lightCurveDir = os.path.dirname(self.csvFilePath)  # This gets the folder where the light-curve.csv is located
            self.detectabilityLogFile = os.path.join(lightCurveDir, 'Detectability')
            if not os.path.exists(self.detectabilityLogFile):
                os.mkdir(self.detectabilityLogFile)
            head, tail = os.path.split(self.csvFilePath)
            name_to_use, _ = os.path.splitext(tail)
            self.detectabilityLogFile = os.path.join(self.detectabilityLogFile, name_to_use)
            self.detectabilityLogFile = self.detectabilityLogFile + '.pyote_detectability.txt'

            # self.normalizationLogFile, _ = os.path.splitext(self.csvFilePath)
            # self.normalizationLogFile = self.normalizationLogFile + '.pyote_normalization.txt'

            lightCurveDir = os.path.dirname(
                self.csvFilePath)  # This gets the folder where the light-curve.csv is located
            self.normalizationLogFile = os.path.join(lightCurveDir, 'Normalization')
            if not os.path.exists(self.normalizationLogFile):
                os.mkdir(self.normalizationLogFile)
            head, tail = os.path.split(self.csvFilePath)
            name_to_use, _ = os.path.splitext(tail)
            self.normalizationLogFile = os.path.join(self.normalizationLogFile, name_to_use)
            self.normalizationLogFile = self.normalizationLogFile + '.pyote_normalization.txt'

            curDateTime = datetime.datetime.today().ctime()
            self.showMsg('')
            self.showMsg(
                '#' * 20 + ' PYOTE ' + version.version() + '  session started: ' + curDateTime + '  ' + '#' * 20)

            self.showMsg('', alternateLogFile=self.detectabilityLogFile)
            self.showMsg(
                '#' * 20 + ' PYOTE ' + version.version() + '  session started: ' + curDateTime + '  ' + '#' * 20,
                alternateLogFile=self.detectabilityLogFile)

            self.showMsg('', alternateLogFile=self.normalizationLogFile)
            self.showMsg(
                '#' * 20 + ' PYOTE ' + version.version() + '  session started: ' + curDateTime + '  ' + '#' * 20,
                alternateLogFile=self.normalizationLogFile)

            # Make the directory 'sticky'
            self.settings.setValue('lightcurvedir', dirpath)
            self.settings.sync()
            self.showMsg('filename: ' + self.csvFilePath, bold=True, color="red")

            # columnPrefix = self.pymovieDataColumnPrefixComboBox.currentText()
            columnPrefix = 'signal'

            try:
                self.droppedFrames = []
                self.cadenceViolation = []
                self.fullDataDictionary = {}
                try:
                    frame, time, value, self.secondary, self.ref2, self.ref3, self.extra, \
                        self.aperture_names, self.headers = \
                        readLightCurve(self.csvFilePath, pymovieColumnType=columnPrefix, pymovieDict=self.fullDataDictionary)
                except ValueError as e:
                    self.showInfo(f'{e}')
                    return
                # self.showMsg(f'If the csv file came from PyMovie - columns with prefix: {columnPrefix} will be read.')
                values = [float(item) for item in value]
                self.yValues = np.array(values)  # yValues = curve to analyze
                self.dataLen = len(self.yValues)
                self.LC1 = np.array(values)

                self.pymovieFileInUse = False

                # Check headers to see if this is a PyMovie file.  Grab the
                # path to video file if it is a PyMovie file
                for header in self.headers:
                    if header.startswith('# PyMovie'):
                        self.pymovieFileInUse = True
                    if header.startswith('# PyMovie') or header.startswith('Limovie'):
                        for line in self.headers:
                            if line.startswith('# source:') or line.startswith('"FileName :'):

                                if line.startswith('# source:'):  # PyMovie format
                                    self.pathToVideo = line.replace('# source:', '', 1).strip()
                                    _, self.ext = os.path.splitext(self.pathToVideo)

                                if line.startswith('"FileName :'):  # Limovie format
                                    self.pathToVideo = line.replace('"FileName :', '', 1).strip()
                                    self.pathToVideo = self.pathToVideo.strip('"')
                                    _, self.ext = os.path.splitext(self.pathToVideo)

                                if os.path.isfile(self.pathToVideo):
                                    # _, self.ext = os.path.splitext(self.pathToVideo)
                                    if self.ext == '.avi':
                                        ans = readAviFile(0, self.pathToVideo)
                                        if not ans['success']:
                                            self.showMsg(
                                                f'Attempt to read .avi file gave errmg: {ans["errmsg"]}',
                                                color='red', bold=True)
                                            self.pathToVideo = None
                                        else:
                                            self.showMsg(f'fourcc code of avi: {ans["fourcc"]}', blankLine=False)
                                            self.showMsg(f'fps: {ans["fps"]}', blankLine=False)
                                            self.showMsg(f'avi contains {ans["num_frames"]} frames')
                                            # Enable frame view controls
                                            self.enableDisableFrameViewControls(state_to_set=True)
                                    elif self.ext == '.ravf':
                                        self.pathToVideo = None
                                    elif self.ext == '.fit':
                                        self.pathToVideo = None
                                    elif self.ext == '.ser':
                                        ans = readSerFile(0, self.pathToVideo)
                                        if not ans['success']:
                                            self.showMsg(
                                                f'Attempt to read .ser file gave errmg: {ans["errmsg"]}',
                                                color='red', bold=True)
                                            self.pathToVideo = None
                                        else:
                                            # Enable frame view controls
                                            self.enableDisableFrameViewControls(state_to_set=True)
                                    elif self.ext == '':
                                        ans = readFitsFile(0, self.pathToVideo)
                                        if not ans['success']:
                                            self.showMsg(
                                                f'Attempt to read FITS folder gave errmg: {ans["errmsg"]}',
                                                color='red', bold=True)
                                            self.pathToVideo = None
                                        else:
                                            # Enable frame view controls
                                            self.showMsg(f'{ans["num_frames"]} .fits files were found in FITS folder')
                                            self.enableDisableFrameViewControls(state_to_set=True)
                                    elif self.ext == '.adv':
                                        self.fillVizieRtabFromHeaders()  # Added 5.6.9
                                        # For now, we assume that .adv files have embedded timestamps and
                                        # so there is no need to display frames for visual OCR verification
                                        self.pathToVideo = None
                                    elif self.ext == '.aav':
                                        ans = readAavFile(0, self.pathToVideo)
                                        if not ans['success']:
                                            self.showMsg(
                                                f'Attempt to read .aav file gave errmg: {ans["errmsg"]}',
                                                color='red', bold=True)
                                            self.pathToVideo = None
                                        else:
                                            # Enable frame view controls
                                            self.enableDisableFrameViewControls(state_to_set=True)
                                    else:
                                        self.showMsg(f'Unexpected file type of {self.ext} found.')
                                else:
                                    self.showMsg(f'video source file {self.pathToVideo} could not be found.')
                                    self.pathToVideo = None

                # Automatically select all points
                # noinspection PyUnusedLocal
                self.yStatus = [INCLUDED for _i in range(self.dataLen)]

                refStar = [float(item) for item in self.secondary]

                vals = [float(item) for item in self.secondary]
                self.LC2 = np.array(vals)

                vals = [float(item) for item in self.ref2]
                self.LC3 = np.array(vals)

                vals = [float(item) for item in self.ref3]
                self.LC4 = np.array(vals)

                # A pymovie csv file can have more than 4 lightcurves.  The 'extra' lightcurves beyond
                # the standard max of 4 are placed in self.extra which is an array of lightcurves
                if self.extra:
                    for i, light_curve in enumerate(self.extra):
                        vals = [float(item) for item in light_curve]
                        self.extra[i] = np.array(vals[:])

                self.yFrame = []  # 5.6.6 change (erases history - problem reading long file after short file - list index exception
                self.initializeTableView()
                self.yOffsetSpinBoxes[0].setEnabled(True)

                # If no timestamps were found in the input file, prompt for manual entry
                if self.timestampListIsEmpty(time):
                    self.showMsg('Manual entry of timestamps is required.',
                                 bold=True)
                    # If the user knew there were no timestamps, the is no
                    # reason to show info box.
                    # if not self.manualTimestampCheckBox.isChecked():
                    if True:
                        self.showInfo('This file does not contain timestamp '
                                      'entries --- manual entry of two '
                                      'timestamps is required.'
                                      '\n\nEnter the timestamp '
                                      'values that the avi '
                                      'processing software (Limovie, Tangra, '
                                      'etc) would have produced '
                                      'had the OCR process not failed using the '
                                      'View frame button to display the frames '
                                      'you want to use for timestamp purposes.\n\n'
                                      'By working in this manner, you can continue '
                                      'processing the file as though OCR had '
                                      'succeeded and then follow the standard '
                                      'procedure for reporting results through '
                                      'the IOTA event reporting spreadsheet ('
                                      'which will make the needed corrections for camera delay and VTI offset).')

                self.showMsg('=' * 20 + ' file header lines ' + '=' * 20, bold=True, blankLine=False)
                for item in self.headers:
                    self.showMsg(item, blankLine=False)
                self.showMsg('=' * 20 + ' end header lines ' + '=' * 20, bold=True)

                if self.ext == ".adv":
                    self.fillVizieRtabFromHeaders()  # Added 5.6.9

                if self.ext == ".ravf":
                    self.fillVizieRtabFromRavfHeaders()  # Added 5.7.2

                self.yTimes = time[:]
                self.yValues = np.array(values)
                self.yValCopy = np.ndarray(shape=(len(self.yValues),))
                np.copyto(self.yValCopy, self.yValues)
                self.yRefStarCopy = np.array(refStar)

                self.dataLen = len(self.yValues)
                self.yFrame = frame[:]

                # Automatically select all points
                # noinspection PyUnusedLocal
                self.yStatus = [INCLUDED for _i in range(self.dataLen)]
                self.left = 0
                self.right = self.dataLen - 1

                self.mainPlot.autoRange()

                self.mainPlot.setMouseEnabled(x=True, y=False)

                self.setDataLimits.setEnabled(True)
                self.writePlot.setEnabled(True)

                self.markDzone.setEnabled(True)
                self.markRzone.setEnabled(True)
                self.singlePointDropButton.setEnabled(True)
                self.markEzone.setEnabled(not self.ne3NotInUseRadioButton.isChecked())

                self.calcFlashEdge.setEnabled(True)
                self.magDropSqwaveEdit.setEnabled(True)
                self.minEventEdit.setEnabled(True)
                self.maxEventEdit.setEnabled(True)
                self.locateEvent.setEnabled(True)

                self.firstPassPenumbralFit = True

                self.doBlockIntegration.setEnabled(True)
                self.blockSizeEdit.setEnabled(True)
                self.startOver.setEnabled(True)

                self.VizieRdict = {
                    "timestamps": None,
                    "yValues": None,
                    "yStatus": None,
                }
                self.timeDelta, self.droppedFrames, self.cadenceViolation, self.errRate, timingReport = \
                    getTimeStepAndOutliers(self.yTimes, self.yValues, self.yStatus, VizieRdict=self.VizieRdict)

                # added in 5.0.1

                if not timingReport == []:
                    if len(timingReport) > 100:
                        self.showInfo(f'{len(timingReport)} timing errors were found. Only the first 100 are printed.\n\n'
                                      f'All will be shown on the plot.')
                        for line in timingReport[0:99]:
                            self.showMsg(line, blankLine=False, bold=True, color='red')
                    else:
                        for line in timingReport:
                            self.showMsg(line, blankLine=False, bold=True, color='red')
                else:
                    self.showMsg('No time step irregularities found.')

                self.fillTableViewOfData()

                self.frameTimeEdit.setText(fp.to_precision(self.timeDelta, 6))

                self.showMsg('timeDelta: ' + fp.to_precision(self.timeDelta, 6) + ' seconds per reading',
                             blankLine=False)
                self.showMsg('timestamp error rate: ' + fp.to_precision(100 *
                                                                        self.errRate, 3) + '%')

                # self.changePrimary()  # Done only to fill in the light-curve name boxes
                self.solution = None
                #     self.newRedrawMainPlot()
                self.mainPlot.autoRange()

                if self.droppedFrames or self.cadenceViolation:
                    self.showTimestampErrors.setEnabled(True)
                    self.showTimestampErrors.setChecked(True)
                self.newRedrawMainPlot()
                self.mainPlot.autoRange()

                # if self.timeDelta == 0.0 and not self.manualTimestampCheckBox.isChecked():
                if self.timeDelta == 0.0:
                    self.showInfo("Analysis of timestamp fields resulted in "
                                  "an invalid timeDelta of 0.0\n\nSuggestion: switch to the Manual timestamps tab"
                                  " and press the 'Manual timestamp entry' button."
                                  "\n\nThis will give you a chance to "
                                  "manually correct the timestamps using "
                                  "the data available in the table in the "
                                  "lower left corner or incorporate flash timing data.")
            except Exception as e:
                self.showMsg(str(e))
                self.showMsg(f'This error may be because the data column name {columnPrefix} does not exist or an unhandled exception.')

    @staticmethod
    def decimalToDMS(dval):
        decimalValue = float(dval)
        deg = int(decimalValue)
        md = abs(decimalValue - deg) * 60
        minutes = int(md)
        sec = (md - minutes) * 60

        # min = int(decimalValue-deg)
        # sec = (decimalValue-deg) * 3600 % 60
        return f'{deg:d}', f'{minutes:d}', f'{sec:0.3f}'

    def fillVizieRtabFromHeaders(self):
        try:
            for header in self.headers:
                if header.startswith("# date at frame"):
                    try:
                        parts = header.split(":")
                        parts = parts[1].split("-")
                        self.vzDateYearSpinner.setValue(int(parts[0]))
                        self.vzDateMonthSpinner.setValue(int(parts[1]))
                        self.vzDateDaySpinner.setValue(int(parts[2]))
                    except Exception:  # noqa
                        self.showMsg("Failed to parse ADV header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#OBSERVER:"):
                    try:
                        parts = header.split(":")
                        self.vzObserverNameEdit.setText(parts[1].strip())
                    except Exception:  # noqa
                        self.showMsg("Failed to parse ADV header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#LATITUDE:"):
                    try:
                        parts = header.split(":")
                        deg, minutes,sec = self.decimalToDMS(parts[1].strip())
                        self.vzSiteLatDegEdit.setText(deg)
                        self.vzSiteLatMinEdit.setText(minutes)
                        self.vzSiteLatSecsEdit.setText(sec)
                    except Exception:  # noqa
                        self.showMsg("Failed to parse ADV header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#LONGITUDE:"):
                    try:
                        parts = header.split(":")
                        deg, minutes, sec = self.decimalToDMS(parts[1].strip())
                        self.vzSiteLongDegEdit.setText(deg)
                        self.vzSiteLongMinEdit.setText(minutes)
                        self.vzSiteLongSecsEdit.setText(sec)
                    except Exception:  # noqa
                        self.showMsg("Failed to parse ADV header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#ALTITUDE:"):
                    try:
                        parts = header.split(":")
                        self.vzSiteAltitudeEdit.setText(parts[1].strip())
                    except Exception:  # noqa
                        self.showMsg("Failed to parse ADV header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#OBJNAME:"):
                    try:
                        parts = header.split(":")
                        parts = parts[1].split("occ.")
                        # print(parts)
                        p1 = parts[0].split(")")
                        self.vzAsteroidNameEdit.setText(p1[1].strip())
                        asteroidId = p1[0].split("(")
                        self.vzAsteroidNumberEdit.setText(asteroidId[1].strip())
                        p2 = parts[1].strip().split(" ")
                        if p2[0] == 'UCAC4':
                            self.vzStarUCAC4Edit.setText(p2[1])
                        elif p2[0] == "Tycho2":
                            self.vzStarTycho2Edit.setText(p2[1])
                        elif p2[0] == 'Hipparcos':
                            self.vzStarHipparcosEdit.setText(p2[1])
                    except Exception:  # noqa
                        self.showMsg("Failed to parse ADV header: " + header, bold=True, color='red')
                    continue

        except Exception:  # noqa
            self.showMsg("Failed to parse some item in ADV meta-data.", bold=True, color='red')
            print("Failed to parse some item in ADV meta-data.")

    def fillVizieRtabFromRavfHeaders(self):
        try:
            for header in self.headers:
                if header.startswith("# date at frame"):
                    try:
                        parts = header.split(":")
                        parts = parts[1].split("-")
                        self.vzDateYearSpinner.setValue(int(parts[0]))
                        self.vzDateMonthSpinner.setValue(int(parts[1]))
                        self.vzDateDaySpinner.setValue(int(parts[2]))
                    except Exception:  # noqa
                        self.showMsg("Failed to parse RAVF header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#OBSERVER:"):
                    try:
                        parts = header.split(":")
                        self.vzObserverNameEdit.setText(parts[1].strip())
                    except Exception:  # noqa
                        self.showMsg("Failed to parse RAVF header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#LATITUDE:"):
                    try:
                        parts = header.split(":")
                        deg, minutes,sec = self.decimalToDMS(parts[1].strip())
                        self.vzSiteLatDegEdit.setText(deg)
                        self.vzSiteLatMinEdit.setText(minutes)
                        self.vzSiteLatSecsEdit.setText(sec)
                    except Exception:  # noqa
                        self.showMsg("Failed to parse RAVF header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#LONGITUDE:"):
                    try:
                        parts = header.split(":")
                        deg, minutes, sec = self.decimalToDMS(parts[1].strip())
                        self.vzSiteLongDegEdit.setText(deg)
                        self.vzSiteLongMinEdit.setText(minutes)
                        self.vzSiteLongSecsEdit.setText(sec)
                    except Exception:  # noqa
                        self.showMsg("Failed to parse RAVF header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#ALTITUDE:"):
                    try:
                        parts = header.split(":")
                        self.vzSiteAltitudeEdit.setText(parts[1].strip())
                    except Exception:  # noqa
                        self.showMsg("Failed to parse RAVF header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#OCCULTATION-OBJECT-NUMBER:"):
                    try:
                        parts = header.split(":")
                        self.vzAsteroidNumberEdit.setText(parts[1].strip())
                    except Exception:  # noqa
                        self.showMsg("Failed to parse RAVF header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#OCCULTATION-OBJECT-NAME:"):
                    try:
                        parts = header.split(":")
                        self.vzAsteroidNameEdit.setText(parts[1].strip())
                    except Exception:  # noqa
                        self.showMsg("Failed to parse ADV header: " + header, bold=True, color='red')
                    continue
                if header.startswith("#OCCULTATION-STAR"):
                    try:
                        parts = header.split(":")
                        p2 = parts[1].strip().split(" ")
                        if p2[0] == 'UCAC4':
                            self.vzStarUCAC4Edit.setText(p2[1])
                        elif p2[0] == "Tycho2":
                            self.vzStarTycho2Edit.setText(p2[1])
                        elif p2[0] == 'Hipparcos':
                            self.vzStarHipparcosEdit.setText(p2[1])
                    except Exception:  # noqa
                        self.showMsg("Failed to parse ADV header: " + header, bold=True, color='red')
                    continue

        except Exception:  # noqa
            self.showMsg("Failed to parse some item in ADV meta-data.", bold=True, color='red')
            print("Failed to parse some item in ADV meta-data.")


    def illustrateTimestampOutliers(self):
        for pos in self.cadenceViolation:
            vLine = pg.InfiniteLine(pos=pos + 0.5, pen=(255, 184, 28))  # golden yellow
            self.mainPlot.addItem(vLine)

        for pos in self.droppedFrames:
            vLine = pg.InfiniteLine(pos=pos + 0.5, pen=(255, 0, 0))  # red
            self.mainPlot.addItem(vLine)

    def prettyPrintCorCoefs(self):
        outStr = 'noise corr coefs: ['

        posCoefs = []
        for coef in self.corCoefs:
            if coef < acfCoefThreshold:
                break
            posCoefs.append(coef)

        for i in range(len(posCoefs) - 1):
            outStr = outStr + fp.to_precision(posCoefs[i], 3) + ', '
        outStr = outStr + fp.to_precision(posCoefs[-1], 3)
        outStr = outStr + ']  (based on ' + str(self.numPtsInCorCoefs) + ' points)'
        outStr = outStr + '  sigmaB: ' + f'{self.sigmaB:0.2f}'
        self.showMsg(outStr)

    def processEventNoise(self, secondPass=False):
        if len(self.selectedPoints) != 2:
            self.showInfo('Exactly two points must be selected for this operation')
            return
        selPts = self.selectedPoints.keys()
        left = int(min(selPts))
        right = int(max(selPts))
        if (right - left) < 9:
            if secondPass:
                self.removePointSelections()
                self.sigmaA = self.sigmaB
                return
            else:
                self.showInfo('At least 10 points must be included.')
                return
        if left < self.left or right > self.right:
            self.showInfo('Selection point(s) outside of included data points')
            self.removePointSelections()
            return
        else:
            self.eventXvals = []
            self.eventYvals = []
            for i in range(left, right + 1):
                self.eventXvals.append(i)
                self.eventYvals.append(self.yValues[i])
            self.showSelectedPoints('Points selected for event noise '
                                    'analysis: ')

        self.removePointSelections()
        _, self.numNApts, self.sigmaA = getCorCoefs(self.eventXvals, self.eventYvals)
        self.showMsg('Event noise analysis done using ' + str(self.numNApts) +
                     ' points ---  sigmaA: ' + fp.to_precision(self.sigmaA, 4))

        self.newRedrawMainPlot()

    def processEventNoiseFromIterativeSolution(self, left, right):

        assert left >= self.left
        assert right <= self.right

        self.eventXvals = []
        self.eventYvals = []

        for i in range(left, right + 1):
            self.eventXvals.append(i)
            self.eventYvals.append(self.yValues[i])

        self.A = np.mean(self.eventYvals)
        self.sigmaA = np.std(self.eventYvals)


    def processBaselineNoise(self, secondPass=False):

        if len(self.selectedPoints) != 2:
            self.showInfo('Exactly two points must be selected for this operation')
            return
        selPts = self.selectedPoints.keys()
        left = int(min(selPts))
        right = int(max(selPts))
        if (right - left) < 14:
            if secondPass:
                self.removePointSelections()
                return
            else:
                self.showInfo('At least 15 points must be included.')
                return
        if left < self.left or right > self.right:
            self.showInfo('Selection point(s) outside of included data points')
            return
        else:
            self.baselineXvals = []
            self.baselineYvals = []
            for i in range(left, right + 1):
                self.baselineXvals.append(i)
                self.baselineYvals.append(self.yValues[i])
            self.showSelectedPoints('Points selected for baseline noise '
                                    'analysis: ')

        self.removePointSelections()

        self.newCorCoefs, self.numNApts, sigB = getCorCoefs(self.baselineXvals, self.baselineYvals)
        self.showMsg('Baseline noise analysis done using ' + str(self.numNApts) +
                     ' baseline points')
        if len(self.corCoefs) == 0:
            self.corCoefs = np.ndarray(shape=(len(self.newCorCoefs),))
            np.copyto(self.corCoefs, self.newCorCoefs)
            self.numPtsInCorCoefs = self.numNApts
            self.sigmaB = sigB
        else:
            totalPoints = self.numNApts + self.numPtsInCorCoefs
            self.corCoefs = (self.corCoefs * self.numPtsInCorCoefs +
                             self.newCorCoefs * self.numNApts) / totalPoints
            self.sigmaB = (self.sigmaB * self.numPtsInCorCoefs +
                           sigB * self.numNApts) / totalPoints
            self.numPtsInCorCoefs = totalPoints

        self.prettyPrintCorCoefs()

        # Try to warn user about the possible need for block integration by testing the lag 1
        # and lag 2 correlation coefficients.  The tests are just guesses on my part, so only
        # warnings are given.  Later, the Cholesky-Decomposition may fail because block integration
        # was really needed.  That is a fatal error but is trapped and the user alerted to the problem

        if len(self.corCoefs) > 1:
            if self.corCoefs[1] >= 0.7 and self.ne3NotInUseRadioButton.isChecked():
                self.showInfo('The auto-correlation coefficient at lag 1 is suspiciously large. '
                              'This may be because the light curve needs some degree of block integration. '
                              'Failure to do a needed block integration allows point-to-point correlations caused by '
                              'the camera integration to artificially induce non-physical correlated noise.')
            elif len(self.corCoefs) > 2:
                if self.corCoefs[2] >= 0.3 and self.ne3NotInUseRadioButton.isChecked():
                    self.showInfo('The auto-correlation coefficient at lag 2 is suspiciously large. '
                                  'This may be because the light curve needs some degree of block integration. '
                                  'Failure to do a needed block integration allows point-to-point correlations caused by '
                                  'the camera integration to artificially induce non-physical correlated noise.')

        if self.sigmaA is None:
            self.sigmaA = self.sigmaB

        self.newRedrawMainPlot()

        self.locateEvent.setEnabled(True)
        self.markDzone.setEnabled(True)
        self.markRzone.setEnabled(True)
        self.singlePointDropButton.setEnabled(True)
        self.markEzone.setEnabled(not self.ne3NotInUseRadioButton.isChecked())

        self.minEventEdit.setEnabled(True)
        self.maxEventEdit.setEnabled(True)
        self.magDropSqwaveEdit.setEnabled(True)

    # def processBaselineNoiseFromIterativeSolution(self, left, right):
    def processBaselineNoiseFromIterativeSolution(self):
        # Calc baseline stats (with D and R points excluded, but only if they are at transition points)
        self.B = np.mean(self.baselineYvals)
        self.sigmaB = np.std(self.baselineYvals)

        if not self.userDeterminedBaselineStats:
            self.newCorCoefs, self.numNApts, sigB = getCorCoefs(self.baselineXvals,
                                                                self.baselineYvals)
            # if sigB is None:
            #     sigB = self.sigmaB

            if len(self.corCoefs) == 0:
                self.corCoefs = np.ndarray(shape=(len(self.newCorCoefs),))
                np.copyto(self.corCoefs, self.newCorCoefs)
                self.numPtsInCorCoefs = self.numNApts
            else:
                totalPoints = self.numNApts + self.numPtsInCorCoefs
                self.corCoefs = (self.corCoefs * self.numPtsInCorCoefs +
                                 self.newCorCoefs * self.numNApts) / totalPoints
                self.numPtsInCorCoefs = totalPoints

    def removePointSelections(self):
        for i, oldStatus in self.selectedPoints.items():
            self.yStatus[i] = oldStatus
        self.selectedPoints = {}

    def disableAllButtons(self):
        self.calcFlashEdge.setEnabled(False)
        self.setDataLimits.setEnabled(False)
        self.doBlockIntegration.setEnabled(False)
        self.blockSizeEdit.setEnabled(False)
        self.locateEvent.setEnabled(False)
        self.fillExcelReportButton.setEnabled(False)
        self.startOver.setEnabled(False)
        self.markDzone.setEnabled(False)
        self.markRzone.setEnabled(False)
        self.singlePointDropButton.setEnabled(False)
        self.markEzone.setEnabled(False)
        self.minEventEdit.setEnabled(False)
        self.maxEventEdit.setEnabled(False)
        self.magDropSqwaveEdit.setEnabled(False)
        self.writeBarPlots.setEnabled(False)
        self.writeCSVButton.setEnabled(False)

    # noinspection PyUnusedLocal
    def restart(self):

        if self.blockSize != 1:
            self.showInfo("This lightcurve has been block integrated.\n\nIf you need to undo block integration,\n"
                          "re-reading the csv is required because\n"
                          "the Start over function is not able undo \nblock integration.\n\n"
                          "Further changes to block integration parameters\nare now locked out.")

        self.userDeterminedBaselineStats = False
        self.userDeterminedEventStats = False
        self.userTrimInEffect = False

        self.selectedPoints = {}

        # We do this to erase any plot info present from an 'other model' fit
        self.modelYsamples = None
        self.modelY = None

        savedFlashEdges = self.flashEdges
        self.initializeVariablesThatDontDependOnAfile()
        self.flashEdges = savedFlashEdges
        self.disableAllButtons()

        if self.blockSize == 1:
            self.blockSizeEdit.setEnabled(True)
            self.doBlockIntegration.setEnabled(True)


        if self.errBarWin:
            self.errBarWin.close()

        self.dataLen = len(self.yTimes)
        self.VizieRdict = {
            "timestamps": None,
            "yValues": None,
            "yStatus": None,
        }
        self.timeDelta, self.droppedFrames, self.cadenceViolation, self.errRate, _ = \
            getTimeStepAndOutliers(self.yTimes, self.yValues, self.yStatus, VizieRdict=self.VizieRdict)
        self.frameTimeEdit.setText(fp.to_precision(self.timeDelta, 6))

        self.fillTableViewOfData()

        # Enable the initial set of buttons (allowed operations)
        self.startOver.setEnabled(True)
        self.setDataLimits.setEnabled(True)

        self.markDzone.setEnabled(True)
        self.markRzone.setEnabled(True)
        self.singlePointDropButton.setEnabled(True)
        self.markEzone.setEnabled(not self.ne3NotInUseRadioButton.isChecked())
        self.locateEvent.setEnabled(True)
        self.minEventEdit.setEnabled(True)
        self.maxEventEdit.setEnabled(True)
        self.magDropSqwaveEdit.setEnabled(True)

        self.clearBaselineRegionsButton.setEnabled(False)
        self.calcStatsFromBaselineRegionsButton.setEnabled(False)

        # Reset the data plot so that all points are visible
        # self.mainPlot.autoRange()

        # Show all data points as INCLUDED
        self.yStatus = [INCLUDED for _i in range(self.dataLen)]

        # Set the 'left' and 'right' edges of 'included' data to 'all'
        self.left = 0
        self.right = self.dataLen - 1

        self.minEventEdit.clear()
        self.maxEventEdit.clear()

        self.bkgndRegionLimits = []

        self.newRedrawMainPlot()
        # self.mainPlot.autoRange()
        # self.showMsg('*' * 20 + ' starting over ' + '*' * 20, color='blue')

    def vizierStartOver(self):
        # Show all data points as INCLUDED
        try:
            if self.yStatus:
                self.yStatus = [INCLUDED for _i in range(self.dataLen)]
            else:
                return
        except AttributeError:
            return

        # Set the 'left' and 'right' edges of 'included' data to 'all'
        self.left = 0
        self.right = self.dataLen - 1

        self.VizieRdict = {
            "timestamps": None,
            "yValues": None,
            "yStatus": None,
        }
        self.timeDelta, self.droppedFrames, self.cadenceViolation, self.errRate, _ = \
            getTimeStepAndOutliers(self.yTimes, self.yValues, self.yStatus, VizieRdict=self.VizieRdict)

        self.newRedrawMainPlot()


    def extendAndDrawModelLightcurve(self):

        if self.modelTimeOffset is None:
            # This done to allow a saved LCP to be reloaded.
            return

        if not self.timestampListIsEmpty(self.yTimes):
            # Compute time duration of lightcurve observation
            tObsStart = convertTimeStringToTime(self.yTimes[0])
            tObsEnd = convertTimeStringToTime(self.yTimes[-1])
            tObsDur = tObsEnd - tObsStart
            if tObsDur < 0:  # The lightcurve acquisition passed through midnight
                tObsDur += 24 * 60 * 60  # And a days worth of seconds

            tModelDur = (self.modelXkm[-1] - self.modelXkm[0]) / self.Lcp.shadow_speed  # in seconds
            self.modelDuration = tModelDur
            # modelLengthKm = self.modelDuration * self.Lcp.shadow_speed

            # Convert model edge D and R locations from km to time - relative to model center
            if self.modelDedgeKm is not None:
                self.modelDedgeSecs = self.modelDedgeKm / self.Lcp.shadow_speed
                self.modelRedgeSecs = self.modelRedgeKm / self.Lcp.shadow_speed

                # Convert edge locations from time in model space to time in observation space
                # DedgeKm = self.modelDedgeKm - self.modelXkm[0]
                # RedgeKm = self.modelRedgeKm - self.modelXkm[0]
                # edgeCenterTime = (DedgeKm + RedgeKm) / 2.0 / self.Lcp.shadow_speed
            else:
                self.modelDedgeSecs = self.modelDuration / 2
                self.modelRedgeSecs = self.modelDuration / 2
                # DedgeKm = modelLengthKm / 2 - self.modelXkm[0]
                # RedgeKm = modelLengthKm / 2 - self.modelXkm[0]
                # edgeCenterTime = (DedgeKm + RedgeKm) / 2.0 / self.Lcp.shadow_speed

            if self.firstContactKm is not None:
                self.firstContactSecs = self.firstContactKm / self.Lcp.shadow_speed
                self.lastContactSecs = self.lastContactKm / self.Lcp.shadow_speed

            # self.modelDedgeSecs += self.modelTimeOffset + edgeCenterTime
            # self.modelRedgeSecs += self.modelTimeOffset + edgeCenterTime

            modelXsec = self.modelXkm / self.Lcp.shadow_speed
            modelXsec -= modelXsec[0]
            modelXsec += self.modelTimeOffset + tObsStart
            modelRdgNum = (modelXsec - tObsStart) / self.Lcp.frame_time
            self.modelPtsXrdgNum = modelRdgNum
            self.modelPtsXsec = modelXsec
            self.modelPtsY = self.modelY

            dEdgeRdgNumOffset = (tModelDur / 2 + self.modelDedgeSecs) / self.Lcp.frame_time
            rEdgeRdgNumOffset = dEdgeRdgNumOffset + (self.modelRedgeSecs - self.modelDedgeSecs) / self.Lcp.frame_time

            if self.firstContactSecs is not None:
                firstContactRdgNumOffset = (tModelDur / 2 + self.firstContactSecs) / self.Lcp.frame_time
                lastContactRdgNumOffset = firstContactRdgNumOffset + (self.lastContactSecs - self.firstContactSecs) / self.Lcp.frame_time

            del modelRdgNum
            del modelXsec
            gc.collect()

            mPen = pg.mkPen(color=(255, 0, 0), width=self.lineWidthSpinner.value())
            self.mainPlot.plot(self.modelPtsXrdgNum, self.modelPtsY, pen=mPen, symbol=None)

            if self.modelDedgeKm is not None:
                # Add D and R edge positions to plot.
                ePen = pg.mkPen(color=(100, 100, 100), style=QtCore.Qt.PenStyle.DashLine,
                                width=self.lineWidthSpinner.value())

                # Convert edge times to reading number units. This only works if there are no dropped readings
                # self.modelDedgeRdgValue = self.modelDedgeSecs * (self.dataLen - 1) / tObsDur
                # self.modelRedgeRdgValue = self.modelRedgeSecs * (self.dataLen - 1) / tObsDur

                self.modelDedgeRdgValue = self.modelPtsXrdgNum[0] + dEdgeRdgNumOffset
                self.modelRedgeRdgValue = self.modelPtsXrdgNum[0] + rEdgeRdgNumOffset

                lo_int = self.Lcp.bottom_ADU
                hi_int = self.Lcp.baseline_ADU
                span = hi_int - lo_int

                if self.firstContactKm is not None:
                    self.firstContactRdgValue = self.modelPtsXrdgNum[0] + firstContactRdgNumOffset  # noqa
                    self.lastContactRdgValue  = self.modelPtsXrdgNum[0] + lastContactRdgNumOffset   # noqa

                    cPen = pg.mkPen(color=(0, 150, 0), style=QtCore.Qt.PenStyle.DashLine,
                                    width=self.lineWidthSpinner.value())
                    fc = self.firstContactRdgValue
                    self.mainPlot.plot([fc, fc], [hi_int - 0.4 * span, hi_int + 0.4 * span], pen=cPen, symbol=None)
                    lc = self.lastContactRdgValue
                    self.mainPlot.plot([lc, lc], [hi_int - 0.4 * span, hi_int + 0.4 * span], pen=cPen, symbol=None)

                D = self.modelDedgeRdgValue
                R = self.modelRedgeRdgValue
                if D >= 0:
                    self.mainPlot.plot([D, D], [lo_int - 0.1 * span, hi_int + 0.1 * span], pen=ePen, symbol=None)
                if R <= self.dataLen - 1:
                    self.mainPlot.plot([R, R], [lo_int - 0.1 * span, hi_int + 0.1 * span], pen=ePen, symbol=None)

            self.modelXvalues = self.modelPtsXrdgNum
            self.modelYvalues = np.array(self.modelY)
        else:
            self.showInfo(f'Timestamps are required for model lightcurves')

    def drawSolution(self):
        def plot(x, y, pen):
            self.mainPlot.plot(x, y, pen=pen, symbol=None)

        def plotDcurve():
            # The units of self.timeDelta are seconds per entry, so the conversion in the next line
            # gets us x_values[] converted to entryNum units.
            x_values = self.underlyingLightcurveAns['time deltas'] / self.timeDelta
            x_values += D
            y_values = self.underlyingLightcurveAns['D curve']
            if self.underlyingLightcurveAns['star D'] is not None:
                z_values = self.underlyingLightcurveAns['star D']
            else:
                z_values = self.underlyingLightcurveAns['raw D']

            x_trimmed = []
            y_trimmed = []
            z_trimmed = []
            for i in range(x_values.size):
                if self.left <= x_values[i] <= max_x:
                    x_trimmed.append(x_values[i])
                    y_trimmed.append(y_values[i])
                    z_trimmed.append(z_values[i])

            # (150, 100, 100) is the brownish color we use to show the underlying lightcurve
            if self.showUnderlyingLightcurveCheckBox.isChecked():
                plot(x_trimmed, z_trimmed, pen=pg.mkPen((150, 100, 100), width=self.lineWidthSpinner.value()))

            if self.exponentialDtheoryPts is None:
                if not self.showCameraResponseCheckBox.isChecked():
                    return
                # Now overplot with the blue camera response curve
                plot(x_trimmed, y_trimmed, pen=pg.mkPen((0, 0, 255), width=self.lineWidthSpinner.value()))
                # Extend camera response to the left and right if necessary...
                if x_trimmed:
                    if x_trimmed[0] > self.left:
                        plot([self.left, x_trimmed[0]], [y_trimmed[0], y_trimmed[0]],
                             pen=pg.mkPen((0, 0, 255), width=self.lineWidthSpinner.value()))
                    if x_trimmed[-1] < max_x:
                        plot([x_trimmed[-1], max_x], [y_trimmed[-1], y_trimmed[-1]],
                             pen=pg.mkPen((0, 0, 255), width=self.lineWidthSpinner.value()))

        def plotRcurve():
            # The units of self.timeDelta are seconds per entry, so the conversion in the next line
            # gets us x_values[] converted to entryNum units.
            x_values = self.underlyingLightcurveAns['time deltas'] / self.timeDelta
            x_values += R
            y_values = self.underlyingLightcurveAns['R curve']
            if self.underlyingLightcurveAns['star R'] is not None:
                z_values = self.underlyingLightcurveAns['star R']
            else:
                z_values = self.underlyingLightcurveAns['raw R']

            x_trimmed = []
            y_trimmed = []
            z_trimmed = []
            for i in range(x_values.size):
                if min_x <= x_values[i] <= self.right:
                    x_trimmed.append(x_values[i])
                    y_trimmed.append(y_values[i])
                    z_trimmed.append(z_values[i])

            # (150, 100, 100) is the brownish color we use to show the underlying lightcurve
            if self.showUnderlyingLightcurveCheckBox.isChecked() and x_trimmed:
                plot(x_trimmed, z_trimmed, pen=pg.mkPen((150, 100, 100), width=self.lineWidthSpinner.value()))

            if self.exponentialRtheoryPts is None and x_trimmed:
                if not self.showCameraResponseCheckBox.isChecked():
                    return
                # Now overplot with the blue camera response curve
                plot(x_trimmed, y_trimmed, pen=pg.mkPen((0, 0, 255), width=self.lineWidthSpinner.value()))
                # Extend camera response to the left and right if necessary...
                if x_trimmed[0] > min_x:
                    plot([min_x, x_trimmed[0]], [y_trimmed[0], y_trimmed[0]],
                         pen=pg.mkPen((0, 0, 255), width=self.lineWidthSpinner.value()))

                if x_trimmed[-1] < self.right:
                    plot([x_trimmed[-1], self.right], [y_trimmed[-1], y_trimmed[-1]],
                         pen=pg.mkPen((0, 0, 255), width=self.lineWidthSpinner.value()))

                if x_trimmed:
                    pass

        def plotGeometricShadowAtD():
            if self.showEdgesCheckBox.isChecked() and self.exponentialDtheoryPts is None:
                pen = pg.mkPen(color=(255, 0, 0), style=QtCore.Qt.PenStyle.DashLine,
                               width=self.lineWidthSpinner.value())
                self.mainPlot.plot([D, D], [lo_int, hi_int], pen=pen, symbol=None)

        def plotGeometricShadowAtR():
            if self.showEdgesCheckBox.isChecked() and self.exponentialRtheoryPts is None:
                pen = pg.mkPen(color=(0, 200, 0), style=QtCore.Qt.PenStyle.DashLine,
                               width=self.lineWidthSpinner.value())
                self.mainPlot.plot([R, R], [lo_int, hi_int], pen=pen, symbol=None)

        hi_int = max(self.yValues[self.left:self.right])
        lo_int = min(self.yValues[self.left:self.right])

        if self.squareWaveRadioButton.isChecked():
            # Plot the camera response for a square wave solution
            hi_int = max(self.yValues[self.left:self.right])
            lo_int = min(self.yValues[self.left:self.right])

            D = self.solution[0]
            R = self.solution[1]
            solutionPen = pg.mkPen((255, 0, 0), width=self.lineWidthSpinner.value())
            dPen = pg.mkPen((100, 100, 100), width=self.lineWidthSpinner.value(),
                            style=QtCore.Qt.PenStyle.DashLine,)
            rPen = pg.mkPen((100, 100, 100), width=self.lineWidthSpinner.value(),
                            style=QtCore.Qt.PenStyle.DashLine,)
            underlyingPen = pg.mkPen((150, 100, 100), width=self.lineWidthSpinner.value())

            if self.magDropSqwaveEdit.text() == '':
                bottom = self.A
            else:
                try:
                    magDrop = float(self.magDropSqwaveEdit.text())
                    bottom = self.calcBottomADU(self.B, magDrop)
                except ValueError as e:
                    self.showInfo(f'Error in expected magDrop entry\n\n'
                                  f'{e}\n\n'
                                  f'... treating it as "no entry"')
                    bottom = self.A

            if self.eventType == 'DandR':
                if self.showCameraResponseCheckBox.isChecked():
                    plot([self.left, D - 1], [self.B, self.B], pen=solutionPen)
                    plot([D-1, D], [self.B, bottom], pen=solutionPen)
                    plot([D, R], [bottom, bottom], pen=solutionPen)
                    plot([R, R+1], [bottom, self.B], pen=solutionPen)
                    plot([R+1, self.right], [self.B, self.B], pen=solutionPen)
                if self.showUnderlyingLightcurveCheckBox.isChecked():
                    plot([self.left, D], [self.B, self.B], pen=underlyingPen)
                    plot([D, D], [self.B, bottom], pen=underlyingPen)
                    plot([D, R], [bottom, bottom], pen=underlyingPen)
                    plot([R, R], [bottom, self.B], pen=underlyingPen)
                    plot([R, self.right], [self.B, self.B], pen=underlyingPen)
                if self.showEdgesCheckBox.isChecked():
                    plot([D, D], [lo_int, hi_int], pen=dPen)
                    plot([R, R], [lo_int, hi_int], pen=rPen)
            elif self.eventType == 'Donly':
                if self.showCameraResponseCheckBox.isChecked():
                    plot([self.left, D - 1], [self.B, self.B], pen=solutionPen)
                    plot([D - 1, D], [self.B, bottom], pen=solutionPen)
                    plot([D, self.right], [bottom, bottom], pen=solutionPen)
                if self.showUnderlyingLightcurveCheckBox.isChecked():
                    plot([self.left, D], [self.B, self.B], pen=underlyingPen)
                    plot([D, D], [self.B, bottom], pen=underlyingPen)
                    plot([D, self.right], [bottom, bottom], pen=underlyingPen)
                if self.showEdgesCheckBox.isChecked():
                    plot([D, D], [lo_int, hi_int], pen=dPen)
            else:
                if self.showCameraResponseCheckBox.isChecked():
                    plot([self.left, R], [bottom, bottom], pen=solutionPen)
                    plot([R, R + 1], [bottom, self.B], pen=solutionPen)
                    plot([R + 1, self.right], [self.B, self.B], pen=solutionPen)
                if self.showUnderlyingLightcurveCheckBox.isChecked():
                    plot([self.left, R], [bottom, bottom], pen=underlyingPen)
                    plot([R, R], [bottom, self.B], pen=underlyingPen)
                    plot([R, self.right], [self.B, self.B], pen=underlyingPen)
                if self.showEdgesCheckBox.isChecked():
                    plot([R, R], [lo_int, hi_int], pen=rPen)


            return
        elif self.exponentialDtheoryPts is None:
            if self.modelY is not None and self.modelXkm is not None:
                ePen = pg.mkPen((0, 255, 0), width=self.lineWidthSpinner.value())
                plot(self.modelXkm, self.modelY, pen=ePen)
                return

        if self.eventType == 'DandR':
            if self.exponentialDtheoryPts is None:
                D = self.solution[0]
                R = self.solution[1]
            else:
                D = self.exponentialDedge
                R = self.exponentialRedge
                D = self.solution[0]
                R = self.solution[1]

            max_x = min_x = (D + R) / 2.0

        elif self.eventType == 'Donly':
            # if self.exponentialDtheoryPts is None:
            D = self.solution[0]

            max_x = self.right
            plotDcurve()
            plotGeometricShadowAtD()

        elif self.eventType == 'Ronly':
            R = self.solution[1]

            min_x = self.left
            plotRcurve()
            plotGeometricShadowAtR()
        else:
            raise Exception('Unrecognized event type of |' + self.eventType + '|')

    def calcNumBandApoints(self):
        if self.eventType == 'Donly':
            self.nBpts = self.solution[0] - self.left
            self.nApts = self.right - self.solution[0] - 1

        if self.eventType == 'Ronly':
            self.nBpts = self.right - self.solution[1]
            self.nApts = self.solution[1] - self.left

        if self.eventType == 'DandR':
            self.nBpts = self.right - self.solution[1] + self.solution[0] - self.left
            self.nApts = self.solution[1] - self.solution[0] - 1

        if self.nBpts < 1:
            self.nBpts = 1

        if self.nApts < 1:
            self.nApts = 1

    def drawEnvelope(self):

        def plotGeometricShadowAtD(d):
            if self.showErrBarsCheckBox.isChecked():
                pen = pg.mkPen(color=(255, 0, 0), style=QtCore.Qt.PenStyle.DotLine, width=self.lineWidthSpinner.value())
                self.mainPlot.plot([d, d], [lo_int, hi_int], pen=pen, symbol=None)

        def plotGeometricShadowAtR(r):
            if self.showErrBarsCheckBox.isChecked():
                pen = pg.mkPen(color=(0, 200, 0), style=QtCore.Qt.PenStyle.DotLine, width=self.lineWidthSpinner.value())
                self.mainPlot.plot([r, r], [lo_int, hi_int], pen=pen, symbol=None)

        if self.solution is None:
            return

        # Make shortened geometric shadow markers to distinguish the error bar versions from the central value
        hi_int = max(self.yValues[self.left:self.right])
        lo_int = min(self.yValues[self.left:self.right])
        delta_int = (hi_int - lo_int) * 0.1
        hi_int -= delta_int
        lo_int += delta_int

        if self.eventType == 'Donly':
            D = self.solution[0]
            Dright = D + self.plusD
            Dleft = D - self.minusD
            plotGeometricShadowAtD(Dright)
            plotGeometricShadowAtD(Dleft)
            return

        if self.eventType == 'Ronly':
            R = self.solution[1]
            Rright = R + self.plusR
            Rleft = R - self.minusR
            plotGeometricShadowAtR(Rright)
            plotGeometricShadowAtR(Rleft)
            return

        if self.eventType == 'DandR':
            R = self.solution[1]
            D = self.solution[0]

            Rright = R + self.plusR
            Rleft = R - self.minusR
            Dright = D + self.plusD
            Dleft = D - self.minusD
            plotGeometricShadowAtR(Rright)
            plotGeometricShadowAtR(Rleft)
            plotGeometricShadowAtD(Dright)
            plotGeometricShadowAtD(Dleft)
            return

    def newRedrawMainPlot(self):
        QtWidgets.QApplication.processEvents()
        self.redrawMainPlot()

    def redrawMainPlot(self):

        if self.right is not None:
            right = min(self.dataLen, self.right + 1)
        else:
            right = self.dataLen
        if self.left is None:
            left = 0
        else:
            left = self.left

        self.mainPlot.clear()

        if self.modelYsamples is not None:
            dotSize = self.dotSizeSpinner.value()
            # Show single square dot at ends of model lightcurve
            # self.mainPlot.plot([self.modelXsamples[0], self.modelXsamples[-1]],
            #                    [self.modelYsamples[0], self.modelYsamples[-1]],
            #                    pen=None, symbol='s',
            #                    symbolBrush=(255, 0, 0), symbolSize=dotSize + 2)

            # Show all of the model lightcurve sample points
            self.mainPlot.plot(self.modelXsamples, self.modelYsamples,
                               pen=None, symbol='o',
                               symbolBrush=(255, 0, 0), symbolSize=dotSize + 2)

        if self.yValues is None:
            return

        if self.showTimestampErrors.checkState():
            self.illustrateTimestampOutliers()

        self.mainPlot.addItem(self.verticalCursor)

        self.targetIndex = 0
        # Find index of target lightcurve
        for i, checkBox in enumerate(self.targetCheckBoxes):
            if checkBox.isChecked():
                self.targetIndex = i
                break

        refIndex = None
        # Find index of reference selection (if any)
        for i, checkBox in enumerate(self.referenceCheckBoxes):
            if checkBox.isChecked():
                refIndex = i
                break

        self.mainPlot.plot(self.yValues + self.yOffsetSpinBoxes[self.targetIndex].value())
        self.mainPlot.getPlotItem().showAxis('bottom')
        self.mainPlot.getPlotItem().showAxis('left')

        dotSize = self.dotSizeSpinner.value()

        minY = np.min(self.yValues + self.yOffsetSpinBoxes[self.targetIndex].value())
        maxY = np.max(self.yValues + self.yOffsetSpinBoxes[self.targetIndex].value())

        try:
            # New code to deal with hit-defect
            if self.targetKey == '':
                self.targetKey = self.lightcurveTitles[0].text()
            hit_defect_flags = self.getHitDefectFlags(self.targetKey)

            # Plot the 'target' lightcurve
            x = [i for i in range(self.dataLen) if self.yStatus[i] == INCLUDED]
            y = [self.yValues[i] for i in range(self.dataLen) if self.yStatus[i] == INCLUDED]
            y = np.array(y) + self.yOffsetSpinBoxes[self.targetIndex].value()

            # if hit_defect_flags:
            #     y_hit = [self.yValues[i] for i in range(self.dataLen) if hit_defect_flags[i] > 0.0]
            #     x_hit = [i for i in range(self.dataLen) if hit_defect_flags[i] > 0.0]
            #     y_hit = np.array(y_hit) + self.yOffsetSpinBoxes[self.targetIndex].value()

            self.mainPlot.plot(x, y, pen=None, symbol='o',
                               symbolBrush=INCLUDED_COLOR, symbolSize=dotSize)
            if hit_defect_flags:
                y_hit = [self.yValues[i] for i in range(self.dataLen) if hit_defect_flags[i] > 0.0]
                x_hit = [i for i in range(self.dataLen) if hit_defect_flags[i] > 0.0]
                y_hit = np.array(y_hit) + self.yOffsetSpinBoxes[self.targetIndex].value()
                self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                   symbolBrush=SELECTED_COLOR, symbolSize=dotSize*2)
                self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                   symbolBrush=INCLUDED_COLOR, symbolSize=dotSize)
            x = [i for i in range(self.dataLen) if self.yStatus[i] == BASELINE]
            y = [self.yValues[i] for i in range(self.dataLen) if self.yStatus[i] == BASELINE]
            y = np.array(y) + self.yOffsetSpinBoxes[self.targetIndex].value()
            self.mainPlot.plot(x, y, pen=None, symbol='o',
                               symbolBrush=BASELINE_COLOR, symbolSize=dotSize)

            x = [i for i in range(self.dataLen) if self.yStatus[i] == EVENT]
            y = [self.yValues[i] for i in range(self.dataLen) if self.yStatus[i] == EVENT]
            y = np.array(y) + self.yOffsetSpinBoxes[self.targetIndex].value()
            self.mainPlot.plot(x, y, pen=None, symbol='o',
                               symbolBrush=EVENT_COLOR, symbolSize=dotSize)

            x = [i for i in range(self.dataLen) if self.yStatus[i] == SELECTED]
            y = [self.yValues[i] for i in range(self.dataLen) if self.yStatus[i] == SELECTED]
            y = np.array(y) + self.yOffsetSpinBoxes[self.targetIndex].value()
            self.mainPlot.plot(x, y, pen=None, symbol='o',
                               symbolBrush=SELECTED_COLOR, symbolSize=dotSize + 4)

            x = [i for i in range(self.dataLen) if self.yStatus[i] == MISSING]
            y = [self.yValues[i] for i in range(self.dataLen) if self.yStatus[i] == MISSING]
            y = np.array(y) + self.yOffsetSpinBoxes[self.targetIndex].value()
            self.mainPlot.plot(x, y, pen=None, symbol='o',
                               symbolBrush=MISSING_COLOR, symbolSize=dotSize + 4)

        except IndexError:
            pass

        if self.exponentialDtheoryPts is not None:
            xVals = [self.exponentialDinitialX]
            for i in range(len(self.exponentialDtheoryPts) - 1):
                xVals.append(xVals[-1] + 1)
            self.mainPlot.plot(xVals, self.exponentialDtheoryPts, pen=None, symbol='o',
                               symbolBrush=(160, 128, 96), symbolSize=dotSize + 4)
            hi_int = max(self.yValues[self.left:self.right])
            lo_int = min(self.yValues[self.left:self.right])
            pen = pg.mkPen(color=(200, 0, 0), style=QtCore.Qt.PenStyle.DashLine, width=self.lineWidthSpinner.value())
            self.mainPlot.plot([self.solution[0], self.solution[0]], [lo_int, hi_int], pen=pen, symbol=None)

        if self.exponentialRtheoryPts is not None:
            xVals = [self.exponentialRinitialX]
            for i in range(len(self.exponentialRtheoryPts) - 1):
                xVals.append(xVals[-1] + 1)
            self.mainPlot.plot(xVals, self.exponentialRtheoryPts, pen=None, symbol='o',
                               symbolBrush=(160, 128, 96), symbolSize=dotSize + 4)
            hi_int = max(self.yValues[self.left:self.right])
            lo_int = min(self.yValues[self.left:self.right])
            pen = pg.mkPen(color=(0, 200, 0), style=QtCore.Qt.PenStyle.DashLine, width=self.lineWidthSpinner.value())
            self.mainPlot.plot([self.solution[1], self.solution[1]], [lo_int, hi_int], pen=pen, symbol=None)

        if len(self.yRefStar) == self.dataLen:
            if not self.skipNormalization and not self.suppressNormalization:
                # Update reference curve smoothing
                if self.smoothingIntervalSpinBox.value() > 0:
                    # Start with pristine (original) values for the curve being analyzed.
                    self.processTargetSelection(self.targetIndex, redraw=False)
                    self.newNormalize()
                else:
                    self.smoothSecondary = []
                    # Return to original values
                    self.processTargetSelection(self.targetIndex, redraw=False)
                    self.fillTableViewOfData()
                    self.normalized = False
                self.skipNormalization = True
                self.newRedrawMainPlot()
            else:
                self.skipNormalization = False
                self.suppressNormalization = False

            # Plot the normalization reference lightcurve
            if refIndex is not None and self.showCheckBoxes[refIndex].isChecked() and self.left is not None:
                minY = min(minY, np.min(self.yRefStar + self.yOffsetSpinBoxes[refIndex].value()))
                maxY = max(maxY, np.max(self.yRefStar + self.yOffsetSpinBoxes[refIndex].value()))
                xOffset = self.xOffsetSpinBoxes[refIndex].value()
                x = [i + xOffset for i in range(self.left, self.right + 1)]
                y = [self.yRefStar[i] for i in range(self.left, self.right + 1)]
                y = np.array(y) + self.yOffsetSpinBoxes[refIndex].value()
                self.mainPlot.plot(x, y)
                self.mainPlot.plot(x, y, pen=None, symbol='o',
                                   symbolBrush=(0, 255, 0), symbolSize=dotSize)

                # Plot the continuous smoothed curve through the reference lightcurve
                if len(self.smoothSecondary) > 0:
                    self.mainPlot.plot(x, self.smoothSecondary + self.yOffsetSpinBoxes[refIndex].value(),
                                       pen=pg.mkPen((100, 100, 100), width=4), symbol=None)

        dotColors = [(255, 0, 0), (160, 32, 255), (80, 208, 255), (96, 255, 128),
                     (255, 224, 32), (255, 160, 16), (160, 128, 96), (64, 64, 64), (255, 208, 160), (0, 128, 0)]

        for i, checkBox in enumerate(self.showCheckBoxes):
            if checkBox.isChecked():
                if self.targetCheckBoxes[i].isChecked():
                    continue
                if self.referenceCheckBoxes[i].isChecked():
                    continue
                if i == 0:
                    hit_defect_flags = self.getHitDefectFlags(self.lightcurveTitles[i].text())
                    self.mainPlot.plot(self.LC1 + self.yOffsetSpinBoxes[i].value())
                    minY = min(minY, np.min(self.LC1) + self.yOffsetSpinBoxes[i].value())
                    maxY = max(maxY, np.max(self.LC1) + self.yOffsetSpinBoxes[i].value())

                    x = [j for j in range(left, right)]
                    y = [self.LC1[j] for j in range(left, right)]
                    y = np.array(y) + self.yOffsetSpinBoxes[0].value()
                    self.mainPlot.plot(x, y, pen=None, symbol='o',
                                       symbolBrush=dotColors[i], symbolSize=dotSize)
                    if hit_defect_flags:
                        y_hit = [self.LC1[j] for j in range(self.dataLen) if hit_defect_flags[j] > 0.0]
                        x_hit = [j for j in range(self.dataLen) if hit_defect_flags[j] > 0.0]
                        y_hit = np.array(y_hit) + self.yOffsetSpinBoxes[i].value()
                        self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                           symbolBrush=SELECTED_COLOR, symbolSize=dotSize * 2)
                        self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                           symbolBrush=dotColors[i], symbolSize=dotSize)
                elif i == 1:
                    hit_defect_flags = self.getHitDefectFlags(self.lightcurveTitles[i].text())
                    self.mainPlot.plot(self.LC2 + self.yOffsetSpinBoxes[i].value())
                    minY = min(minY, np.min(self.LC2) + self.yOffsetSpinBoxes[i].value())
                    maxY = max(maxY, np.max(self.LC2) + self.yOffsetSpinBoxes[i].value())
                    x = [j for j in range(left, right)]
                    y = [self.LC2[j] for j in range(left, right)]
                    y = np.array(y) + self.yOffsetSpinBoxes[1].value()
                    self.mainPlot.plot(x, y, pen=None, symbol='o',
                                       symbolBrush=dotColors[i], symbolSize=dotSize)
                    if hit_defect_flags:
                        y_hit = [self.LC2[j] for j in range(self.dataLen) if hit_defect_flags[j] > 0.0]
                        x_hit = [j for j in range(self.dataLen) if hit_defect_flags[j] > 0.0]
                        y_hit = np.array(y_hit) + self.yOffsetSpinBoxes[i].value()
                        self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                           symbolBrush=SELECTED_COLOR, symbolSize=dotSize * 2)
                        self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                           symbolBrush=dotColors[i], symbolSize=dotSize)
                elif i == 2:
                    hit_defect_flags = self.getHitDefectFlags(self.lightcurveTitles[i].text())
                    self.mainPlot.plot(self.LC3 + self.yOffsetSpinBoxes[i].value())
                    minY = min(minY, np.min(self.LC3) + self.yOffsetSpinBoxes[i].value())
                    maxY = max(maxY, np.max(self.LC3) + self.yOffsetSpinBoxes[i].value())
                    x = [j for j in range(left, right)]
                    y = [self.LC3[i] for i in range(left, right)]
                    y = np.array(y) + self.yOffsetSpinBoxes[i].value()
                    self.mainPlot.plot(x, y, pen=None, symbol='o',
                                       symbolBrush=dotColors[i], symbolSize=dotSize)
                    if hit_defect_flags:
                        y_hit = [self.LC3[j] for j in range(self.dataLen) if hit_defect_flags[j] > 0.0]
                        x_hit = [j for j in range(self.dataLen) if hit_defect_flags[j] > 0.0]
                        y_hit = np.array(y_hit) + self.yOffsetSpinBoxes[i].value()
                        self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                           symbolBrush=SELECTED_COLOR, symbolSize=dotSize * 2)
                        self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                           symbolBrush=dotColors[i], symbolSize=dotSize)
                elif i == 3:
                    hit_defect_flags = self.getHitDefectFlags(self.lightcurveTitles[i].text())
                    self.mainPlot.plot(self.LC4 + self.yOffsetSpinBoxes[i].value())
                    minY = min(minY, np.min(self.LC4) + self.yOffsetSpinBoxes[i].value())
                    maxY = max(maxY, np.max(self.LC4) + self.yOffsetSpinBoxes[i].value())
                    x = [j for j in range(left, right)]
                    y = [self.LC4[j] for j in range(left, right)]
                    y = np.array(y) + self.yOffsetSpinBoxes[3].value()
                    self.mainPlot.plot(x, y, pen=None, symbol='o',
                                       symbolBrush=dotColors[i], symbolSize=dotSize)
                    if hit_defect_flags:
                        y_hit = [self.LC4[j] for j in range(self.dataLen) if hit_defect_flags[j] > 0.0]
                        x_hit = [j for j in range(self.dataLen) if hit_defect_flags[j] > 0.0]
                        y_hit = np.array(y_hit) + self.yOffsetSpinBoxes[i].value()
                        self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                           symbolBrush=SELECTED_COLOR, symbolSize=dotSize * 2)
                        self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                           symbolBrush=dotColors[i], symbolSize=dotSize)
                else:
                    hit_defect_flags = self.getHitDefectFlags(self.lightcurveTitles[i].text())
                    self.mainPlot.plot(self.extra[i - 4] + self.yOffsetSpinBoxes[i].value())
                    minY = min(minY, np.min(self.extra[i - 4] + self.yOffsetSpinBoxes[i].value()))
                    maxY = max(maxY, np.max(self.extra[i - 4] + self.yOffsetSpinBoxes[i].value()))
                    x = [j for j in range(left, right)]
                    y = [self.extra[i - 4][j] for j in range(left, right)]
                    y = np.array(y) + self.yOffsetSpinBoxes[i].value()
                    self.mainPlot.plot(x, y, pen=None, symbol='o',
                                       symbolBrush=dotColors[i], symbolSize=dotSize)
                    if hit_defect_flags:
                        y_hit = [self.extra[i-4][j] for j in range(self.dataLen) if hit_defect_flags[j] > 0.0]
                        x_hit = [j for j in range(self.dataLen) if hit_defect_flags[j] > 0.0]
                        y_hit = np.array(y_hit) + self.yOffsetSpinBoxes[i].value()
                        self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                           symbolBrush=SELECTED_COLOR, symbolSize=dotSize * 2)
                        self.mainPlot.plot(x_hit, y_hit, pen=None, symbol='o',
                                           symbolBrush=dotColors[i], symbolSize=dotSize)

        self.mainPlot.setYRange(minY, maxY)

        if self.dRegion is not None:
            self.mainPlot.addItem(self.dRegion)
        if self.rRegion is not None:
            self.mainPlot.addItem(self.rRegion)
        if self.bkgndRegions is not []:
            for region in self.bkgndRegions:
                self.mainPlot.addItem(region)

        if self.exponentialDtheoryPts is None and not self.squareWaveRadioButton.isChecked():
            if self.modelXkm is not None and self.modelY is not None:
                self.extendAndDrawModelLightcurve()

        if self.solution:
            self.drawSolution()

        if self.minusD is not None or self.minusR is not None:
            # We have data for drawing an envelope
            self.drawEnvelope()

    def getHitDefectFlags(self, light_curve_name):

        hit_name = ''
        hit_defect_flags = []
        if light_curve_name.startswith('signal') or light_curve_name.startswith('appsum'):
            parts = light_curve_name.split('-', 1)
            hit_name = 'hit-defect-' + parts[1]
        if hit_name:
            if hit_name in self.fullDataDictionary.keys():
                hit_defect_flags = self.fullDataDictionary[hit_name]
        return hit_defect_flags

    def showSelectedPoints(self, header):
        selPts = list(self.selectedPoints.keys())
        selPts.sort()
        self.showMsg(header + str(selPts))

    def doTrim(self):
        if self.dataLen is None:
            return

        if len(self.selectedPoints) != 0:
            if len(self.selectedPoints) != 2:
                self.showInfo('Exactly two points must be selected for a trim operation')
                return
            self.showSelectedPoints('Data trimmed/selected using points: ')
            selPts = list(self.selectedPoints.keys())
            selPts.sort()
            self.left = selPts[0]
            self.right = selPts[1]
            self.userTrimInEffect = True
        else:
            # self.showInfo('All points will be selected (because no trim points specified)')
            self.showMsg('All data points were selected')
            self.left = 0
            self.right = self.dataLen - 1

        for i in range(0, self.left):
            self.yStatus[i] = EXCLUDED
        for i in range(min(self.dataLen, self.right + 1), self.dataLen):
            self.yStatus[i] = EXCLUDED
        for i in range(self.left, min(self.dataLen, self.right + 1)):
            self.yStatus[i] = INCLUDED

        if len(self.smoothSecondary) > 0:
            self.smoothSecondary = []
            if self.smoothingIntervalSpinBox.value() > 0:
                # Start with pristine (original) values for the curve being analyzed.
                self.processTargetSelection(self.targetIndex, redraw=False)
                self.newNormalize()

        self.selectedPoints = {}
        self.newRedrawMainPlot()
        self.doBlockIntegration.setEnabled(False)
        self.blockSizeEdit.setEnabled(False)
        self.mainPlot.autoRange()


def main(csv_file_path=None):
    # csv_file_path gets filled in by PyMovie

    os.environ['QT_MAC_WANTS_LAYER'] = '1'  # This line needed when Mac updated to Big Sur

    import traceback
    app = PyQt5.QtWidgets.QApplication(sys.argv)

    print(f'PyOTE Version: {version.version()}')

    if sys.platform == 'linux':
        # PyQt5.QtWidgets.QApplication.setStyle('macintosh')
        PyQt5.QtWidgets.QApplication.setStyle('windows')
        print(f'os: Linux')
    elif sys.platform == 'darwin':
        # PyQt5.QtWidgets.QApplication.setStyle('macintosh')
        PyQt5.QtWidgets.QApplication.setStyle('windows')
        print(f'os: MacOS')
    else:
        print(f'os: Windows')
        PyQt5.QtWidgets.QApplication.setStyle('windows')
        app.setStyleSheet("QTabWidget, QComboBox, QLabel, QTableWidget, QTextEdit, QDoubleSpinBox, QSpinBox,"
                          "QProgressBar, QAbstractButton, QPushButton, QToolButton, QCheckBox, "
                          "QRadioButton, QLineEdit {font-size: 8pt}")

    # Save the current/proper sys.excepthook object
    saved_excepthook = sys.excepthook

    def exception_hook(exctype, value, tb):
        print('')
        print('=' * 30)
        print(value)
        print('=' * 30)
        print('')

        traceback.print_tb(tb)
        # Call the usual exception processor
        saved_excepthook(exctype, value, tb)

    sys.excepthook = exception_hook

    form = SimplePlot(csv_file_path)
    form.show()
    app.exec_()
    sys.exit()


if __name__ == '__main__':
    main()
